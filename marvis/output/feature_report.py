"""Standalone feature-analysis Excel report (FEATURE phase, form A).

Writes the per-feature metrics computed by ``compute_feature_metrics`` into a
downloadable workbook, mirroring the model-report download pipeline. Missing
metrics render as "n/a" rather than blank, so a sheet is never silently empty.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from marvis.artifacts import TransactionalArtifactStore
from marvis.output.xlsx_safety import safe_xlsx_cell


# Columns that are useful for every report, independently of the metric
# selection.  Metric columns are assembled from the keys that actually ride in
# the rows; this keeps an explicitly unchecked metric out of the workbook.
_BASE_COLUMNS: list[tuple[str, str]] = [
    ("feature", "特征"),
    ("recommendation", "Agent建议"),
    ("recommendation_reason", "推荐原因"),
]

# (trigger keys, columns).  A group is appended only when at least one trigger
# key is present. Coverage is deliberately grouped because selecting
# ``coverage`` computes its whole quality profile in one pass.
_METRIC_COLUMN_GROUPS: list[tuple[set[str], list[tuple[str, str]]]] = [
    ({"iv"}, [("iv", "IV")]),
    ({"ks"}, [("ks", "KS")]),
    ({"auc"}, [("auc", "AUC")]),
    ({"psi", "psi_reason"}, [("psi", "PSI"), ("psi_reason", "PSI说明")]),
    (
        {
            "coverage",
            "missing_rate",
            "mode_rate",
            "zero_rate",
            "valid_count",
            "unique_count",
            "unique_rate",
            "mean",
            "std",
            "min",
            "q25",
            "median",
            "q75",
            "max",
        },
        [
            ("coverage", "覆盖率"),
            ("missing_rate", "缺失率"),
            ("mode_rate", "单一值率"),
            ("zero_rate", "零值率"),
            ("valid_count", "有效样本数"),
            ("unique_count", "唯一值数"),
            ("unique_rate", "唯一值率"),
            ("mean", "均值"),
            ("std", "标准差"),
            ("min", "最小值"),
            ("q25", "P25"),
            ("median", "中位数"),
            ("q75", "P75"),
            ("max", "最大值"),
        ],
    ),
    ({"lift_top_bin"}, [("lift_top_bin", "头部lift")]),
    (
        {"lift_head_5", "lift_head_10", "lift_tail_5", "lift_tail_10", "lift_reason"},
        [
            ("lift_head_5", "头部lift5%"),
            ("lift_head_10", "头部lift10%"),
            ("lift_tail_5", "尾部lift5%"),
            ("lift_tail_10", "尾部lift10%"),
            ("lift_reason", "Lift说明"),
        ],
    ),
    ({"importance"}, [("importance", "重要性")]),
    (
        {"psi_month_first", "psi_month_first_reason", "psi_month_first_series"},
        [("psi_month_first", "月度PSI(首月基准)"), ("psi_month_first_reason", "月度PSI(首月)说明")],
    ),
    (
        {"psi_month_last", "psi_month_last_reason", "psi_month_last_series"},
        [("psi_month_last", "月度PSI(末月基准)"), ("psi_month_last_reason", "月度PSI(末月)说明")],
    ),
    (
        {"psi_month_previous", "psi_month_previous_reason", "psi_month_previous_series"},
        [("psi_month_previous", "月度PSI(逐月环比)"), ("psi_month_previous_reason", "月度PSI(环比)说明")],
    ),
    (
        {"psi_split", "psi_split_reason", "psi_split_series"},
        [("psi_split", "样本集PSI"), ("psi_split_reason", "样本集PSI说明")],
    ),
    (
        {
            "business_meaning",
            "expected_direction",
            "actual_direction",
            "meaning_consistency",
            "meaning_consistency_reason",
            "meaning_judgement_source",
        },
        [
            ("business_meaning", "业务含义"),
            ("expected_direction", "预期方向"),
            ("actual_direction", "实际方向"),
            ("meaning_consistency", "含义方向一致性"),
            ("meaning_consistency_reason", "含义方向说明"),
            ("meaning_judgement_source", "判断来源"),
        ],
    ),
]

_PSI_VIEWS: list[tuple[str, str]] = [
    ("psi_month_first", "月度PSI(首月基准)"),
    ("psi_month_last", "月度PSI(末月基准)"),
    ("psi_month_previous", "月度PSI(逐月环比)"),
    ("psi_split", "样本集PSI"),
]

# Kept as a compatibility alias for consumers that imported the old private
# constant in local extensions.
_COLUMNS: list[tuple[str, str]] = [
    *_BASE_COLUMNS,
    ("iv", "IV"),
    ("ks", "KS"),
    ("auc", "AUC"),
    ("psi", "PSI"),
    ("psi_reason", "PSI说明"),
    ("missing_rate", "缺失率"),
    ("mode_rate", "单一值率"),
    ("zero_rate", "零值率"),
    ("valid_count", "有效样本数"),
    ("unique_count", "唯一值数"),
    ("unique_rate", "唯一值率"),
    ("lift_top_bin", "头部lift"),
    ("mean", "均值"),
    ("std", "标准差"),
    ("min", "最小值"),
    ("q25", "P25"),
    ("median", "中位数"),
    ("q75", "P75"),
    ("max", "最大值"),
]


def render_feature_report(
    metrics: list[dict],
    out_path: Path,
    *,
    collinear: dict | None = None,
    binning: list[dict] | None = None,
) -> Path:
    out_path = Path(out_path)
    artifact = TransactionalArtifactStore(out_path.parent).stage(out_path.name)
    rows = [item for item in (metrics or []) if isinstance(item, dict)]
    columns = _selected_columns(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "特征指标"
    sheet.append([label for _key, label in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for item in rows:
        sheet.append([_cell(item.get(key)) for key, _label in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, (_key, label) in enumerate(columns, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = (
            42 if label in {"推荐原因", "PSI说明", "Lift说明"} else max(12, min(len(label) * 2 + 4, 24))
        )
    # Optional collinear / VIF sheet — written only when the VIF metric was selected.
    if isinstance(collinear, dict):
        _append_collinear_sheet(workbook, collinear)
    if _has_psi_series(rows):
        _append_psi_detail_sheet(workbook, rows)
    if binning:
        _append_binning_sheet(workbook, binning)
    try:
        workbook.save(artifact.path)
        artifact.promote()
        artifact.commit()
    except Exception:
        artifact.rollback()
        raise
    return artifact.final_path


def _selected_columns(rows: list[dict]) -> list[tuple[str, str]]:
    present = {str(key) for row in rows for key in row}
    columns = list(_BASE_COLUMNS)
    for triggers, group in _METRIC_COLUMN_GROUPS:
        if present.intersection(triggers):
            columns.extend(group)
    return columns


def _has_psi_series(rows: list[dict]) -> bool:
    return any(
        isinstance(row.get(f"{view}_series"), list)
        for row in rows
        for view, _label in _PSI_VIEWS
    )


def _append_psi_detail_sheet(workbook: Workbook, rows: list[dict]) -> None:
    sheet = workbook.create_sheet("PSI明细")
    columns = [
        ("feature", "特征"),
        ("view", "PSI口径"),
        ("base", "基准"),
        ("compare", "对比"),
        ("psi", "PSI"),
    ]
    sheet.append([label for _key, label in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        feature = row.get("feature")
        for view, view_label in _PSI_VIEWS:
            series = row.get(f"{view}_series")
            if not isinstance(series, list):
                continue
            for point in series:
                if not isinstance(point, dict):
                    continue
                sheet.append([
                    _cell(feature),
                    view_label,
                    _cell(point.get("base")),
                    _cell(point.get("compare")),
                    _cell(point.get("psi")),
                ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, (_key, label) in enumerate(columns, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(
            14,
            min(len(label) * 2 + 8, 30),
        )


_BINNING_COLUMNS: list[tuple[str, str]] = [
    ("feature", "特征"),
    ("requested_bins", "请求箱数"),
    ("actual_bins", "实际箱数"),
    ("bin_index", "箱号"),
    ("risk_rank", "风险排序"),
    ("interval", "区间"),
    ("count", "样本数"),
    ("bad_count", "坏样本数"),
    ("good_count", "好样本数"),
    ("bad_rate", "坏率"),
    ("cumulative_bad_rate", "累计坏率"),
    ("lift", "单箱Lift"),
    ("cumulative_lift", "累计Lift"),
    ("ks", "KS"),
    ("woe", "WOE"),
    ("iv_contribution", "IV贡献"),
    ("total_iv", "总IV"),
    ("direction", "风险方向"),
    ("degraded_reason", "分箱说明"),
]


def _append_binning_sheet(workbook: Workbook, binning: list[dict]) -> None:
    sheet = workbook.create_sheet("分箱分析")
    sheet.append([label for _key, label in _BINNING_COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for feature_result in binning:
        if not isinstance(feature_result, dict):
            continue
        common = {
            key: feature_result.get(key)
            for key in ("feature", "requested_bins", "actual_bins", "total_iv", "direction", "degraded_reason")
        }
        for row in feature_result.get("rows") or []:
            if not isinstance(row, dict):
                continue
            values = {**common, **row}
            sheet.append([_cell(values.get(key)) for key, _label in _BINNING_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, (_key, label) in enumerate(_BINNING_COLUMNS, start=1):
        width = 38 if label in {"区间", "分箱说明"} else max(11, min(len(label) * 2 + 4, 20))
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width


def _append_collinear_sheet(workbook: Workbook, collinear: dict) -> None:
    sheet = workbook.create_sheet("共线性(VIF)")
    sheet.append(["特征", "VIF"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for feat, value in (collinear.get("vif") or {}).items():
        sheet.append([_cell(feat), _cell(value)])
    pairs = [p for p in (collinear.get("collinear_pairs") or []) if isinstance(p, (list, tuple)) and len(p) >= 3]
    if pairs:
        sheet.append([])
        header = sheet.max_row + 1
        sheet.append(["特征A", "特征B", "相关系数"])
        for cell in sheet[header]:
            cell.font = Font(bold=True)
        for pair in pairs:
            sheet.append([_cell(pair[0]), _cell(pair[1]), _cell(pair[2])])


def _cell(value):
    if value is None:
        return "n/a"
    if isinstance(value, float):
        return round(value, 6)
    if isinstance(value, dict):
        message = value.get("message")
        if message not in (None, ""):
            value = str(message)
        else:
            value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, (list, tuple, set)):
        value = json.dumps(list(value), ensure_ascii=False, default=str)
    return safe_xlsx_cell(value)


__all__ = ["render_feature_report"]
