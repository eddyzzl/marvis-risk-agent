"""Excel renderer for deterministic VTG-terminal and profitability results.

This module is intentionally presentation-only: every number arrives through
``RiskAnalysisReportPayload``.  Calculation and validation remain in the builtin
``risk_analysis`` pack so the workbook cannot silently diverge from tool output.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from marvis.output.styles import FONT_NAME


RISK_ANALYSIS_REPORT_SHEETS = ["结论摘要", "明细结果", "口径与假设", "数据质量"]

_TITLE_FILL = "163A5F"
_SECTION_FILL = "D9EAF7"
_HEADER_FILL = "1F4E78"
_HEADER_FONT = "FFFFFF"
_WARN_FILL = "FFF2CC"
_FAIL_FILL = "F4CCCC"
_PASS_FILL = "D9EAD3"
_BORDER_COLOR = "C9D2DC"
_EXCEL_CELL_TEXT_MAX_CHARS = 32_767

_ANALYSIS_LABELS = {
    "vtg_terminal": "VTG 终值与年化不良测算",
    "profitability": "收益测算",
}

_LABELS = {
    "product": "产品",
    "cohort": "Cohort",
    "asset_class": "资产类别",
    "as_of_period": "数据期间",
    "as_of_date": "数据截面",
    "scenario": "场景",
    "amount_unit": "金额单位",
    "row_count": "行数",
    "asset_class_count": "资产类别数",
    "analysis_slice_count": "分析切片数",
    "product_count": "产品数",
    "cohort_count": "Cohort 数",
    "negative_product_count": "负净收益产品数",
    "disbursement_amount": "放款金额",
    "total_disbursement_amount": "放款金额合计",
    "avg_daily_balance": "日均余额",
    "total_avg_daily_balance": "日均余额合计",
    "turnover": "周转次数",
    "day_count_basis": "年化天数基础",
    "mob_days": "MOB 天数",
    "mob_balance_rate": "MOB 平均日余额率",
    "mob_balance_amount": "MOB 平均日余额金额",
    "portfolio_turnover": "组合周转次数",
    "mob14_bad_rate": "MOB14 不良率",
    "weighted_mob14_bad_rate": "加权 MOB14 不良率",
    "terminal_bad_rate": "终值不良率",
    "weighted_terminal_bad_rate": "加权终值不良率",
    "terminal_bad_rate_source": "终值来源",
    "terminal_method": "终值方法",
    "auxiliary_terminal_bad_rate": "辅助终值不良率",
    "observed_annualized_bad_rate": "观察年化不良率",
    "annualized_bad_rate": "年化不良率",
    "highest_annualized_bad_rate": "最高年化不良率",
    "highest_annualized_product": "最高年化不良产品",
    "highest_annualized_cohort": "最高年化不良 Cohort",
    "long_term_recovery_rate": "长期回收率",
    "supplied_long_term_recovery_rate": "输入长期回收率",
    "realized_long_term_recovery_rate": "实现长期回收率",
    "previous_mob14_bad_rate": "上期 MOB14 不良率",
    "previous_terminal_bad_rate": "上期终值不良率",
    "previous_turnover": "上期周转次数",
    "previous_turnover_source": "上期周转来源",
    "previous_annualized_bad_rate": "上期年化不良率",
    "previous_annualized_bad_rate_source": "上期年化不良来源",
    "previous_disbursement_amount": "上期放款金额",
    "previous_avg_daily_balance": "上期日均余额",
    "annualized_bad_rate_change": "年化不良率变化",
    "weight": "权重",
    "weight_basis": "权重口径",
    "weight_sum": "权重合计",
    "customer_rate": "对客利率",
    "terminal_vintage_rate": "终值 Vintage 率",
    "risk_turnover": "风险周转次数",
    "loss_timing_factor": "损失时点系数",
    "profit_share_ratio": "合同分润比例",
    "customer_stage_count": "客户阶段数",
    "transaction_weight_sum": "阶段交易权重合计",
    "customer_stage_provenance": "客户阶段计算追溯",
    "per_application_cost": "单笔申请成本",
    "credit_approval_rate": "授信通过率",
    "draw_initiation_rate": "用信发起率",
    "draw_approval_rate": "用信通过率",
    "average_ticket": "件均金额",
    "data_annualization_factor": "数据成本年化系数",
    "tax_method": "税费方法",
    "tax_inclusive_divisor": "含税除数",
    "tax_combined_rate": "综合税费率",
    "interest_loss_rate": "息费损失率",
    "interest_loss_rate_source": "息费损失来源",
    "revenue_share_rate": "分润成本率",
    "risk_cost_rate": "风险成本率",
    "risk_cost_rate_source": "风险成本来源",
    "acquisition_cost_rate": "获客成本率",
    "acquisition_cost_rate_source": "获客成本来源",
    "data_cost_rate": "数据成本率",
    "data_cost_rate_source": "数据成本来源",
    "payment_cost_rate": "支付成本率",
    "collection_cost_rate": "催收成本率",
    "funding_cost_rate": "资金成本率",
    "other_cost_rate": "其他成本率",
    "tax_rate": "税率",
    "tax_rate_source": "税费来源",
    "fixed_income_yield": "类固收收益率",
    "total_cost_rate": "总成本率",
    "net_yield": "净收益率",
    "lowest_net_yield": "最低净收益率",
    "lowest_net_yield_product": "最低净收益产品",
    "lowest_net_yield_as_of_period": "最低净收益期间",
    "lowest_net_yield_scenario": "最低净收益场景",
    "highest_net_yield": "最高净收益率",
    "highest_net_yield_product": "最高净收益产品",
    "highest_net_yield_as_of_period": "最高净收益期间",
    "highest_net_yield_scenario": "最高净收益场景",
    "max_cost_rate": "最大成本率",
    "max_cost_component": "最大成本项",
    "max_cost_product": "最大成本产品",
    "max_cost_as_of_period": "最大成本期间",
    "max_cost_scenario": "最大成本场景",
    "largest_scenario_net_yield_spread": "最大场景净收益差",
    "largest_scenario_net_yield_spread_product": "最大场景差产品",
    "largest_scenario_net_yield_spread_as_of_period": "最大场景差期间",
    "largest_scenario_net_yield_spread_high_scenario": "高净收益场景",
    "largest_scenario_net_yield_spread_low_scenario": "低净收益场景",
}


@dataclass(frozen=True)
class RiskAnalysisReportPayload:
    analysis_kind: str
    column_map: dict[str, str]
    product_scope: list[str]
    as_of_period: str
    headline_metrics: dict[str, Any]
    key_points: list[str]
    red_flags: list[str]
    assumptions: list[str]
    source_row_count: int
    row_count: int
    detail_rows: list[dict[str, Any]] = field(default_factory=list)
    summary_rows: list[dict[str, Any]] = field(default_factory=list)
    formula_definitions: list[dict[str, str]] = field(default_factory=list)
    data_quality: list[dict[str, str]] = field(default_factory=list)


def render_risk_analysis_report(
    payload: RiskAnalysisReportPayload,
    out_path: Path,
) -> Path:
    """Render a calculation payload to ``out_path`` without recalculation."""

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = _ANALYSIS_LABELS.get(
        payload.analysis_kind, payload.analysis_kind
    )
    workbook.properties.subject = "MARVIS 风险分析确定性报表"
    _write_conclusions(workbook, payload)
    _write_details(workbook, payload)
    _write_assumptions(workbook, payload)
    _write_data_quality(workbook, payload)
    workbook.save(out_path)
    return out_path


def _write_conclusions(workbook: Workbook, payload: RiskAnalysisReportPayload) -> None:
    sheet = workbook.create_sheet("结论摘要")
    _prepare_sheet(sheet)
    sheet.merge_cells("A1:D1")
    title = _ANALYSIS_LABELS.get(payload.analysis_kind, payload.analysis_kind)
    sheet["A1"] = f"{title}报告"
    _style_title(sheet["A1"])
    sheet["A2"] = "分析类型"
    sheet["B2"] = title
    sheet["C2"] = "源数据行数"
    sheet["D2"] = payload.source_row_count
    sheet["A3"] = "产品范围"
    sheet["B3"] = _safe_excel_text("、".join(payload.product_scope) or "未提供")
    sheet["C3"] = "数据期间"
    sheet["D3"] = _cell_value(payload.as_of_period)
    sheet["A4"] = "结果行数"
    sheet["B4"] = payload.row_count
    for cell in (
        sheet["A2"],
        sheet["C2"],
        sheet["A3"],
        sheet["C3"],
        sheet["A4"],
    ):
        cell.font = Font(name=FONT_NAME, bold=True, color="334155")
        cell.fill = PatternFill("solid", fgColor=_SECTION_FILL)

    row = 5
    row = _write_section_title(sheet, row, "核心指标")
    sheet.cell(row=row, column=1, value="指标")
    sheet.cell(row=row, column=2, value="取值")
    _style_header_row(sheet, row, 2)
    for key, value in payload.headline_metrics.items():
        row += 1
        sheet.cell(row=row, column=1, value=_label(key))
        value_cell = sheet.cell(row=row, column=2, value=_cell_value(value))
        _apply_number_format(value_cell, key)

    row += 2
    row = _write_section_title(sheet, row, "关键结论")
    for index, point in enumerate(payload.key_points, start=1):
        sheet.cell(row=row, column=1, value=index)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        sheet.cell(row=row, column=2, value=_cell_value(point))
        sheet.cell(row=row, column=2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        row += 1

    row += 1
    row = _write_section_title(sheet, row, "风险提示")
    flags = payload.red_flags or ["未发现需要单独提示的业务或数据质量红旗。"]
    for index, flag in enumerate(flags, start=1):
        sheet.cell(row=row, column=1, value=index)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cell = sheet.cell(row=row, column=2, value=_cell_value(flag))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if payload.red_flags:
            cell.fill = PatternFill("solid", fgColor=_WARN_FILL)
        row += 1

    sheet.freeze_panes = "A5"
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 24
    _style_used_range(sheet)


def _write_details(workbook: Workbook, payload: RiskAnalysisReportPayload) -> None:
    sheet = workbook.create_sheet("明细结果")
    _prepare_sheet(sheet)
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "产品汇总"
    _style_title(sheet["A1"])
    row = 3
    row = _write_dict_table(sheet, payload.summary_rows, start_row=row)
    row += 2
    sheet.cell(row=row, column=1, value="逐行计算结果")
    _style_section_cell(sheet.cell(row=row, column=1))
    detail_header_row = row + 1
    detail_end_row = _write_dict_table(
        sheet, payload.detail_rows, start_row=detail_header_row
    )
    if payload.detail_rows:
        last_col = get_column_letter(len(_headers(payload.detail_rows)))
        sheet.auto_filter.ref = f"A{detail_header_row}:{last_col}{detail_end_row - 1}"
        sheet.freeze_panes = f"A{detail_header_row + 1}"
        _add_negative_yield_rules(
            sheet, payload.detail_rows, detail_header_row, detail_end_row - 1
        )
    else:
        sheet.freeze_panes = "A3"
    _autofit_columns(sheet)
    _style_used_range(sheet)


def _write_assumptions(workbook: Workbook, payload: RiskAnalysisReportPayload) -> None:
    sheet = workbook.create_sheet("口径与假设")
    _prepare_sheet(sheet)
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "字段映射、计算口径与假设"
    _style_title(sheet["A1"])
    headers = ["类别", "项目", "公式/来源", "说明"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column, value=header)
    _style_header_row(sheet, 3, len(headers))
    row = 4
    for canonical, source in payload.column_map.items():
        sheet.append(
            [
                "字段映射",
                _cell_value(canonical),
                _cell_value(source),
                _cell_value(f"{_label(canonical)} <- {source}"),
            ]
        )
        row += 1
    for item in payload.formula_definitions:
        sheet.append(
            [
                "计算公式",
                _cell_value(item.get("metric")),
                _cell_value(item.get("formula")),
                _cell_value(item.get("note")),
            ]
        )
        row += 1
    for index, assumption in enumerate(payload.assumptions, start=1):
        sheet.append(["业务假设", f"假设 {index}", "", _cell_value(assumption)])
        row += 1
    sheet.auto_filter.ref = f"A3:D{max(3, row - 1)}"
    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 58
    sheet.column_dimensions["D"].width = 62
    _style_used_range(sheet)


def _write_data_quality(workbook: Workbook, payload: RiskAnalysisReportPayload) -> None:
    sheet = workbook.create_sheet("数据质量")
    _prepare_sheet(sheet)
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "数据质量检查与业务红旗"
    _style_title(sheet["A1"])
    for column, header in enumerate(("状态", "检查项", "结果说明"), start=1):
        sheet.cell(row=3, column=column, value=header)
    _style_header_row(sheet, 3, 3)
    row = 4
    for check in payload.data_quality:
        status = str(check.get("status") or "PASS").upper()
        sheet.cell(row=row, column=1, value=status)
        sheet.cell(row=row, column=2, value=_cell_value(check.get("check")))
        sheet.cell(row=row, column=3, value=_cell_value(check.get("detail")))
        _style_status_cell(sheet.cell(row=row, column=1), status)
        row += 1
    if payload.red_flags:
        for flag in payload.red_flags:
            sheet.cell(row=row, column=1, value="WARN")
            sheet.cell(row=row, column=2, value="业务红旗")
            sheet.cell(row=row, column=3, value=_cell_value(flag))
            _style_status_cell(sheet.cell(row=row, column=1), "WARN")
            row += 1
    else:
        sheet.cell(row=row, column=1, value="PASS")
        sheet.cell(row=row, column=2, value="业务红旗")
        sheet.cell(row=row, column=3, value="未发现。")
        _style_status_cell(sheet.cell(row=row, column=1), "PASS")
        row += 1
    sheet.auto_filter.ref = f"A3:C{row - 1}"
    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 86
    _style_used_range(sheet)


def _write_dict_table(sheet, rows: list[dict[str, Any]], *, start_row: int) -> int:
    if not rows:
        sheet.cell(row=start_row, column=1, value="无数据")
        return start_row + 1
    headers = _headers(rows)
    for column, key in enumerate(headers, start=1):
        sheet.cell(row=start_row, column=column, value=_label(key))
    _style_header_row(sheet, start_row, len(headers))
    row_index = start_row + 1
    for item in rows:
        for column, key in enumerate(headers, start=1):
            cell = sheet.cell(
                row=row_index, column=column, value=_cell_value(item.get(key))
            )
            _apply_number_format(cell, key)
        row_index += 1
    return row_index


def _headers(rows: list[dict[str, Any]]) -> list[str]:
    result: list[str] = []
    for row in rows:
        for key in row:
            if key not in result:
                result.append(key)
    return result


def _prepare_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0


def _style_title(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=_TITLE_FILL)
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.parent.row_dimensions[cell.row].height = 28


def _style_section_cell(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=_SECTION_FILL)
    cell.font = Font(name=FONT_NAME, bold=True, color="1F2937")


def _write_section_title(sheet, row: int, title: str) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = sheet.cell(row=row, column=1, value=title)
    _style_section_cell(cell)
    return row + 1


def _style_header_row(sheet, row: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.font = Font(name=FONT_NAME, bold=True, color=_HEADER_FONT)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=Side(style="thin", color=_BORDER_COLOR))
    sheet.row_dimensions[row].height = 24


def _style_status_cell(cell, status: str) -> None:
    fill = (
        _PASS_FILL
        if status == "PASS"
        else _FAIL_FILL
        if status == "FAIL"
        else _WARN_FILL
    )
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=FONT_NAME, bold=True)
    cell.alignment = Alignment(horizontal="center")


def _style_used_range(sheet) -> None:
    thin = Side(style="hair", color=_BORDER_COLOR)
    for row in sheet.iter_rows(
        min_row=1,
        max_row=max(1, sheet.max_row),
        min_col=1,
        max_col=max(1, sheet.max_column),
    ):
        for cell in row:
            if cell.value is None:
                continue
            if cell.row != 1 and cell.font == Font():
                cell.font = Font(name=FONT_NAME, size=10, color="111827")
            elif cell.row != 1 and not cell.font.name:
                cell.font = Font(
                    name=FONT_NAME,
                    size=cell.font.sz or 10,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                )
            alignment = copy(cell.alignment)
            alignment.vertical = "top"
            alignment.wrap_text = True
            cell.alignment = alignment
            if not cell.border.bottom.style:
                cell.border = Border(bottom=thin)


def _autofit_columns(sheet) -> None:
    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        max_length = 0
        for cell in sheet[letter]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 11), 36)


def _apply_number_format(cell, key: str) -> None:
    if not isinstance(cell.value, (int, float)) or isinstance(cell.value, bool):
        return
    normalized = str(key).lower()
    if normalized in {"weight", "weight_sum"} or normalized.endswith(
        ("_rate", "_yield", "_change", "_ratio", "_spread")
    ):
        cell.number_format = "0.00%;[Red](0.00%);-"
    elif "turnover" in normalized:
        cell.number_format = "0.00x;[Red](0.00x);-"
    elif any(token in normalized for token in ("amount", "balance")):
        cell.number_format = "#,##0.00;[Red](#,##0.00);-"
    elif "count" in normalized:
        cell.number_format = "#,##0;[Red](#,##0);-"
    else:
        cell.number_format = "#,##0.0000;[Red](#,##0.0000);-"


def _add_negative_yield_rules(
    sheet,
    rows: list[dict[str, Any]],
    header_row: int,
    end_row: int,
) -> None:
    if end_row <= header_row:
        return
    headers = _headers(rows)
    for key in ("net_yield", "fixed_income_yield"):
        if key not in headers:
            continue
        column = get_column_letter(headers.index(key) + 1)
        sheet.conditional_formatting.add(
            f"{column}{header_row + 1}:{column}{end_row}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor=_FAIL_FILL),
            ),
        )


def _cell_value(value: Any) -> Any:
    if isinstance(value, str):
        return _safe_excel_text(value)
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if hasattr(value, "item"):
        return _cell_value(value.item())
    if isinstance(value, (list, tuple)):
        return _safe_excel_text("、".join(str(item) for item in value))
    if isinstance(value, dict):
        return _safe_excel_text(json.dumps(value, ensure_ascii=False, sort_keys=True))
    return _safe_excel_text(str(value))


def _safe_excel_text(value: str) -> str:
    """Keep uploaded labels/text from becoming formulas in the XLSX output."""

    text = ILLEGAL_CHARACTERS_RE.sub(" ", str(value))
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text[: _EXCEL_CELL_TEXT_MAX_CHARS - 1]
    return text[:_EXCEL_CELL_TEXT_MAX_CHARS]


def _label(key: str) -> str:
    return _LABELS.get(str(key), str(key))


__all__ = [
    "RISK_ANALYSIS_REPORT_SHEETS",
    "RiskAnalysisReportPayload",
    "render_risk_analysis_report",
]
