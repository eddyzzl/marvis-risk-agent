"""S3 组合分析报告拼装 (对照 model_report.py 先例).

PortfolioReportPayload dataclass -> render_portfolio_report(payload, path)：
用 openpyxl 写多 sheet（组合概览/桶迁徙/逐月流量/细分画像/稳定性趋势/预期损失/
数据质量红旗汇总），经 TransactionalArtifactStore stage/promote/commit 落盘。

报告只搬运前序步骤已持久化的数字，不重算（INV-1）——payload 各字段直接来自
flow/migration/segment/trend/el 步骤的输出 dict。
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from marvis.artifacts import TransactionalArtifactStore


PORTFOLIO_REPORT_SHEETS = [
    "组合概览",
    "桶迁徙",
    "逐月流量",
    "细分画像",
    "稳定性趋势",
    "预期损失",
    "数据质量红旗",
]


@dataclass(frozen=True)
class PortfolioReportPayload:
    project_meta: dict = field(default_factory=dict)
    flow: dict | None = None
    migration: dict | None = None
    segment: dict | None = None
    trend: dict | None = None
    expected_loss: dict | None = None
    red_flags: list[dict] = field(default_factory=list)


def render_portfolio_report(payload: PortfolioReportPayload, out_path: Path) -> Path:
    out_path = Path(out_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_overview(workbook, payload)
    _write_migration(workbook, payload.migration)
    _write_flow(workbook, payload.flow)
    _write_segment(workbook, payload.segment)
    _write_trend(workbook, payload.trend)
    _write_expected_loss(workbook, payload.expected_loss)
    _write_red_flags(workbook, payload.red_flags)
    artifact = TransactionalArtifactStore(out_path.parent).stage(out_path.name)
    try:
        workbook.save(artifact.path)
        final_path = artifact.promote()
        artifact.commit()
        return final_path
    except Exception:
        artifact.rollback()
        raise


def _write_overview(workbook: Workbook, payload: PortfolioReportPayload) -> None:
    sheet = workbook.create_sheet("组合概览")
    rows: list[tuple] = [("项目元数据", "")]
    rows.extend((str(key), _cell(value)) for key, value in payload.project_meta.items())
    if payload.expected_loss:
        rows.append(("预期损失", ""))
        rows.append(("total_el", _cell(payload.expected_loss.get("total_el"))))
        assumptions = payload.expected_loss.get("assumptions") or {}
        for key, value in assumptions.items():
            rows.append((f"假设.{key}", _cell(value)))
    if payload.segment:
        conc = payload.segment.get("concentration") or {}
        rows.append(("细分集中度", ""))
        for key, value in conc.items():
            rows.append((f"concentration.{key}", _cell(value)))
    rows.append(("数据质量红旗数", _cell(len(payload.red_flags))))
    _write_rows(sheet, rows)


def _write_migration(workbook: Workbook, migration: dict | None) -> None:
    sheet = workbook.create_sheet("桶迁徙")
    if not migration:
        sheet["A1"] = "无数据"
        return
    heat_table = migration.get("heat_table") or []
    _write_dict_table(sheet, heat_table)


def _write_flow(workbook: Workbook, flow: dict | None) -> None:
    sheet = workbook.create_sheet("逐月流量")
    if not flow:
        sheet["A1"] = "无数据"
        return
    _write_dict_table(sheet, flow.get("net_flows") or [])


def _write_segment(workbook: Workbook, segment: dict | None) -> None:
    sheet = workbook.create_sheet("细分画像")
    if not segment:
        sheet["A1"] = "无数据"
        return
    _write_dict_table(sheet, segment.get("segments") or [])


def _write_trend(workbook: Workbook, trend: dict | None) -> None:
    sheet = workbook.create_sheet("稳定性趋势")
    if not trend:
        sheet["A1"] = "无数据（未提供 experiment_id，趋势步已剪除）"
        return
    _write_dict_table(sheet, trend.get("trend") or [])


def _write_expected_loss(workbook: Workbook, expected_loss: dict | None) -> None:
    sheet = workbook.create_sheet("预期损失")
    if not expected_loss:
        sheet["A1"] = "无数据"
        return
    _write_dict_table(sheet, expected_loss.get("el_by_month") or [])
    start = sheet.max_row + 2
    sheet.cell(row=start, column=1, value="链式吸收概率")
    _write_dict_table(sheet, expected_loss.get("chain") or [], start_row=start + 1)


def _write_red_flags(workbook: Workbook, red_flags: list[dict]) -> None:
    sheet = workbook.create_sheet("数据质量红旗")
    if not red_flags:
        sheet["A1"] = "无红旗"
        return
    _write_dict_table(sheet, red_flags)


def _write_rows(sheet, rows: list[tuple]) -> None:
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=_cell(value))
    _style_header(sheet)


def _write_dict_table(sheet, rows: list[dict], *, start_row: int = 1) -> None:
    if not rows:
        sheet.cell(row=start_row, column=1, value="无数据")
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=start_row, column=col_index, value=str(header))
    for row_index, row in enumerate(rows, start=start_row + 1):
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=col_index, value=_cell(row.get(header)))
    _style_header(sheet, row=start_row)


def _cell(value: Any):
    if isinstance(value, (str, int, float)) or value is None:
        return value
    if isinstance(value, (list, dict, tuple)):
        return str(value)
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def _style_header(sheet, *, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def payload_to_dict(payload: PortfolioReportPayload) -> dict:
    return asdict(payload)


__all__ = [
    "PORTFOLIO_REPORT_SHEETS",
    "PortfolioReportPayload",
    "payload_to_dict",
    "render_portfolio_report",
]
