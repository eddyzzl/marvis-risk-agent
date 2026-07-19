"""Deterministic XLSX evidence projection for automatic rule-tree assets.

The workbook contains no formulas and performs no metric calculation.  Every
strategy fact comes from the strict, self-authenticating automatic-tree asset;
the only additional content is a fixed inventory describing the six supported
delivery formats.  OOXML member order, timestamps, and ZIP metadata are
canonicalized so the same asset always produces identical bytes.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import datetime
from io import BytesIO
import json
import re
import unicodedata
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage

from marvis.output.automatic_tree_visual import (
    AUTOMATIC_TREE_PNG_RENDERER_VERSION,
    render_automatic_tree_png,
)
from marvis.packs.strategy.automatic_tree_asset import (
    validate_automatic_tree_asset,
)
from marvis.packs.strategy.errors import StrategyError


AUTOMATIC_TREE_REPORT_SCHEMA_VERSION = "strategy.automatic-tree-report.v1"
AUTOMATIC_TREE_REPORT_SHEET_NAMES = (
    "Summary",
    "Tree",
    "Nodes",
    "Leaf Rules",
    "Diagnostics",
    "Lineage",
    "Delivery",
)

_FIXED_WORKBOOK_DATETIME = datetime(2000, 1, 1)
_FIXED_ZIP_DATETIME = (1980, 1, 1, 0, 0, 0)
_FORMULA_PREFIXES = frozenset("=+-@")
_ILLEGAL_XLSX_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_CORE_MODIFIED_TIMESTAMP = re.compile(
    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)"
)
_MAX_XLSX_CELL_CHARACTERS = 32_767
_MAX_EMBEDDED_IMAGE_WIDTH = 1800

_METRIC_FIELDS = (
    "total",
    "good",
    "bad",
    "bad_rate",
    "share",
    "bad_capture",
    "lift",
    "loan_amount_total",
    "loan_amount_coverage_count",
    "loan_amount_coverage",
    "loan_amount_coverage_rate",
    "overdue_amount_total",
    "overdue_amount_coverage_count",
    "overdue_amount_coverage",
    "overdue_amount_coverage_rate",
    "amount_pair_coverage_count",
    "amount_pair_coverage",
    "amount_pair_coverage_rate",
    "paired_loan_amount_total",
    "paired_overdue_amount_total",
    "overdue_rate",
)


class AutomaticTreeReportError(StrategyError):
    """A validated automatic-tree asset cannot be safely rendered to XLSX."""


def render_automatic_tree_report_xlsx(asset: Mapping[str, Any]) -> bytes:
    """Return a deterministic, formula-safe XLSX projection of ``asset``."""

    canonical = validate_automatic_tree_asset(asset)
    tree_png = render_automatic_tree_png(canonical)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "MARVIS"
    workbook.properties.lastModifiedBy = "MARVIS"
    workbook.properties.title = "Automatic Rule Tree Evidence"
    workbook.properties.description = AUTOMATIC_TREE_REPORT_SCHEMA_VERSION
    workbook.properties.created = _FIXED_WORKBOOK_DATETIME
    workbook.properties.modified = _FIXED_WORKBOOK_DATETIME
    try:
        _write_summary(workbook, canonical)
        _write_tree(workbook, canonical, tree_png)
        _write_nodes(workbook, canonical)
        _write_leaf_rules(workbook, canonical)
        _write_diagnostics(workbook, canonical)
        _write_lineage(workbook, canonical)
        _write_delivery(workbook)
        raw = BytesIO()
        workbook.save(raw)
    finally:
        workbook.close()
    return _canonicalize_xlsx_bytes(raw.getvalue())


def _write_summary(workbook: Workbook, asset: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("Summary")
    _prepare_table_sheet(sheet, freeze="A2")
    _append_row(sheet, ("Field", "Value"))
    tree_result = asset["tree_result"]
    tree = tree_result["tree"]
    training = tree_result["training"]
    preprocessing = tree_result["preprocessing"]
    search = tree_result["search"]
    lifecycle = asset["lifecycle"]
    rows = (
        ("report_schema_version", AUTOMATIC_TREE_REPORT_SCHEMA_VERSION),
        ("png_renderer_version", AUTOMATIC_TREE_PNG_RENDERER_VERSION),
        ("asset_schema_version", asset["schema_version"]),
        ("asset_type", asset["asset_type"]),
        ("asset_id", asset["asset_id"]),
        ("asset_hash", asset["asset_hash"]),
        ("producer_version", asset["producer_version"]),
        ("candidate_id", asset["candidate_evidence"]["candidate_id"]),
        ("candidate_evidence_hash", asset["candidate_evidence"]["evidence_hash"]),
        ("candidate_stage", lifecycle["candidate_stage"]),
        ("observation_stage", lifecycle["observation_stage"]),
        ("validation_status", lifecycle["validation_status"]),
        ("tree_schema_version", tree_result["schema_version"]),
        ("tree_id", tree["tree_id"]),
        ("tree_root_node_id", tree["root_node_id"]),
        ("tree_result_hash", tree_result["result_hash"]),
        ("tree_node_count", tree["node_count"]),
        ("tree_leaf_count", tree["leaf_count"]),
        ("tree_leaf_ids", _canonical_json_text(tree["leaf_ids"])),
        ("training_row_count", training["row_count"]),
        ("target_col", training["target_col"]),
        ("feature_order", _canonical_json_text(training["feature_order"])),
        ("sample_weight", _canonical_json_text(training["sample_weight"])),
        ("loan_amount_col", training["loan_amount_col"]),
        ("overdue_amount_col", training["overdue_amount_col"]),
        ("seed", training["seed"]),
        ("sklearn_version", training["sklearn_version"]),
        ("cart", _canonical_json_text(training["cart"])),
        ("preprocessing_missing_policy", preprocessing["missing_policy"]),
        ("preprocessing_medians", _canonical_json_text(preprocessing["medians"])),
        ("search_method", search["method"]),
        ("search_truncated", search["truncated"]),
        (
            "search_cutpoint_evaluations_upper_bound",
            search["cutpoint_evaluations_upper_bound"],
        ),
        ("search_tie_break", _canonical_json_text(search["tie_break"])),
        ("directions", _canonical_json_text(tree_result["directions"])),
        ("budgets", _canonical_json_text(tree_result["budgets"])),
        ("checks", _canonical_json_text(tree_result["checks"])),
    )
    for row in rows:
        _append_row(sheet, row)
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 110


def _write_tree(
    workbook: Workbook,
    asset: Mapping[str, Any],
    tree_png: bytes,
) -> None:
    sheet = workbook.create_sheet("Tree")
    sheet.sheet_view.showGridLines = False
    _append_row(sheet, ("Automatic Rule Tree",))
    _append_row(sheet, ("Asset ID", asset["asset_id"]))
    _append_row(sheet, ("Asset Hash", asset["asset_hash"]))
    _append_row(sheet, ("PNG Renderer", AUTOMATIC_TREE_PNG_RENDERER_VERSION))
    image_stream = BytesIO(tree_png)
    image = WorksheetImage(image_stream)
    if image.width > _MAX_EMBEDDED_IMAGE_WIDTH:
        scale = _MAX_EMBEDDED_IMAGE_WIDTH / image.width
        image.width = _MAX_EMBEDDED_IMAGE_WIDTH
        image.height = int(round(image.height * scale))
    sheet.add_image(image, "A5")
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 90


def _write_nodes(workbook: Workbook, asset: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("Nodes")
    _prepare_table_sheet(sheet, freeze="A2")
    headers = (
        "Node ID",
        "Kind",
        "Depth",
        "Path JSON",
        "Feature",
        "Threshold",
        "Sklearn Threshold",
        "Threshold Adjustment",
        "Missing Child",
        "Left Child ID",
        "Right Child ID",
        "Rule ID",
        "Direction Expected",
        "Direction Status",
        "Direction Basis",
        "Direction Bad Rate Delta",
        "Direction Left JSON",
        "Direction Right JSON",
        *_metric_headers(),
    )
    _append_row(sheet, headers)
    nodes = asset["tree_result"]["tree"]["nodes"]
    for node in nodes:
        diagnostic = node.get("direction_diagnostic")
        _append_row(
            sheet,
            (
                node["node_id"],
                node["kind"],
                node["depth"],
                _canonical_json_text(node["path"]),
                node.get("feature"),
                node.get("threshold"),
                node.get("sklearn_threshold"),
                node.get("threshold_adjustment"),
                node.get("missing_child"),
                node.get("left_child_id"),
                node.get("right_child_id"),
                node.get("rule_id"),
                None if diagnostic is None else diagnostic["expected_direction"],
                None if diagnostic is None else diagnostic["status"],
                None if diagnostic is None else diagnostic["basis"],
                None if diagnostic is None else diagnostic["primary_bad_rate_delta"],
                None
                if diagnostic is None
                else _canonical_json_text(diagnostic["left"]),
                None
                if diagnostic is None
                else _canonical_json_text(diagnostic["right"]),
                *_metric_values(node["metrics"]),
            ),
        )


def _write_leaf_rules(workbook: Workbook, asset: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("Leaf Rules")
    _prepare_table_sheet(sheet, freeze="A2")
    headers = (
        "Leaf ID",
        "Fragment ID",
        "Fragment Hash",
        "Rule ID",
        "Condition JSON",
        "Clauses JSON",
        "Requirements JSON",
        "Effect ID",
        *_metric_headers(),
    )
    _append_row(sheet, headers)
    rules = asset["tree_result"]["rules"]
    rule_by_leaf_id = {rule["leaf_id"]: rule for rule in rules}
    for fragment in asset["fragments"]:
        rule = rule_by_leaf_id[fragment["leaf_id"]]
        _append_row(
            sheet,
            (
                fragment["leaf_id"],
                fragment["fragment_id"],
                fragment["fragment_hash"],
                fragment["rule_id"],
                _canonical_json_text(fragment["condition"]),
                _canonical_json_text(rule["clauses"]),
                _canonical_json_text(fragment["requirements"]),
                fragment["effect_id"],
                *_metric_values(fragment["metrics"]),
            ),
        )


def _write_diagnostics(workbook: Workbook, asset: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("Diagnostics")
    _prepare_table_sheet(sheet, freeze="A2")
    _append_row(
        sheet,
        (
            "Record Type",
            "Code",
            "Node ID",
            "Feature",
            "Expected Direction",
            "Basis",
            "Primary Bad Rate Delta",
        ),
    )
    diagnostics = asset["diagnostics"]
    for violation in diagnostics["direction_violations"]:
        _append_row(
            sheet,
            (
                "direction_violation_detail",
                None,
                violation["node_id"],
                violation["feature"],
                violation["expected_direction"],
                violation["basis"],
                violation["primary_bad_rate_delta"],
            ),
        )
    for red_flag in diagnostics["red_flags"]:
        _append_row(
            sheet,
            (
                "red_flag",
                red_flag["code"],
                red_flag["node_id"],
                red_flag["feature"],
                red_flag["expected_direction"],
                None,
                None,
            ),
        )


def _write_lineage(workbook: Workbook, asset: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("Lineage")
    _prepare_table_sheet(sheet, freeze="A2")
    _append_row(sheet, ("Field", "Value"))
    lifecycle = asset["lifecycle"]
    identity = asset["identity"]
    candidate = asset["candidate_evidence"]
    tree_result = asset["tree_result"]
    rows: list[tuple[object, object]] = [
        ("asset.schema_version", asset["schema_version"]),
        ("asset.asset_type", asset["asset_type"]),
        ("asset.asset_id", asset["asset_id"]),
        ("asset.asset_hash", asset["asset_hash"]),
        ("asset.producer_version", asset["producer_version"]),
        ("candidate.candidate_id", candidate["candidate_id"]),
        ("candidate.evidence_hash", candidate["evidence_hash"]),
        ("lifecycle.candidate_stage", lifecycle["candidate_stage"]),
        ("lifecycle.observation_stage", lifecycle["observation_stage"]),
        ("lifecycle.validation_status", lifecycle["validation_status"]),
        ("tree.schema_version", tree_result["schema_version"]),
        ("tree.tree_id", tree_result["tree"]["tree_id"]),
        ("tree.result_hash", tree_result["result_hash"]),
    ]
    for field in (
        "task_id",
        "dataset_id",
        "dataset_content_hash",
        "workspace_revision",
        "workspace_generation",
        "semantic_mapping_hash",
        "registry_metadata_hash",
        "sample_context_hash",
    ):
        rows.append((f"identity.{field}", identity[field]))
    rows.extend(
        (f"source_ref[{index}]", source_ref)
        for index, source_ref in enumerate(asset["source_refs"])
    )
    for row in rows:
        _append_row(sheet, row)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 100


def _write_delivery(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Delivery")
    _prepare_table_sheet(sheet, freeze="A2")
    _append_row(sheet, ("Format", "Suggested File Name", "Purpose"))
    rows = (
        ("JSON", "automatic_tree.json", "Canonical validated automatic-tree asset"),
        ("Python", "automatic_tree.py", "Typed-DSL tree scorer for Python delivery"),
        ("SQL", "automatic_tree.sql", "SQL CASE projection for supported semantics"),
        ("SVG", "automatic_tree.svg", "Scalable topology and evidence visual"),
        ("PNG", "automatic_tree.png", "Raster topology for reports and previews"),
        ("XLSX", "automatic_tree.xlsx", "Evidence workbook, image, and lineage"),
    )
    for row in rows:
        _append_row(sheet, row)
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 60


def _metric_headers() -> tuple[str, ...]:
    unweighted = tuple(f"Unweighted {field}" for field in _METRIC_FIELDS)
    weighted = tuple(f"Weighted {field}" for field in _METRIC_FIELDS)
    return (*unweighted, "Weighted Status", *weighted)


def _metric_values(metrics: Mapping[str, Any]) -> tuple[object, ...]:
    unweighted = metrics["unweighted"]
    weighted = metrics["weighted"]
    weighted_values = (
        tuple(weighted.get(field) for field in _METRIC_FIELDS)
        if weighted["status"] == "available"
        else (None,) * len(_METRIC_FIELDS)
    )
    return (
        *(unweighted.get(field) for field in _METRIC_FIELDS),
        weighted["status"],
        *weighted_values,
    )


def _prepare_table_sheet(sheet: Any, *, freeze: str) -> None:
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False


def _append_row(sheet: Any, values: Sequence[object]) -> None:
    sheet.append([_xlsx_cell(value) for value in values])


def _xlsx_cell(value: object) -> object:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, float):
        return repr(value)
    if isinstance(value, int):
        return str(value) if len(str(abs(value))) > 15 else value
    if not isinstance(value, str):
        raise AutomaticTreeReportError(
            f"report cell contains unsupported {type(value).__name__}"
        )
    text = value
    if _looks_like_formula(text):
        text = "'" + text
    text = _ILLEGAL_XLSX_CONTROL.sub(
        lambda match: f"\\u{ord(match.group(0)):04x}", text
    )
    if len(text) > _MAX_XLSX_CELL_CHARACTERS:
        raise AutomaticTreeReportError(
            "report cell exceeds Excel's 32767 character limit"
        )
    return text


def _looks_like_formula(value: str) -> bool:
    for character in value:
        if character.isspace() or unicodedata.category(character).startswith("C"):
            continue
        return character in _FORMULA_PREFIXES
    return False


def _canonical_json_text(value: object) -> str:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
    except (TypeError, ValueError) as exc:
        raise AutomaticTreeReportError(
            f"automatic tree report value is not canonical JSON: {exc}"
        ) from exc


def _canonicalize_xlsx_bytes(raw: bytes) -> bytes:
    source = BytesIO(raw)
    destination = BytesIO()
    with ZipFile(source, "r") as input_archive:
        members = sorted(input_archive.infolist(), key=lambda item: item.filename)
        with ZipFile(
            destination,
            "w",
            compression=ZIP_DEFLATED,
            compresslevel=9,
            allowZip64=True,
        ) as output_archive:
            for source_member in members:
                member = ZipInfo(source_member.filename, date_time=_FIXED_ZIP_DATETIME)
                member.compress_type = ZIP_DEFLATED
                member.create_system = 0
                member.external_attr = 0
                member.internal_attr = 0
                member.comment = b""
                member.extra = b""
                payload = input_archive.read(source_member.filename)
                if source_member.filename == "docProps/core.xml":
                    payload, replacements = _CORE_MODIFIED_TIMESTAMP.subn(
                        rb"\g<1>2000-01-01T00:00:00Z\g<2>", payload
                    )
                    if replacements != 1:
                        raise AutomaticTreeReportError(
                            "generated workbook has an invalid modified timestamp contract"
                        )
                output_archive.writestr(
                    member,
                    payload,
                    compress_type=ZIP_DEFLATED,
                    compresslevel=9,
                )
    return destination.getvalue()


__all__ = [
    "AUTOMATIC_TREE_REPORT_SCHEMA_VERSION",
    "AUTOMATIC_TREE_REPORT_SHEET_NAMES",
    "AutomaticTreeReportError",
    "render_automatic_tree_report_xlsx",
]
