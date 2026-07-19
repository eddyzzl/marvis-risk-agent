"""Deterministic, evidence-only exports for the Candidate Lab.

The renderer consumes the existing strategy candidate and univariate contracts.
It deliberately does not calculate, round, rank, or otherwise reinterpret any
metric.  JSON is a canonical envelope over the two validated inputs; XLSX is a
human-readable projection of that same envelope.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import datetime
from io import BytesIO
import json
import math
import re
import unicodedata
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook

from marvis.feature.univariate import SCHEMA_VERSION as UNIVARIATE_SCHEMA_VERSION
from marvis.packs.strategy.candidate_evidence import validate_candidate_evidence
from marvis.packs.strategy.dsl import canonicalize_expression
from marvis.packs.strategy.errors import StrategyError


REPORT_SCHEMA_VERSION = "strategy.candidate-report.v1"
REPORT_SHEET_NAMES = (
    "Summary",
    "Rankings",
    "Bins",
    "Metrics",
    "Red Flags",
    "Lineage",
)

_TOP_LEVEL_FIELDS = frozenset(
    {
        "schema_version",
        "target",
        "target_definition",
        "row_count",
        "feature_count",
        "parameters",
        "features",
        "rankings",
        "resource_budget",
    }
)
_PARAMETER_FIELDS = frozenset(
    {
        "bin_count",
        "smoothing",
        "min_bin_pct",
        "seed",
        "loan_amount",
        "overdue_amount",
    }
)
_FEATURE_FIELDS = frozenset(
    {
        "feature",
        "feature_type",
        "row_count",
        "missing_rate",
        "sentinel_values",
        "methods",
    }
)
_METHOD_FIELDS = frozenset(
    {
        "method",
        "requested_method",
        "actual_method",
        "status",
        "evidence",
        "metrics",
        "bins",
    }
)
_METHOD_METRIC_FIELDS = frozenset(
    {"iv", "ks", "auc", "risk_direction", "missing_rate", "amount_metrics"}
)
_RANKING_FIELDS = frozenset({"feature", "method", "iv", "ks", "auc"})
_RESOURCE_BUDGET_FIELDS = frozenset(
    {
        "max_rows",
        "max_features",
        "max_bins",
        "max_categories",
        "rows_used",
        "features_used",
        "method_runs",
        "truncated",
    }
)
_AMOUNT_METRIC_FIELDS = frozenset({"loan_amount", "overdue_amount", "overdue_rate"})
_BIN_REQUIRED_FIELDS = frozenset(
    {
        "index",
        "id",
        "kind",
        "condition",
        "count",
        "share",
        "good",
        "bad",
        "bad_rate",
        "woe",
        "iv_contribution",
        "lift",
        "cumulative_ks",
        "amount_metrics",
    }
)
_BIN_OPTIONAL_FIELDS = frozenset(
    {"lower", "upper", "include_lower", "include_upper", "value"}
)
_METHODS = frozenset(
    {"equal_frequency", "equal_width", "chimerge", "tree", "categorical"}
)
_BIN_KINDS = frozenset({"numeric_interval", "category", "sentinel", "missing"})
_FIXED_WORKBOOK_DATETIME = datetime(2000, 1, 1)
_FIXED_ZIP_DATETIME = (1980, 1, 1, 0, 0, 0)
_FORMULA_PREFIXES = frozenset("=+-@")
_ILLEGAL_XLSX_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_CORE_MODIFIED_TIMESTAMP = re.compile(
    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)"
)
_MAX_XLSX_CELL_CHARACTERS = 32_767


class StrategyCandidateReportError(StrategyError):
    """Candidate evidence cannot be represented by the report contract."""


def render_strategy_candidate_bundle(
    evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> dict[str, bytes]:
    """Return canonical JSON and deterministic XLSX bytes for one input pair."""

    payload = _validated_report_payload(evidence, analysis)
    return {
        "json": _canonical_json_bytes(payload),
        "xlsx": _render_xlsx(payload),
    }


def canonical_strategy_candidate_report_json(
    evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> bytes:
    """Return the canonical UTF-8 JSON envelope for the validated inputs."""

    return _canonical_json_bytes(_validated_report_payload(evidence, analysis))


def render_strategy_candidate_report_xlsx(
    evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> bytes:
    """Return a deterministic, formula-safe XLSX projection of the inputs."""

    return _render_xlsx(_validated_report_payload(evidence, analysis))


def _validated_report_payload(
    evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> dict[str, Any]:
    normalized_evidence = validate_candidate_evidence(evidence)
    normalized_analysis = _validate_univariate_analysis(analysis)
    # Python structural equality conflates JSON-distinct values such as
    # ``false``/``0`` and ``1``/``1.0``.  Bind the two contracts by their
    # canonical JSON bytes so type drift cannot join evidence A to analysis B.
    if _canonical_json_bytes(normalized_evidence["analysis"]) != _canonical_json_bytes(
        normalized_analysis
    ):
        raise StrategyCandidateReportError(
            "candidate evidence analysis does not match the univariate analysis"
        )
    return {
        "schema_version": REPORT_SCHEMA_VERSION,
        "candidate_evidence": normalized_evidence,
        "univariate_analysis": normalized_analysis,
    }


def _validate_univariate_analysis(payload: Mapping[str, Any]) -> dict[str, Any]:
    if not isinstance(payload, Mapping):
        raise StrategyCandidateReportError("univariate analysis must be an object")
    _require_exact_fields(payload, _TOP_LEVEL_FIELDS, "univariate analysis")
    normalized = _detached_json_object(payload, "univariate analysis")

    if normalized["schema_version"] != UNIVARIATE_SCHEMA_VERSION:
        raise StrategyCandidateReportError(
            f"univariate analysis schema_version must be {UNIVARIATE_SCHEMA_VERSION}"
        )
    _required_text(normalized["target"], "univariate analysis.target")
    _validate_target_definition(normalized["target_definition"])
    row_count = _positive_int(normalized["row_count"], "univariate analysis.row_count")
    feature_count = _positive_int(
        normalized["feature_count"], "univariate analysis.feature_count"
    )
    _validate_parameters(normalized["parameters"])
    available_metrics = _validate_features(
        normalized["features"], row_count=row_count, feature_count=feature_count
    )
    _validate_rankings(normalized["rankings"], available_metrics=available_metrics)
    _validate_resource_budget(
        normalized["resource_budget"],
        row_count=row_count,
        feature_count=feature_count,
        method_count=sum(len(feature["methods"]) for feature in normalized["features"]),
    )
    return normalized


def _validate_target_definition(value: object) -> None:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(
            "univariate analysis.target_definition must be an object"
        )
    _require_exact_fields(
        value, frozenset({"good", "bad"}), "univariate analysis.target_definition"
    )
    if not _is_int(value["good"]) or not _is_int(value["bad"]):
        raise StrategyCandidateReportError(
            "univariate analysis.target_definition must map good to 0 and bad to 1"
        )
    if value["good"] != 0 or value["bad"] != 1:
        raise StrategyCandidateReportError(
            "univariate analysis.target_definition must map good to 0 and bad to 1"
        )


def _validate_parameters(value: object) -> None:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(
            "univariate analysis.parameters must be an object"
        )
    _require_exact_fields(value, _PARAMETER_FIELDS, "univariate analysis.parameters")
    bin_count = _positive_int(
        value["bin_count"], "univariate analysis.parameters.bin_count"
    )
    if not 3 <= bin_count <= 20:
        raise StrategyCandidateReportError(
            "univariate analysis.parameters.bin_count must be between 3 and 20"
        )
    _positive_finite_number(
        value["smoothing"], "univariate analysis.parameters.smoothing"
    )
    _rate(value["min_bin_pct"], "univariate analysis.parameters.min_bin_pct", upper=0.5)
    seed = _non_negative_int(value["seed"], "univariate analysis.parameters.seed")
    if seed > 2_147_483_647:
        raise StrategyCandidateReportError(
            "univariate analysis.parameters.seed is outside the supported range"
        )
    for field in ("loan_amount", "overdue_amount"):
        item = value[field]
        if item is not None:
            _required_text(item, f"univariate analysis.parameters.{field}")


def _validate_features(
    value: object, *, row_count: int, feature_count: int
) -> dict[tuple[str, str], Mapping[str, Any]]:
    features = _array(value, "univariate analysis.features", required=True)
    if len(features) != feature_count:
        raise StrategyCandidateReportError(
            "univariate analysis.feature_count does not match features"
        )
    names: set[str] = set()
    available_metrics: dict[tuple[str, str], Mapping[str, Any]] = {}
    for feature_index, feature in enumerate(features):
        path = f"univariate analysis.features[{feature_index}]"
        if not isinstance(feature, Mapping):
            raise StrategyCandidateReportError(f"{path} must be an object")
        _require_exact_fields(feature, _FEATURE_FIELDS, path)
        name = _required_text(feature["feature"], f"{path}.feature")
        if name in names:
            raise StrategyCandidateReportError(
                "univariate analysis.features contains duplicate feature names"
            )
        names.add(name)
        feature_type = feature["feature_type"]
        if feature_type not in {"numeric", "categorical"}:
            raise StrategyCandidateReportError(
                f"{path}.feature_type must be numeric or categorical"
            )
        if _positive_int(feature["row_count"], f"{path}.row_count") != row_count:
            raise StrategyCandidateReportError(
                f"{path}.row_count does not match row_count"
            )
        _rate(feature["missing_rate"], f"{path}.missing_rate")
        sentinels = _array(feature["sentinel_values"], f"{path}.sentinel_values")
        for sentinel_index, sentinel in enumerate(sentinels):
            _json_scalar(sentinel, f"{path}.sentinel_values[{sentinel_index}]")
            if feature_type == "numeric" and (
                isinstance(sentinel, bool) or not isinstance(sentinel, int | float)
            ):
                raise StrategyCandidateReportError(
                    f"{path}.sentinel_values must be numeric for a numeric feature"
                )
        sentinel_keys = [
            _canonical_json_text(float(item) if feature_type == "numeric" else item)
            for item in sentinels
        ]
        if len(set(sentinel_keys)) != len(sentinel_keys):
            raise StrategyCandidateReportError(
                f"{path}.sentinel_values must not contain duplicates"
            )
        methods = _array(feature["methods"], f"{path}.methods", required=True)
        method_names: set[str] = set()
        for method_index, method in enumerate(methods):
            method_name, metrics = _validate_method(
                method,
                path=f"{path}.methods[{method_index}]",
                row_count=row_count,
                feature_name=name,
                feature_type=str(feature_type),
                sentinel_values=sentinels,
            )
            if method_name in method_names:
                raise StrategyCandidateReportError(
                    f"{path}.methods contains duplicate method names"
                )
            method_names.add(method_name)
            if metrics is not None:
                available_metrics[(name, method_name)] = metrics
    return available_metrics


def _validate_method(
    value: object,
    *,
    path: str,
    row_count: int,
    feature_name: str,
    feature_type: str,
    sentinel_values: Sequence[object],
) -> tuple[str, Mapping[str, Any] | None]:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(f"{path} must be an object")
    _require_exact_fields(value, _METHOD_FIELDS, path)
    method = _required_text(value["method"], f"{path}.method")
    if method not in _METHODS:
        raise StrategyCandidateReportError(f"{path}.method is unsupported")
    if value["requested_method"] != method:
        raise StrategyCandidateReportError(f"{path}.requested_method must match method")

    status = value["status"]
    bins = _array(value["bins"], f"{path}.bins")
    if status == "available":
        if value["actual_method"] != method:
            raise StrategyCandidateReportError(
                f"{path}.actual_method must match an available requested method"
            )
        evidence = _array(value["evidence"], f"{path}.evidence")
        for index, item in enumerate(evidence):
            if not isinstance(item, Mapping):
                raise StrategyCandidateReportError(
                    f"{path}.evidence[{index}] must be an object"
                )
        metrics = _validate_method_metrics(value["metrics"], path=f"{path}.metrics")
        if not bins:
            raise StrategyCandidateReportError(
                f"{path}.bins must not be empty when available"
            )
        _validate_bins(
            bins,
            path=f"{path}.bins",
            row_count=row_count,
            method_ks=float(metrics["ks"]),
            feature_name=feature_name,
            feature_type=feature_type,
            sentinel_values=sentinel_values,
        )
        return method, metrics
    if status == "unavailable":
        if value["actual_method"] is not None:
            raise StrategyCandidateReportError(
                f"{path}.actual_method must be null when unavailable"
            )
        if not isinstance(value["evidence"], Mapping):
            raise StrategyCandidateReportError(
                f"{path}.evidence must be an object when unavailable"
            )
        _required_text(value["evidence"].get("kind"), f"{path}.evidence.kind")
        if value["metrics"] is not None:
            raise StrategyCandidateReportError(
                f"{path}.metrics must be null when unavailable"
            )
        if bins:
            raise StrategyCandidateReportError(
                f"{path}.bins must be empty when unavailable"
            )
        return method, None
    raise StrategyCandidateReportError(
        f"{path}.status must be available or unavailable"
    )


def _validate_method_metrics(value: object, *, path: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(f"{path} must be an object")
    _require_exact_fields(value, _METHOD_METRIC_FIELDS, path)
    _finite_number(value["iv"], f"{path}.iv")
    _rate(value["ks"], f"{path}.ks")
    _rate(value["auc"], f"{path}.auc")
    _required_text(value["risk_direction"], f"{path}.risk_direction")
    _rate(value["missing_rate"], f"{path}.missing_rate")
    _validate_amount_metrics(value["amount_metrics"], path=f"{path}.amount_metrics")
    return value


def _validate_bins(
    bins: Sequence[object],
    *,
    path: str,
    row_count: int,
    method_ks: float,
    feature_name: str,
    feature_type: str,
    sentinel_values: Sequence[object],
) -> None:
    ids: set[str] = set()
    counted_rows = 0
    cumulative_values: list[float] = []
    for expected_index, item in enumerate(bins):
        item_path = f"{path}[{expected_index}]"
        if not isinstance(item, Mapping):
            raise StrategyCandidateReportError(f"{item_path} must be an object")
        fields = set(item)
        missing = _BIN_REQUIRED_FIELDS - fields
        unknown = fields - (_BIN_REQUIRED_FIELDS | _BIN_OPTIONAL_FIELDS)
        if missing or unknown:
            _raise_field_error(item_path, missing=missing, unknown=unknown)
        if _non_negative_int(item["index"], f"{item_path}.index") != expected_index:
            raise StrategyCandidateReportError(f"{item_path}.index is not canonical")
        bin_id = _required_text(item["id"], f"{item_path}.id")
        if bin_id in ids:
            raise StrategyCandidateReportError(f"{path} contains duplicate bin ids")
        ids.add(bin_id)
        kind = item["kind"]
        if kind not in _BIN_KINDS:
            raise StrategyCandidateReportError(f"{item_path}.kind is unsupported")
        if not isinstance(item["condition"], Mapping):
            raise StrategyCandidateReportError(
                f"{item_path}.condition must be an object"
            )
        try:
            canonical_condition = canonicalize_expression(item["condition"])
        except StrategyError as exc:
            raise StrategyCandidateReportError(
                f"{item_path}.condition is not canonical Strategy DSL"
            ) from exc
        if _canonical_json_text(canonical_condition) != _canonical_json_text(
            item["condition"]
        ):
            raise StrategyCandidateReportError(
                f"{item_path}.condition is not canonical Strategy DSL"
            )
        expected_condition = _expected_bin_condition(
            item,
            feature_name=feature_name,
            feature_type=feature_type,
            sentinel_values=sentinel_values,
            path=item_path,
        )
        if _canonical_json_text(canonical_condition) != _canonical_json_text(
            expected_condition
        ):
            raise StrategyCandidateReportError(
                f"{item_path}.condition does not match its bin definition"
            )
        count = _non_negative_int(item["count"], f"{item_path}.count")
        share = _rate(item["share"], f"{item_path}.share")
        good = _non_negative_int(item["good"], f"{item_path}.good")
        bad = _non_negative_int(item["bad"], f"{item_path}.bad")
        if good + bad != count:
            raise StrategyCandidateReportError(
                f"{item_path}.count must equal good plus bad"
            )
        if not math.isclose(share, count / row_count, rel_tol=0.0, abs_tol=1e-15):
            raise StrategyCandidateReportError(
                f"{item_path}.share does not match count and row_count"
            )
        counted_rows += count
        if item["bad_rate"] is not None:
            bad_rate = _rate(item["bad_rate"], f"{item_path}.bad_rate")
            if count == 0 or not math.isclose(
                bad_rate,
                bad / count,
                rel_tol=0.0,
                abs_tol=1e-15,
            ):
                raise StrategyCandidateReportError(
                    f"{item_path}.bad_rate does not match bad and count"
                )
        elif count:
            raise StrategyCandidateReportError(
                f"{item_path}.bad_rate must be present for a non-empty bin"
            )
        _finite_number(item["woe"], f"{item_path}.woe")
        _finite_number(item["iv_contribution"], f"{item_path}.iv_contribution")
        if item["lift"] is not None:
            _finite_number(item["lift"], f"{item_path}.lift")
        cumulative_values.append(
            _rate(item["cumulative_ks"], f"{item_path}.cumulative_ks")
        )
        _validate_amount_metrics(
            item["amount_metrics"], path=f"{item_path}.amount_metrics"
        )
        _validate_bin_shape(item, kind=kind, path=item_path)
    if counted_rows != row_count:
        raise StrategyCandidateReportError(
            f"{path} bin counts do not cover the analysis row_count"
        )
    if not math.isclose(
        max(cumulative_values, default=0.0),
        method_ks,
        rel_tol=0.0,
        abs_tol=1e-15,
    ):
        raise StrategyCandidateReportError(
            f"{path} cumulative KS does not match the method KS"
        )


def _expected_bin_condition(
    item: Mapping[str, Any],
    *,
    feature_name: str,
    feature_type: str,
    sentinel_values: Sequence[object],
    path: str,
) -> dict[str, Any]:
    kind = item["kind"]
    normalized_sentinels = (
        [float(value) for value in sentinel_values]
        if feature_type == "numeric"
        else list(sentinel_values)
    )
    if kind == "missing":
        return {"op": "is_null", "field": feature_name}
    if kind in {"category", "sentinel"}:
        if kind == "category" and feature_type != "categorical":
            raise StrategyCandidateReportError(
                f"{path} category bin requires a categorical feature"
            )
        if "value" not in item:
            raise StrategyCandidateReportError(f"{path}.value is required")
        value_key = _canonical_json_text(item["value"])
        sentinel_keys = {_canonical_json_text(value) for value in normalized_sentinels}
        if (kind == "sentinel") != (value_key in sentinel_keys):
            raise StrategyCandidateReportError(
                f"{path}.kind does not match the feature sentinel values"
            )
        condition = {
            "op": "compare",
            "field": feature_name,
            "operator": "==",
            "value": item["value"],
            "missing": "no_match",
        }
        if feature_type == "categorical":
            condition["coercion"] = "strict"
        return canonicalize_expression(condition)
    if kind != "numeric_interval":
        raise StrategyCandidateReportError(f"{path}.kind is unsupported")
    if feature_type != "numeric":
        raise StrategyCandidateReportError(
            f"{path} numeric interval requires a numeric feature"
        )
    if item.get("include_lower") is not True or item.get("include_upper") is not False:
        raise StrategyCandidateReportError(
            f"{path} numeric interval must be left-closed and right-open"
        )
    args: list[dict[str, Any]] = []
    if item.get("lower") is not None:
        args.append(
            {
                "op": "compare",
                "field": feature_name,
                "operator": ">=",
                "value": item["lower"],
                "missing": "no_match",
            }
        )
    if item.get("upper") is not None:
        args.append(
            {
                "op": "compare",
                "field": feature_name,
                "operator": "<",
                "value": item["upper"],
                "missing": "no_match",
            }
        )
    if normalized_sentinels:
        args.append(
            {
                "op": "compare",
                "field": feature_name,
                "operator": "not_in",
                "value": normalized_sentinels,
                "missing": "no_match",
            }
        )
    if not args:
        return {"op": "is_not_null", "field": feature_name}
    if len(args) == 1:
        return canonicalize_expression(args[0])
    return canonicalize_expression({"op": "and", "args": args})


def _validate_bin_shape(item: Mapping[str, Any], *, kind: str, path: str) -> None:
    if kind == "numeric_interval":
        required = {"lower", "upper", "include_lower", "include_upper"}
        if not required <= set(item) or "value" in item:
            raise StrategyCandidateReportError(
                f"{path} has invalid numeric interval fields"
            )
        for field in ("lower", "upper"):
            if item[field] is not None:
                _finite_number(item[field], f"{path}.{field}")
        for field in ("include_lower", "include_upper"):
            if not isinstance(item[field], bool):
                raise StrategyCandidateReportError(f"{path}.{field} must be a boolean")
        return
    if kind in {"category", "sentinel"}:
        if "value" not in item or set(item) & {
            "lower",
            "upper",
            "include_lower",
            "include_upper",
        }:
            raise StrategyCandidateReportError(f"{path} has invalid scalar bin fields")
        _json_scalar(item["value"], f"{path}.value")
        return
    if set(item) & _BIN_OPTIONAL_FIELDS:
        raise StrategyCandidateReportError(f"{path} has invalid missing-bin fields")


def _validate_amount_metrics(value: object, *, path: str) -> None:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(f"{path} must be an object")
    _require_exact_fields(value, _AMOUNT_METRIC_FIELDS, path)
    _validate_amount_measure(value["loan_amount"], path=f"{path}.loan_amount")
    _validate_amount_measure(value["overdue_amount"], path=f"{path}.overdue_amount")
    _validate_overdue_rate(value["overdue_rate"], path=f"{path}.overdue_rate")


def _validate_amount_measure(value: object, *, path: str) -> None:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(f"{path} must be an object")
    status = value.get("status")
    if status == "available":
        _require_exact_fields(
            value, frozenset({"status", "sum", "covered_count", "coverage_rate"}), path
        )
        number = _finite_number(value["sum"], f"{path}.sum")
        if number < 0:
            raise StrategyCandidateReportError(f"{path}.sum must be non-negative")
        _non_negative_int(value["covered_count"], f"{path}.covered_count")
        _rate(value["coverage_rate"], f"{path}.coverage_rate")
        return
    if status == "unavailable":
        allowed = frozenset({"status", "reason", "coverage_rate"})
        fields = set(value)
        missing = {"status", "reason"} - fields
        unknown = fields - allowed
        if missing or unknown:
            _raise_field_error(path, missing=missing, unknown=unknown)
        _required_text(value["reason"], f"{path}.reason")
        if "coverage_rate" in value:
            _rate(value["coverage_rate"], f"{path}.coverage_rate")
        return
    raise StrategyCandidateReportError(f"{path}.status is unsupported")


def _validate_overdue_rate(value: object, *, path: str) -> None:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(f"{path} must be an object")
    status = value.get("status")
    if status == "available":
        _require_exact_fields(
            value, frozenset({"status", "value", "paired_count"}), path
        )
        number = _finite_number(value["value"], f"{path}.value")
        if number < 0:
            raise StrategyCandidateReportError(f"{path}.value must be non-negative")
        _non_negative_int(value["paired_count"], f"{path}.paired_count")
        return
    if status in {"unavailable", "not_applicable"}:
        _require_exact_fields(value, frozenset({"status", "reason"}), path)
        _required_text(value["reason"], f"{path}.reason")
        return
    raise StrategyCandidateReportError(f"{path}.status is unsupported")


def _validate_rankings(
    value: object,
    *,
    available_metrics: Mapping[tuple[str, str], Mapping[str, Any]],
) -> None:
    rankings = _array(value, "univariate analysis.rankings")
    seen: set[tuple[str, str]] = set()
    for index, ranking in enumerate(rankings):
        path = f"univariate analysis.rankings[{index}]"
        if not isinstance(ranking, Mapping):
            raise StrategyCandidateReportError(f"{path} must be an object")
        _require_exact_fields(ranking, _RANKING_FIELDS, path)
        identity = (
            _required_text(ranking["feature"], f"{path}.feature"),
            _required_text(ranking["method"], f"{path}.method"),
        )
        if identity in seen:
            raise StrategyCandidateReportError(
                "univariate analysis.rankings contains duplicate feature/method pairs"
            )
        seen.add(identity)
        metrics = available_metrics.get(identity)
        if metrics is None:
            raise StrategyCandidateReportError(
                f"{path} does not reference an available feature method"
            )
        for metric in ("iv", "ks", "auc"):
            _finite_number(ranking[metric], f"{path}.{metric}")
            if ranking[metric] != metrics[metric]:
                raise StrategyCandidateReportError(
                    f"{path}.{metric} does not match the source method metric"
                )
    if seen != set(available_metrics):
        raise StrategyCandidateReportError(
            "univariate analysis.rankings must cover every available feature method"
        )


def _validate_resource_budget(
    value: object, *, row_count: int, feature_count: int, method_count: int
) -> None:
    if not isinstance(value, Mapping):
        raise StrategyCandidateReportError(
            "univariate analysis.resource_budget must be an object"
        )
    _require_exact_fields(
        value, _RESOURCE_BUDGET_FIELDS, "univariate analysis.resource_budget"
    )
    for field in (
        "max_rows",
        "max_features",
        "max_bins",
        "max_categories",
        "rows_used",
        "features_used",
        "method_runs",
    ):
        _non_negative_int(value[field], f"univariate analysis.resource_budget.{field}")
    if not isinstance(value["truncated"], bool):
        raise StrategyCandidateReportError(
            "univariate analysis.resource_budget.truncated must be a boolean"
        )
    expected = {
        "rows_used": row_count,
        "features_used": feature_count,
        "method_runs": method_count,
    }
    for field, expected_value in expected.items():
        if value[field] != expected_value:
            raise StrategyCandidateReportError(
                f"univariate analysis.resource_budget.{field} does not match the result"
            )


def _render_xlsx(payload: Mapping[str, Any]) -> bytes:
    evidence = payload["candidate_evidence"]
    analysis = payload["univariate_analysis"]
    assert isinstance(evidence, Mapping)
    assert isinstance(analysis, Mapping)

    workbook = Workbook(write_only=True)
    workbook.properties.creator = "MARVIS"
    workbook.properties.lastModifiedBy = "MARVIS"
    workbook.properties.title = "Strategy Candidate Report"
    workbook.properties.description = REPORT_SCHEMA_VERSION
    workbook.properties.created = _FIXED_WORKBOOK_DATETIME
    workbook.properties.modified = _FIXED_WORKBOOK_DATETIME
    try:
        _write_summary(workbook, evidence=evidence, analysis=analysis)
        _write_rankings(workbook, analysis=analysis)
        _write_bins(workbook, analysis=analysis)
        _write_metrics(workbook, evidence=evidence, analysis=analysis)
        _write_red_flags(workbook, evidence=evidence, analysis=analysis)
        _write_lineage(workbook, evidence=evidence, analysis=analysis)
        raw = BytesIO()
        workbook.save(raw)
    finally:
        workbook.close()
    return _canonicalize_xlsx_bytes(raw.getvalue())


def _write_summary(
    workbook: Workbook, *, evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> None:
    sheet = workbook.create_sheet("Summary")
    _append_row(sheet, ("Field", "Value"))
    generation = evidence["generation"]
    assert isinstance(generation, Mapping)
    rows = (
        ("report_schema_version", REPORT_SCHEMA_VERSION),
        ("candidate_id", evidence["candidate_id"]),
        ("candidate_type", evidence["candidate_type"]),
        ("effect_stage", evidence["effect_stage"]),
        ("validation_status", evidence["validation_status"]),
        ("producer_version", evidence["producer_version"]),
        ("evidence_hash", evidence["evidence_hash"]),
        ("target", analysis["target"]),
        ("row_count", analysis["row_count"]),
        ("feature_count", analysis["feature_count"]),
        ("target_definition", _canonical_json_text(analysis["target_definition"])),
        ("generation_parameters", _canonical_json_text(generation["parameters"])),
        ("generation_seed", generation["seed"]),
        ("generation_budget", generation["budget"]),
        ("generation_truncated", generation["truncated"]),
        ("univariate_parameters", _canonical_json_text(analysis["parameters"])),
        ("resource_budget", _canonical_json_text(analysis["resource_budget"])),
    )
    for row in rows:
        _append_row(sheet, row)


def _write_rankings(workbook: Workbook, *, analysis: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("Rankings")
    _append_row(sheet, ("Feature", "Method", "IV", "KS", "AUC"))
    rankings = analysis["rankings"]
    assert isinstance(rankings, Sequence)
    for item in rankings:
        assert isinstance(item, Mapping)
        _append_row(
            sheet,
            (item["feature"], item["method"], item["iv"], item["ks"], item["auc"]),
        )


_BIN_HEADERS = (
    "Feature",
    "Feature Type",
    "Method",
    "Method Status",
    "Bin Index",
    "Bin ID",
    "Bin Kind",
    "Lower",
    "Upper",
    "Include Lower",
    "Include Upper",
    "Value",
    "Condition JSON",
    "Count",
    "Share",
    "Good",
    "Bad",
    "Bad Rate",
    "WOE",
    "IV Contribution",
    "Lift",
    "Cumulative KS",
    "Amount Metrics JSON",
)


def _write_bins(workbook: Workbook, *, analysis: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("Bins")
    _append_row(sheet, _BIN_HEADERS)
    features = analysis["features"]
    assert isinstance(features, Sequence)
    for feature in features:
        assert isinstance(feature, Mapping)
        methods = feature["methods"]
        assert isinstance(methods, Sequence)
        for method in methods:
            assert isinstance(method, Mapping)
            bins = method["bins"]
            assert isinstance(bins, Sequence)
            if not bins:
                _append_row(
                    sheet,
                    (
                        feature["feature"],
                        feature["feature_type"],
                        method["method"],
                        method["status"],
                        *([None] * (len(_BIN_HEADERS) - 4)),
                    ),
                )
                continue
            for item in bins:
                assert isinstance(item, Mapping)
                value = item.get("value")
                _append_row(
                    sheet,
                    (
                        feature["feature"],
                        feature["feature_type"],
                        method["method"],
                        method["status"],
                        item["index"],
                        item["id"],
                        item["kind"],
                        item.get("lower"),
                        item.get("upper"),
                        item.get("include_lower"),
                        item.get("include_upper"),
                        value,
                        _canonical_json_text(item["condition"]),
                        item["count"],
                        item["share"],
                        item["good"],
                        item["bad"],
                        item["bad_rate"],
                        item["woe"],
                        item["iv_contribution"],
                        item["lift"],
                        item["cumulative_ks"],
                        _canonical_json_text(item["amount_metrics"]),
                    ),
                )


_METRIC_HEADERS = (
    "Source",
    "Feature",
    "Method",
    "Metric",
    "Dimension",
    "Status",
    "Value",
    "Details JSON",
)


def _write_metrics(
    workbook: Workbook, *, evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> None:
    sheet = workbook.create_sheet("Metrics")
    _append_row(sheet, _METRIC_HEADERS)
    metrics = evidence["metrics"]
    assert isinstance(metrics, Sequence)
    for item in metrics:
        assert isinstance(item, Mapping)
        _append_row(
            sheet,
            (
                "candidate_evidence",
                None,
                None,
                item["metric_name"],
                item["dimension"],
                item["status"],
                item["value"],
                None,
            ),
        )

    features = analysis["features"]
    assert isinstance(features, Sequence)
    for feature in features:
        assert isinstance(feature, Mapping)
        methods = feature["methods"]
        assert isinstance(methods, Sequence)
        for method in methods:
            assert isinstance(method, Mapping)
            if method["status"] == "unavailable":
                _append_row(
                    sheet,
                    (
                        "univariate_analysis",
                        feature["feature"],
                        method["method"],
                        None,
                        None,
                        "unavailable",
                        None,
                        _canonical_json_text(method["evidence"]),
                    ),
                )
                continue
            method_metrics = method["metrics"]
            assert isinstance(method_metrics, Mapping)
            for metric_name in ("iv", "ks", "auc", "risk_direction", "missing_rate"):
                _append_row(
                    sheet,
                    (
                        "univariate_analysis",
                        feature["feature"],
                        method["method"],
                        metric_name,
                        None,
                        method["status"],
                        method_metrics[metric_name],
                        None,
                    ),
                )
            amount_metrics = method_metrics["amount_metrics"]
            assert isinstance(amount_metrics, Mapping)
            for metric_name in ("loan_amount", "overdue_amount", "overdue_rate"):
                amount = amount_metrics[metric_name]
                assert isinstance(amount, Mapping)
                value_key = "sum" if metric_name != "overdue_rate" else "value"
                detail = {
                    key: item
                    for key, item in amount.items()
                    if key not in {"status", value_key}
                }
                _append_row(
                    sheet,
                    (
                        "univariate_analysis",
                        feature["feature"],
                        method["method"],
                        f"amount_metrics.{metric_name}",
                        None,
                        amount["status"],
                        amount.get(value_key),
                        _canonical_json_text(detail) if detail else None,
                    ),
                )


_RED_FLAG_HEADERS = (
    "Source",
    "Feature",
    "Method",
    "Kind",
    "Severity",
    "Message",
    "Details JSON",
)


def _write_red_flags(
    workbook: Workbook, *, evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> None:
    sheet = workbook.create_sheet("Red Flags")
    _append_row(sheet, _RED_FLAG_HEADERS)
    red_flags = evidence["red_flags"]
    assert isinstance(red_flags, Sequence)
    for flag in red_flags:
        _append_row(sheet, ("candidate_evidence", None, None, flag, None, None, None))

    features = analysis["features"]
    assert isinstance(features, Sequence)
    for feature in features:
        assert isinstance(feature, Mapping)
        methods = feature["methods"]
        assert isinstance(methods, Sequence)
        for method in methods:
            assert isinstance(method, Mapping)
            evidence_items = method["evidence"]
            if isinstance(evidence_items, Mapping):
                evidence_items = [evidence_items]
            assert isinstance(evidence_items, Sequence)
            for item in evidence_items:
                assert isinstance(item, Mapping)
                _append_row(
                    sheet,
                    (
                        "univariate_analysis",
                        feature["feature"],
                        method["method"],
                        item.get("kind"),
                        item.get("severity"),
                        item.get("message"),
                        _canonical_json_text(item),
                    ),
                )


def _write_lineage(
    workbook: Workbook, *, evidence: Mapping[str, Any], analysis: Mapping[str, Any]
) -> None:
    sheet = workbook.create_sheet("Lineage")
    _append_row(sheet, ("Field", "Value"))
    identity = evidence["identity"]
    assert isinstance(identity, Mapping)
    rows: list[tuple[object, object]] = [
        ("candidate_id", evidence["candidate_id"]),
        ("evidence_hash", evidence["evidence_hash"]),
        ("candidate_schema_version", evidence["schema_version"]),
        ("univariate_schema_version", analysis["schema_version"]),
    ]
    for field in (
        "task_id",
        "dataset_id",
        "dataset_content_hash",
        "workspace_revision",
        "workspace_generation",
        "semantic_mapping_hash",
    ):
        rows.append((f"identity.{field}", identity[field]))
    source_refs = evidence["source_refs"]
    assert isinstance(source_refs, Sequence)
    rows.extend(
        (f"source_ref[{index}]", item) for index, item in enumerate(source_refs)
    )
    for row in rows:
        _append_row(sheet, row)


def _append_row(sheet: Any, values: Sequence[object]) -> None:
    sheet.append([_xlsx_cell(value) for value in values])


def _xlsx_cell(value: object) -> object:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, float):
        # openpyxl formats numeric floats with ``%.16g``; some legitimate
        # evidence values then reload one ULP away from the kernel output.
        # XLSX is a presentation artifact (canonical JSON remains numeric), so
        # preserve the exact IEEE-754 value as shortest round-trip-safe text.
        return repr(value)
    if isinstance(value, int):
        # Excel stores only 15 significant decimal digits.  Keep large lineage
        # revisions and exact counts as text instead of silently rounding them.
        return str(value) if len(str(abs(value))) > 15 else value
    if not isinstance(value, str):
        raise StrategyCandidateReportError(
            f"report cell contains unsupported {type(value).__name__}"
        )
    text = value
    if _looks_like_formula(text):
        text = "'" + text
    text = _ILLEGAL_XLSX_CONTROL.sub(
        lambda match: f"\\u{ord(match.group(0)):04x}", text
    )
    if len(text) > _MAX_XLSX_CELL_CHARACTERS:
        raise StrategyCandidateReportError(
            "report cell exceeds Excel's 32767 character limit"
        )
    return text


def _looks_like_formula(value: str) -> bool:
    for character in value:
        if character.isspace() or unicodedata.category(character).startswith("C"):
            continue
        return character in _FORMULA_PREFIXES
    return False


def _canonicalize_xlsx_bytes(raw: bytes) -> bytes:
    source = BytesIO(raw)
    destination = BytesIO()
    with ZipFile(source, "r") as input_archive:
        members = sorted(input_archive.infolist(), key=lambda item: item.filename)
        with ZipFile(
            destination,
            "w",
            compression=ZIP_DEFLATED,
            compresslevel=9,
            allowZip64=True,
        ) as output_archive:
            for source_member in members:
                member = ZipInfo(source_member.filename, date_time=_FIXED_ZIP_DATETIME)
                member.compress_type = ZIP_DEFLATED
                member.create_system = 0
                member.external_attr = 0
                member.internal_attr = 0
                member.comment = b""
                member.extra = b""
                payload = input_archive.read(source_member.filename)
                if source_member.filename == "docProps/core.xml":
                    payload, replacements = _CORE_MODIFIED_TIMESTAMP.subn(
                        rb"\g<1>2000-01-01T00:00:00Z\g<2>", payload
                    )
                    if replacements != 1:
                        raise StrategyCandidateReportError(
                            "generated workbook has an invalid modified timestamp contract"
                        )
                output_archive.writestr(
                    member,
                    payload,
                    compress_type=ZIP_DEFLATED,
                    compresslevel=9,
                )
    return destination.getvalue()


def _detached_json_object(value: Mapping[str, Any], field_name: str) -> dict[str, Any]:
    _validate_json_value(value, path=field_name)
    restored = json.loads(_canonical_json_text(value))
    assert isinstance(restored, dict)
    return restored


def _validate_json_value(value: object, *, path: str) -> None:
    if value is None or isinstance(value, str | bool | int):
        return
    if isinstance(value, float):
        if not math.isfinite(value):
            raise StrategyCandidateReportError(f"{path} contains a non-finite number")
        return
    if isinstance(value, Mapping):
        for key, item in value.items():
            if not isinstance(key, str):
                raise StrategyCandidateReportError(f"{path} contains a non-string key")
            _validate_json_value(item, path=f"{path}.{key}")
        return
    if isinstance(value, Sequence) and not isinstance(value, str | bytes | bytearray):
        for index, item in enumerate(value):
            _validate_json_value(item, path=f"{path}[{index}]")
        return
    raise StrategyCandidateReportError(
        f"{path} contains unsupported JSON value {type(value).__name__}"
    )


def _canonical_json_bytes(value: object) -> bytes:
    return _canonical_json_text(value).encode("utf-8")


def _canonical_json_text(value: object) -> str:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
    except (TypeError, ValueError) as exc:
        raise StrategyCandidateReportError(
            f"report value is not canonical JSON: {exc}"
        ) from exc


def _array(value: object, field_name: str, *, required: bool = False) -> list[Any]:
    if isinstance(value, str | bytes | bytearray) or not isinstance(value, list):
        raise StrategyCandidateReportError(f"{field_name} must be an array")
    if required and not value:
        raise StrategyCandidateReportError(f"{field_name} must not be empty")
    return value


def _json_scalar(value: object, field_name: str) -> None:
    if value is None or isinstance(value, str | bool):
        return
    _finite_number(value, field_name)


def _required_text(value: object, field_name: str) -> str:
    if not isinstance(value, str) or not value:
        raise StrategyCandidateReportError(f"{field_name} must be non-empty text")
    return value


def _finite_number(value: object, field_name: str) -> int | float:
    if isinstance(value, bool) or not isinstance(value, int | float):
        raise StrategyCandidateReportError(f"{field_name} must be a finite number")
    if not math.isfinite(value):
        raise StrategyCandidateReportError(f"{field_name} must be a finite number")
    return value


def _positive_finite_number(value: object, field_name: str) -> int | float:
    number = _finite_number(value, field_name)
    if number <= 0:
        raise StrategyCandidateReportError(f"{field_name} must be positive")
    return number


def _rate(value: object, field_name: str, *, upper: float = 1.0) -> int | float:
    number = _finite_number(value, field_name)
    if not 0 <= number <= upper:
        raise StrategyCandidateReportError(f"{field_name} is outside its valid range")
    return number


def _is_int(value: object) -> bool:
    return isinstance(value, int) and not isinstance(value, bool)


def _non_negative_int(value: object, field_name: str) -> int:
    if not _is_int(value) or value < 0:
        raise StrategyCandidateReportError(
            f"{field_name} must be a non-negative integer"
        )
    return value


def _positive_int(value: object, field_name: str) -> int:
    if not _is_int(value) or value < 1:
        raise StrategyCandidateReportError(f"{field_name} must be a positive integer")
    return value


def _require_exact_fields(
    value: Mapping[str, Any], expected: frozenset[str], field_name: str
) -> None:
    if any(not isinstance(key, str) for key in value):
        raise StrategyCandidateReportError(f"{field_name} keys must be strings")
    missing = expected - set(value)
    unknown = set(value) - expected
    if missing or unknown:
        _raise_field_error(field_name, missing=missing, unknown=unknown)


def _raise_field_error(
    field_name: str, *, missing: set[str] | frozenset[str], unknown: set[str]
) -> None:
    details = []
    if missing:
        details.append("missing: " + ", ".join(sorted(missing)))
    if unknown:
        details.append("unknown: " + ", ".join(sorted(unknown)))
    raise StrategyCandidateReportError(
        f"{field_name} fields are invalid ({'; '.join(details)})"
    )


__all__ = [
    "REPORT_SCHEMA_VERSION",
    "REPORT_SHEET_NAMES",
    "StrategyCandidateReportError",
    "canonical_strategy_candidate_report_json",
    "render_strategy_candidate_bundle",
    "render_strategy_candidate_report_xlsx",
]
