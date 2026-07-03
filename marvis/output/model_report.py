from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from marvis.artifacts import TransactionalArtifactStore
from marvis.packs.modeling.report_compute import ReportSectionStatus


MODEL_REPORT_SHEETS = [
    "汇总",
    "样本分析",
    "Vintage",
    "特征重要性",
    "评分卡",
    "评分分段",
    "概率校准",
    "oot分箱评估_十分箱",
    "单变量分析",
    "压力测试",
]


@dataclass(frozen=True)
class ModelReportPayload:
    project_meta: dict
    dataset_split: list[dict]
    stability: list[dict]
    sample_analysis: list[dict] | None
    vintage: dict | None
    feature_importance: list[dict]
    scorecard_table: list[dict] = field(default_factory=list)
    score_bands: list[dict] = field(default_factory=list)
    calibration: list[dict] = field(default_factory=list)
    univariate: list[dict] = field(default_factory=list)
    oot_bin_table: list[dict] = field(default_factory=list)
    stress_product_removal: dict = field(default_factory=dict)
    stress_low_pricing: dict | None = None
    narratives: dict = field(default_factory=dict)
    section_status: list[ReportSectionStatus] = field(default_factory=list)


def render_model_report(payload: ModelReportPayload, out_path: Path) -> Path:
    out_path = Path(out_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_summary(workbook, payload)
    _write_section_sheet(
        workbook,
        "样本分析",
        payload.sample_analysis,
        _unavailable_reason(payload, "sample_analysis"),
    )
    _write_section_sheet(
        workbook,
        "Vintage",
        _vintage_rows(payload.vintage),
        _unavailable_reason(payload, "vintage"),
    )
    _write_section_sheet(workbook, "特征重要性", payload.feature_importance, None)
    _write_section_sheet(workbook, "评分卡", payload.scorecard_table, None)
    _write_score_band_sheet(workbook, payload.score_bands)
    _write_section_sheet(workbook, "概率校准", payload.calibration, None)
    _write_section_sheet(
        workbook,
        "oot分箱评估_十分箱",
        payload.oot_bin_table,
        _unavailable_reason(payload, "amount_bin"),
    )
    _write_section_sheet(workbook, "单变量分析", payload.univariate, None)
    _write_stress_sheet(workbook, payload)
    artifact = TransactionalArtifactStore(out_path.parent).stage(out_path.name)
    try:
        workbook.save(artifact.path)
        final_path = artifact.promote()
        artifact.commit()
        return final_path
    except Exception:
        artifact.rollback()
        raise


def _write_summary(workbook: Workbook, payload: ModelReportPayload) -> None:
    sheet = workbook.create_sheet("汇总")
    rows = [
        ("一、建模背景", ""),
        ("项目元数据", ""),
        *[(str(key), _cell(value)) for key, value in payload.project_meta.items()],
        ("数据集划分", ""),
        *_dict_rows(payload.dataset_split),
        ("稳定性指标", ""),
        *_dict_rows(payload.stability),
        ("二、样本分析结论", str(payload.narratives.get("sample", ""))),
        ("三、Vintage分析结论", str(payload.narratives.get("vintage", ""))),
        ("四、模型结论", str(payload.narratives.get("model", ""))),
        ("五、使用产品清单", _product_list_summary(payload)),
        ("六、压力测试", str(payload.narratives.get("stress", ""))),
    ]
    _write_rows(sheet, rows)


def _write_section_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict] | None,
    unavailable_reason: str | None,
) -> None:
    sheet = workbook.create_sheet(title)
    if unavailable_reason:
        sheet["A1"] = f"无业务数据（{unavailable_reason}）"
        sheet["A1"].font = Font(bold=True, color="9C0006")
        sheet["A1"].fill = PatternFill("solid", fgColor="FFC7CE")
        return
    _write_dict_table(sheet, rows or [])


def _write_score_band_sheet(workbook: Workbook, rows: list[dict]) -> None:
    """DOM-5: score-band sheet with a caption documenting the shared bin-edge basis
    and cumulation direction, followed by the table itself. Bin edges are computed
    once on train (see tools.py::_score_band_rows) so the caption states that
    explicitly instead of leaving readers to infer it from the data."""
    sheet = workbook.create_sheet("评分分段")
    if not rows:
        sheet["A1"] = "无数据"
        return
    edges_source = rows[0].get("bin_edges_source") or "train"
    direction = rows[0].get("cum_direction") or "higher_is_riskier"
    direction_label = "分数越高风险越高（PD）" if direction == "higher_is_riskier" else "分数越高风险越低（评分卡分数）"
    cum_reading = (
        "累计列自高分向低分累计（先批核低分/低风险箱）"
        if direction == "higher_is_riskier"
        else "累计列自低分向高分累计（先批核低分箱，即先批核高风险，累计口径随分数上升而放宽）"
    )
    example = _score_band_worked_example(rows)
    splits_present = "/".join(dict.fromkeys(row.get("split") for row in rows if row.get("split")))
    sheet["A1"] = (
        f"分箱边界口径：等频分箱边界在 {edges_source} 集上确定，{splits_present} 共用同一组边界；"
        f"{direction_label}；{cum_reading}。"
    )
    sheet["A1"].font = Font(bold=True)
    if example:
        sheet["A2"] = example
        sheet["A2"].font = Font(italic=True)
    _write_dict_table(sheet, rows, start_row=4)


def _score_band_worked_example(rows: list[dict]) -> str:
    """A single worked-reading row (DOM-5 how-to-fix #4): picks the OOT split's
    boundary of the cutoff-side band closest to 50% cumulative pass rate as a
    concrete "cutoff -> approval rate / bad rate" example for the report reader."""
    oot_rows = [row for row in rows if row.get("split") == "oot" and row.get("cum_count_pct") is not None]
    candidates = oot_rows or [row for row in rows if row.get("cum_count_pct") is not None]
    if not candidates:
        return ""
    closest = min(candidates, key=lambda row: abs(row["cum_count_pct"] - 0.5))
    cutoff = closest.get("score_upper") if closest.get("cum_direction") == "higher_is_riskier" else closest.get("score_lower")
    cum_pct = closest.get("cum_count_pct")
    cum_bad = closest.get("cum_bad_rate")
    if cutoff is None or cum_pct is None:
        return ""
    bad_text = f"{cum_bad:.2%}" if cum_bad is not None else "n/a"
    return (
        f"示例读法（{closest.get('split')}）：cutoff≈{cutoff:.4g} → 通过率约 {cum_pct:.2%}，"
        f"累计坏账率约 {bad_text}"
    )


def _write_stress_sheet(workbook: Workbook, payload: ModelReportPayload) -> None:
    sheet = workbook.create_sheet("压力测试")
    sheet["A1"] = "6.1 产品缺失"
    _write_dict_table(sheet, _dict_payload_rows(payload.stress_product_removal), start_row=2)
    start_row = sheet.max_row + 2
    sheet.cell(row=start_row, column=1, value="6.2 低定价人群占比提升")
    reason = _unavailable_reason(payload, "low_pricing")
    if reason:
        sheet.cell(row=start_row + 1, column=1, value=f"无业务数据（{reason}）")
    else:
        _write_dict_table(
            sheet,
            _dict_payload_rows(payload.stress_low_pricing or {}),
            start_row=start_row + 1,
        )


def _write_rows(sheet, rows: list[tuple]) -> None:
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=_cell(value))
    _style_header(sheet)


def _write_dict_table(sheet, rows: list[dict], *, start_row: int = 1) -> None:
    if not rows:
        sheet.cell(row=start_row, column=1, value="无数据")
        return
    headers = list(rows[0].keys())
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=start_row, column=col_index, value=str(header))
    for row_index, row in enumerate(rows, start=start_row + 1):
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=col_index, value=_cell(row.get(header)))
    _style_header(sheet, row=start_row)


def _dict_rows(rows: list[dict]) -> list[tuple]:
    out = []
    for row in rows:
        out.extend((str(key), _cell(value)) for key, value in row.items())
    return out


def _vintage_rows(vintage: dict | None) -> list[dict] | None:
    if vintage is None:
        return None
    headers = vintage.get("headers") or []
    counts = vintage.get("counts") or {}
    amounts = vintage.get("amounts") or {}
    rows = []
    for cohort, values in (vintage.get("curves") or {}).items():
        row = {"放款月": cohort}
        if cohort in counts:
            row["放款笔数"] = counts[cohort]
        amount = amounts.get(cohort)
        if isinstance(amount, dict):
            row["放款金额"] = amount.get("total")
            row["件均金额"] = amount.get("average")
        for header, value in zip(headers, values, strict=False):
            row[str(header)] = value
        rows.append(row)
    return rows


def _dict_payload_rows(payload: dict) -> list[dict]:
    rows = []
    for key, value in payload.items():
        if isinstance(value, dict):
            rows.append({"项目": key, **value})
        else:
            rows.append({"项目": key, "值": _cell(value)})
    return rows


def _unavailable_reason(payload: ModelReportPayload, section: str) -> str | None:
    for status in payload.section_status:
        if status.section == section and not status.available:
            return status.reason or "缺少业务数据"
    return None


def _product_list_summary(payload: ModelReportPayload) -> str:
    unavailable = _unavailable_reason(payload, "product_list")
    if unavailable:
        return unavailable
    products = []
    seen = set()
    for row in payload.feature_importance:
        product = row.get("产品名称")
        if not product:
            continue
        vendor = row.get("厂商名称")
        label = f"{product}（{vendor}）" if vendor else str(product)
        if label in seen:
            continue
        seen.add(label)
        products.append(label)
    return "；".join(products)


def _cell(value: Any):
    if isinstance(value, (str, int, float)) or value is None:
        return value
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def _style_header(sheet, *, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def payload_to_dict(payload: ModelReportPayload) -> dict:
    return {
        **asdict(payload),
        "section_status": [asdict(status) for status in payload.section_status],
    }


__all__ = ["MODEL_REPORT_SHEETS", "ModelReportPayload", "payload_to_dict", "render_model_report"]
