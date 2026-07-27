"""Fixed-skeleton model report for NON-binary targets.

The binary credit-risk sections are kept as explicit n/a sheets (rather than
deleted) so report consumers see the same workbook shape while regression /
multiclass metrics live in an extra ``模型指标`` sheet.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from marvis.artifacts import TransactionalArtifactStore
from marvis.output.model_report import MODEL_REPORT_SHEETS
from marvis.output.xlsx_safety import safe_xlsx_cell

# (display label, ModelMetrics field stem) per target type. The stem is prefixed with
# the split (train_/test_/oot_) to read the matching ModelMetrics attribute.
_METRIC_SPECS: dict[str, list[tuple[str, str]]] = {
    "continuous": [("RMSE", "rmse"), ("MAE", "mae"), ("R²", "r2")],
    "multiclass": [("macro-AUC", "macro_auc"), ("logloss", "logloss"), ("准确率", "accuracy")],
}
_TARGET_LABEL = {"continuous": "回归(连续型)", "multiclass": "多分类"}


def render_minimal_model_report(experiment, out_path: Path, *, artifact=None) -> Path:
    out_path = Path(out_path)
    config = experiment.config
    metrics = experiment.metrics
    target_type = getattr(config, "target_type", "binary")

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("汇总")
    summary["A1"] = "模型开发报告(精简)"
    summary["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("目标列", config.target_col),
        ("目标类型", _TARGET_LABEL.get(target_type, target_type)),
        ("算法", experiment.recipe_id),
        ("特征数", len(config.features)),
        ("说明", "非二分类任务:不含坏率/Vintage/OOT分箱/压力测试等二分类信贷专用分析。"),
    ]
    for offset, (key, value) in enumerate(summary_rows, start=3):
        summary[f"A{offset}"] = key
        summary[f"B{offset}"] = safe_xlsx_cell(value)

    for title in MODEL_REPORT_SHEETS:
        if title == "汇总":
            continue
        sheet = workbook.create_sheet(title)
        sheet["A1"] = "n/a"
        sheet["B1"] = "非二分类不适用"
        sheet["A1"].font = Font(bold=True)

    sheet = workbook.create_sheet("模型指标")
    for col, header in zip("ABCD", ("指标", "train", "test", "oot")):
        cell = sheet[f"{col}1"]
        cell.value = header
        cell.font = Font(bold=True)
    specs = _METRIC_SPECS.get(target_type, [("KS", "ks"), ("AUC", "auc")])
    for row, (label, stem) in enumerate(specs, start=2):
        sheet[f"A{row}"] = label
        for col, split in zip("BCD", ("train", "test", "oot")):
            value = getattr(metrics, f"{split}_{stem}", None) if metrics is not None else None
            sheet[f"{col}{row}"] = "n/a" if value is None else round(float(value), 6)

    if target_type == "multiclass":
        detail = workbook.create_sheet("多分类明细")
        headers = ("数据集", "类别", "召回率", "精确率", "样本数")
        for column, header in enumerate(headers, start=1):
            cell = detail.cell(row=1, column=column, value=header)
            cell.font = Font(bold=True)
        per_class = (
            artifact.params.get("per_class")
            if artifact is not None and isinstance(getattr(artifact, "params", None), dict)
            else {}
        )
        output_row = 2
        for split in ("train", "test", "oot"):
            split_rows = per_class.get(split) if isinstance(per_class, dict) else None
            if not isinstance(split_rows, dict):
                continue
            for class_name, row in split_rows.items():
                row = row if isinstance(row, dict) else {}
                detail.cell(row=output_row, column=1, value=split)
                detail.cell(
                    row=output_row,
                    column=2,
                    value=safe_xlsx_cell(class_name),
                )
                detail.cell(
                    row=output_row,
                    column=3,
                    value="n/a" if row.get("recall") is None else round(float(row["recall"]), 6),
                )
                detail.cell(
                    row=output_row,
                    column=4,
                    value="n/a" if row.get("precision") is None else round(float(row["precision"]), 6),
                )
                detail.cell(row=output_row, column=5, value=int(row.get("support") or 0))
                output_row += 1
        if output_row == 2:
            detail["A2"] = "n/a"
            detail["B2"] = "模型产物未记录逐类别指标"

    artifact = TransactionalArtifactStore(out_path.parent).stage(out_path.name)
    try:
        workbook.save(artifact.path)
        final_path = artifact.promote()
        artifact.commit()
        return final_path
    except Exception:
        artifact.rollback()
        raise


__all__ = ["render_minimal_model_report"]
