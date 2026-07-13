from __future__ import annotations

from collections.abc import Iterator, Sequence
import csv
from dataclasses import dataclass
import io
from itertools import product
import math
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import iterparse
from zipfile import BadZipFile, ZipFile

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet._reader import WorkSheetParser
import xlrd

from marvis.validation.input_contracts import (
    FEATURE_METADATA_SCHEMA,
    FeatureMetadataResolution,
    FeatureMetadataRow,
    MetadataCoverage,
    PmmlInputManifest,
    StressUnit,
)


FEATURE_ALIASES = (
    "feature",
    "特征名",
    "特征名称",
    "指标英文",
    "feature_name",
    "特征英文名",
    "var",
)
CATEGORY_ALIASES = (
    "category",
    "类别",
    "分类",
    "数据源",
    "来源",
    "source",
    "特征分类",
    "产品名称",
    "product",
    "厂商名称",
    "厂商",
    "供应商",
    "特征产品",
    "特征信源",
)
IMPORTANCE_ALIASES = (
    "importance",
    "feature_importance",
    "gain",
    "权重",
    "特征重要性",
    "score",
)

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")
DEFAULT_MAX_METADATA_BYTES = 128 * 1024 * 1024
DEFAULT_MAX_METADATA_ROWS = 200_000
DEFAULT_MAX_METADATA_COLUMNS = 512
MAX_XLS_METADATA_BYTES = 32 * 1024 * 1024
MAX_METADATA_CELL_CHARS = 32_768
MAX_METADATA_DECODED_CHARS = 128 * 1024 * 1024
MAX_PARQUET_MEMBER_DECODED_BYTES = 256 * 1024 * 1024
MAX_PARQUET_TOTAL_DECODED_BYTES = 512 * 1024 * 1024
PARQUET_BATCH_ROWS = 8_192
MAX_XLSX_ARCHIVE_ENTRIES = 10_000
MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
MAX_XLSX_EAGER_XML_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 250.0
MAX_ALIAS_COMBINATIONS = 1_024
MAX_ALIAS_ROW_EVALUATIONS = 2_000_000
MAX_INSPECTION_WORK_UNITS = 2_000_000
MAX_MERGED_EXPANDED_CELLS = 2_000_000
MAX_SELECTIONS = MAX_ALIAS_COMBINATIONS
MAX_DIAGNOSTICS = 64
MAX_DIAGNOSTIC_CHARS = 500
MAX_MERGED_RANGES = 10_000
_STREAM_CHUNK_BYTES = 64 * 1024
_MAX_XML_PROLOG_BYTES = 64 * 1024

_FEATURE_ALIAS_SET = frozenset(FEATURE_ALIASES)
_CATEGORY_ALIAS_SET = frozenset(CATEGORY_ALIASES)
_IMPORTANCE_ALIAS_SET = frozenset(IMPORTANCE_ALIASES)
_ALL_ALIAS_SET = _FEATURE_ALIAS_SET | _CATEGORY_ALIAS_SET | _IMPORTANCE_ALIAS_SET


@dataclass(frozen=True)
class FeatureMetadataSelection:
    sheet_name: str | None
    feature_col: str
    category_col: str
    importance_col: str


@dataclass(frozen=True)
class FeatureMetadataInspection:
    path: str
    selections: tuple[FeatureMetadataSelection, ...]
    blocking_errors: tuple[str, ...]
    inspection_complete: bool = True

    def only_valid_selection(self) -> FeatureMetadataSelection:
        if self.blocking_errors:
            raise ValueError(
                _bounded_text("; ".join(self.blocking_errors), MAX_DIAGNOSTIC_CHARS)
            )
        if len(self.selections) != 1:
            raise ValueError("feature metadata selection requires user confirmation")
        return self.selections[0]


class FeatureMetadataInspectionIncomplete(ValueError):
    """Candidate inspection could not finish within the supported safety bounds."""


@dataclass(frozen=True)
class _Table:
    sheet_name: str | None
    columns: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]


@dataclass(frozen=True)
class _ExcelSheetOutcome:
    table: _Table | None
    diagnostics: tuple[str, ...]
    structural_error: str | None = None


@dataclass(frozen=True)
class _MergedRange:
    min_col: int
    min_row: int
    max_col: int
    max_row: int


@dataclass
class _InspectionLedger:
    work_limit: int
    merged_cell_limit: int
    work_used: int = 0
    expanded_cells: int = 0

    def charge_work(self, units: int) -> None:
        if units < 0 or units > self.work_limit - self.work_used:
            raise _InspectionBudgetError(
                "feature metadata inspection work limit exceeded"
            )
        self.work_used += units

    def charge_merged_cells(self, cells: int) -> None:
        if cells < 0 or cells > self.merged_cell_limit - self.expanded_cells:
            raise _InspectionBudgetError(
                "feature metadata merged expansion limit exceeded"
            )
        self.expanded_cells += cells


@dataclass(frozen=True)
class _ManifestContext:
    model_features: tuple[str, ...]
    model_feature_set: frozenset[str]
    unsupported_by_feature: dict[str, str]
    stress_units: dict[str, tuple[str, ...]]
    work_size: int


class _MetadataLimitError(FeatureMetadataInspectionIncomplete):
    pass


class _InspectionBudgetError(FeatureMetadataInspectionIncomplete):
    pass


class _BoundedBinaryReader(io.RawIOBase):
    def __init__(self, handle: Any, *, limit: int) -> None:
        self._handle = handle
        self._limit = limit
        self._consumed = 0

    def readable(self) -> bool:
        return True

    def readinto(self, buffer: Any) -> int:
        remaining = self._limit + 1 - self._consumed
        if remaining <= 0:
            raise _MetadataLimitError("feature metadata exceeds byte limit")
        payload = self._handle.read(min(len(buffer), remaining))
        count = len(payload)
        if count:
            buffer[:count] = payload
            self._consumed += count
            if self._consumed > self._limit:
                raise _MetadataLimitError("feature metadata exceeds byte limit")
        return count


class _CountingTextIterator:
    def __init__(self, handle: Any, *, limit: int) -> None:
        self._handle = handle
        self._limit = limit
        self._decoded = 0

    def __iter__(self) -> _CountingTextIterator:
        return self

    def __next__(self) -> str:
        line = next(self._handle)
        self._decoded += len(line)
        if self._decoded > self._limit:
            raise _MetadataLimitError(
                "feature metadata exceeds decoded character limit"
            )
        return line


def inspect_feature_metadata(
    path: Path,
    manifest: PmmlInputManifest,
    *,
    max_bytes: int = DEFAULT_MAX_METADATA_BYTES,
    max_rows: int = DEFAULT_MAX_METADATA_ROWS,
    max_columns: int = DEFAULT_MAX_METADATA_COLUMNS,
    max_diagnostics: int = MAX_DIAGNOSTICS,
    max_diagnostic_chars: int = MAX_DIAGNOSTIC_CHARS,
) -> FeatureMetadataInspection:
    """Inspect bounded metadata candidates without choosing among aliases."""

    try:
        return _inspect_feature_metadata(
            path,
            manifest,
            max_bytes=max_bytes,
            max_rows=max_rows,
            max_columns=max_columns,
            max_diagnostics=max_diagnostics,
            max_diagnostic_chars=max_diagnostic_chars,
        )
    except ValueError as exc:
        limit = (
            min(max_diagnostic_chars, MAX_DIAGNOSTIC_CHARS)
            if isinstance(max_diagnostic_chars, int)
            and not isinstance(max_diagnostic_chars, bool)
            and max_diagnostic_chars > 0
            else MAX_DIAGNOSTIC_CHARS
        )
        bounded = _bounded_text(str(exc), limit)
        if isinstance(exc, FeatureMetadataInspectionIncomplete):
            raise FeatureMetadataInspectionIncomplete(bounded) from None
        raise ValueError(bounded) from None


def _inspect_feature_metadata(
    path: Path,
    manifest: PmmlInputManifest,
    *,
    max_bytes: int,
    max_rows: int,
    max_columns: int,
    max_diagnostics: int,
    max_diagnostic_chars: int,
) -> FeatureMetadataInspection:

    selected_path = _validated_path(path)
    _validate_limits(
        max_bytes=max_bytes,
        max_rows=max_rows,
        max_columns=max_columns,
        max_diagnostics=max_diagnostics,
        max_diagnostic_chars=max_diagnostic_chars,
    )
    _validate_file_before_read(selected_path, max_bytes=max_bytes)
    ledger = _InspectionLedger(
        work_limit=MAX_INSPECTION_WORK_UNITS,
        merged_cell_limit=MAX_MERGED_EXPANDED_CELLS,
    )
    manifest_context = _build_manifest_context(manifest, ledger=ledger)
    resolution_cache: dict[
        tuple[tuple[str, ...], tuple[str, ...], tuple[float, ...]], str | None
    ] = {}

    selections: list[FeatureMetadataSelection] = []
    diagnostics: list[str] = []
    structural_errors: list[str] = []
    incomplete_diagnostics: list[str] = []
    inspection_complete = True
    suffix = selected_path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        outcomes = _iter_excel_candidate_outcomes(
            selected_path,
            max_rows=max_rows,
            max_columns=max_columns,
            ledger=ledger,
        )
        for outcome in outcomes:
            _extend_diagnostics(
                diagnostics,
                outcome.diagnostics,
                limit=max_diagnostics,
                char_limit=max_diagnostic_chars,
            )
            if outcome.structural_error:
                structural_errors.append(outcome.structural_error)
            if outcome.table is None:
                continue
            valid, rejected, selection_complete = _column_selections(
                outcome.table,
                manifest_context,
                ledger=ledger,
                resolution_cache=resolution_cache,
                max_rejections=max_diagnostics,
                diagnostic_char_limit=max_diagnostic_chars,
            )
            inspection_complete = inspection_complete and selection_complete
            if not selection_complete:
                incomplete_diagnostics.extend(rejected)
            if len(selections) + len(valid) > MAX_SELECTIONS:
                raise FeatureMetadataInspectionIncomplete(
                    "feature metadata selection limit exceeded"
                )
            selections.extend(valid)
            _extend_diagnostics(
                diagnostics,
                rejected,
                limit=max_diagnostics,
                char_limit=max_diagnostic_chars,
            )
    else:
        table = _read_single_table(
            selected_path,
            max_bytes=max_bytes,
            max_rows=max_rows,
            max_columns=max_columns,
        )
        valid, rejected, selection_complete = _column_selections(
            table,
            manifest_context,
            ledger=ledger,
            resolution_cache=resolution_cache,
            max_rejections=max_diagnostics,
            diagnostic_char_limit=max_diagnostic_chars,
        )
        inspection_complete = inspection_complete and selection_complete
        if not selection_complete:
            incomplete_diagnostics.extend(rejected)
        if len(valid) > MAX_SELECTIONS:
            raise FeatureMetadataInspectionIncomplete(
                "feature metadata selection limit exceeded"
            )
        selections.extend(valid)
        _extend_diagnostics(
            diagnostics,
            rejected,
            limit=max_diagnostics,
            char_limit=max_diagnostic_chars,
        )

    if selections:
        return FeatureMetadataInspection(
            str(selected_path),
            tuple(selections),
            tuple(incomplete_diagnostics[:max_diagnostics]),
            inspection_complete,
        )
    if structural_errors:
        raise ValueError(_bounded_text(structural_errors[0], max_diagnostic_chars))
    if not diagnostics:
        diagnostics.append("feature metadata has no candidate table")
    return FeatureMetadataInspection(
        str(selected_path),
        (),
        tuple(diagnostics[:max_diagnostics]),
        inspection_complete,
    )


def normalize_feature_metadata(
    path: Path,
    *,
    selection: FeatureMetadataSelection,
    manifest: PmmlInputManifest,
    max_bytes: int = DEFAULT_MAX_METADATA_BYTES,
    max_rows: int = DEFAULT_MAX_METADATA_ROWS,
    max_columns: int = DEFAULT_MAX_METADATA_COLUMNS,
) -> FeatureMetadataResolution:
    """Normalize one confirmed metadata selection into a complete resolution."""

    try:
        return _normalize_feature_metadata(
            path,
            selection=selection,
            manifest=manifest,
            max_bytes=max_bytes,
            max_rows=max_rows,
            max_columns=max_columns,
        )
    except ValueError as exc:
        raise _bounded_value_error(exc, limit=MAX_DIAGNOSTIC_CHARS) from None


def _normalize_feature_metadata(
    path: Path,
    *,
    selection: FeatureMetadataSelection,
    manifest: PmmlInputManifest,
    max_bytes: int,
    max_rows: int,
    max_columns: int,
) -> FeatureMetadataResolution:

    selected_path = _validated_path(path)
    _validate_limits(
        max_bytes=max_bytes,
        max_rows=max_rows,
        max_columns=max_columns,
        max_diagnostics=MAX_DIAGNOSTICS,
        max_diagnostic_chars=MAX_DIAGNOSTIC_CHARS,
    )
    _validate_file_before_read(selected_path, max_bytes=max_bytes)
    ledger = _InspectionLedger(
        work_limit=MAX_INSPECTION_WORK_UNITS,
        merged_cell_limit=MAX_MERGED_EXPANDED_CELLS,
    )
    manifest_context = _build_manifest_context(manifest, ledger=ledger)
    _validate_selection(selection)
    table = _read_selected_table(
        selected_path,
        selection=selection,
        max_bytes=max_bytes,
        max_rows=max_rows,
        max_columns=max_columns,
        ledger=ledger,
    )
    rows = _normalize_rows(table, selection)
    merged = _merge_identical_and_reject_conflicts(rows)
    return _resolve_against_manifest(merged, manifest_context)


def _validated_path(path: Path) -> Path:
    selected = Path(path).expanduser().resolve()
    if not selected.is_file():
        raise ValueError("feature metadata file does not exist")
    return selected


def _validate_limits(
    *,
    max_bytes: int,
    max_rows: int,
    max_columns: int,
    max_diagnostics: int,
    max_diagnostic_chars: int,
) -> None:
    for value, label in (
        (max_bytes, "metadata byte"),
        (max_columns, "metadata column"),
        (max_diagnostics, "metadata diagnostic"),
        (max_diagnostic_chars, "metadata diagnostic character"),
    ):
        if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
            raise ValueError(f"{label} limit must be a positive integer")
    if isinstance(max_rows, bool) or not isinstance(max_rows, int) or max_rows < 0:
        raise ValueError("metadata row limit must be a non-negative integer")


def _validate_file_before_read(path: Path, *, max_bytes: int) -> None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        # The CSV stream itself enforces max_bytes + 1 without a TOCTOU stat.
        return
    if suffix == ".xls":
        _enforce_byte_limit(
            path,
            min(max_bytes, MAX_XLS_METADATA_BYTES),
            message="feature metadata exceeds XLS byte limit",
        )
        return
    _enforce_byte_limit(path, max_bytes)
    if suffix == ".xlsx":
        _validate_xlsx_archive(path)


def _enforce_byte_limit(
    path: Path, limit: int, *, message: str = "feature metadata exceeds byte limit"
) -> None:
    try:
        size = path.stat().st_size
    except OSError as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot read feature metadata file"
        ) from exc
    if size > limit:
        raise FeatureMetadataInspectionIncomplete(message)


def _validate_xlsx_archive(path: Path) -> None:
    try:
        with ZipFile(path, "r") as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise FeatureMetadataInspectionIncomplete(
                    "feature metadata XLSX exceeds archive entry limit"
                )
            total_uncompressed = 0
            for member in members:
                if member.flag_bits & 0x1:
                    raise FeatureMetadataInspectionIncomplete(
                        "encrypted feature metadata XLSX is not supported"
                    )
                if member.file_size > MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES:
                    raise FeatureMetadataInspectionIncomplete(
                        "feature metadata XLSX member exceeds uncompressed byte limit"
                    )
                is_xml = member.filename.endswith((".xml", ".rels"))
                is_worksheet = member.filename.startswith("xl/worksheets/")
                if (
                    is_xml
                    and not is_worksheet
                    and member.file_size > MAX_XLSX_EAGER_XML_UNCOMPRESSED_BYTES
                ):
                    raise FeatureMetadataInspectionIncomplete(
                        "feature metadata XLSX eager XML member exceeds byte limit"
                    )
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES:
                    raise FeatureMetadataInspectionIncomplete(
                        "feature metadata XLSX exceeds total uncompressed byte limit"
                    )
                if member.file_size:
                    ratio = member.file_size / max(member.compress_size, 1)
                    if ratio > MAX_XLSX_COMPRESSION_RATIO:
                        raise FeatureMetadataInspectionIncomplete(
                            "feature metadata XLSX member exceeds compression ratio limit"
                        )
            for member in members:
                if not member.filename.endswith((".xml", ".rels")):
                    continue
                with archive.open(member, "r") as source:
                    _reject_unsafe_xml_declarations(source)
    except ValueError:
        raise
    except (BadZipFile, OSError) as exc:
        raise ValueError("invalid feature metadata XLSX archive") from exc


def _reject_unsafe_xml_declarations(source: Any) -> None:
    payload = b""
    while len(payload) < _MAX_XML_PROLOG_BYTES:
        chunk = source.read(
            min(_STREAM_CHUNK_BYTES, _MAX_XML_PROLOG_BYTES - len(payload))
        )
        if not chunk:
            break
        payload += chunk
        probe = payload.upper()
        if b"<!DOCTYPE" in probe or b"<!ENTITY" in probe:
            raise ValueError(
                "unsafe feature metadata XML contains DOCTYPE or ENTITY"
            )
        if _xml_root_started(payload):
            return
    raise FeatureMetadataInspectionIncomplete(
        "feature metadata XML prolog exceeds safety limit"
    )


def _xml_root_started(payload: bytes) -> bool:
    value = payload
    if value.startswith(b"\xef\xbb\xbf"):
        value = value[3:]
    while True:
        value = value.lstrip(b" \t\r\n")
        if value.startswith(b"<?"):
            end = value.find(b"?>", 2)
            if end < 0:
                return False
            value = value[end + 2 :]
            continue
        if value.startswith(b"<!--"):
            end = value.find(b"-->", 4)
            if end < 0:
                return False
            value = value[end + 3 :]
            continue
        if value.startswith(b"<!"):
            raise ValueError("unsafe feature metadata XML declaration")
        return (
            len(value) >= 2
            and value[0] == ord("<")
            and (chr(value[1]).isalpha() or value[1] == ord("_"))
        )


def _read_single_table(
    path: Path, *, max_bytes: int, max_rows: int, max_columns: int
) -> _Table:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_table(
            path,
            max_bytes=max_bytes,
            max_rows=max_rows,
            max_columns=max_columns,
        )
    if suffix == ".parquet":
        return _read_parquet_table(path, max_rows=max_rows, max_columns=max_columns)
    if suffix == ".feather":
        raise ValueError(
            "Feather feature metadata is not safely bounded; convert it to CSV or Parquet"
        )
    raise ValueError(f"unsupported feature metadata format: {suffix}")


def _read_selected_table(
    path: Path,
    *,
    selection: FeatureMetadataSelection,
    max_bytes: int,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> _Table:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        if selection.sheet_name is None:
            raise ValueError("Excel feature metadata selection requires a sheet name")
        return _read_selected_xlsx_sheet(
            path,
            selection=selection,
            max_rows=max_rows,
            max_columns=max_columns,
            ledger=ledger,
        )
    if suffix == ".xls":
        if selection.sheet_name is None:
            raise ValueError("Excel feature metadata selection requires a sheet name")
        return _read_selected_xls_sheet(
            path,
            selection=selection,
            max_rows=max_rows,
            max_columns=max_columns,
            ledger=ledger,
        )
    if selection.sheet_name is not None:
        raise ValueError("non-Excel feature metadata cannot select a sheet")
    return _read_single_table(
        path,
        max_bytes=max_bytes,
        max_rows=max_rows,
        max_columns=max_columns,
    )


def _read_csv_table(
    path: Path, *, max_bytes: int, max_rows: int, max_columns: int
) -> _Table:
    failures: list[Exception] = []
    for encoding in CSV_ENCODINGS:
        try:
            return _read_csv_stream(
                path,
                encoding=encoding,
                max_bytes=max_bytes,
                max_rows=max_rows,
                max_columns=max_columns,
            )
        except _MetadataLimitError:
            raise
        except (UnicodeDecodeError, csv.Error) as exc:
            failures.append(exc)
            continue
    del failures
    raise ValueError("feature metadata CSV must use UTF-8 or GB18030 encoding")


def _read_csv_stream(
    path: Path,
    *,
    encoding: str,
    max_bytes: int,
    max_rows: int,
    max_columns: int,
) -> _Table:
    try:
        with path.open("rb") as binary:
            bounded = _BoundedBinaryReader(binary, limit=max_bytes)
            with io.BufferedReader(bounded) as buffered:
                with io.TextIOWrapper(
                    buffered,
                    encoding=encoding,
                    errors="strict",
                    newline="",
                ) as text:
                    counted = _CountingTextIterator(
                        text, limit=MAX_METADATA_DECODED_CHARS
                    )
                    reader = csv.reader(counted, strict=True)
                    try:
                        raw_header = next(reader)
                    except StopIteration as exc:
                        raise ValueError(
                            "feature metadata CSV has no header row"
                        ) from exc
                    columns = _validate_csv_headers(
                        raw_header, max_columns=max_columns
                    )
                    selected_indices = tuple(
                        index
                        for index, value in enumerate(columns)
                        if value in _ALL_ALIAS_SET
                    )
                    projected_columns = tuple(
                        columns[index] for index in selected_indices
                    )
                    if not _has_required_metadata_aliases(projected_columns):
                        return _Table(None, projected_columns, ())
                    rows: list[tuple[Any, ...]] = []
                    for index, row in enumerate(reader, start=1):
                        if index > max_rows:
                            raise _MetadataLimitError(
                                "feature metadata exceeds row limit"
                            )
                        if len(row) != len(columns):
                            raise ValueError(
                                "feature metadata CSV contains ragged rows"
                            )
                        _validate_cell_values(row)
                        projected = tuple(row[item] for item in selected_indices)
                        if projected and not _row_is_blank(projected):
                            rows.append(projected)
                    return _Table(None, projected_columns, tuple(rows))
    except _MetadataLimitError:
        raise
    except OSError as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot read feature metadata CSV"
        ) from exc


def _validate_csv_headers(
    raw_headers: Sequence[Any], *, max_columns: int
) -> tuple[str, ...]:
    if len(raw_headers) > max_columns:
        raise FeatureMetadataInspectionIncomplete(
            "feature metadata exceeds column limit"
        )
    headers = tuple(_header_text(value) for value in raw_headers)
    if not headers:
        raise ValueError("feature metadata contains no columns")
    if any(not value for value in headers):
        raise ValueError("feature metadata contains a blank column name")
    duplicates = _duplicates(headers)
    if duplicates:
        raise ValueError(
            "feature metadata contains duplicate column names: "
            + _bounded_join(duplicates, char_limit=MAX_DIAGNOSTIC_CHARS - 50)
        )
    return headers


def _iter_excel_candidate_outcomes(
    path: Path,
    *,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> Iterator[_ExcelSheetOutcome]:
    if path.suffix.lower() == ".xlsx":
        yield from _iter_xlsx_candidate_outcomes(
            path, max_rows=max_rows, max_columns=max_columns, ledger=ledger
        )
        return
    yield from _iter_xls_candidate_outcomes(
        path, max_rows=max_rows, max_columns=max_columns, ledger=ledger
    )


def _iter_xlsx_candidate_outcomes(
    path: Path,
    *,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> Iterator[_ExcelSheetOutcome]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot read feature metadata workbook"
        ) from exc
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.reset_dimensions()
            try:
                raw_header = _xlsx_header(worksheet, max_columns=max_columns)
            except FeatureMetadataInspectionIncomplete:
                raise
            except ValueError as exc:
                message = f"sheet {sheet_name}: {exc}"
                yield _ExcelSheetOutcome(None, (), message)
                continue
            outcome = _excel_header_outcome(raw_header, sheet_name=sheet_name)
            if outcome is not None:
                yield outcome
                continue
            try:
                table = _xlsx_candidate_table(
                    worksheet,
                    raw_header=raw_header,
                    sheet_name=sheet_name,
                    max_rows=max_rows,
                    max_columns=max_columns,
                    ledger=ledger,
                )
            except FeatureMetadataInspectionIncomplete:
                raise
            except ValueError as exc:
                message = f"sheet {sheet_name}: {exc}"
                yield _ExcelSheetOutcome(None, (), message)
                continue
            yield _ExcelSheetOutcome(table, ())
    finally:
        workbook.close()


def _xlsx_header(worksheet: Any, *, max_columns: int) -> tuple[Any, ...]:
    try:
        with worksheet._get_source() as source:  # noqa: SLF001
            parser = WorkSheetParser(
                source,
                worksheet._shared_strings,  # noqa: SLF001
                data_only=worksheet.parent.data_only,
                epoch=worksheet.parent.epoch,
                date_formats=worksheet.parent._date_formats,  # noqa: SLF001
                timedelta_formats=worksheet.parent._timedelta_formats,  # noqa: SLF001
            )
            parsed = next(parser.parse(), None)
    except Exception as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot read feature metadata sheet header"
        ) from exc
    if parsed is None:
        return ()
    row_index, cells = parsed
    if row_index != 1:
        return ()
    raw = [None] * min(
        max((cell["column"] for cell in cells), default=0), max_columns
    )
    for cell in cells:
        column = int(cell["column"])
        value = cell["value"]
        if column > max_columns:
            if _header_text(value) in _ALL_ALIAS_SET:
                raise FeatureMetadataInspectionIncomplete(
                    "feature metadata exceeds column limit"
                )
            continue
        raw[column - 1] = value
    return _trim_trailing_blank_cells(tuple(raw))


def _excel_header_outcome(
    raw_header: tuple[Any, ...], *, sheet_name: str
) -> _ExcelSheetOutcome | None:
    headers = tuple(_header_text(value) for value in raw_header)
    alias_headers = tuple(value for value in headers if value in _ALL_ALIAS_SET)
    duplicate_aliases = _duplicates(alias_headers)
    if duplicate_aliases:
        message = (
            f"sheet {sheet_name}: duplicate alias column names: "
            + _bounded_join(
                duplicate_aliases, char_limit=MAX_DIAGNOSTIC_CHARS - 80
            )
        )
        return _ExcelSheetOutcome(None, (message,))
    feature_matches, category_matches, importance_matches = _alias_matches(headers)
    missing = []
    if not feature_matches:
        missing.append("feature")
    if not category_matches:
        missing.append("category")
    if not importance_matches:
        missing.append("importance")
    if missing:
        return _ExcelSheetOutcome(
            None,
            (f"sheet {sheet_name}: missing {'/'.join(missing)} column alias",),
        )
    return None


def _xlsx_candidate_table(
    worksheet: Any,
    *,
    raw_header: tuple[Any, ...],
    sheet_name: str,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> _Table:
    headers, selected_indices = _project_alias_headers(
        raw_header, max_columns=max_columns
    )
    max_selected_col = max(selected_indices) + 1
    raw_rows: list[list[Any]] = []
    iterator = worksheet.iter_rows(
        min_row=2,
        max_col=max_selected_col,
        values_only=True,
    )
    for index, row in enumerate(iterator):
        if index >= max_rows:
            raise FeatureMetadataInspectionIncomplete(
                "feature metadata exceeds row limit"
            )
        raw_rows.append(list(row))
    selected_columns = frozenset(index + 1 for index in selected_indices)
    merged_ranges = _xlsx_merged_ranges(
        worksheet,
        selected_columns=selected_columns,
        max_rows=max_rows,
        max_columns=max_columns,
    )
    _expand_declared_merges(
        raw_header=list(raw_header[:max_selected_col]),
        raw_rows=raw_rows,
        merged_ranges=merged_ranges,
        selected_columns=selected_columns,
        ledger=ledger,
    )
    projected = _project_nonblank_rows(raw_rows, selected_indices)
    return _Table(sheet_name, headers, projected)


def _xlsx_merged_ranges(
    worksheet: Any,
    *,
    selected_columns: frozenset[int],
    max_rows: int,
    max_columns: int,
) -> tuple[_MergedRange, ...]:
    ranges: list[_MergedRange] = []
    try:
        with worksheet._get_source() as source:  # noqa: SLF001
            for _event, element in iterparse(source, events=("end",)):
                if element.tag.rsplit("}", 1)[-1] != "mergeCell":
                    element.clear()
                    continue
                reference = element.attrib.get("ref", "")
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(reference)
                except (TypeError, ValueError) as exc:
                    raise ValueError("invalid merged-cell range") from exc
                if not any(min_col <= column <= max_col for column in selected_columns):
                    element.clear()
                    continue
                if max_col > max_columns:
                    raise FeatureMetadataInspectionIncomplete(
                        "feature metadata exceeds column limit"
                    )
                if max_row - 1 > max_rows:
                    raise FeatureMetadataInspectionIncomplete(
                        "feature metadata exceeds row limit"
                    )
                ranges.append(_MergedRange(min_col, min_row, max_col, max_row))
                if len(ranges) > MAX_MERGED_RANGES:
                    raise FeatureMetadataInspectionIncomplete(
                        "feature metadata merged-cell range limit exceeded"
                    )
                element.clear()
    except ValueError:
        raise
    except Exception as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot inspect feature metadata merged cells"
        ) from exc
    return tuple(ranges)


def _read_selected_xlsx_sheet(
    path: Path,
    *,
    selection: FeatureMetadataSelection,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> _Table:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot read feature metadata workbook"
        ) from exc
    try:
        if selection.sheet_name not in workbook.sheetnames:
            raise ValueError("selected feature metadata sheet does not exist")
        worksheet = workbook[selection.sheet_name]
        worksheet.reset_dimensions()
        raw_header = _xlsx_header(worksheet, max_columns=max_columns)
        _require_selected_headers(raw_header, selection)
        return _xlsx_candidate_table(
            worksheet,
            raw_header=raw_header,
            sheet_name=selection.sheet_name,
            max_rows=max_rows,
            max_columns=max_columns,
            ledger=ledger,
        )
    finally:
        workbook.close()


def _iter_xls_candidate_outcomes(
    path: Path,
    *,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> Iterator[_ExcelSheetOutcome]:
    try:
        workbook = xlrd.open_workbook(path, on_demand=True, formatting_info=True)
    except Exception as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot read feature metadata workbook"
        ) from exc
    try:
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            raw_header = (
                _trim_trailing_blank_cells(tuple(sheet.row_values(0)))
                if sheet.nrows
                else ()
            )
            outcome = _excel_header_outcome(raw_header, sheet_name=sheet_name)
            if outcome is not None:
                yield outcome
                continue
            try:
                table = _xls_candidate_table(
                    sheet,
                    raw_header=raw_header,
                    sheet_name=sheet_name,
                    max_rows=max_rows,
                    max_columns=max_columns,
                    ledger=ledger,
                )
            except FeatureMetadataInspectionIncomplete:
                raise
            except ValueError as exc:
                message = f"sheet {sheet_name}: {exc}"
                yield _ExcelSheetOutcome(None, (), message)
                continue
            yield _ExcelSheetOutcome(table, ())
    finally:
        workbook.release_resources()


def _xls_candidate_table(
    sheet: Any,
    *,
    raw_header: tuple[Any, ...],
    sheet_name: str,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> _Table:
    headers, selected_indices = _project_alias_headers(
        raw_header, max_columns=max_columns
    )
    data_row_count = max(int(sheet.nrows) - 1, 0)
    if data_row_count > max_rows:
        raise FeatureMetadataInspectionIncomplete(
            "feature metadata exceeds row limit"
        )
    max_selected_col = max(selected_indices) + 1
    raw_rows = [
        list(sheet.row_values(row_index, 0, max_selected_col))
        for row_index in range(1, int(sheet.nrows))
    ]
    selected_columns = frozenset(index + 1 for index in selected_indices)
    merged_ranges = _xls_merged_ranges(
        sheet,
        selected_columns=selected_columns,
        max_rows=max_rows,
        max_columns=max_columns,
    )
    _expand_declared_merges(
        raw_header=list(raw_header[:max_selected_col]),
        raw_rows=raw_rows,
        merged_ranges=merged_ranges,
        selected_columns=selected_columns,
        ledger=ledger,
    )
    projected = _project_nonblank_rows(raw_rows, selected_indices)
    return _Table(sheet_name, headers, projected)


def _xls_merged_ranges(
    sheet: Any,
    *,
    selected_columns: frozenset[int],
    max_rows: int,
    max_columns: int,
) -> tuple[_MergedRange, ...]:
    ranges: list[_MergedRange] = []
    for row_low, row_high, col_low, col_high in sheet.merged_cells:
        merged = _MergedRange(col_low + 1, row_low + 1, col_high, row_high)
        if not any(
            merged.min_col <= column <= merged.max_col
            for column in selected_columns
        ):
            continue
        if merged.max_col > max_columns:
            raise FeatureMetadataInspectionIncomplete(
                "feature metadata exceeds column limit"
            )
        if merged.max_row - 1 > max_rows:
            raise FeatureMetadataInspectionIncomplete(
                "feature metadata exceeds row limit"
            )
        ranges.append(merged)
        if len(ranges) > MAX_MERGED_RANGES:
            raise FeatureMetadataInspectionIncomplete(
                "feature metadata merged-cell range limit exceeded"
            )
    return tuple(ranges)


def _read_selected_xls_sheet(
    path: Path,
    *,
    selection: FeatureMetadataSelection,
    max_rows: int,
    max_columns: int,
    ledger: _InspectionLedger,
) -> _Table:
    try:
        workbook = xlrd.open_workbook(path, on_demand=True, formatting_info=True)
    except Exception as exc:
        raise FeatureMetadataInspectionIncomplete(
            "cannot read feature metadata workbook"
        ) from exc
    try:
        if selection.sheet_name not in workbook.sheet_names():
            raise ValueError("selected feature metadata sheet does not exist")
        sheet = workbook.sheet_by_name(selection.sheet_name)
        raw_header = (
            _trim_trailing_blank_cells(tuple(sheet.row_values(0)))
            if sheet.nrows
            else ()
        )
        _require_selected_headers(raw_header, selection)
        return _xls_candidate_table(
            sheet,
            raw_header=raw_header,
            sheet_name=selection.sheet_name,
            max_rows=max_rows,
            max_columns=max_columns,
            ledger=ledger,
        )
    finally:
        workbook.release_resources()


def _project_alias_headers(
    raw_header: tuple[Any, ...], *, max_columns: int
) -> tuple[tuple[str, ...], tuple[int, ...]]:
    selected_indices = tuple(
        index
        for index, value in enumerate(raw_header)
        if _header_text(value) in _ALL_ALIAS_SET
    )
    if not selected_indices:
        raise ValueError("feature metadata has no alias columns")
    if max(selected_indices) >= max_columns:
        raise FeatureMetadataInspectionIncomplete(
            "feature metadata exceeds column limit"
        )
    headers = tuple(_header_text(raw_header[index]) for index in selected_indices)
    duplicates = _duplicates(headers)
    if duplicates:
        raise ValueError(
            "feature metadata contains duplicate alias column names: "
            + _bounded_join(duplicates, char_limit=MAX_DIAGNOSTIC_CHARS - 70)
        )
    return headers, selected_indices


def _project_nonblank_rows(
    raw_rows: list[list[Any]], selected_indices: tuple[int, ...]
) -> tuple[tuple[Any, ...], ...]:
    rows: list[tuple[Any, ...]] = []
    decoded_chars = 0
    for raw_row in raw_rows:
        projected = tuple(
            raw_row[index] if index < len(raw_row) else None
            for index in selected_indices
        )
        _validate_cell_values(projected)
        decoded_chars += _decoded_character_count(projected)
        if decoded_chars > MAX_METADATA_DECODED_CHARS:
            raise FeatureMetadataInspectionIncomplete(
                "feature metadata exceeds decoded character limit"
            )
        if _row_is_blank(projected):
            continue
        rows.append(projected)
    return tuple(rows)


def _expand_declared_merges(
    *,
    raw_header: list[Any],
    raw_rows: list[list[Any]],
    merged_ranges: tuple[_MergedRange, ...],
    selected_columns: frozenset[int],
    ledger: _InspectionLedger,
) -> None:
    rows = [raw_header, *raw_rows]
    for merged in merged_ranges:
        anchor_row = merged.min_row - 1
        anchor_col = merged.min_col - 1
        if anchor_row >= len(rows) or anchor_col >= len(rows[anchor_row]):
            continue
        anchor = rows[anchor_row][anchor_col]
        columns = tuple(
            column
            for column in selected_columns
            if merged.min_col <= column <= merged.max_col
        )
        if not columns:
            continue
        last_row = min(merged.max_row, len(rows))
        if last_row < merged.min_row:
            continue
        ledger.charge_merged_cells(
            (last_row - merged.min_row + 1) * len(columns)
        )
        for excel_row in range(merged.min_row, last_row + 1):
            row_index = excel_row - 1
            row = rows[row_index]
            for excel_col in columns:
                col_index = excel_col - 1
                if col_index >= len(row):
                    row.extend([None] * (col_index + 1 - len(row)))
                if _is_blank(row[col_index]):
                    row[col_index] = anchor


def _read_parquet_table(path: Path, *, max_rows: int, max_columns: int) -> _Table:
    try:
        parquet = pq.ParquetFile(path)
        raw_headers = tuple(parquet.schema_arrow.names)
        headers = _validate_arrow_headers(raw_headers, max_columns=max_columns)
        projected = tuple(value for value in headers if value in _ALL_ALIAS_SET)
        if not _has_required_metadata_aliases(projected):
            return _Table(None, projected, ())
        row_count = int(parquet.metadata.num_rows)
        if row_count > max_rows:
            raise FeatureMetadataInspectionIncomplete(
                "feature metadata exceeds row limit"
            )
        _validate_parquet_role_types(parquet.schema_arrow, projected)
        _validate_parquet_projection_size(parquet, projected)
        rows: list[tuple[Any, ...]] = []
        decoded_chars = 0
        for batch in parquet.iter_batches(
            batch_size=PARQUET_BATCH_ROWS, columns=list(projected)
        ):
            projected_rows = _arrow_batch_rows(batch, projected)
            decoded_chars += sum(
                _decoded_character_count(row) for row in projected_rows
            )
            if decoded_chars > MAX_METADATA_DECODED_CHARS:
                raise FeatureMetadataInspectionIncomplete(
                    "feature metadata exceeds decoded character limit"
                )
            rows.extend(projected_rows)
        return _Table(None, projected, tuple(rows))
    except ValueError:
        raise
    except (OSError, pa.ArrowException) as exc:
        raise ValueError("invalid feature metadata Parquet file") from exc


def _has_required_metadata_aliases(columns: Sequence[str]) -> bool:
    column_set = frozenset(columns)
    return bool(
        column_set & _FEATURE_ALIAS_SET
        and column_set & _CATEGORY_ALIAS_SET
        and column_set & _IMPORTANCE_ALIAS_SET
    )


def _validate_parquet_role_types(
    schema: pa.Schema, projected: tuple[str, ...]
) -> None:
    for name in projected:
        field_type = schema.field(name).type
        if name in _FEATURE_ALIAS_SET:
            role = "feature"
            allowed = _is_arrow_text_scalar(field_type)
        elif name in _CATEGORY_ALIAS_SET:
            role = "category"
            allowed = _is_arrow_text_scalar(field_type)
        else:
            role = "importance"
            allowed = _is_arrow_importance_scalar(field_type)
        if not allowed:
            raise ValueError(
                f"feature metadata Parquet {role} column type is unsupported"
            )


def _is_arrow_text_scalar(value: pa.DataType) -> bool:
    if pa.types.is_string(value) or pa.types.is_large_string(value):
        return True
    return pa.types.is_dictionary(value) and (
        pa.types.is_string(value.value_type)
        or pa.types.is_large_string(value.value_type)
    )


def _is_arrow_importance_scalar(value: pa.DataType) -> bool:
    return (
        pa.types.is_integer(value)
        or pa.types.is_floating(value)
        or pa.types.is_decimal(value)
        or _is_arrow_text_scalar(value)
    )


def _validate_parquet_projection_size(
    parquet: pq.ParquetFile, projected: tuple[str, ...]
) -> None:
    indices = [parquet.schema_arrow.get_field_index(name) for name in projected]
    total = 0
    for row_group_index in range(parquet.metadata.num_row_groups):
        row_group = parquet.metadata.row_group(row_group_index)
        for column_index in indices:
            size = int(row_group.column(column_index).total_uncompressed_size)
            if size > MAX_PARQUET_MEMBER_DECODED_BYTES:
                raise FeatureMetadataInspectionIncomplete(
                    "feature metadata Parquet column chunk exceeds decoded byte limit"
                )
            total += size
            if total > MAX_PARQUET_TOTAL_DECODED_BYTES:
                raise FeatureMetadataInspectionIncomplete(
                    "feature metadata Parquet projection exceeds decoded byte limit"
                )


def _validate_arrow_headers(
    raw_headers: Sequence[Any], *, max_columns: int
) -> tuple[str, ...]:
    if len(raw_headers) > max_columns:
        raise FeatureMetadataInspectionIncomplete(
            "feature metadata exceeds column limit"
        )
    headers = tuple(_header_text(value) for value in raw_headers)
    if any(not value for value in headers):
        raise ValueError("feature metadata contains a blank column name")
    duplicates = _duplicates(headers)
    if duplicates:
        raise ValueError(
            "feature metadata contains duplicate column names: "
            + _bounded_join(duplicates, char_limit=MAX_DIAGNOSTIC_CHARS - 50)
        )
    return headers


def _arrow_batch_rows(
    batch: pa.RecordBatch, columns: tuple[str, ...]
) -> list[tuple[Any, ...]]:
    values = [batch.column(index).to_pylist() for index in range(len(columns))]
    rows: list[tuple[Any, ...]] = []
    for row_index in range(batch.num_rows):
        projected = tuple(column[row_index] for column in values)
        _validate_cell_values(projected)
        if not _row_is_blank(projected):
            rows.append(projected)
    return rows


def _column_selections(
    table: _Table,
    manifest_context: _ManifestContext,
    *,
    ledger: _InspectionLedger,
    resolution_cache: dict[
        tuple[tuple[str, ...], tuple[str, ...], tuple[float, ...]], str | None
    ],
    max_rejections: int,
    diagnostic_char_limit: int,
) -> tuple[list[FeatureMetadataSelection], list[str], bool]:
    # Coverage is resolved after the three physical columns have been paired.
    # A partial importance table may still be usable when omitted PMML features
    # can inherit one unambiguous category from their feature namespace.
    feature_matches, category_matches, importance_matches = _alias_matches(table.columns)
    missing = []
    if not feature_matches:
        missing.append("feature column alias")
    if not category_matches:
        missing.append("category column alias")
    if not importance_matches:
        missing.append("importance column alias")
    label = _sheet_label(table.sheet_name)
    if missing:
        return [], [f"{label}: missing " + ", ".join(missing)], True
    combination_count = (
        len(feature_matches) * len(category_matches) * len(importance_matches)
    )
    if combination_count > MAX_ALIAS_COMBINATIONS:
        return [], [f"{label}: alias combination limit exceeded"], False
    if combination_count * len(table.rows) > MAX_ALIAS_ROW_EVALUATIONS:
        return [], [f"{label}: alias row evaluation limit exceeded"], False
    ledger.charge_work(
        combination_count
        * (
            len(table.rows)
            + len(manifest_context.model_features)
            + len(manifest_context.stress_units)
        )
    )

    feature_cache = {
        column: _normalize_text_column(table, column=column, role="feature")
        for column in feature_matches
    }
    category_cache = {
        column: _normalize_text_column(table, column=column, role="category")
        for column in category_matches
    }
    importance_cache = {
        column: _normalize_importance_column(table, column=column)
        for column in importance_matches
    }
    valid: list[FeatureMetadataSelection] = []
    rejected: list[str] = []
    for feature_col, category_col, importance_col in product(
        feature_matches, category_matches, importance_matches
    ):
        if len({feature_col, category_col, importance_col}) != 3:
            continue
        selection = FeatureMetadataSelection(
            table.sheet_name,
            feature_col,
            category_col,
            importance_col,
        )
        feature_values, feature_error = feature_cache[feature_col]
        category_values, category_error = category_cache[category_col]
        importance_values, importance_error = importance_cache[importance_col]
        cache_error = feature_error or category_error or importance_error
        if cache_error:
            if len(rejected) < max_rejections:
                _append_rejection(
                    rejected,
                    f"{label} [{feature_col}/{category_col}/{importance_col}]: "
                    f"{cache_error}",
                    limit=max_rejections,
                    char_limit=diagnostic_char_limit,
                )
            continue
        resolution_key = (feature_values, category_values, importance_values)
        if resolution_key not in resolution_cache:
            try:
                rows = _rows_from_normalized_columns(
                    table,
                    feature_values=feature_values,
                    category_values=category_values,
                    importance_values=importance_values,
                )
                merged = _merge_identical_and_reject_conflicts(rows)
                _resolve_against_manifest(merged, manifest_context)
            except FeatureMetadataInspectionIncomplete:
                raise
            except ValueError as exc:
                resolution_cache[resolution_key] = _bounded_text(
                    str(exc), MAX_DIAGNOSTIC_CHARS
                )
            else:
                resolution_cache[resolution_key] = None
        resolution_error = resolution_cache[resolution_key]
        if resolution_error:
            if len(rejected) < max_rejections:
                _append_rejection(
                    rejected,
                    f"{label} [{feature_col}/{category_col}/{importance_col}]: "
                    f"{resolution_error}",
                    limit=max_rejections,
                    char_limit=diagnostic_char_limit,
                )
            continue
        if len(valid) >= MAX_SELECTIONS:
            return [], [f"{label}: alias selection limit exceeded"], False
        valid.append(selection)
    return valid, rejected, True


def _alias_matches(
    columns: Sequence[str],
) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    feature = tuple(value for value in columns if value in _FEATURE_ALIAS_SET)
    category = tuple(value for value in columns if value in _CATEGORY_ALIAS_SET)
    importance = tuple(value for value in columns if value in _IMPORTANCE_ALIAS_SET)
    return feature, category, importance


def _normalize_text_column(
    table: _Table, *, column: str, role: str
) -> tuple[tuple[str, ...], str | None]:
    index = table.columns.index(column)
    values: list[str] = []
    for row_index, row in enumerate(table.rows, start=2):
        try:
            value = _exact_text(row[index])
        except FeatureMetadataInspectionIncomplete:
            raise
        except ValueError as exc:
            return (), str(exc)
        if not value:
            return (), f"{role} is blank at row {row_index}"
        values.append(value)
    if not values:
        return (), "feature metadata contains no rows"
    return tuple(values), None


def _normalize_importance_column(
    table: _Table, *, column: str
) -> tuple[tuple[float, ...], str | None]:
    index = table.columns.index(column)
    values = [row[index] for row in table.rows]
    if not values:
        return (), "feature metadata contains no rows"
    for row_index, value in enumerate(values, start=2):
        try:
            _validate_cell_value(value)
        except FeatureMetadataInspectionIncomplete:
            raise
        except ValueError as exc:
            return (), str(exc)
        if _is_blank(value) or isinstance(value, bool):
            return (), f"importance is blank at row {row_index}"
    try:
        converted = pd.to_numeric(pd.Series(values, dtype="object"), errors="raise")
        normalized = tuple(float(value) for value in converted.tolist())
    except (TypeError, ValueError, OverflowError):
        return (), "importance contains a non-numeric value"
    if any(not math.isfinite(value) for value in normalized):
        return (), "importance contains a non-finite value"
    return normalized, None


def _rows_from_normalized_columns(
    table: _Table,
    *,
    feature_values: tuple[str, ...],
    category_values: tuple[str, ...],
    importance_values: tuple[float, ...],
) -> list[FeatureMetadataRow]:
    if not (
        len(feature_values) == len(category_values) == len(importance_values)
    ):
        raise ValueError("normalized metadata column lengths do not match")
    return [
        FeatureMetadataRow(
            feature=feature,
            category=category,
            importance=importance,
            source_sheet=table.sheet_name,
            in_pmml=False,
        )
        for feature, category, importance in zip(
            feature_values, category_values, importance_values, strict=True
        )
    ]


def _append_rejection(
    destination: list[str], value: str, *, limit: int, char_limit: int
) -> None:
    if len(destination) >= limit:
        return
    destination.append(_bounded_text(value, char_limit))


def _normalize_rows(
    table: _Table, selection: FeatureMetadataSelection
) -> list[FeatureMetadataRow]:
    _validate_selection(selection)
    index_by_name = {name: index for index, name in enumerate(table.columns)}
    missing = [
        name
        for name in (
            selection.feature_col,
            selection.category_col,
            selection.importance_col,
        )
        if name not in index_by_name
    ]
    if missing:
        raise ValueError(
            "selected metadata columns are missing: "
            + _bounded_join(missing, char_limit=MAX_DIAGNOSTIC_CHARS - 50)
        )
    feature_values, feature_error = _normalize_text_column(
        table, column=selection.feature_col, role="feature"
    )
    category_values, category_error = _normalize_text_column(
        table, column=selection.category_col, role="category"
    )
    importance_values, importance_error = _normalize_importance_column(
        table, column=selection.importance_col
    )
    error = feature_error or category_error or importance_error
    if error:
        raise ValueError(error)
    return _rows_from_normalized_columns(
        table,
        feature_values=feature_values,
        category_values=category_values,
        importance_values=importance_values,
    )


def _merge_identical_and_reject_conflicts(
    rows: list[FeatureMetadataRow],
) -> list[FeatureMetadataRow]:
    merged: list[FeatureMetadataRow] = []
    by_feature: dict[str, FeatureMetadataRow] = {}
    for row in rows:
        existing = by_feature.get(row.feature)
        if existing is None:
            by_feature[row.feature] = row
            merged.append(row)
            continue
        if (
            existing.category != row.category
            or existing.importance != row.importance
        ):
            raise ValueError(
                "conflicting feature metadata for "
                + _bounded_text(row.feature, 200)
            )
    return merged


def _build_manifest_context(
    manifest: PmmlInputManifest, *, ledger: _InspectionLedger
) -> _ManifestContext:
    model_features = _normalized_manifest_features(manifest.model_features)
    unsupported_by_feature: dict[str, str] = {}
    for diagnostic in manifest.unsupported_derivations:
        feature, separator, _detail = diagnostic.partition(":")
        feature_name = _exact_text(feature)
        if separator and feature_name:
            unsupported_by_feature.setdefault(feature_name, diagnostic)

    source_units: dict[str, StressUnit] = {}
    normalized_units: dict[str, tuple[str, ...]] = {}
    raw_field_count = 0
    for unit in manifest.stress_units:
        feature = _exact_text(unit.model_feature)
        if not feature:
            raise ValueError("stress unit has an empty model feature")
        existing = source_units.get(feature)
        if existing is not None and existing != unit:
            raise ValueError(
                f"conflicting stress unit for {_bounded_text(feature, 200)}"
            )
        raw_fields = tuple(_exact_text(value) for value in unit.raw_input_fields)
        source_units[feature] = unit
        normalized_units[feature] = raw_fields
        raw_field_count += len(raw_fields)

    work_size = len(model_features) + len(manifest.stress_units) + raw_field_count
    ledger.charge_work(work_size)
    return _ManifestContext(
        model_features=model_features,
        model_feature_set=frozenset(model_features),
        unsupported_by_feature=unsupported_by_feature,
        stress_units=normalized_units,
        work_size=work_size,
    )


def _resolve_against_manifest(
    rows: list[FeatureMetadataRow], context: _ManifestContext
) -> FeatureMetadataResolution:
    model_features = context.model_features
    by_feature = {row.feature: row for row in rows}
    missing = [feature for feature in model_features if feature not in by_feature]
    unresolved: list[str] = []
    for feature in missing:
        inferred = _inferred_metadata_row(feature, rows)
        if inferred is None:
            unresolved.append(feature)
            continue
        rows.append(inferred)
        by_feature[feature] = inferred
    if unresolved:
        raise ValueError(
            "missing PMML feature metadata: "
            + _bounded_join(unresolved, char_limit=300)
        )
    ordered: list[FeatureMetadataRow] = []
    model_set = context.model_feature_set
    for feature in model_features:
        row = by_feature[feature]
        ordered.append(
            FeatureMetadataRow(
                row.feature,
                row.category,
                row.importance,
                row.source_sheet,
                True,
            )
        )
    extras = tuple(row.feature for row in rows if row.feature not in model_set)
    for row in rows:
        if row.feature in model_set:
            continue
        ordered.append(
            FeatureMetadataRow(
                row.feature,
                row.category,
                row.importance,
                row.source_sheet,
                False,
            )
        )

    category_fields = _resolve_stress_fields(
        rows_by_feature=by_feature,
        context=context,
    )
    return FeatureMetadataResolution(
        schema_version=FEATURE_METADATA_SCHEMA,
        rows=tuple(ordered),
        coverage=MetadataCoverage(1.0, 1.0, 1.0, 1.0),
        per_category_raw_fields=category_fields,
        extra_features=extras,
        conflicts=(),
    )


def _inferred_metadata_row(
    feature: str, rows: Sequence[FeatureMetadataRow]
) -> FeatureMetadataRow | None:
    for namespace in _metadata_namespace_keys(feature):
        siblings = [
            row
            for row in rows
            if namespace in _metadata_namespace_keys(row.feature)
        ]
        categories = {row.category for row in siblings}
        if len(categories) != 1:
            continue
        sheets = {row.source_sheet for row in siblings}
        return FeatureMetadataRow(
            feature=feature,
            category=next(iter(categories)),
            importance=0.0,
            source_sheet=next(iter(sheets)) if len(sheets) == 1 else None,
            in_pmml=True,
        )
    return None


def _metadata_namespace_keys(feature: str) -> tuple[str, ...]:
    namespace, separator, _remainder = feature.partition("_")
    if not separator or not namespace:
        return ()
    keys = [namespace]
    family = "".join(character for character in namespace if character.isalpha())[:2]
    if len(family) == 2 and family != namespace:
        keys.append(family)
    return tuple(keys)


def _resolve_stress_fields(
    *,
    rows_by_feature: dict[str, FeatureMetadataRow],
    context: _ManifestContext,
) -> dict[str, tuple[str, ...]]:
    per_category: dict[str, list[str]] = {}
    raw_categories: dict[str, str] = {}
    for feature in context.model_features:
        if feature in context.unsupported_by_feature:
            raise ValueError(
                "unsupported stress unit for "
                f"{_bounded_text(feature, 200)}: "
                f"{_bounded_text(context.unsupported_by_feature[feature], 250)}"
            )
        raw_fields = context.stress_units.get(feature)
        if raw_fields is None:
            raise ValueError(
                f"missing stress unit for {_bounded_text(feature, 200)}"
            )
        if not raw_fields or any(not value for value in raw_fields):
            raise ValueError(
                f"empty stress unit for {_bounded_text(feature, 200)}"
            )
        category = rows_by_feature[feature].category
        fields = per_category.setdefault(category, [])
        for raw_field in raw_fields:
            previous = raw_categories.get(raw_field)
            if previous is not None and previous != category:
                raise ValueError(
                    "raw stress field "
                    f"{_bounded_text(raw_field, 200)} category conflict: "
                    f"{_bounded_text(previous, 100)} vs "
                    f"{_bounded_text(category, 100)}"
                )
            raw_categories.setdefault(raw_field, category)
            if raw_field not in fields:
                fields.append(raw_field)
    return {category: tuple(fields) for category, fields in per_category.items()}


def _normalized_manifest_features(features: Sequence[str]) -> tuple[str, ...]:
    normalized = tuple(_exact_text(value) for value in features)
    if any(not value for value in normalized):
        raise ValueError("PMML model feature is blank after trimming")
    duplicates = _duplicates(normalized)
    if duplicates:
        raise ValueError(
            "PMML model features duplicate after trimming: "
            + _bounded_join(duplicates, char_limit=300)
        )
    return normalized


def _validate_selection(selection: FeatureMetadataSelection) -> None:
    if not isinstance(selection, FeatureMetadataSelection):
        raise TypeError("selection must be a FeatureMetadataSelection")
    columns = (
        selection.feature_col,
        selection.category_col,
        selection.importance_col,
    )
    if any(not isinstance(value, str) or not value.strip() for value in columns):
        raise ValueError("selected metadata columns must be non-empty strings")
    if any(len(value) > MAX_METADATA_CELL_CHARS for value in columns):
        raise FeatureMetadataInspectionIncomplete(
            "selected metadata column exceeds cell length limit"
        )
    if len(set(columns)) != 3:
        raise ValueError("one physical metadata column cannot serve multiple roles")


def _require_selected_headers(
    raw_header: tuple[Any, ...], selection: FeatureMetadataSelection
) -> None:
    headers = tuple(_header_text(value) for value in raw_header)
    for selected in (
        selection.feature_col,
        selection.category_col,
        selection.importance_col,
    ):
        count = headers.count(selected)
        if count == 0:
            raise ValueError(
                "selected metadata column is missing: "
                + _bounded_text(selected, 200)
            )
        if count > 1:
            raise ValueError(
                "selected metadata column is duplicated: "
                + _bounded_text(selected, 200)
            )


def _header_text(value: Any) -> str:
    return _exact_text(value)


def _exact_text(value: Any) -> str:
    _validate_cell_value(value)
    if _is_blank(value):
        return ""
    return str(value).strip()


def _validate_cell_values(values: Sequence[Any]) -> None:
    for value in values:
        _validate_cell_value(value)


def _validate_cell_value(value: Any) -> None:
    if isinstance(value, str) and len(value) > MAX_METADATA_CELL_CHARS:
        raise FeatureMetadataInspectionIncomplete(
            "feature metadata cell length exceeds limit"
        )


def _decoded_character_count(values: Sequence[Any]) -> int:
    return sum(len(value) for value in values if isinstance(value, str))


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if bool(pd.isna(value)):
            return True
    except (TypeError, ValueError):
        pass
    return isinstance(value, str) and not value.strip()


def _row_is_blank(values: Sequence[Any]) -> bool:
    return all(_is_blank(value) for value in values)


def _trim_trailing_blank_cells(values: tuple[Any, ...]) -> tuple[Any, ...]:
    end = len(values)
    while end and _is_blank(values[end - 1]):
        end -= 1
    return values[:end]


def _duplicates(values: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for value in values:
        if value in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(value)
    return duplicates


def _sheet_label(sheet_name: str | None) -> str:
    return "metadata file" if sheet_name is None else f"sheet {sheet_name}"


def _extend_diagnostics(
    destination: list[str],
    values: Sequence[str],
    *,
    limit: int,
    char_limit: int,
) -> None:
    for value in values:
        if len(destination) >= limit:
            return
        bounded = _bounded_text(str(value), char_limit)
        if bounded not in destination:
            destination.append(bounded)


def _bounded_text(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    if limit <= 3:
        return value[:limit]
    return value[: limit - 3] + "..."


def _bounded_value_error(error: ValueError, *, limit: int) -> ValueError:
    return ValueError(_bounded_text(str(error), limit))


def _bounded_join(values: Sequence[str], *, char_limit: int) -> str:
    result = ""
    for value in values:
        candidate = value if not result else f"{result}, {value}"
        if len(candidate) > char_limit:
            return _bounded_text(candidate, char_limit)
        result = candidate
    return result
