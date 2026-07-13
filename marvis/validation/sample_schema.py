from __future__ import annotations

from collections.abc import Iterator, Sequence
import csv
from dataclasses import dataclass
from hashlib import sha256
import io
from pathlib import Path
import re
from typing import Any

import pandas as pd
import pyarrow as pa
import pyarrow.ipc as ipc
import pyarrow.parquet as pq
from openpyxl import load_workbook
import xlrd

from marvis.settings import DEFAULT_MAX_EXCEL_ROWS, DEFAULT_MAX_EXCEL_UPLOAD_BYTES
from marvis.validation.input_contracts import SampleSchema


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")
PREVIEW_ROWS = 20
MAX_CSV_PREVIEW_BYTES = 2 * 1024 * 1024
HASH_CHUNK_BYTES = 1024 * 1024
_SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")
_UTF8_BOM = b"\xef\xbb\xbf"


@dataclass(frozen=True)
class _ExcelSheetScan:
    data_row_counts: dict[str, int]
    nonempty_sheets: tuple[str, ...]
    overflow_sheets: frozenset[str]


def inspect_sample_schema(
    path: Path,
    *,
    max_excel_upload_bytes: int = DEFAULT_MAX_EXCEL_UPLOAD_BYTES,
    max_excel_rows: int = DEFAULT_MAX_EXCEL_ROWS,
) -> SampleSchema:
    """Inspect a validation sample without materializing the complete dataset."""

    selected_path = Path(path).expanduser().resolve()
    suffix = selected_path.suffix.lower()
    if suffix == ".csv":
        return _inspect_csv(selected_path)
    if suffix == ".parquet":
        return _inspect_parquet(selected_path)
    if suffix == ".feather":
        return _inspect_feather(selected_path)
    if suffix in {".xlsx", ".xls"}:
        return _inspect_excel(
            selected_path,
            max_upload_bytes=max_excel_upload_bytes,
            max_rows=max_excel_rows,
        )
    raise ValueError(f"unsupported validation sample format: {suffix}")


def iter_sample_projection(
    path: Path,
    *,
    columns: tuple[str, ...],
    chunk_size: int,
    schema: SampleSchema | None = None,
    max_excel_upload_bytes: int = DEFAULT_MAX_EXCEL_UPLOAD_BYTES,
    max_excel_rows: int = DEFAULT_MAX_EXCEL_ROWS,
) -> Iterator[pd.DataFrame]:
    """Yield requested columns in stable order using format-native batches."""

    if chunk_size <= 0:
        raise ValueError("sample chunk size must be positive")
    _validate_requested_columns(columns)
    selected_path = Path(path).expanduser().resolve()
    suffix = selected_path.suffix.lower()
    if schema is None:
        selected_schema = inspect_sample_schema(
            selected_path,
            max_excel_upload_bytes=max_excel_upload_bytes,
            max_excel_rows=max_excel_rows,
        )
    else:
        if not isinstance(schema, SampleSchema):
            raise TypeError("schema must be a SampleSchema")
        _validate_supplied_schema(selected_path, schema)
        selected_schema = schema

    missing = [name for name in columns if name not in selected_schema.columns]
    if missing:
        raise ValueError("sample projection missing columns: " + ", ".join(missing))

    ordered = list(columns)
    if suffix == ".csv":
        for frame in pd.read_csv(
            selected_path,
            usecols=ordered,
            encoding=selected_schema.encoding,
            chunksize=chunk_size,
        ):
            yield frame.loc[:, ordered]
        return

    if suffix == ".parquet":
        parquet = pq.ParquetFile(selected_path)
        for batch in parquet.iter_batches(batch_size=chunk_size, columns=ordered):
            yield batch.to_pandas().loc[:, ordered]
        return

    if suffix == ".feather":
        with pa.memory_map(str(selected_path), "r") as source:
            reader = ipc.RecordBatchFileReader(source)
            indices = [reader.schema.get_field_index(name) for name in ordered]
            for batch_index in range(reader.num_record_batches):
                batch = reader.get_batch(batch_index).select(indices)
                for start in range(0, batch.num_rows, chunk_size):
                    yield batch.slice(start, chunk_size).to_pandas().loc[:, ordered]
        return

    if suffix in {".xlsx", ".xls"}:
        _validate_positive_limit(max_excel_upload_bytes, "Excel byte")
        _validate_positive_limit(max_excel_rows, "Excel row")
        _enforce_excel_file_size(selected_path, max_excel_upload_bytes)
        if not selected_schema.sheet_name:
            raise ValueError("Excel sample schema has no selected sheet name")
        if (
            selected_schema.row_count is not None
            and selected_schema.row_count > max_excel_rows
        ):
            raise ValueError("Excel sample exceeds row limit")
        scan = _scan_excel_sheets(selected_path, max_rows=max_excel_rows)
        if selected_schema.sheet_name in scan.overflow_sheets:
            raise ValueError("Excel sample exceeds row limit")
        current_row_count = scan.data_row_counts.get(selected_schema.sheet_name)
        if current_row_count is None:
            raise ValueError("selected Excel sample sheet is no longer present")
        if current_row_count > max_excel_rows:
            raise ValueError("Excel sample exceeds row limit")
        if (
            selected_schema.row_count is not None
            and current_row_count != selected_schema.row_count
        ):
            raise ValueError("Excel sample row count no longer matches inspected schema")
        raw_frame = pd.read_excel(
            selected_path,
            sheet_name=selected_schema.sheet_name,
            header=None,
        )
        if raw_frame.empty:
            raise ValueError("selected Excel sample sheet is empty")
        current_columns = _validate_columns(raw_frame.iloc[0].tolist())
        if current_columns != selected_schema.columns:
            raise ValueError("Excel sample columns no longer match inspected schema")
        frame = raw_frame.iloc[1:].copy()
        if len(frame) > max_excel_rows:
            raise ValueError("Excel sample exceeds row limit after loading")
        frame.columns = list(current_columns)
        frame = frame.loc[:, ordered]
        for start in range(0, len(frame), chunk_size):
            yield frame.iloc[start : start + chunk_size].loc[:, ordered].copy()
        return

    raise ValueError(f"unsupported validation sample format: {suffix}")


def _inspect_csv(path: Path) -> SampleSchema:
    digest = _sha256_file(path)
    preview_bytes, truncated = _read_bounded_preview(
        path, MAX_CSV_PREVIEW_BYTES
    )
    preview_bytes = _complete_csv_record_prefix(
        preview_bytes, truncated=truncated
    )
    failures: list[Exception] = []
    for encoding in CSV_ENCODINGS:
        try:
            text = preview_bytes.decode(encoding, errors="strict")
            rows = _csv_preview_rows(text, PREVIEW_ROWS)
        except (UnicodeDecodeError, csv.Error) as exc:
            failures.append(exc)
            continue
        if not rows:
            raise ValueError("validation sample CSV has no header row")
        columns = _validate_columns(rows[0])
        data_rows = [row for row in rows[1:] if row]
        for row in data_rows:
            if len(row) != len(columns):
                raise ValueError("validation sample CSV preview contains ragged rows")
        preview = pd.DataFrame(data_rows, columns=list(columns)).convert_dtypes()
        return SampleSchema(
            path=str(path),
            columns=columns,
            dtypes=_pandas_dtypes(preview, columns),
            row_count=None,
            preview_row_count=len(data_rows),
            encoding=encoding,
            sha256=digest,
            sheet_name=None,
        )
    del failures
    raise ValueError("validation sample CSV must use UTF-8 or GB18030 encoding")


def _inspect_parquet(path: Path) -> SampleSchema:
    digest = _sha256_file(path)
    try:
        parquet = pq.ParquetFile(path)
        arrow_schema = parquet.schema_arrow
        columns = _validate_columns(arrow_schema.names)
        row_count = parquet.metadata.num_rows
    except (OSError, pa.ArrowException, ValueError) as exc:
        raise ValueError("invalid validation sample Parquet file") from exc
    return SampleSchema(
        path=str(path),
        columns=columns,
        dtypes={
            column: str(field.type)
            for column, field in zip(columns, arrow_schema, strict=True)
        },
        row_count=row_count,
        preview_row_count=0,
        encoding=None,
        sha256=digest,
        sheet_name=None,
    )


def _inspect_feather(path: Path) -> SampleSchema:
    digest = _sha256_file(path)
    try:
        with pa.memory_map(str(path), "r") as source:
            reader = ipc.RecordBatchFileReader(source)
            arrow_schema = reader.schema
            columns = _validate_columns(arrow_schema.names)
            row_count = sum(
                reader.get_batch(index).num_rows
                for index in range(reader.num_record_batches)
            )
    except (OSError, pa.ArrowException, ValueError) as exc:
        raise ValueError("invalid validation sample Feather file") from exc
    return SampleSchema(
        path=str(path),
        columns=columns,
        dtypes={
            column: str(field.type)
            for column, field in zip(columns, arrow_schema, strict=True)
        },
        row_count=row_count,
        preview_row_count=0,
        encoding=None,
        sha256=digest,
        sheet_name=None,
    )


def _inspect_excel(
    path: Path, *, max_upload_bytes: int, max_rows: int
) -> SampleSchema:
    _validate_positive_limit(max_upload_bytes, "Excel byte")
    _validate_positive_limit(max_rows, "Excel row")
    _enforce_excel_file_size(path, max_upload_bytes)
    scan = _scan_excel_sheets(path, max_rows=max_rows)
    if not scan.nonempty_sheets:
        raise ValueError("样本工作簿没有非空 sheet")
    if len(scan.nonempty_sheets) > 1:
        raise ValueError(
            "样本工作簿包含多个非空 sheet，请另存为单 sheet 样本后重新选择"
        )
    sheet_name = scan.nonempty_sheets[0]
    if sheet_name in scan.overflow_sheets:
        raise ValueError("Excel sample exceeds row limit")
    row_count = scan.data_row_counts.get(sheet_name)
    if row_count is None:
        raise ValueError("无法读取样本工作簿 sheet 行数")
    digest = _sha256_file(path)
    try:
        frame = pd.read_excel(
            path,
            sheet_name=sheet_name,
            header=None,
            nrows=PREVIEW_ROWS + 1,
            dtype=object,
        )
    except Exception as exc:
        raise ValueError("无法读取样本工作簿") from exc
    columns = _validate_columns(frame.iloc[0].tolist())
    preview = frame.iloc[1:].copy()
    preview.columns = list(columns)
    preview = preview.dropna(how="all").convert_dtypes()
    return SampleSchema(
        path=str(path),
        columns=columns,
        dtypes=_pandas_dtypes(preview, columns),
        row_count=row_count,
        preview_row_count=len(preview),
        encoding=None,
        sha256=digest,
        sheet_name=sheet_name,
    )


def _validate_positive_limit(value: int, label: str) -> None:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ValueError(f"{label} limit must be a positive integer")


def _enforce_excel_file_size(path: Path, limit: int) -> None:
    try:
        size = path.stat().st_size
    except OSError as exc:
        raise ValueError("无法读取样本工作簿") from exc
    if size > limit:
        raise ValueError("Excel sample exceeds size limit")


def _scan_excel_sheets(path: Path, *, max_rows: int) -> _ExcelSheetScan:
    """Stream actual worksheet rows, never trusting XLSX dimension metadata."""

    try:
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                counts: dict[str, int] = {}
                nonempty: list[str] = []
                overflow: set[str] = set()
                for name in workbook.sheetnames:
                    worksheet = workbook[name]
                    # read_only worksheets otherwise trust the optional
                    # <dimension> element, which is attacker-controlled.
                    worksheet.reset_dimensions()
                    total_rows = 0
                    has_content = False
                    for row in worksheet.iter_rows(values_only=True):
                        total_rows += 1
                        if any(not _is_blank(value) for value in row):
                            has_content = True
                        if total_rows > max_rows + 1:
                            overflow.add(name)
                            # A row beyond the cap is enough to reject this
                            # sheet; do not stream the remainder.
                            has_content = True
                            break
                    counts[name] = max(total_rows - 1, 0)
                    if has_content:
                        nonempty.append(name)
                return _ExcelSheetScan(
                    data_row_counts=counts,
                    nonempty_sheets=tuple(nonempty),
                    overflow_sheets=frozenset(overflow),
                )
            finally:
                workbook.close()
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            counts = {}
            nonempty = []
            overflow = set()
            for name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(name)
                total_rows = int(sheet.nrows)
                counts[name] = max(total_rows - 1, 0)
                if total_rows > max_rows + 1:
                    overflow.add(name)
                # xlrd derives nrows/ncols from actual BIFF cell records rather
                # than a separately declared worksheet dimension.
                if total_rows > 0 and int(sheet.ncols) > 0:
                    nonempty.append(name)
            return _ExcelSheetScan(
                data_row_counts=counts,
                nonempty_sheets=tuple(nonempty),
                overflow_sheets=frozenset(overflow),
            )
        finally:
            workbook.release_resources()
    except Exception as exc:
        raise ValueError("无法读取样本工作簿") from exc


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(HASH_CHUNK_BYTES), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_bounded_preview(path: Path, limit: int) -> tuple[bytes, bool]:
    with path.open("rb") as handle:
        payload = handle.read(limit + 1)
    return payload[:limit], len(payload) > limit


def _complete_csv_record_prefix(payload: bytes, *, truncated: bool) -> bytes:
    """Return only complete raw CSV records within the byte preview.

    UTF-8 and GB18030 preserve ASCII quote/CR/LF bytes, so record boundaries
    can be found before decoding without ever accepting a partial multibyte or
    quoted record. Doubled quotes are treated as escaped quotes.
    """

    in_quotes = False
    at_field_start = True
    record_ends: list[int] = []
    # A UTF-8 BOM is encoding metadata only at absolute file start. Keep its
    # bytes in the returned prefix for utf-8-sig decoding, but do not let it
    # consume the first field-start state before an opening quote.
    index = len(_UTF8_BOM) if payload.startswith(_UTF8_BOM) else 0
    while index < len(payload):
        value = payload[index]
        if in_quotes:
            if value == 0x22 and index + 1 < len(payload) and payload[index + 1] == 0x22:
                index += 2
                continue
            if value == 0x22:
                in_quotes = False
            index += 1
            continue
        if value == 0x22 and at_field_start:  # opening quote
            in_quotes = True
            at_field_start = False
            index += 1
            continue
        if value == 0x2C:  # comma
            at_field_start = True
        elif value == 0x0D:  # CR or CRLF
            if index + 1 < len(payload) and payload[index + 1] == 0x0A:
                record_ends.append(index + 2)
                at_field_start = True
                index += 2
                continue
            if not (truncated and index + 1 == len(payload)):
                record_ends.append(index + 1)
                at_field_start = True
        elif value == 0x0A:  # LF
            record_ends.append(index + 1)
            at_field_start = True
        else:
            # With skipinitialspace=False, any byte (including a space or a
            # quote after data) means the field has already started.
            at_field_start = False
        index += 1

    if not truncated:
        if in_quotes:
            raise ValueError("validation sample CSV has an unterminated quoted field")
        return payload
    if not record_ends:
        raise ValueError("validation sample CSV header exceeds preview byte limit")
    return payload[: record_ends[-1]]


def _csv_preview_rows(text: str, limit: int) -> list[list[str]]:
    reader = csv.reader(io.StringIO(text, newline=""), strict=True)
    rows: list[list[str]] = []
    for row in reader:
        rows.append(row)
        if len(rows) >= limit + 1:
            break
    return rows


def _validate_columns(values: Sequence[Any]) -> tuple[str, ...]:
    raw_names = [None if _is_blank(value) else str(value) for value in values]
    occupied = {name for name in raw_names if name is not None}
    columns: list[str] = []
    for index, name in enumerate(raw_names):
        if name is None:
            candidate = f"__marvis_unnamed_column_{index}__"
            suffix = 1
            while candidate in occupied:
                candidate = f"__marvis_unnamed_column_{index}_{suffix}__"
                suffix += 1
            name = candidate
            occupied.add(name)
        columns.append(name)
    if not columns:
        raise ValueError("validation sample contains no columns")
    seen: set[str] = set()
    duplicates: list[str] = []
    for name in columns:
        if name in seen and name not in duplicates:
            duplicates.append(name)
        seen.add(name)
    if duplicates:
        raise ValueError(
            "validation sample contains duplicate column names: "
            + ", ".join(duplicates)
        )
    return tuple(columns)


def _validate_requested_columns(columns: tuple[str, ...]) -> None:
    if any(not isinstance(name, str) or not name for name in columns):
        raise ValueError("sample projection columns must be non-empty strings")
    if len(set(columns)) != len(columns):
        raise ValueError("sample projection contains duplicate requested columns")


def _validate_supplied_schema(path: Path, schema: SampleSchema) -> None:
    try:
        schema_path = Path(schema.path).expanduser().resolve()
    except (OSError, ValueError) as exc:
        raise ValueError("sample schema path is invalid") from exc
    if schema_path != path:
        raise ValueError("sample schema does not match sample path")
    if not _SHA256_PATTERN.fullmatch(schema.sha256):
        raise ValueError("sample schema has invalid SHA-256")
    _validate_columns(schema.columns)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if bool(pd.isna(value)):
            return True
    except (TypeError, ValueError):
        pass
    return isinstance(value, str) and not value.strip()


def _pandas_dtypes(
    frame: pd.DataFrame, columns: tuple[str, ...]
) -> dict[str, str]:
    if frame.empty:
        return {name: "object" for name in columns}
    return {name: str(frame[name].dtype) for name in columns}
