from __future__ import annotations

from contextlib import contextmanager
from itertools import islice
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator, Literal
from uuid import uuid4

import pandas as pd
import xlrd
from openpyxl import load_workbook

from marvis.data.errors import DataIngestError, DatasetTooLargeError


MAX_HEADER_ROWS = 5
PREVIEW_ROWS = 25

# GAP-1: Excel stores "Number"-typed cells as IEEE754 doubles at the file-format
# level -- if a long id (e.g. an 18-digit national id) was entered into a Number
# cell rather than a Text cell, its trailing digits are already lost before this
# code ever reads the file; there is no way to recover the original digits. The
# best we can do is *detect* an integer-shaped float at or above this magnitude
# and flag the column so the user knows to re-export/re-enter it as text.
LONG_ID_FLOAT_THRESHOLD = 1e15
_OLE_COMPOUND_MAGIC = bytes.fromhex("d0cf11e0a1b11ae1")
_ZIP_CONTAINER_MAGICS = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
ExcelFormat = Literal["xls", "xlsx"]


def detect_excel_container_format_from_prefix(prefix: bytes) -> ExcelFormat | None:
    """Cheaply classify Excel-capable containers from their leading bytes.

    This is deliberately *not* workbook validation.  Callers use it to spot a
    binary container early (including during a streaming upload), then use
    :func:`require_excel_format` before routing the file to an Excel parser.
    Keeping these two decisions separate prevents a damaged OLE/ZIP payload
    from being treated as CSV merely because workbook validation failed.
    """

    if prefix.startswith(_OLE_COMPOUND_MAGIC):
        return "xls"
    if prefix.startswith(_ZIP_CONTAINER_MAGICS):
        return "xlsx"
    return None


def detect_excel_container_format(path: Path) -> ExcelFormat | None:
    """Return cheap OLE/ZIP container evidence without parsing a workbook."""

    try:
        with Path(path).open("rb") as workbook_file:
            return detect_excel_container_format_from_prefix(
                workbook_file.read(len(_OLE_COMPOUND_MAGIC))
            )
    except OSError:
        return None


def detect_excel_format(path: Path) -> ExcelFormat | None:
    """Return the verified workbook family instead of trusting its suffix."""

    container_format = detect_excel_container_format(path)
    if container_format == "xlsx" and is_xlsx_workbook(path):
        return "xlsx"
    if container_format == "xls" and is_xls_workbook(path):
        return "xls"
    return None


def is_excel_workbook(path: Path) -> bool:
    return detect_excel_format(path) is not None


def require_excel_format(path: Path) -> ExcelFormat:
    """Return a validated Excel format or reject the payload fail-closed."""

    workbook_format = detect_excel_format(path)
    if workbook_format is None:
        raise DataIngestError(
            "file is not a readable BIFF .xls or OOXML .xlsx/.xlsm workbook"
        )
    return workbook_format


def is_xlsx_workbook(path: Path) -> bool:
    """Return whether ``path`` is a readable OOXML Excel workbook.

    Some upstream export/download systems preserve Excel bytes while assigning
    a ``.csv`` filename.  Both ZIP membership and an openpyxl parse operate on
    binary handles, so suffixes do not participate in the decision.  A generic
    or damaged ZIP is deliberately rejected.
    """

    path = Path(path)
    if detect_excel_container_format(path) != "xlsx":
        return False
    try:
        with path.open("rb") as workbook_file:
            with zipfile.ZipFile(workbook_file) as archive:
                members = set(archive.namelist())
        if not {"[Content_Types].xml", "xl/workbook.xml"}.issubset(members):
            return False
        with _open_xlsx_workbook(path) as workbook:
            return bool(workbook.sheetnames)
    except Exception:
        return False


def is_xls_workbook(path: Path) -> bool:
    """Return whether ``path`` is a readable BIFF workbook in an OLE container."""

    path = Path(path)
    if detect_excel_container_format(path) != "xls":
        return False
    try:
        with _open_xls_workbook(path) as workbook:
            return bool(workbook.sheet_names())
    except Exception:
        return False


@contextmanager
def _open_xlsx_workbook(path: Path) -> Iterator[object]:
    """Open OOXML by handle and close both workbook and file on every path."""

    with Path(path).open("rb") as workbook_file:
        workbook = load_workbook(workbook_file, read_only=True, data_only=True)
        try:
            yield workbook
        finally:
            workbook.close()


@contextmanager
def _open_xls_workbook(path: Path) -> Iterator[object]:
    """Open BIFF content and always release xlrd's on-demand resources."""

    workbook = xlrd.open_workbook(Path(path), on_demand=True)
    try:
        yield workbook
    finally:
        workbook.release_resources()


@dataclass(frozen=True)
class IngestReport:
    sheet: str
    header_rows: int
    data_start_row: int
    flattened_columns: list[str]
    original_shape: tuple[int, int]
    suspected_truncated_id_columns: tuple[str, ...] = field(default_factory=tuple)


def list_sheets(path: Path) -> list[str]:
    path = Path(path)
    workbook_format = require_excel_format(path)
    try:
        if workbook_format == "xls":
            with _open_xls_workbook(path) as workbook:
                return list(workbook.sheet_names())
        with _open_xlsx_workbook(path) as workbook:
            return list(workbook.sheetnames)
    except Exception as exc:
        raise DataIngestError(f"cannot list Excel sheets: {exc}") from exc


def probe_sheet_row_count(
    path: Path,
    sheet: str,
    *,
    stop_after: int | None = None,
) -> int | None:
    """Stream rows up to ``stop_after`` without trusting OOXML dimensions.

    Exporters sometimes write an understated ``<dimension>`` (for example
    ``A1:A1`` even though later rows exist).  ``reset_dimensions()`` makes the
    read-only worksheet inspect actual row records.  Iteration stops as soon as
    the caller's rejection threshold is reached, so the guard cannot become an
    unbounded pre-read for an oversized workbook.  A returned ``stop_after``
    therefore means "at least this many rows" rather than an exact total.
    """

    if stop_after is not None and stop_after < 1:
        raise DataIngestError("stop_after must be at least 1")
    path = Path(path)
    workbook_format = require_excel_format(path)
    try:
        if workbook_format == "xls":
            with _open_xls_workbook(path) as workbook:
                row_count = int(workbook.sheet_by_name(sheet).nrows)
                return min(row_count, stop_after) if stop_after is not None else row_count
        with _open_xlsx_workbook(path) as workbook:
            worksheet = workbook[sheet]
            worksheet.reset_dimensions()
            row_count = 0
            for row_count, _row in enumerate(worksheet.iter_rows(), start=1):
                if stop_after is not None and row_count >= stop_after:
                    break
            return row_count
    except Exception as exc:
        raise DataIngestError(f"cannot probe Excel sheet {sheet}: {exc}") from exc


def detect_header_rows(raw: pd.DataFrame) -> int:
    if raw.empty:
        raise DataIngestError("cannot detect headers for an empty sheet")
    limit = min(MAX_HEADER_ROWS, len(raw))
    for index in range(1, limit):
        if _looks_like_data_row(raw.iloc[index]):
            return max(index, 1)
    return 1


def flatten_headers(raw: pd.DataFrame, header_rows: int) -> tuple[pd.DataFrame, list[str]]:
    if raw.empty or raw.dropna(how="all").empty:
        raise DataIngestError("cannot flatten headers for an empty sheet")
    if header_rows < 1 or header_rows > len(raw):
        raise DataIngestError("header_rows is outside the sheet bounds")

    header_block = raw.iloc[:header_rows].copy()
    header_block = header_block.ffill(axis=1)
    flattened_columns = []
    for column_index in range(header_block.shape[1]):
        parts = [
            _header_part(header_block.iloc[row_index, column_index])
            for row_index in range(header_rows)
        ]
        parts = _dedupe_consecutive([part for part in parts if part])
        flattened_columns.append("_".join(parts) or f"col_{column_index}")
    flattened_columns = _disambiguate_duplicates(flattened_columns)

    data = raw.iloc[header_rows:].dropna(how="all").reset_index(drop=True)
    data.columns = flattened_columns
    return data, flattened_columns


def ingest_sheet(
    path: Path,
    sheet: str,
    out_dir: Path,
    *,
    header_rows: int | None = None,
    max_rows: int | None = None,
) -> tuple[Path, IngestReport]:
    path = Path(path)
    workbook_format = require_excel_format(path)
    if header_rows is not None and header_rows < 1:
        raise DataIngestError("header_rows must be at least 1")
    if max_rows is not None and max_rows < 1:
        raise DataIngestError("max_rows must be at least 1")

    probed_rows = None
    if max_rows is not None:
        header_budget = header_rows or MAX_HEADER_ROWS
        probe_limit = max_rows + header_budget + 1
        probed_rows = probe_sheet_row_count(path, sheet, stop_after=probe_limit)
        if probed_rows is not None and probed_rows >= probe_limit:
            raise DatasetTooLargeError(
                reason=f"Excel 工作表 {sheet!r} 使用区行数超过上限",
                limit=max_rows,
                actual=probed_rows,
                unit="rows",
            )

    try:
        preview = _read_excel_frame(
            path,
            workbook_format=workbook_format,
            sheet=sheet,
            nrows=PREVIEW_ROWS,
        )
    except Exception as exc:
        raise DataIngestError(f"cannot read sheet {sheet}: {exc}") from exc

    if preview.empty or preview.dropna(how="all").empty:
        raise DataIngestError(f"sheet is empty: {sheet}")
    resolved_header_rows = header_rows or detect_header_rows(preview)
    if (
        max_rows is not None
        and probed_rows is not None
        and probed_rows > max_rows + resolved_header_rows
    ):
        raise DatasetTooLargeError(
            reason=f"Excel 工作表 {sheet!r} 数据行数超过上限",
            limit=max_rows,
            actual=probed_rows - resolved_header_rows,
            unit="rows",
        )

    full_read_limit = (
        max_rows + resolved_header_rows + 1 if max_rows is not None else None
    )
    try:
        full = _read_excel_frame(
            path,
            workbook_format=workbook_format,
            sheet=sheet,
            nrows=full_read_limit,
        )
    except Exception as exc:
        raise DataIngestError(f"cannot read sheet {sheet}: {exc}") from exc

    if full.empty or full.dropna(how="all").empty:
        raise DataIngestError(f"sheet is empty: {sheet}")
    data, flattened_columns = flatten_headers(full, resolved_header_rows)
    if data.empty:
        raise DataIngestError(f"sheet has no data rows: {sheet}")
    if max_rows is not None and len(data) > max_rows:
        raise DatasetTooLargeError(
            reason=f"Excel 工作表 {sheet!r} 数据行数超过上限",
            limit=max_rows,
            actual=len(data),
            unit="rows",
        )

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / new_excel_artifact_name(sheet)
    data.to_parquet(out_path, index=False)
    report = IngestReport(
        sheet=sheet,
        header_rows=resolved_header_rows,
        data_start_row=resolved_header_rows,
        flattened_columns=flattened_columns,
        original_shape=tuple(int(value) for value in full.shape),
        suspected_truncated_id_columns=_suspected_truncated_id_columns(data),
    )
    return out_path, report


def _read_excel_frame(
    path: Path,
    *,
    workbook_format: ExcelFormat,
    sheet: str,
    nrows: int | None,
) -> pd.DataFrame:
    """Read Excel content without letting an engine bypass ``nrows``.

    pandas/openpyxl applies ``nrows`` only after openpyxl has loaded a workbook
    when ``read_only=False``.  OOXML therefore stays on openpyxl's streaming
    worksheet iterator and is sliced before DataFrame construction.  BIFF uses
    xlrd through pandas, whose ``nrows`` remains explicit and bounded.
    """

    if workbook_format == "xlsx":
        with _open_xlsx_workbook(path) as workbook:
            worksheet = workbook[sheet]
            worksheet.reset_dimensions()
            rows = worksheet.iter_rows(values_only=True)
            if nrows is not None:
                rows = islice(rows, nrows)
            return pd.DataFrame(rows)

    with Path(path).open("rb") as workbook_file:
        return pd.read_excel(
            workbook_file,
            sheet_name=sheet,
            header=None,
            nrows=nrows,
            engine="xlrd",
        )


def _suspected_truncated_id_columns(data: pd.DataFrame) -> tuple[str, ...]:
    """Flag columns whose values look like a long id that Excel already stored
    as a Number cell (precision unrecoverable -- see LONG_ID_FLOAT_THRESHOLD).

    openpyxl/pandas read a Number cell as a Python float, but an integer-valued
    float (e.g. 1.1e17) is round-tripped back to a plain ``int`` by the time it
    reaches this DataFrame -- so both float and int cells are inspected here;
    what matters is the *magnitude*, which already reflects any precision lost
    at the Excel-file-format level before this code ever ran.
    """
    flagged: list[str] = []
    for column in data.columns:
        values = data[column].dropna()
        if values.empty:
            continue
        numeric_like = [value for value in values if isinstance(value, (int, float))]
        if not numeric_like:
            continue
        long_integer_shaped = [
            value for value in numeric_like
            if float(value).is_integer() and abs(value) >= LONG_ID_FLOAT_THRESHOLD
        ]
        if long_integer_shaped and len(long_integer_shaped) / len(numeric_like) >= 0.9:
            flagged.append(str(column))
    return tuple(flagged)


def _looks_like_data_row(row: pd.Series) -> bool:
    values = [value for value in row.tolist() if pd.notna(value) and str(value).strip()]
    if not values:
        return False
    typed = 0
    for value in values:
        if isinstance(value, pd.Timestamp):
            typed += 1
            continue
        if pd.to_numeric(pd.Series([value]), errors="coerce").notna().iloc[0]:
            typed += 1
            continue
        if pd.to_datetime(pd.Series([value]), errors="coerce").notna().iloc[0]:
            typed += 1
    return typed / len(values) >= 0.5


def _header_part(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text or text.lower().startswith("unnamed:"):
        return ""
    return text


def _dedupe_consecutive(parts: list[str]) -> list[str]:
    deduped = []
    for part in parts:
        if not deduped or deduped[-1] != part:
            deduped.append(part)
    return deduped


def _disambiguate_duplicates(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for name in names:
        count = seen.get(name, 0) + 1
        seen[name] = count
        result.append(name if count == 1 else f"{name}_{count}")
    return result


def _safe_sheet_name(sheet: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(sheet))
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = re.sub(r"\.{2,}", "_", cleaned).strip(" ._")
    return cleaned or "sheet"


def new_excel_artifact_name(sheet: str) -> str:
    """Return a safe immutable parquet name unique across import calls."""

    return f"{_safe_sheet_name(sheet)}_{uuid4().hex}.parquet"


__all__ = [
    "LONG_ID_FLOAT_THRESHOLD",
    "MAX_HEADER_ROWS",
    "PREVIEW_ROWS",
    "IngestReport",
    "detect_header_rows",
    "detect_excel_container_format",
    "detect_excel_container_format_from_prefix",
    "detect_excel_format",
    "flatten_headers",
    "ingest_sheet",
    "is_excel_workbook",
    "is_xls_workbook",
    "is_xlsx_workbook",
    "list_sheets",
    "new_excel_artifact_name",
    "probe_sheet_row_count",
    "require_excel_format",
]
