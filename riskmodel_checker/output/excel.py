from __future__ import annotations

from pathlib import Path
import re
from typing import Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from riskmodel_checker.output.image_render import render_roc_ks_graph
from riskmodel_checker.output.styles import (
    BORDER_COLOR,
    BRAND_HEADER_FILL,
    BRAND_HEADER_FONT_COLOR,
    FONT_NAME,
    FONT_SIZE_PT,
    ks_delta_cell_color,
    status_cell_color,
)
from riskmodel_checker.validation.results import (
    BinRow,
    ValidationResults,
)

_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
_MAX_SHEET_TITLE_LENGTH = 31


def write_validation_excel(results: ValidationResults, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_overview(workbook, results)
    _write_basic_info(workbook, results)
    _write_monthly_distribution(workbook, results)
    _write_hyperparameters(workbook, results)
    _write_feature_importance(workbook, results)
    _write_effectiveness_overall(workbook, results)
    _write_psi_stability(workbook, results)
    _write_roc_ks_images(workbook, results, output_path.parent / "excel_images")
    for split in ("train", "test", "oot"):
        _write_bins(
            workbook,
            f"分箱_{split}",
            results.effectiveness.bin_tables.get(split, []),
            first_header=f"{split}(独立分箱)",
        )
    _write_monthly_effectiveness(workbook, results)
    _write_stress_summary(workbook, results)
    for category_result in results.stress_test.per_category:
        sheet_name = f"压力测试_分箱_{category_result.category}"
        _write_bins(
            workbook,
            sheet_name,
            category_result.bin_table,
            first_header=category_result.category,
        )

    workbook.save(output_path)
    return output_path


def _write_overview(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("验证总览")
    summary = results.reproducibility.summary
    rows = [
        ("项目", "取值"),
        ("模型", f"{results.model_name} {results.model_version}"),
        ("算法", results.algorithm),
        ("抽样行数", results.reproducibility.sample_size),
        ("对齐行数", summary.match_count),
        ("差异行数", summary.mismatch_count),
        ("最大绝对差", summary.max_abs_diff),
        ("可复现性状态", summary.status.value),
    ]
    _write_rows(sheet, rows, header_rows=1)
    sheet.cell(row=len(rows), column=2).fill = PatternFill(
        start_color=status_cell_color(summary.status),
        end_color=status_cell_color(summary.status),
        fill_type="solid",
    )


def _write_basic_info(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("样本基本信息")
    total_count = sum(row.sample_count for row in results.basic_info.split_summary)
    rows: list[tuple] = [("数据集", "时间范围", "样本量", "样本占比", "坏样本量", "逾期率")]
    rows.extend(
        (
            row.split,
            _period_text(row.period_start, row.period_end, default="-"),
            row.sample_count,
            _ratio(row.sample_count, total_count),
            row.bad_count,
            row.bad_rate,
        )
        for row in results.basic_info.split_summary
    )
    _write_rows(
        sheet,
        rows,
        header_rows=1,
        percent_columns={3, 5},
        data_bar_columns={2: "5A8AC6", 5: "F8696B"},
    )


def _write_monthly_distribution(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("样本逐月分布")
    total_count = sum(row.sample_count for row in results.basic_info.monthly_distribution)
    rows: list[tuple] = [("月份", "样本量", "样本占比", "坏样本量", "逾期率")]
    rows.extend(
        (row.month, row.sample_count, _ratio(row.sample_count, total_count), row.bad_count, row.bad_rate)
        for row in results.basic_info.monthly_distribution
    )
    _write_rows(
        sheet,
        rows,
        header_rows=1,
        percent_columns={2, 4},
        data_bar_columns={1: "5A8AC6", 4: "F8696B"},
    )


def _write_hyperparameters(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("模型超参")
    rows: list[tuple] = [("参数", "取值")]
    rows.extend((str(key), str(value))
                for key, value in results.basic_info.hyperparameters.items())
    _write_rows(sheet, rows, header_rows=1)


def _write_feature_importance(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("特征重要性")
    rows: list[tuple] = [("排名", "特征", "类别", "重要性")]
    rows.extend((row.rank, row.feature, row.category, row.importance)
                for row in results.basic_info.feature_importance)
    _write_rows(sheet, rows, header_rows=1, decimal_columns={3})


def _write_effectiveness_overall(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("模型效果")
    split_summary = {row.split: row for row in results.basic_info.split_summary}
    rows: list[tuple] = [(
        "数据集", "时间范围", "样本量", "逾期率", "坏样本量",
        "KS(%)", "AUC(%)", "5%头部lift", "5%尾部lift", "PSI",
    )]
    rows.extend(
        (
            row.split,
            _period_text(
                split_summary.get(row.split).period_start if split_summary.get(row.split) else "",
                split_summary.get(row.split).period_end if split_summary.get(row.split) else "",
                default="-",
            ),
            row.sample_count,
            row.bad_rate,
            _bad_count(row),
            _pct_point(row.ks),
            _pct_point(row.auc),
            _optional_number(row.head_lift_5pct),
            _optional_number(row.tail_lift_5pct),
            "BASE" if row.split == "train" else row.psi_vs_train,
        )
        for row in results.effectiveness.overall
    )
    _write_rows(
        sheet,
        rows,
        header_rows=1,
        percent_columns={3},
        number_formats={5: "0.0", 6: "0.0", 7: "0.00", 8: "0.00", 9: "0.000"},
        data_bar_columns={5: "63BE7B"},
    )


def _write_psi_stability(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("PSI稳定性")
    rows: list[tuple] = [(
        "分箱", "train+test样本数", "train+test占比", "oot样本数", "oot占比", "PSI",
    )]
    rows.extend(
        (
            row.bin_label,
            row.expected_count,
            row.expected_pct,
            row.actual_count,
            row.actual_pct,
            row.psi,
        )
        for row in results.effectiveness.psi_stability_table
    )
    _write_rows(
        sheet,
        rows,
        header_rows=1,
        percent_columns={2, 4},
        number_formats={5: "0.0000"},
        data_bar_columns={1: "5A8AC6", 3: "5A8AC6"},
        color_scale_columns={5},
    )


def _write_roc_ks_images(workbook: Workbook, results: ValidationResults, image_dir: Path) -> None:
    sheet = workbook.create_sheet("ROC_KS曲线")
    image_dir.mkdir(parents=True, exist_ok=True)
    for index, split in enumerate(("train", "test", "oot")):
        start_row = index * 32 + 1
        sheet.cell(row=start_row, column=1, value=f"{split} ROC曲线和KS曲线")
        sheet.cell(row=start_row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_PT, bold=True)
        image_path = render_roc_ks_graph(
            results.effectiveness.roc_ks_curves.get(split),
            image_dir / f"roc_ks_graph_{split}.png",
            title_prefix=split,
        )
        image = OpenpyxlImage(str(image_path))
        image.width = 720
        image.height = 480
        sheet.add_image(image, f"A{start_row + 1}")
    sheet.column_dimensions["A"].width = 100


def _write_monthly_effectiveness(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("逐月效果")
    by_month: dict[str, dict[str, object]] = {}
    for ks_row in results.effectiveness.monthly_ks:
        by_month.setdefault(ks_row.month, {})["KS"] = ks_row.ks
        by_month[ks_row.month]["样本数"] = ks_row.sample_count
        by_month[ks_row.month]["坏样本量"] = ks_row.bad_count
        by_month[ks_row.month]["逾期率"] = ks_row.bad_rate
        by_month[ks_row.month]["AUC"] = ks_row.auc
        by_month[ks_row.month]["5%头部lift"] = ks_row.head_lift_5pct
        by_month[ks_row.month]["5%尾部lift"] = ks_row.tail_lift_5pct
    for psi_row in results.effectiveness.monthly_psi:
        by_month.setdefault(psi_row.month, {})["PSI"] = psi_row.psi_vs_train
        by_month[psi_row.month]["PSI(首月基准)"] = psi_row.psi_first_month
        by_month[psi_row.month]["PSI(尾月基准)"] = psi_row.psi_last_month
        by_month[psi_row.month]["PSI(较上一有样本月)"] = psi_row.psi_mom
        by_month[psi_row.month]["PSI参考月"] = _psi_reference_month_text(
            psi_row.psi_mom_reference_month,
            has_calendar_gap=psi_row.psi_mom_has_calendar_gap,
        )
    months = sorted(by_month)
    first_month = months[0] if months else ""
    last_month = months[-1] if months else ""
    rows: list[tuple] = [(
        "月份", "样本量", "逾期率", "坏样本量", "KS(%)", "AUC(%)",
        "5%头部lift", "5%尾部lift", "PSI(首月基准)", "PSI(尾月基准)",
        "PSI(较上一有样本月)", "PSI参考月",
    )]
    rows.extend(
        (
            month,
            data.get("样本数", 0),
            data.get("逾期率", 0.0),
            data.get("坏样本量", 0),
            _pct_point(float(data.get("KS", 0.0))),
            _pct_point(float(data.get("AUC", 0.0))),
            _optional_number(data.get("5%头部lift")),
            _optional_number(data.get("5%尾部lift")),
            "BASE" if month == first_month else _optional_number(data.get("PSI(首月基准)")),
            "BASE" if month == last_month else _optional_number(data.get("PSI(尾月基准)")),
            "-" if month == first_month else _optional_number(data.get("PSI(较上一有样本月)")),
            "-" if month == first_month else data.get("PSI参考月", ""),
        )
        for month, data in sorted(by_month.items())
    )
    _write_rows(
        sheet,
        rows,
        header_rows=1,
        percent_columns={2},
        number_formats={4: "0.0", 5: "0.0", 6: "0.00", 7: "0.00", 8: "0.000", 9: "0.000", 10: "0.000"},
        data_bar_columns={4: "63BE7B"},
    )


def _write_stress_summary(workbook: Workbook, results: ValidationResults) -> None:
    sheet = workbook.create_sheet("压力测试_汇总")
    rows: list[tuple] = [(
        "类别",
        "状态",
        "置 -9999 特征数",
        "KS_baseline",
        "KS_after",
        "KS_delta",
        "PSI",
        "错误",
    )]
    baseline_ks = results.stress_test.baseline.ks
    rows.extend(
        (item.category, _stress_status_label(item.status), len(item.dropped_features), baseline_ks,
         item.ks_after if item.ks_after is not None else "",
         item.ks_delta if item.ks_delta is not None else "",
         item.psi_vs_baseline if item.psi_vs_baseline is not None else "",
         item.error or "")
        for item in results.stress_test.per_category
    )
    _write_rows(sheet, rows, header_rows=1, decimal_columns={3, 4, 5, 6})
    # color KS_delta cells based on threshold
    for row_index, item in enumerate(results.stress_test.per_category, start=2):
        if item.ks_delta is None:
            continue
        color = ks_delta_cell_color(item.ks_delta)
        if color:
            sheet.cell(row=row_index, column=6).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid",
            )


def _stress_status_label(status: str) -> str:
    return {
        "completed": "完成",
        "skipped": "跳过",
        "error": "异常",
        "partial": "部分完成",
        "failed": "失败",
    }.get(str(status or ""), str(status or ""))


def _write_bins(
    workbook: Workbook,
    sheet_name: str,
    bins: list[BinRow],
    *,
    first_header: str = "分箱",
) -> None:
    sheet = workbook.create_sheet(_safe_sheet_title(workbook, sheet_name))
    rows: list[tuple] = [(
        first_header, "样本总数", "累计占比", "逾期数量", "逾期率",
        "累计逾期率", "单组lift", "累计lift", "ks",
    )]
    rows.extend(_reference_bin_rows(bins))
    _write_rows(
        sheet,
        rows,
        header_rows=1,
        percent_columns={2, 4, 5},
        number_formats={6: "0.00", 7: "0.00", 8: "0.0000"},
        color_scale_columns={4},
        data_bar_columns={7: "63BE7B"},
    )


def _write_rows(
    sheet,
    rows: Iterable[tuple],
    *,
    header_rows: int,
    percent_columns: set[int] | None = None,
    decimal_columns: set[int] | None = None,
    number_formats: dict[int, str] | None = None,
    data_bar_columns: dict[int, str] | None = None,
    color_scale_columns: set[int] | None = None,
) -> None:
    percent_columns = percent_columns or set()
    decimal_columns = decimal_columns or set()
    number_formats = number_formats or {}
    data_bar_columns = data_bar_columns or {}
    color_scale_columns = color_scale_columns or set()
    rows = list(rows)
    if not rows:
        return

    header_fill = PatternFill(
        start_color=BRAND_HEADER_FILL,
        end_color=BRAND_HEADER_FILL,
        fill_type="solid",
    )
    header_font = Font(
        name=FONT_NAME, size=FONT_SIZE_PT, bold=True,
        color=BRAND_HEADER_FONT_COLOR,
    )
    body_font = Font(name=FONT_NAME, size=FONT_SIZE_PT)
    thin = Side(border_style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    column_count = max(len(row) for row in rows)
    for row_index, row_values in enumerate(rows, start=1):
        for column_index in range(column_count):
            value = row_values[column_index] if column_index < len(row_values) else ""
            cell = sheet.cell(row=row_index, column=column_index + 1, value=value)
            cell.font = header_font if row_index <= header_rows else body_font
            cell.alignment = center
            cell.border = border
            if row_index <= header_rows:
                cell.fill = header_fill
            else:
                if column_index in percent_columns and isinstance(value, (int, float)):
                    cell.number_format = "0.00%"
                elif column_index in number_formats and isinstance(value, (int, float)):
                    cell.number_format = number_formats[column_index]
                elif column_index in decimal_columns and isinstance(value, (int, float)):
                    cell.number_format = "0.0000"

    _apply_reference_conditional_formatting(
        sheet,
        start_row=header_rows + 1,
        end_row=len(rows),
        data_bar_columns=data_bar_columns,
        color_scale_columns=color_scale_columns,
    )

    for column_index in range(column_count):
        sheet.column_dimensions[get_column_letter(column_index + 1)].width = 14


def _apply_reference_conditional_formatting(
    sheet,
    *,
    start_row: int,
    end_row: int,
    data_bar_columns: dict[int, str],
    color_scale_columns: set[int],
) -> None:
    if end_row < start_row:
        return
    for column_index in color_scale_columns:
        column = get_column_letter(column_index + 1)
        sheet.conditional_formatting.add(
            f"{column}{start_row}:{column}{end_row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )
    for column_index, color in data_bar_columns.items():
        column = get_column_letter(column_index + 1)
        sheet.conditional_formatting.add(
            f"{column}{start_row}:{column}{end_row}",
            DataBarRule(
                start_type="percentile", start_value=0,
                end_type="percentile", end_value=100,
                color=color,
                showValue=True,
            ),
        )


def _reference_bin_rows(bins: list[BinRow]) -> list[tuple]:
    total = sum(row.sample_count for row in bins)
    total_bad = sum(row.bad_count for row in bins)
    overall_bad_rate = _ratio(total_bad, total)
    cumulative_count = 0
    cumulative_bad = 0
    rows: list[tuple] = []
    for row in bins:
        cumulative_count += row.sample_count
        cumulative_bad += row.bad_count
        cumulative_bad_rate = _ratio(cumulative_bad, cumulative_count)
        rows.append((
            _score_interval(row.score_lower, row.score_upper),
            row.sample_count,
            _ratio(cumulative_count, total),
            row.bad_count,
            row.bad_rate,
            cumulative_bad_rate,
            row.lift,
            _ratio(cumulative_bad_rate, overall_bad_rate),
            row.ks,
        ))
    return rows


def _score_interval(lower: float, upper: float) -> str:
    return f"[{_compact_number(lower)},{_compact_number(upper)}]"


def _compact_number(value: float) -> str:
    if value == float("inf"):
        return "inf"
    if value == float("-inf"):
        return "-inf"
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.3f}".rstrip("0").rstrip(".")


def _period_text(start: str, end: str, *, default: str) -> str:
    if not start and not end:
        return default
    if not start:
        return str(end)
    if not end:
        return str(start)
    return str(start) if start == end else f"{start}-{end}"


def _ratio(numerator: float, denominator: float) -> float:
    return float(numerator / denominator) if denominator else 0.0


def _pct_point(value: float) -> float:
    return round(float(value) * 100, 1)


def _bad_count(row) -> int:
    if row.bad_count is not None:
        return int(row.bad_count)
    return int(round(row.sample_count * row.bad_rate))


def _optional_number(value) -> float | str:
    return "" if value is None else float(value)


def _psi_reference_month_text(month: str, *, has_calendar_gap: bool) -> str:
    if not month:
        return ""
    return f"{month}(跨月)" if has_calendar_gap else str(month)


def _safe_sheet_title(workbook: Workbook, title: str) -> str:
    base = _INVALID_SHEET_TITLE_CHARS.sub("_", str(title)).strip().strip("'") or "Sheet"
    candidate = base[:_MAX_SHEET_TITLE_LENGTH]
    if candidate not in workbook.sheetnames:
        return candidate
    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{base[:_MAX_SHEET_TITLE_LENGTH - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        index += 1
