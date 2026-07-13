from dataclasses import replace
from pathlib import Path
import re

from openpyxl import load_workbook
import pytest

from marvis.output.excel import write_validation_excel
from marvis.validation.results import (
    BasicInfoResult,
    BinRow,
    ConsistencyStatus,
    ConsistencySummary,
    EffectivenessResult,
    FeatureImportanceRow,
    MonthlyKsRow,
    MonthlyPsiRow,
    MonthlyRow,
    OverallRow,
    PmmlScoringResult,
    PsiStabilityRow,
    ReproducibilityResult,
    RocKsCurve,
    ScoreCompareRow,
    SplitRow,
    StressBaseline,
    StressCategoryResult,
    StressTestResult,
    ValidationResults,
)


def _make_results() -> ValidationResults:
    repro = ReproducibilityResult(
        sample_size=2, seed=42,
        rows=[
            ScoreCompareRow(0, 0.1, 0.1, 0.1),
            ScoreCompareRow(1, 0.5, 0.5, 0.5),
        ],
        summary=ConsistencySummary(2, 0, 0.0, ConsistencyStatus.PASS),
    )
    basic = BasicInfoResult(
        sample_period=("202503", "202505"),
        split_summary=[
            SplitRow("train", 100, 10, 0.1, "202503", "202503"),
            SplitRow("test", 50, 5, 0.1, "202504", "202504"),
            SplitRow("oot", 50, 5, 0.1, "202505", "202505"),
        ],
        monthly_distribution=[MonthlyRow("202503", 50, 5, 0.1)],
        hyperparameters={"learning_rate": 0.05, "max_depth": 5},
        feature_importance=[
            FeatureImportanceRow(1, "x1", 0.8, category="征信"),
            FeatureImportanceRow(2, "x2", 0.2, category="行为"),
        ],
    )
    bin_row = BinRow(1, 0.0, 0.5, 50, 5, 0.1, 0.5, 0.5, 1.0, 0.0)
    eff = EffectivenessResult(
        overall=[
            OverallRow("train", 0.3, 0.0, 100, 0.1, bad_count=10, auc=0.75,
                       head_lift_5pct=0.2, tail_lift_5pct=2.0),
            OverallRow("test", 0.28, 0.05, 50, 0.1, bad_count=5, auc=0.72,
                       head_lift_5pct=0.3, tail_lift_5pct=1.9),
            OverallRow("oot", 0.25, 0.12, 50, 0.1, bad_count=5, auc=0.70,
                       head_lift_5pct=0.4, tail_lift_5pct=1.8),
        ],
        bin_tables={"train": [bin_row], "test": [bin_row], "oot": [bin_row]},
        monthly_ks=[MonthlyKsRow("202503", 0.3, 50, bad_count=5, bad_rate=0.1,
                                 auc=0.75, head_lift_5pct=0.2, tail_lift_5pct=2.0)],
        monthly_psi=[MonthlyPsiRow("202503", 0.05, psi_first_month=0.0,
                                   psi_last_month=0.0, psi_mom=None)],
        psi_stability_table=[
            PsiStabilityRow("[0,0.5]", 150, 1.0, 50, 1.0, 0.0),
        ],
        roc_ks_curves={
            split: RocKsCurve(
                split=split,
                fpr=[0.0, 0.5, 1.0],
                tpr=[0.0, 0.8, 1.0],
                ks_curve=[0.0, 0.3, 0.0],
                ks=0.3,
                population_at_ks=0.5,
            )
            for split in ("train", "test", "oot")
        },
    )
    stress = StressTestResult(
        baseline=StressBaseline(ks=0.25, sample_count=50, bin_table=[bin_row]),
        per_category=[
            StressCategoryResult(
                category="征信", dropped_features=["x1"],
                ks_after=0.20, ks_delta=-0.05, psi_vs_baseline=0.08,
                bin_table=[bin_row], error=None,
            ),
        ],
    )
    return ValidationResults(
        model_name="A卡", model_version="v1",
        algorithm="lgb", target_type="binary",
        schema_version="marvis.validation_results.v1",
        reproducibility=repro, basic_info=basic,
        effectiveness=eff, stress_test=stress,
    )


def _make_pmml_results() -> ValidationResults:
    legacy = _make_results()
    scoring = PmmlScoringResult(
        schema_version="marvis.pmml_scoring.v1",
        cache_key="c" * 64,
        pmml_sha256="d" * 64,
        sample_sha256="e" * 64,
        engine="pypmml-pmml4s-batch",
        engine_version="1.5.8",
        output_field="probability_1",
        input_row_count=3,
        success_count=3,
        failure_count=0,
        null_count=0,
        non_finite_count=0,
        elapsed_seconds=0.1,
        rows_per_second=30.0,
        chunk_size=2,
        required_input_count=2,
        missing_inputs=[],
        score_artifact_path="pmml_scores.parquet",
        score_artifact_sha256="a" * 64,
        status="pass",
        bounded_errors=[],
    )
    return replace(
        legacy,
        schema_version="marvis.validation_results.v2",
        pmml_scoring=scoring,
        reproducibility=None,
    )


def test_write_excel_includes_all_required_sheets(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)

    wb = load_workbook(output, data_only=True)
    expected = {
        "验证总览", "样本基本信息", "样本逐月分布", "模型超参",
        "特征重要性", "模型效果", "分箱_train", "分箱_test", "分箱_oot",
        "PSI稳定性", "ROC_KS曲线", "逐月效果", "压力测试_汇总", "压力测试_分箱_征信",
    }
    assert expected.issubset(set(wb.sheetnames))


def test_overview_sheet_contains_status_text(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    wb = load_workbook(output, data_only=True)
    overview = wb["验证总览"]
    values = [cell.value for row in overview.iter_rows() for cell in row]
    assert any(isinstance(v, str) and "pass" in v.lower() for v in values)


def test_feature_importance_sheet_includes_category(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)

    wb = load_workbook(output, data_only=True)
    sheet = wb["特征重要性"]

    assert [cell.value for cell in sheet[1]] == ["排名", "特征", "类别", "重要性"]
    assert sheet["C2"].value == "征信"
    assert sheet["C3"].value == "行为"


def test_stress_summary_labels_negative_9999_sentinel(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    wb = load_workbook(output, data_only=True)

    header_values = [cell.value for cell in wb["压力测试_汇总"][1]]

    assert "置 -9999 特征数" in header_values
    assert "置 null 特征数" not in header_values


def test_stress_summary_includes_category_coverage_row(tmp_path: Path):
    base = _make_results()
    results = replace(
        base,
        stress_test=replace(
            base.stress_test,
            status="partial",
            unclassified_features=["BH_A044_C0580"],
        ),
    )
    output = tmp_path / "out.xlsx"

    write_validation_excel(results, output)

    sheet = load_workbook(output, data_only=True)["压力测试_汇总"]
    values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    assert values[1][0] == "分类覆盖"
    assert values[1][1] == "部分完成"
    assert "未分类特征 1 个：BH_A044_C0580" in values[1][-1]


def test_header_style_applied(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    wb = load_workbook(output)
    header = wb["样本基本信息"].cell(row=1, column=1)
    assert header.fill.fgColor.rgb.endswith("C00000")
    assert header.font.bold is True


def test_reference_sample_and_effect_headers_and_number_formats(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    wb = load_workbook(output)

    sample_sheet = wb["样本基本信息"]
    assert [cell.value for cell in sample_sheet[1][:6]] == [
        "数据集", "时间范围", "样本量", "样本占比", "坏样本量", "逾期率",
    ]
    assert sample_sheet.cell(row=2, column=4).number_format == "0.00%"
    assert sample_sheet.cell(row=2, column=6).number_format == "0.00%"

    effect_sheet = wb["模型效果"]
    assert [cell.value for cell in effect_sheet[1][:10]] == [
        "数据集", "时间范围", "样本量", "逾期率", "坏样本量",
        "KS(%)", "AUC(%)", "5%头部lift", "5%尾部lift", "PSI",
    ]
    assert effect_sheet.cell(row=2, column=4).number_format == "0.00%"
    assert effect_sheet.cell(row=2, column=6).number_format == "0.0"
    assert effect_sheet.cell(row=2, column=7).number_format == "0.0"
    assert effect_sheet.cell(row=2, column=8).number_format == "0.00"
    assert effect_sheet.cell(row=2, column=10).value == "BASE"

    monthly_sheet = wb["逐月效果"]
    assert [cell.value for cell in monthly_sheet[1][:12]] == [
        "月份", "样本量", "逾期率", "坏样本量", "KS(%)", "AUC(%)",
        "5%头部lift", "5%尾部lift", "PSI(首月基准)", "PSI(尾月基准)",
        "PSI(较上一有样本月)", "PSI参考月",
    ]


def test_effectiveness_sheet_preserves_zero_bad_count(tmp_path: Path):
    results = _make_results()
    overall = list(results.effectiveness.overall)
    overall[0] = replace(overall[0], bad_count=0, bad_rate=0.004)
    results = replace(
        results,
        effectiveness=replace(results.effectiveness, overall=overall),
    )
    output = tmp_path / "out.xlsx"

    write_validation_excel(results, output)

    wb = load_workbook(output, data_only=True)
    sheet = wb["模型效果"]
    assert sheet.cell(row=2, column=5).value == 0


def test_reference_bin_table_columns_and_formats(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    wb = load_workbook(output)
    sheet = wb["分箱_oot"]

    assert [cell.value for cell in sheet[1][:9]] == [
        "oot(独立分箱)", "样本总数", "累计占比", "逾期数量", "逾期率",
        "累计逾期率", "单组lift", "累计lift", "ks",
    ]
    assert sheet.cell(row=2, column=1).value == "[0,0.5]"
    assert sheet.cell(row=2, column=3).number_format == "0.00%"
    assert sheet.cell(row=2, column=5).number_format == "0.00%"
    assert sheet.cell(row=2, column=6).number_format == "0.00%"
    assert sheet.cell(row=2, column=7).number_format == "0.00"
    assert sheet.cell(row=2, column=8).number_format == "0.00"
    assert sheet.cell(row=2, column=9).number_format == "0.0000"


def test_new_report_image_materials_are_in_excel(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    wb = load_workbook(output)

    psi_sheet = wb["PSI稳定性"]
    assert [cell.value for cell in psi_sheet[1][:6]] == [
        "分箱", "train+test样本数", "train+test占比", "oot样本数", "oot占比", "PSI",
    ]
    assert psi_sheet.cell(row=2, column=3).number_format == "0.00%"
    assert psi_sheet.cell(row=2, column=5).number_format == "0.00%"
    assert psi_sheet.cell(row=2, column=6).number_format == "0.0000"

    pressure_sheet = wb["压力测试_分箱_征信"]
    assert pressure_sheet.cell(row=1, column=1).value == "征信"
    assert len(wb["ROC_KS曲线"]._images) == 3


def test_write_excel_rolls_back_existing_file_and_images_when_save_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    previous_excel = output.read_bytes()
    image_dir = tmp_path / "excel_images"
    previous_images = {
        item.name: item.read_bytes()
        for item in image_dir.glob("*.png")
    }
    assert previous_images

    def fail_save(self, filename):
        raise RuntimeError("save failed")

    monkeypatch.setattr("openpyxl.workbook.workbook.Workbook.save", fail_save)

    with pytest.raises(RuntimeError, match="save failed"):
        write_validation_excel(_make_results(), output)

    assert output.read_bytes() == previous_excel
    assert {
        item.name: item.read_bytes()
        for item in image_dir.glob("*.png")
    } == previous_images
    assert not (tmp_path / ".staging").exists()
    assert not list(tmp_path.glob(".out.xlsx.*.bak"))
    assert not list(tmp_path.glob(".excel_images.*.bak"))


def test_reference_conditional_formatting_ranges_are_applied(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    write_validation_excel(_make_results(), output)
    wb = load_workbook(output)

    bin_ranges = {str(item.sqref) for item in wb["分箱_train"].conditional_formatting}
    psi_ranges = {str(item.sqref) for item in wb["PSI稳定性"].conditional_formatting}

    assert {"E2", "H2"}.issubset(bin_ranges)
    assert {"B2", "D2", "F2"}.issubset(psi_ranges)


def test_stress_category_sheet_titles_are_excel_safe_and_unique(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    results = _make_results()
    category = "京东/携程:长类别名超过三十一字符用于验证sheet名"
    category_rows = [
        replace(results.stress_test.per_category[0], category=category),
        replace(results.stress_test.per_category[0], category=category),
    ]
    results = replace(
        results,
        stress_test=replace(results.stress_test, per_category=category_rows),
    )

    write_validation_excel(results, output)

    wb = load_workbook(output, data_only=True)
    stress_sheets = [
        name for name in wb.sheetnames if name.startswith("压力测试_分箱_京东_携程")
    ]
    assert len(stress_sheets) == 2
    assert len(set(stress_sheets)) == 2
    assert all(len(name) <= 31 for name in stress_sheets)
    assert all(not re.search(r"[\[\]:*?/\\]", name) for name in stress_sheets)
