from __future__ import annotations

from copy import deepcopy
from io import BytesIO

from openpyxl import load_workbook
import pytest

from marvis.output.strategy_report_bundle import (
    STRATEGY_REPORT_SHEET_TITLES,
    StrategyReportOutputError,
    render_strategy_report_bundle,
    render_strategy_report_bundle_docx,
    render_strategy_report_bundle_markdown,
    render_strategy_report_bundle_xlsx,
)
from marvis.packs.strategy.report_bundle import (
    REPORT_CORE_SHEET_KEYS,
    build_named_report_field,
    canonical_strategy_report_bundle_json,
)
from tests.test_strategy_report_bundle import (
    _bundle,
    _present,
    _sections,
    _source,
)


def _workbook(raw: bytes):
    return load_workbook(BytesIO(raw), data_only=False, read_only=False)


def _find_row(sheet, first_value):
    for row in sheet.iter_rows():
        if row[0].value == first_value:
            return row
    raise AssertionError(f"row {first_value!r} not found in {sheet.title}")


def test_all_four_projections_are_deterministic_and_share_one_manifest():
    bundle = _bundle()

    first = render_strategy_report_bundle(bundle)
    second = render_strategy_report_bundle(bundle)

    assert first == second
    assert first["json"] == canonical_strategy_report_bundle_json(bundle).encode(
        "utf-8"
    )
    assert first["xlsx"].startswith(b"PK")
    assert first["docx"] == render_strategy_report_bundle_docx(bundle)
    assert first["docx"].startswith(b"PK")
    markdown = first["markdown"].decode("utf-8")
    assert markdown.startswith("# 风险策略迭代评审\n")
    assert "开发回测" in markdown
    assert "空白不等于 0" in markdown
    assert bundle["content_sha256"] in markdown


def test_workbook_has_stable_core_sheet_keys_and_human_titles():
    workbook = _workbook(render_strategy_report_bundle_xlsx(_bundle()))
    try:
        assert workbook.sheetnames == list(REPORT_CORE_SHEET_KEYS)
        for key in REPORT_CORE_SHEET_KEYS:
            assert workbook[key]["A1"].value == STRATEGY_REPORT_SHEET_TITLES[key]
        assert workbook.properties.creator == "MARVIS"
        assert workbook.properties.description == "strategy.report-output.v2"
    finally:
        workbook.close()


def test_blank_zero_maturity_and_formula_safety_survive_xlsx_projection():
    sections = _sections()
    sections[0]["summary_fields"].extend(
        [
            build_named_report_field(
                field_id="formula_like_text",
                label="外部说明",
                field=_present(
                    "  =HYPERLINK(\"https://invalid.example\")",
                    source=_source("report_context", "formula-text"),
                    origin="user",
                ),
            ),
            build_named_report_field(
                field_id="markdown_text",
                label="富文本",
                field=_present(
                    "<b>a|b</b>",
                    source=_source("report_context", "markdown-text"),
                    origin="user",
                ),
            ),
        ]
    )
    bundle = _bundle(sections=sections)
    raw = render_strategy_report_bundle_xlsx(bundle)
    workbook = _workbook(raw)
    try:
        current = workbook["01_current_state"]
        approval = _find_row(current, "approval_rate")
        unavailable = _find_row(current, "business_background")
        formula = _find_row(current, "formula_like_text")
        assert approval[2].value == 0
        assert approval[3].value == "已有可信值"
        assert unavailable[2].value is None
        assert unavailable[3].value == "暂缺"
        assert formula[2].value.startswith("'  =HYPERLINK")
        assert formula[2].data_type == "s"

        impact = workbook["08_impact"]
        monthly = _find_row(impact, "2026-06")
        assert monthly[4].value == 0
        assert monthly[4].number_format == "0.0000%"
        assert monthly[8].value is None
        assert monthly[9].value == "尚未成熟"
        assert monthly[8].number_format == "General"

        assert all(
            cell.data_type != "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        summary_approval = _find_row(workbook["00_summary"], "approval_rate")
        assert summary_approval[2].value == 0
    finally:
        workbook.close()

    markdown = render_strategy_report_bundle_markdown(bundle).decode("utf-8")
    assert "&lt;b&gt;a\\|b&lt;/b&gt;" in markdown
    assert "| 通过率 | 0 | 已有可信值 |" in markdown
    assert "| 业务背景 |  | 暂缺 |" in markdown
    assert r"| 风险率 \[%\]状态 |" in markdown


def test_nonzero_percentage_has_the_same_human_display_across_formats():
    sections = _sections()
    impact_ref = _source("strategy_impact", "impact-1", "b" * 64)
    sections[5]["tables"][0]["rows"][0]["cells"]["approval_rate"] = _present(
        0.1234,
        source=impact_ref,
    )
    bundle = _bundle(sections=sections)

    markdown = render_strategy_report_bundle_markdown(bundle).decode("utf-8")
    assert "| 12.3400% | 已有可信值 |" in markdown

    workbook = _workbook(render_strategy_report_bundle_xlsx(bundle))
    try:
        monthly = _find_row(workbook["08_impact"], "2026-06")
        assert monthly[4].value == 0.1234
        assert monthly[4].number_format == "0.0000%"
    finally:
        workbook.close()


def test_explicit_appendix_sheet_key_is_stable_and_rendered_after_core_sheets():
    sections = _sections()
    appendix = deepcopy(sections[5]["tables"][0])
    appendix["table_id"] = "score_bins_appendix"
    appendix["sheet_key"] = "appendix_score_bins"
    sections[5]["tables"].append(appendix)
    bundle = _bundle(sections=sections)

    workbook = _workbook(render_strategy_report_bundle_xlsx(bundle))
    try:
        assert workbook.sheetnames == [
            *REPORT_CORE_SHEET_KEYS,
            "appendix_score_bins",
        ]
        assert workbook["appendix_score_bins"]["A1"].value == (
            "分析附件：appendix_score_bins"
        )
        assert _find_row(workbook["appendix_score_bins"], "2026-06")
    finally:
        workbook.close()


def test_xlsx_projection_fails_closed_instead_of_truncating_a_cell():
    sections = _sections()
    sections[0]["summary_fields"].append(
        build_named_report_field(
            field_id="too_long",
            label="超长说明",
            field=_present(
                "x" * 32_768,
                source=_source("report_context", "too-long"),
                origin="user",
            ),
        )
    )
    bundle = _bundle(sections=sections)

    with pytest.raises(StrategyReportOutputError, match="32767"):
        render_strategy_report_bundle_xlsx(bundle)


def test_control_characters_are_escaped_in_workbook_properties_and_markdown():
    bundle = _bundle(title="风险策略\x01评审\x1b\uffff")

    workbook = _workbook(render_strategy_report_bundle_xlsx(bundle))
    try:
        assert workbook.properties.title == "风险策略\\u0001评审\\u001b\\uffff"
        title_row = _find_row(workbook["00_summary"], "报告标题")
        assert title_row[1].value == "风险策略\\u0001评审\\u001b\\uffff"
    finally:
        workbook.close()

    markdown = render_strategy_report_bundle_markdown(bundle).decode("utf-8")
    assert markdown.startswith("# 风险策略\\u0001评审\\u001b\\uffff\n")
    assert "\x01" not in markdown
    assert "\x1b" not in markdown
    assert "\uffff" not in markdown


def test_markdown_projection_neutralizes_links_images_and_inline_markup():
    malicious = "![pixel](https://invalid.example/pixel)`code`_*"
    sections = _sections()
    sections[0]["summary_fields"].append(
        build_named_report_field(
            field_id="external_text",
            label="外部说明",
            field=_present(
                malicious,
                source=_source("report_context", "external-markdown"),
                origin="user",
            ),
        )
    )
    bundle = _bundle(title=malicious, sections=sections)

    markdown = render_strategy_report_bundle_markdown(bundle).decode("utf-8")
    escaped = (
        r"\!\[pixel\]\(https://invalid.example/pixel\)"
        r"\`code\`\_\*"
    )
    assert markdown.startswith(f"# {escaped}\n")
    assert escaped in markdown
    assert "![pixel](" not in markdown
