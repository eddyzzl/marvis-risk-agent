from __future__ import annotations

from contextlib import contextmanager
from io import BytesIO
import json
from pathlib import Path
import sys
import threading

from openpyxl import load_workbook
import pandas as pd
import pytest

from marvis.data.backend import DataBackend
from marvis.data.registry import DatasetRegistry
from marvis.data.workspace import (
    DataSemanticMapping,
    DataWorkspaceDraft,
    data_semantic_mapping_hash,
)
from marvis.db import DatasetRepository, PluginRepository, TaskRepository, init_db
from marvis.db_schema import connect
from marvis.domain import TaskCreate
from marvis.files import sha256_file
from marvis.packs.strategy.candidate_evidence import validate_candidate_evidence
from marvis.packs.strategy import tools as strategy_tools
from marvis.packs.strategy.errors import StrategyError
from marvis.plugins.contracts import ToolContext
from marvis.plugins.loader import load_builtin_packs
from marvis.plugins.manifest import ToolRef
from marvis.plugins.registry import PluginRegistry, ToolRegistry
from marvis.plugins.runner import ToolRunner
from marvis.repositories.data_workspace import DataWorkspaceRepository
from marvis.repositories.task_artifacts import TaskArtifactRepository
from marvis.settings import build_settings


def _runtime(
    tmp_path: Path,
    *,
    target_bad_value: int = 1,
    with_split: bool = False,
    maturity_status: str = "confirmed_matured",
    sample_amount_fields: bool = True,
):
    settings = build_settings(tmp_path / "workspace")
    init_db(settings.db_path)
    plugin_repo = PluginRepository(settings.db_path)
    plugin_registry = PluginRegistry(plugin_repo)
    load_builtin_packs(
        plugin_registry,
        Path(__file__).parents[1] / "marvis" / "packs",
    )
    runner = ToolRunner(
        ToolRegistry(plugin_registry),
        plugin_repo,
        python_executable=sys.executable,
        datasets_root=settings.datasets_dir,
        workspace=settings.workspace,
    )
    registry = DatasetRegistry(
        DatasetRepository(settings.db_path),
        DataBackend(settings.datasets_dir),
        settings.datasets_dir,
    )
    task_repo = TaskRepository(settings.db_path)
    task = task_repo.create_task(
        TaskCreate(
            model_name="univariate-candidate",
            model_version="dev",
            validator="qa",
            source_dir=str(tmp_path / "source"),
            task_type="strategy",
            target_col="bad",
        )
    )
    other_task = task_repo.create_task(
        TaskCreate(
            model_name="foreign",
            model_version="dev",
            validator="qa",
            source_dir=str(tmp_path / "foreign"),
            task_type="strategy",
            target_col="bad",
        )
    )
    bad_one_labels = [0, 0, 0, 1, 0, 1, 0, 1, 0, 1, 1, 1]
    frame_data = {
        "customer_id": [f"C{index:03d}" for index in range(12)],
        "score": [-9999, 100, 150, 200, 250, 300, 350, 400, 450, 500, 550, 600],
        "constant": [7] * 12,
        "segment": ["UNKNOWN", "A", "A", "B", "B", "C"] * 2,
        "loan_amount": [100, 120, 130, 140, 150, 160, 170, 180, 190, 200, 210, 220],
        "overdue_amount": [0, 0, 5, 0, 10, 0, 15, 0, 20, 0, 25, 30],
        "bad": (
            bad_one_labels
            if target_bad_value == 1
            else [1 - value for value in bad_one_labels]
        ),
    }
    if with_split:
        frame_data["sample_split"] = ["dev"] * 8 + ["valid"] * 2 + ["oot"] * 2
    frame = pd.DataFrame(frame_data)
    source = tmp_path / "candidate.parquet"
    frame.to_parquet(source, index=False)
    dataset = registry.register_existing(
        source,
        task_id=task.id,
        role="derived",
    )
    workspace_repo = DataWorkspaceRepository(settings.db_path)
    activated = workspace_repo.save(
        task.id,
        DataWorkspaceDraft(
            active_dataset_id=dataset.id,
            active_dataset_content_hash=dataset.content_hash,
        ),
        expected_revision=0,
    )
    mapping = DataSemanticMapping(
        target_col="bad",
        field_roles={
            "customer_id": "id",
            "score": "score",
            "constant": "date",
            "segment": "categorical",
            "loan_amount": "loan_amount",
            "overdue_amount": "overdue_amount",
            "bad": "target",
            **({"sample_split": "segment"} if with_split else {}),
        },
        business_names={"score": "风险评分", "segment": "客群"},
    )
    workspace = workspace_repo.save(
        task.id,
        DataWorkspaceDraft(
            active_dataset_id=dataset.id,
            active_dataset_content_hash=dataset.content_hash,
            semantic_mapping=mapping,
        ),
        expected_revision=activated.revision,
    )
    sample_design_ref = _materialize_sample_design_ref(
        settings,
        task,
        dataset,
        workspace,
        mapping,
        target_bad_value=target_bad_value,
        with_split=with_split,
        maturity_status=maturity_status,
        sample_amount_fields=sample_amount_fields,
    )
    return (
        settings,
        runner,
        registry,
        task,
        other_task,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    )


def _inputs(dataset, workspace, mapping, sample_design_ref) -> dict:
    return {
        "dataset_id": dataset.id,
        "expected_content_hash": dataset.content_hash,
        "workspace_revision": workspace.revision,
        "analysis_generation": workspace.analysis_generation,
        "semantic_mapping_hash": data_semantic_mapping_hash(mapping),
        "target_col": "bad",
        "sample_design_ref": sample_design_ref,
        "features": [],
        "methods": [],
        "bin_count": 3,
        "min_bin_pct": 0.02,
        "sentinel_values": [-9999, "UNKNOWN"],
    }


def _tool_context(settings, task) -> ToolContext:
    return ToolContext(
        task_id=task.id,
        seed=0,
        datasets_root=settings.datasets_dir,
        workspace=settings.workspace,
    )


def _materialize_sample_design_ref(
    settings,
    task,
    dataset,
    workspace,
    mapping,
    *,
    target_bad_value: int = 1,
    with_split: bool = False,
    maturity_status: str = "confirmed_matured",
    sample_amount_fields: bool = True,
) -> dict[str, str]:
    request = {
        "dataset_id": dataset.id,
        "expected_dataset_content_hash": dataset.content_hash,
        "workspace_revision": workspace.revision,
        "workspace_generation": workspace.analysis_generation,
        "semantic_mapping_hash": data_semantic_mapping_hash(mapping),
        "target_col": "bad",
        "target_bad_value": target_bad_value,
        "performance_window_status": "provided",
        "performance_window_days": 30,
        "observation_window_status": "provided",
        "observation_window_start": "2026-01-01",
        "observation_window_end": "2026-01-31",
        "maturity_status": maturity_status,
        "drop_nan_labels": False,
    }
    if sample_amount_fields:
        request.update(
            {
                "loan_amount_col": "loan_amount",
                "overdue_amount_col": "overdue_amount",
            }
        )
    if with_split:
        request.update(
            {
                "split_col": "sample_split",
                "development_values": ["dev"],
                "validation_values": ["valid"],
                "oot_values": ["oot"],
            }
        )
    output = strategy_tools.tool_materialize_sample_design(
        request,
        _tool_context(settings, task),
    )
    return {
        "artifact_id": output["artifact"]["artifact_id"],
        "artifact_content_hash": output["artifact"]["content_hash"],
        "sample_design_id": output["sample_design_id"],
        "sample_design_content_hash": output["content_hash"],
        "partition": "development",
    }


def _candidate_artifacts(settings, task) -> list[dict]:
    return [
        record
        for record in TaskArtifactRepository(settings.db_path).list_for_task(task.id)
        if record["kind"] in {"strategy_candidate_json", "strategy_candidate_xlsx"}
    ]


def test_univariate_tool_uses_active_workspace_and_writes_stable_reports(
    tmp_path: Path,
) -> None:
    (
        settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    inputs = _inputs(dataset, workspace, mapping, sample_design_ref)

    first = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        inputs,
        task_id=task.id,
    )
    repeated = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        inputs,
        task_id=task.id,
    )

    assert first.ok, first.error
    assert repeated.ok, repeated.error
    assert first.output == repeated.output
    output = first.output
    assert output["schema_version"] == "strategy.univariate-candidate-tool.v1"
    assert output["validation_status"] == "unvalidated"
    assert output["feature_count"] == 2
    assert output["available_method_count"] == 5
    assert {row["feature"] for row in output["rankings"]} == {"score", "segment"}
    assert "customer_id" not in {
        item["feature"] for item in output["candidate_evidence"]["analysis"]["features"]
    }
    evidence = validate_candidate_evidence(output["candidate_evidence"])
    assert evidence["candidate_id"] == output["candidate_id"]
    assert evidence["identity"]["workspace_revision"] == workspace.revision
    assert evidence["generation"]["parameters"]["method_mode"] == "type_aware_auto"
    assert evidence["generation"]["parameters"]["sample_design_ref"] == (
        sample_design_ref
    )
    assert any(
        sample_design_ref["artifact_id"] in source_ref
        and sample_design_ref["artifact_content_hash"] in source_ref
        and sample_design_ref["sample_design_id"] in source_ref
        and sample_design_ref["sample_design_content_hash"] in source_ref
        and '"partition":"development"' in source_ref
        for source_ref in evidence["source_refs"]
    )
    assert evidence["generation"]["parameters"]["loan_amount_col"] == "loan_amount"
    assert any(
        metric["metric_name"].endswith(".hit_rate")
        and metric["dimension"] == "loan_amount"
        and metric["status"] == "observed"
        for metric in evidence["metrics"]
    )
    assert any("sentinel_incompatible:score" in flag for flag in output["red_flags"])
    assert len(output["artifacts"]) == 2
    assert all("path" not in artifact for artifact in output["artifacts"])

    records = _candidate_artifacts(settings, task)
    assert len(records) == 2
    assert {record["kind"] for record in records} == {
        "strategy_candidate_json",
        "strategy_candidate_xlsx",
    }
    json_record = next(record for record in records if record["kind"].endswith("json"))
    report = json.loads(Path(json_record["path"]).read_bytes())
    assert report["candidate_evidence"] == evidence
    assert report["univariate_analysis"] == evidence["analysis"]
    xlsx_record = next(record for record in records if record["kind"].endswith("xlsx"))
    workbook = load_workbook(
        BytesIO(Path(xlsx_record["path"]).read_bytes()),
        read_only=True,
    )
    assert workbook.sheetnames == [
        "Summary",
        "Rankings",
        "Bins",
        "Metrics",
        "Red Flags",
        "Lineage",
    ]


def test_univariate_omitted_amount_fields_follow_sample_design_not_semantic_roles(
    tmp_path: Path,
) -> None:
    (
        _settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path, sample_amount_fields=False)

    result = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        _inputs(dataset, workspace, mapping, sample_design_ref),
        task_id=task.id,
    )

    assert result.ok, result.error
    parameters = result.output["candidate_evidence"]["generation"]["parameters"]
    assert parameters["loan_amount_col"] is None
    assert parameters["overdue_amount_col"] is None
    assert not any(
        metric["dimension"] in {"loan_amount", "overdue_amount"}
        and metric["status"] == "observed"
        for metric in result.output["candidate_evidence"]["metrics"]
    )


def test_univariate_tool_filters_development_and_normalizes_reverse_bad_label(
    tmp_path: Path,
) -> None:
    bad_one = _runtime(tmp_path / "bad-one", target_bad_value=1, with_split=True)
    bad_zero = _runtime(tmp_path / "bad-zero", target_bad_value=0, with_split=True)

    outputs = []
    for fixture in (bad_one, bad_zero):
        (
            settings,
            _runner,
            _registry,
            task,
            _other,
            dataset,
            workspace,
            mapping,
            sample_design_ref,
        ) = fixture
        outputs.append(
            strategy_tools.tool_analyze_univariate_candidates(
                _inputs(dataset, workspace, mapping, sample_design_ref),
                _tool_context(settings, task),
            )
        )

    first = outputs[0]["candidate_evidence"]
    reversed_polarity = outputs[1]["candidate_evidence"]
    assert first["analysis"] == reversed_polarity["analysis"]
    assert first["metrics"] == reversed_polarity["metrics"]
    assert first["analysis"]["row_count"] == 8
    assert first["candidate_id"] != reversed_polarity["candidate_id"]
    assert first["generation"]["parameters"]["sample_design_ref"] != (
        reversed_polarity["generation"]["parameters"]["sample_design_ref"]
    )


def test_univariate_tool_resets_duplicate_source_index_after_partition_binding(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path, with_split=True)
    original = DataBackend.read_frame

    def duplicate_index(self, path, *, columns=None, nrows=None):
        frame = original(self, path, columns=columns, nrows=nrows)
        frame.index = [0] * len(frame)
        return frame

    monkeypatch.setattr(DataBackend, "read_frame", duplicate_index)

    output = strategy_tools.tool_analyze_univariate_candidates(
        _inputs(dataset, workspace, mapping, sample_design_ref),
        _tool_context(settings, task),
    )

    assert output["candidate_evidence"]["analysis"]["row_count"] == 8


def test_univariate_tool_requires_sample_design_ref(tmp_path: Path) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    inputs = _inputs(dataset, workspace, mapping, sample_design_ref)
    del inputs["sample_design_ref"]

    with pytest.raises(StrategyError, match="sample_design_ref is required"):
        strategy_tools.tool_analyze_univariate_candidates(
            inputs,
            _tool_context(settings, task),
        )


def test_univariate_tool_rejects_every_sample_design_ref_drift(tmp_path: Path) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    mutations = (
        {**sample_design_ref, "artifact_id": "0" * 64},
        {**sample_design_ref, "artifact_content_hash": "0" * 64},
        {**sample_design_ref, "sample_design_id": "strategy-sample-design-forged"},
        {**sample_design_ref, "sample_design_content_hash": "0" * 64},
        {**sample_design_ref, "partition": "validation"},
    )

    for forged_ref in mutations:
        with pytest.raises(StrategyError):
            strategy_tools.tool_analyze_univariate_candidates(
                {
                    **_inputs(dataset, workspace, mapping, sample_design_ref),
                    "sample_design_ref": forged_ref,
                },
                _tool_context(settings, task),
            )


def test_univariate_tool_rejects_sample_policy_and_amount_mismatch(
    tmp_path: Path,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    base = _inputs(dataset, workspace, mapping, sample_design_ref)

    with pytest.raises(StrategyError, match="drop_nan_labels"):
        strategy_tools.tool_analyze_univariate_candidates(
            {**base, "drop_nan_labels": True},
            _tool_context(settings, task),
        )
    with pytest.raises(StrategyError, match="loan_amount_col"):
        strategy_tools.tool_analyze_univariate_candidates(
            {**base, "loan_amount_col": "score"},
            _tool_context(settings, task),
        )


def test_univariate_tool_rejects_non_mature_sample_design(tmp_path: Path) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path, maturity_status="unknown")

    with pytest.raises(StrategyError, match="confirmed_matured"):
        strategy_tools.tool_analyze_univariate_candidates(
            _inputs(dataset, workspace, mapping, sample_design_ref),
            _tool_context(settings, task),
        )


def test_univariate_tool_rejects_foreign_stale_and_nonactive_bindings(
    tmp_path: Path,
) -> None:
    (
        _settings,
        runner,
        registry,
        task,
        other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    inputs = _inputs(dataset, workspace, mapping, sample_design_ref)

    foreign = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        inputs,
        task_id=other.id,
    )
    assert foreign.ok is False
    assert "dataset not found" in (foreign.error or "")

    stale = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        {**inputs, "workspace_revision": workspace.revision - 1},
        task_id=task.id,
    )
    assert stale.ok is False
    assert "changed after user confirmation" in (stale.error or "")

    other_source = tmp_path / "other.parquet"
    pd.DataFrame({"score": [1, 2], "bad": [0, 1]}).to_parquet(
        other_source,
        index=False,
    )
    nonactive = registry.register_existing(
        other_source,
        task_id=task.id,
        role="sample",
    )
    wrong_binding = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        {
            **inputs,
            "dataset_id": nonactive.id,
            "expected_content_hash": nonactive.content_hash,
        },
        task_id=task.id,
    )
    assert wrong_binding.ok is False
    assert "not the active" in (wrong_binding.error or "")


def test_univariate_tool_rejects_explicit_sensitive_candidate(tmp_path: Path) -> None:
    (
        _settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)

    result = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        {
            **_inputs(dataset, workspace, mapping, sample_design_ref),
            "features": ["customer_id"],
        },
        task_id=task.id,
    )

    assert result.ok is False
    assert "personal-data" in (result.error or "")


def test_univariate_tool_does_not_allow_workspace_to_downgrade_sensitive_role(
    tmp_path: Path,
) -> None:
    (
        settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        _mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    with connect(settings.db_path) as conn:
        row = conn.execute(
            "SELECT columns_json FROM datasets WHERE id = ?",
            (dataset.id,),
        ).fetchone()
        profiles = json.loads(row["columns_json"])
        next(item for item in profiles if item["name"] == "customer_id")[
            "semantic_role"
        ] = "id"
        conn.execute(
            "UPDATE datasets SET columns_json = ? WHERE id = ?",
            (
                json.dumps(profiles, ensure_ascii=False, separators=(",", ":")),
                dataset.id,
            ),
        )
    downgraded_mapping = DataSemanticMapping(
        target_col="bad",
        field_roles={
            "customer_id": "categorical",
            "score": "score",
            "constant": "date",
            "segment": "categorical",
            "loan_amount": "loan_amount",
            "overdue_amount": "overdue_amount",
            "bad": "target",
        },
    )
    downgraded_workspace = DataWorkspaceRepository(settings.db_path).save(
        task.id,
        DataWorkspaceDraft(
            active_dataset_id=dataset.id,
            active_dataset_content_hash=dataset.content_hash,
            semantic_mapping=downgraded_mapping,
        ),
        expected_revision=workspace.revision,
    )
    sample_design_ref = _materialize_sample_design_ref(
        settings,
        task,
        dataset,
        downgraded_workspace,
        downgraded_mapping,
    )

    result = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        {
            **_inputs(
                dataset,
                downgraded_workspace,
                downgraded_mapping,
                sample_design_ref,
            ),
            "features": ["customer_id"],
        },
        task_id=task.id,
    )

    assert result.ok is False
    assert "personal-data" in (result.error or "")


def test_univariate_tool_projects_only_selected_and_required_columns(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    original = DataBackend.read_frame
    reads: list[list[str] | None] = []

    def track_read(self, path, *, columns=None, nrows=None):
        reads.append(None if columns is None else list(columns))
        return original(self, path, columns=columns, nrows=nrows)

    monkeypatch.setattr(DataBackend, "read_frame", track_read)

    result = strategy_tools.tool_analyze_univariate_candidates(
        {
            **_inputs(dataset, workspace, mapping, sample_design_ref),
            "features": ["score"],
            "sentinel_values": [],
        },
        _tool_context(settings, task),
    )

    assert result["feature_count"] == 1
    assert reads == [["score", "loan_amount", "overdue_amount", "bad"]]


def test_univariate_tool_auto_discovers_features_before_projected_read(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    original = DataBackend.read_frame
    reads: list[list[str] | None] = []

    def track_read(self, path, *, columns=None, nrows=None):
        reads.append(None if columns is None else list(columns))
        return original(self, path, columns=columns, nrows=nrows)

    monkeypatch.setattr(DataBackend, "read_frame", track_read)

    result = strategy_tools.tool_analyze_univariate_candidates(
        _inputs(dataset, workspace, mapping, sample_design_ref),
        _tool_context(settings, task),
    )

    assert result["feature_count"] == 2
    assert reads == [["score", "segment", "loan_amount", "overdue_amount", "bad"]]


def test_univariate_tool_rejects_sensitive_candidate_before_reading_values(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)

    def fail_read(*args, **kwargs):
        pytest.fail("sensitive feature values must not be read")

    monkeypatch.setattr(DataBackend, "read_frame", fail_read)

    with pytest.raises(StrategyError, match="personal-data"):
        strategy_tools.tool_analyze_univariate_candidates(
            {
                **_inputs(dataset, workspace, mapping, sample_design_ref),
                "features": ["customer_id"],
            },
            _tool_context(settings, task),
        )


def test_univariate_tool_rolls_back_both_artifacts_when_second_registration_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    original = TaskArtifactRepository.register_on_connection
    calls = 0

    def fail_second(self, conn, **kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("injected second artifact failure")
        return original(self, conn, **kwargs)

    monkeypatch.setattr(
        TaskArtifactRepository,
        "register_on_connection",
        fail_second,
    )

    with pytest.raises(RuntimeError, match="injected second artifact failure"):
        strategy_tools.tool_analyze_univariate_candidates(
            _inputs(dataset, workspace, mapping, sample_design_ref),
            _tool_context(settings, task),
        )
    assert _candidate_artifacts(settings, task) == []
    candidate_dir = settings.tasks_dir / task.id / "strategy_candidates"
    assert not candidate_dir.exists() or not any(candidate_dir.rglob("*"))


def test_univariate_writer_reauthenticates_sample_design_inside_transaction(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    original = strategy_tools._require_univariate_candidate_binding_on_connection

    def remove_sample_design_before_registration(conn, **kwargs):
        original(conn, **kwargs)
        conn.execute(
            "DELETE FROM task_artifacts WHERE task_id = ? AND id = ?",
            (task.id, sample_design_ref["artifact_id"]),
        )

    monkeypatch.setattr(
        strategy_tools,
        "_require_univariate_candidate_binding_on_connection",
        remove_sample_design_before_registration,
    )

    with pytest.raises(StrategyError, match="sample-design artifact disappeared"):
        strategy_tools.tool_analyze_univariate_candidates(
            _inputs(dataset, workspace, mapping, sample_design_ref),
            _tool_context(settings, task),
        )

    assert _candidate_artifacts(settings, task) == []
    # The injected deletion happened in the writer transaction and rolls back
    # together with the blocked candidate registration.
    assert any(
        record["id"] == sample_design_ref["artifact_id"]
        for record in TaskArtifactRepository(settings.db_path).list_for_task(task.id)
    )


def test_failed_registration_rolls_back_before_identical_peer_promotes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    inputs = _inputs(dataset, workspace, mapping, sample_design_ref)
    failing_db_exited = threading.Event()
    release_failing_writer = threading.Event()
    original_register = TaskArtifactRepository.register_on_connection
    original_transaction = TaskArtifactRepository.transaction
    failing_registration_calls = 0

    def fail_second_for_first_writer(self, conn, **kwargs):
        nonlocal failing_registration_calls
        if threading.current_thread().name == "failing-writer":
            failing_registration_calls += 1
            if failing_registration_calls == 2:
                raise RuntimeError("injected post-promotion registration failure")
        return original_register(self, conn, **kwargs)

    @contextmanager
    def pause_failed_writer_after_db_exit(self):
        try:
            with original_transaction(self) as conn:
                yield conn
        finally:
            if threading.current_thread().name == "failing-writer":
                failing_db_exited.set()
                if not release_failing_writer.wait(timeout=10):
                    raise RuntimeError("timed out waiting to release failed writer")

    monkeypatch.setattr(
        TaskArtifactRepository,
        "register_on_connection",
        fail_second_for_first_writer,
    )
    monkeypatch.setattr(
        TaskArtifactRepository,
        "transaction",
        pause_failed_writer_after_db_exit,
    )
    failures: dict[str, BaseException] = {}
    outputs: dict[str, dict] = {}

    def invoke(name: str) -> None:
        try:
            outputs[name] = strategy_tools.tool_analyze_univariate_candidates(
                inputs,
                _tool_context(settings, task),
            )
        except BaseException as exc:  # captured for assertions in the main thread
            failures[name] = exc

    failing = threading.Thread(
        target=invoke,
        args=("failing",),
        name="failing-writer",
    )
    peer = threading.Thread(target=invoke, args=("peer",), name="peer-writer")
    failing.start()
    assert failing_db_exited.wait(timeout=10)
    peer.start()
    peer.join(timeout=10)
    assert not peer.is_alive()
    assert "peer" not in failures
    assert outputs["peer"]["validation_status"] == "unvalidated"

    # The failed writer resumes only after the peer has committed identical
    # final paths. A rollback deferred until after DB exit would delete them.
    release_failing_writer.set()
    failing.join(timeout=10)

    assert not failing.is_alive()
    assert isinstance(failures.get("failing"), RuntimeError)
    records = _candidate_artifacts(settings, task)
    assert len(records) == 2
    for record in records:
        path = Path(record["path"])
        assert path.is_file()
        assert sha256_file(path) == record["content_hash"]


def test_univariate_tool_rechecks_source_after_analysis(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        runner,
        registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    original = strategy_tools.render_strategy_candidate_bundle

    def mutate_source(evidence, analysis):
        bundle = original(evidence, analysis)
        registry.resolve_path(dataset.id).write_bytes(b"out-of-band drift")
        return bundle

    monkeypatch.setattr(
        strategy_tools,
        "render_strategy_candidate_bundle",
        mutate_source,
    )

    with pytest.raises(StrategyError, match="source dataset changed"):
        strategy_tools.tool_analyze_univariate_candidates(
            _inputs(dataset, workspace, mapping, sample_design_ref),
            _tool_context(settings, task),
        )
    assert _candidate_artifacts(settings, task) == []


def test_univariate_tool_preserves_concurrent_workspace_update_and_writes_nothing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    original = strategy_tools.render_strategy_candidate_bundle
    repository = DataWorkspaceRepository(settings.db_path)
    changed = None

    def mutate_workspace(evidence, analysis):
        nonlocal changed
        bundle = original(evidence, analysis)
        changed_mapping = DataSemanticMapping(
            target_col="bad",
            field_roles=dict(mapping.field_roles),
            business_names={"score": "更新后的风险评分"},
        )
        changed = repository.save(
            task.id,
            DataWorkspaceDraft(
                active_dataset_id=dataset.id,
                active_dataset_content_hash=dataset.content_hash,
                semantic_mapping=changed_mapping,
            ),
            expected_revision=workspace.revision,
        )
        return bundle

    monkeypatch.setattr(
        strategy_tools,
        "render_strategy_candidate_bundle",
        mutate_workspace,
    )

    with pytest.raises(
        StrategyError,
        match="(workspace changed during analysis|DataWorkspace binding changed)",
    ):
        strategy_tools.tool_analyze_univariate_candidates(
            _inputs(dataset, workspace, mapping, sample_design_ref),
            _tool_context(settings, task),
        )
    assert _candidate_artifacts(settings, task) == []
    assert changed is not None
    assert repository.get_or_default(task.id).revision == changed.revision


def test_univariate_tool_rejects_all_unavailable_requested_methods(
    tmp_path: Path,
) -> None:
    (
        settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)

    result = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        {
            **_inputs(dataset, workspace, mapping, sample_design_ref),
            "features": ["constant"],
            "methods": ["tree"],
        },
        task_id=task.id,
    )

    assert result.ok is False
    assert "no available candidate method" in (result.error or "")
    assert _candidate_artifacts(settings, task) == []


def test_explicit_numeric_methods_keep_categorical_equal_value_candidates(
    tmp_path: Path,
) -> None:
    (
        _settings,
        runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)

    result = runner.invoke(
        ToolRef("strategy", "analyze_univariate_candidates"),
        {
            **_inputs(dataset, workspace, mapping, sample_design_ref),
            "features": ["score", "segment"],
            "methods": ["equal_width"],
        },
        task_id=task.id,
    )

    assert result.ok, result.error
    methods = {
        feature["feature"]: [
            (method["method"], method["status"]) for method in feature["methods"]
        ]
        for feature in result.output["candidate_evidence"]["analysis"]["features"]
    }
    assert methods == {
        "score": [("equal_width", "available")],
        "segment": [("categorical", "available")],
    }


def test_univariate_tool_rejects_combined_work_beyond_budget(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    monkeypatch.setattr(strategy_tools, "_UNIVARIATE_MAX_EVALUATED_CELLS", 1)
    monkeypatch.setattr(
        strategy_tools,
        "_univariate_sentinel_observed",
        lambda *args, **kwargs: pytest.fail(
            "sentinel discovery must not run before the combined budget check"
        ),
    )

    with pytest.raises(StrategyError, match="combined row/bin work budget"):
        strategy_tools.tool_analyze_univariate_candidates(
            {
                **_inputs(dataset, workspace, mapping, sample_design_ref),
                "sentinel_values": list(range(20)),
            },
            _tool_context(settings, task),
        )

    assert _candidate_artifacts(settings, task) == []


def test_identical_writers_serialize_before_promoting_final_artifacts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        settings,
        _runner,
        _registry,
        task,
        _other,
        dataset,
        workspace,
        mapping,
        sample_design_ref,
    ) = _runtime(tmp_path)
    inputs = _inputs(dataset, workspace, mapping, sample_design_ref)
    late_writer_entered = threading.Event()
    release_late_writer = threading.Event()
    committing_writer_rendered = threading.Event()
    committing_writer_promoted = threading.Event()
    original_require = (
        strategy_tools._require_univariate_candidate_binding_on_connection
    )
    original_render = strategy_tools.render_strategy_candidate_bundle
    original_promote = strategy_tools.ArtifactUnitOfWork.promote_all

    def gated_require(conn, **kwargs):
        if threading.current_thread().name == "late-writer":
            late_writer_entered.set()
            if not release_late_writer.wait(timeout=10):
                raise RuntimeError("timed out waiting to release late writer")
            raise StrategyError("injected late candidate binding failure")
        return original_require(conn, **kwargs)

    def tracked_render(evidence, analysis):
        bundle = original_render(evidence, analysis)
        if threading.current_thread().name == "committing-writer":
            committing_writer_rendered.set()
        return bundle

    def tracked_promote(self):
        result = original_promote(self)
        if threading.current_thread().name == "committing-writer":
            committing_writer_promoted.set()
        return result

    monkeypatch.setattr(
        strategy_tools,
        "_require_univariate_candidate_binding_on_connection",
        gated_require,
    )
    monkeypatch.setattr(
        strategy_tools,
        "render_strategy_candidate_bundle",
        tracked_render,
    )
    monkeypatch.setattr(
        strategy_tools.ArtifactUnitOfWork,
        "promote_all",
        tracked_promote,
    )
    failures: dict[str, BaseException] = {}
    outputs: dict[str, dict] = {}

    def invoke(name: str) -> None:
        try:
            outputs[name] = strategy_tools.tool_analyze_univariate_candidates(
                inputs,
                _tool_context(settings, task),
            )
        except BaseException as exc:  # captured for assertions in the main thread
            failures[name] = exc

    late = threading.Thread(target=invoke, args=("late",), name="late-writer")
    committing = threading.Thread(
        target=invoke,
        args=("committing",),
        name="committing-writer",
    )
    late.start()
    assert late_writer_entered.wait(timeout=10)
    committing.start()
    assert committing_writer_rendered.wait(timeout=10)
    # The peer may finish computation/staging, but it must not replace a final
    # path until the current SQLite writer transaction releases its lock.
    assert not committing_writer_promoted.wait(timeout=1)
    release_late_writer.set()
    late.join(timeout=10)
    committing.join(timeout=10)

    assert not late.is_alive()
    assert not committing.is_alive()
    assert isinstance(failures.get("late"), StrategyError)
    assert "committing" not in failures
    assert outputs["committing"]["validation_status"] == "unvalidated"
    records = _candidate_artifacts(settings, task)
    assert len(records) == 2
    for record in records:
        path = Path(record["path"])
        assert path.is_file()
        assert sha256_file(path) == record["content_hash"]
