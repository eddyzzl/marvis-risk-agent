from __future__ import annotations

import copy
from io import BytesIO
import json
from zipfile import ZipFile

from openpyxl import load_workbook
import pandas as pd
from PIL import Image
import pytest

from marvis.feature.weighted_rule_tree import build_weighted_rule_tree
from marvis.output.automatic_tree_report import (
    AUTOMATIC_TREE_REPORT_SHEET_NAMES,
    render_automatic_tree_report_xlsx,
)
from marvis.output.automatic_tree_visual import AUTOMATIC_TREE_PNG_RENDERER_VERSION
from marvis.packs.strategy.automatic_tree_asset import (
    AutomaticTreeAssetError,
    build_automatic_tree_asset,
)


HASH_A = "a" * 64
HASH_B = "b" * 64
HASH_C = "c" * 64
HASH_D = "d" * 64


def _reverse_mapping_order(value):
    if isinstance(value, dict):
        return {
            key: _reverse_mapping_order(item)
            for key, item in reversed(list(value.items()))
        }
    if isinstance(value, list):
        return [_reverse_mapping_order(item) for item in value]
    return value


def _asset(
    *, formula_payloads: bool = False, weighted_amount_metrics: bool = False
) -> dict:
    feature = "=formula-feature" if formula_payloads else "风险特征"
    frame_data = {
        feature: [0.0, 1.0, 2.0, 3.0, 4.0, 5.0],
        "bad": [0, 0, 0, 1, 1, 1],
    }
    if weighted_amount_metrics:
        frame_data.update(
            {
                "weight": [1.0, 1.5, 2.0, 2.5, 3.0, 3.5],
                "loan_amount": [100.0, 150.0, 200.0, 250.0, 300.0, 350.0],
                "overdue_amount": [0.0, 0.0, 0.0, 25.0, 45.0, 70.0],
            }
        )
    frame = pd.DataFrame(frame_data)
    tree = build_weighted_rule_tree(
        frame,
        feature_cols=[feature],
        target_col="bad",
        sample_weight_col="weight" if weighted_amount_metrics else None,
        directions={feature: "decreasing"},
        max_depth=1,
        min_leaf_count=1,
        loan_amount_col="loan_amount" if weighted_amount_metrics else None,
        overdue_amount_col="overdue_amount" if weighted_amount_metrics else None,
    )
    return build_automatic_tree_asset(
        tree,
        task_id="\x01=2+2" if formula_payloads else "task-automatic-tree-report",
        dataset_id="-danger" if formula_payloads else "dataset-labelled",
        dataset_content_hash=HASH_A,
        workspace_revision=3,
        workspace_generation=7,
        semantic_mapping_hash=HASH_B,
        registry_metadata_hash=HASH_C,
        sample_context_hash=HASH_D,
        source_refs=(
            ["@SUM(1,1)", "workspace:report:3"]
            if formula_payloads
            else ["dataset:dataset-labelled", "workspace:report:3"]
        ),
    )


def _rows_by_key(sheet) -> dict[object, object]:
    return {
        row[0].value: row[1].value
        for row in sheet.iter_rows(min_row=2)
        if row[0].value is not None
    }


def test_report_has_exact_sheets_projected_content_and_embedded_tree_png() -> None:
    asset = _asset()

    raw = render_automatic_tree_report_xlsx(asset)
    workbook = load_workbook(BytesIO(raw), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == AUTOMATIC_TREE_REPORT_SHEET_NAMES
        summary = _rows_by_key(workbook["Summary"])
        assert summary["asset_id"] == asset["asset_id"]
        assert summary["asset_hash"] == asset["asset_hash"]
        assert summary["tree_node_count"] == asset["tree_result"]["tree"]["node_count"]
        assert summary["tree_leaf_count"] == asset["tree_result"]["tree"]["leaf_count"]

        assert len(workbook["Tree"]._images) == 1
        image = workbook["Tree"]._images[0]
        image_bytes = image._data()
        assert image_bytes.startswith(b"\x89PNG\r\n\x1a\n")
        with Image.open(BytesIO(image_bytes)) as png:
            assert png.info["MARVIS Renderer"] == AUTOMATIC_TREE_PNG_RENDERER_VERSION

        node_rows = list(workbook["Nodes"].iter_rows(min_row=2, values_only=True))
        assert len(node_rows) == asset["tree_result"]["tree"]["node_count"]
        leaf_rows = list(workbook["Leaf Rules"].iter_rows(min_row=2, values_only=True))
        assert len(leaf_rows) == asset["tree_result"]["tree"]["leaf_count"]
        leaf_headers = [
            cell.value for cell in next(workbook["Leaf Rules"].iter_rows(max_row=1))
        ]
        first_leaf = dict(zip(leaf_headers, leaf_rows[0], strict=True))
        first_fragment = asset["fragments"][0]
        first_rule = asset["tree_result"]["rules"][0]
        assert first_leaf["Rule ID"] == first_fragment["rule_id"]
        assert first_leaf["Condition JSON"] == json.dumps(
            first_fragment["condition"],
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        assert first_leaf["Clauses JSON"] == json.dumps(
            first_rule["clauses"],
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        assert first_leaf["Requirements JSON"] == "[]"
        assert first_leaf["Effect ID"] == first_fragment["effect_id"]
        assert (
            first_leaf["Unweighted total"]
            == first_fragment["metrics"]["unweighted"]["total"]
        )
        metric_fields = (
            "total",
            "good",
            "bad",
            "bad_rate",
            "share",
            "bad_capture",
            "lift",
            "loan_amount_total",
            "loan_amount_coverage_count",
            "loan_amount_coverage",
            "loan_amount_coverage_rate",
            "overdue_amount_total",
            "overdue_amount_coverage_count",
            "overdue_amount_coverage",
            "overdue_amount_coverage_rate",
            "amount_pair_coverage_count",
            "amount_pair_coverage",
            "amount_pair_coverage_rate",
            "paired_loan_amount_total",
            "paired_overdue_amount_total",
            "overdue_rate",
        )
        expected_leaf = {
            "Leaf ID": first_fragment["leaf_id"],
            "Fragment ID": first_fragment["fragment_id"],
            "Fragment Hash": first_fragment["fragment_hash"],
            "Rule ID": first_fragment["rule_id"],
            "Condition JSON": json.dumps(
                first_fragment["condition"],
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
            "Clauses JSON": json.dumps(
                first_rule["clauses"],
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
            "Requirements JSON": "[]",
            "Effect ID": first_fragment["effect_id"],
            **{
                f"Unweighted {field}": _expected_xlsx_value(
                    first_fragment["metrics"]["unweighted"].get(field)
                )
                for field in metric_fields
            },
            "Weighted Status": first_fragment["metrics"]["weighted"]["status"],
            **{f"Weighted {field}": None for field in metric_fields},
        }
        assert first_leaf == expected_leaf

        delivery = list(workbook["Delivery"].iter_rows(min_row=2, values_only=True))
        assert [row[0] for row in delivery] == [
            "JSON",
            "Python",
            "SQL",
            "SVG",
            "PNG",
            "XLSX",
        ]
        assert all(len(row) == 3 for row in delivery)
    finally:
        workbook.close()


def test_report_is_formula_injection_safe_and_replaces_illegal_controls() -> None:
    raw = render_automatic_tree_report_xlsx(_asset(formula_payloads=True))
    workbook = load_workbook(BytesIO(raw), data_only=False)
    try:
        text_cells = [
            cell
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        assert any(cell.value == "'=formula-feature" for cell in text_cells)
        assert any(cell.value == "'@SUM(1,1)" for cell in text_cells)
        assert any("\\u0001" in cell.value for cell in text_cells)
        assert all(cell.data_type != "f" for cell in text_cells)
        assert not any(
            any(
                ord(character) < 32 and character not in "\t\n\r"
                for character in cell.value
            )
            for cell in text_cells
        )
    finally:
        workbook.close()


def test_report_projects_weighted_and_amount_metrics_without_recalculation() -> None:
    asset = _asset(weighted_amount_metrics=True)
    raw = render_automatic_tree_report_xlsx(asset)
    workbook = load_workbook(BytesIO(raw), data_only=False)
    try:
        rows = list(workbook["Nodes"].iter_rows(values_only=True))
        headers = rows[0]
        root_row = dict(zip(headers, rows[1], strict=True))
        root_metrics = asset["tree_result"]["tree"]["nodes"][0]["metrics"]

        assert root_row["Unweighted loan_amount_total"] == repr(
            root_metrics["unweighted"]["loan_amount_total"]
        )
        assert root_row["Unweighted overdue_rate"] == repr(
            root_metrics["unweighted"]["overdue_rate"]
        )
        assert root_row["Weighted Status"] == root_metrics["weighted"]["status"]
        assert root_row["Weighted total"] == repr(root_metrics["weighted"]["total"])
        assert root_row["Weighted bad_rate"] == repr(
            root_metrics["weighted"]["bad_rate"]
        )
    finally:
        workbook.close()


def test_report_repeats_byte_for_byte_with_fixed_zip_metadata_and_order() -> None:
    asset = _asset()
    reordered = _reverse_mapping_order(asset)

    first = render_automatic_tree_report_xlsx(asset)
    second = render_automatic_tree_report_xlsx(asset)

    assert first == second
    assert first == render_automatic_tree_report_xlsx(reordered)
    with ZipFile(BytesIO(first)) as archive:
        members = archive.infolist()
        assert [member.filename for member in members] == sorted(
            member.filename for member in members
        )
        assert all(member.date_time == (1980, 1, 1, 0, 0, 0) for member in members)


def test_report_bytes_ignore_environment_font_paths(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    asset = _asset()
    baseline = render_automatic_tree_report_xlsx(asset)
    monkeypatch.setenv("FONTCONFIG_FILE", "/nonexistent/marvis-fonts.conf")
    monkeypatch.setenv("FONTCONFIG_PATH", "/nonexistent/marvis-fonts")

    assert render_automatic_tree_report_xlsx(asset) == baseline


def test_report_fails_closed_on_tampered_asset() -> None:
    tampered = copy.deepcopy(_asset())
    tampered["identity"]["workspace_revision"] += 1

    with pytest.raises(AutomaticTreeAssetError):
        render_automatic_tree_report_xlsx(tampered)


def _expected_xlsx_value(value: object) -> object:
    return repr(value) if isinstance(value, float) else value
