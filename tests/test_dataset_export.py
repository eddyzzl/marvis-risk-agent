from __future__ import annotations

import csv
import json
from datetime import date
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from openpyxl import load_workbook

import marvis.data.dataset_export as dataset_export
from marvis.data.dataset_export import (
    DatasetExportBudgetError,
    DatasetExportConfig,
    DatasetExportConfigError,
    DatasetExportInputError,
    export_dataset,
)


def _write_parquet(path: Path, table: pa.Table) -> None:
    pq.write_table(table, path, row_group_size=2)


def _representative_table() -> pa.Table:
    return pa.Table.from_arrays(
        [
            pa.array(["张三", " \t=2+2", None], type=pa.string()),
            pa.array(["0013800000000", "+8613900000000", "普通"], type=pa.string()),
            pa.array([9007199254740993, 15, None], type=pa.uint64()),
            pa.array(
                [Decimal("123.45"), Decimal("9999999999999999.99"), None],
                type=pa.decimal128(20, 2),
            ),
            pa.array([date(2026, 7, 1), date(2026, 7, 2), None], type=pa.date32()),
            pa.array([1.5, float("nan"), float("inf")], type=pa.float64()),
        ],
        names=["客户姓名", "手机号", "order_id", "金额", "申请日", "ratio"],
    )


def test_csv_export_is_streamed_safe_exact_and_json_evidence_is_finite(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.parquet"
    output = tmp_path / "export.csv"
    _write_parquet(source, _representative_table())

    def forbid_full_table(*_args, **_kwargs):
        raise AssertionError("full-table parquet reads are forbidden")

    monkeypatch.setattr(pq, "read_table", forbid_full_table)
    result = export_dataset(
        source,
        output,
        format="csv",
        temp_directory=tmp_path / "scratch",
        text_columns=("手机号",),
        config=DatasetExportConfig(batch_size=2),
    )

    assert output.read_bytes().startswith(b"\xef\xbb\xbf")
    with output.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[0] == ["客户姓名", "手机号", "order_id", "金额", "申请日", "ratio"]
    assert rows[1] == [
        "张三",
        "0013800000000",
        "'9007199254740993",
        "123.45",
        "2026-07-01",
        "1.5",
    ]
    assert rows[2] == [
        "' \t=2+2",
        "'+8613900000000",
        "15",
        "'9999999999999999.99",
        "2026-07-02",
        "NaN",
    ]
    assert rows[3] == ["", "普通", "", "", "", "Infinity"]

    assert result["schema_version"] == "dataset-export-result.v1"
    assert result["producer"] == {
        "name": "marvis.data.dataset_export",
        "version": "1",
    }
    assert result["output"]["format"] == "csv"
    assert result["output"]["row_count"] == 3
    assert result["output"]["column_count"] == 6
    assert result["output"]["size_bytes"] == output.stat().st_size
    assert len(result["output"]["content_hash"]) == 64
    assert result["output"]["hash_algorithm"] == "sha256"
    assert [column["name"] for column in result["output"]["columns"]] == rows[0]
    assert result["options"]["encoding"] == "utf-8-sig"
    assert result["options"]["null_representation"] == "empty_cell"
    assert result["options"]["text_columns"] == ["手机号"]
    assert result["safety"]["formula_cells_escaped"] == 2
    assert result["safety"]["large_integer_cells_as_text"] == 1
    assert result["safety"]["high_precision_decimal_cells_as_text"] == 1
    assert result["safety"]["non_finite_cells_as_text"] == 2
    json.dumps(result, allow_nan=False)


def test_csv_export_is_byte_deterministic_and_preserves_batch_order(tmp_path: Path) -> None:
    source = tmp_path / "source.parquet"
    first = tmp_path / "first.csv"
    second = tmp_path / "second.csv"
    _write_parquet(source, pa.table({"x": list(range(9)), "label": list("abcdefghi")}))

    first_result = export_dataset(
        source,
        first,
        format="csv",
        temp_directory=tmp_path / "scratch",
        config=DatasetExportConfig(batch_size=2),
    )
    second_result = export_dataset(
        source,
        second,
        format="csv",
        temp_directory=tmp_path / "scratch",
        config=DatasetExportConfig(batch_size=3),
    )

    assert first.read_bytes() == second.read_bytes()
    assert first_result["output"]["content_hash"] == second_result["output"]["content_hash"]
    with first.open("r", encoding="utf-8-sig", newline="") as handle:
        assert [int(row[0]) for row in list(csv.reader(handle))[1:]] == list(range(9))


def test_xlsx_export_uses_text_cells_for_sensitive_and_precision_risky_values(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.parquet"
    output = tmp_path / "export.xlsx"
    table = pa.Table.from_arrays(
        [
            pa.array([" \n@SUM(A1:A2)", "李四"], type=pa.string()),
            pa.array(["0013800000000", None], type=pa.string()),
            pa.array([9007199254740993, None], type=pa.uint64()),
            pa.array([Decimal("1234567890123456.78"), None], type=pa.decimal128(20, 2)),
            pa.array([date(2026, 7, 19), None], type=pa.date32()),
        ],
        names=["=姓名", "phone", "id", "金额", "日期"],
    )
    _write_parquet(source, table)

    result = export_dataset(
        source,
        output,
        format="xlsx",
        temp_directory=tmp_path / "scratch",
        text_columns=("phone", "id"),
        config=DatasetExportConfig(batch_size=1),
    )

    workbook = load_workbook(output, read_only=True, data_only=False)
    sheet = workbook["data"]
    rows = list(sheet.iter_rows(values_only=False))
    assert [cell.value for cell in rows[0]] == ["'=姓名", "phone", "id", "金额", "日期"]
    assert rows[1][0].value == "' \n@SUM(A1:A2)"
    assert rows[1][0].data_type == "s"
    assert rows[1][1].value == "0013800000000"
    assert rows[1][1].data_type == "s"
    assert rows[1][2].value == "9007199254740993"
    assert rows[1][2].data_type == "s"
    assert rows[1][3].value == "1234567890123456.78"
    assert rows[1][3].data_type == "s"
    assert rows[1][4].value.date() == date(2026, 7, 19)
    assert [cell.value for cell in rows[2]] == ["李四"]
    assert result["options"]["workbook_mode"] == "write_only"
    assert result["safety"]["formula_cells_escaped"] == 2
    assert result["safety"]["large_integer_cells_as_text"] == 1
    assert result["safety"]["high_precision_decimal_cells_as_text"] == 1
    json.dumps(result, allow_nan=False)


def test_formula_guard_detects_leading_control_characters_in_csv(tmp_path: Path) -> None:
    source = tmp_path / "source.parquet"
    output = tmp_path / "export.csv"
    _write_parquet(source, pa.table({"value": ["\x01=cmd", "\r+cmd", "safe"]}))

    result = export_dataset(
        source,
        output,
        format="csv",
        temp_directory=tmp_path / "scratch",
    )

    with output.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[1][0] == "'\x01=cmd"
    assert rows[2][0] == "'\r+cmd"
    assert rows[3][0] == "safe"
    assert result["safety"]["formula_cells_escaped"] == 2


@pytest.mark.parametrize("format_name,suffix", [("csv", ".csv"), ("xlsx", ".xlsx")])
def test_export_rejects_existing_output_without_changing_it(
    tmp_path: Path,
    format_name: str,
    suffix: str,
) -> None:
    source = tmp_path / "source.parquet"
    output = tmp_path / f"existing{suffix}"
    _write_parquet(source, pa.table({"x": [1]}))
    output.write_bytes(b"keep-me")

    with pytest.raises(DatasetExportInputError, match="already exists"):
        export_dataset(
            source,
            output,
            format=format_name,
            temp_directory=tmp_path / "scratch",
        )

    assert output.read_bytes() == b"keep-me"


def test_export_requires_readable_parquet_matching_format_and_known_text_columns(
    tmp_path: Path,
) -> None:
    csv_source = tmp_path / "source.csv"
    csv_source.write_text("x\n1\n", encoding="utf-8")
    with pytest.raises(DatasetExportInputError, match="normalized Parquet"):
        export_dataset(
            csv_source,
            tmp_path / "out.csv",
            format="csv",
            temp_directory=tmp_path / "scratch",
        )

    source = tmp_path / "source.parquet"
    _write_parquet(source, pa.table({"x": [1]}))
    with pytest.raises(DatasetExportInputError, match="suffix"):
        export_dataset(
            source,
            tmp_path / "out.xlsx",
            format="csv",
            temp_directory=tmp_path / "scratch",
        )
    with pytest.raises(DatasetExportInputError, match="unknown text column"):
        export_dataset(
            source,
            tmp_path / "out.csv",
            format="csv",
            temp_directory=tmp_path / "scratch",
            text_columns=("missing",),
        )
    with pytest.raises(DatasetExportInputError, match="csv or xlsx"):
        export_dataset(
            source,
            tmp_path / "out.csv",
            format="json",
            temp_directory=tmp_path / "scratch",
        )


def test_budget_failures_happen_without_output_or_scratch_residue(tmp_path: Path) -> None:
    source = tmp_path / "source.parquet"
    output = tmp_path / "out.csv"
    _write_parquet(source, pa.table({"x": [1, 2], "y": [3, 4]}))

    with pytest.raises(DatasetExportBudgetError) as rows_error:
        export_dataset(
            source,
            output,
            format="csv",
            temp_directory=tmp_path / "scratch",
            config=DatasetExportConfig(max_rows=1),
        )
    assert rows_error.value.to_detail() == {
        "kind": "dataset_export_budget_exceeded",
        "dimension": "rows",
        "actual": 2,
        "limit": 1,
    }
    assert not output.exists()

    with pytest.raises(DatasetExportBudgetError) as columns_error:
        export_dataset(
            source,
            output,
            format="csv",
            temp_directory=tmp_path / "scratch",
            config=DatasetExportConfig(max_columns=1),
        )
    assert columns_error.value.dimension == "columns"
    assert not output.exists()
    assert not list(tmp_path.glob(".out.csv.*.tmp"))


def test_xlsx_excel_row_limit_counts_header_and_never_truncates(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.parquet"
    output = tmp_path / "out.xlsx"
    _write_parquet(source, pa.table({"x": [1]}))

    class OversizedParquetFile:
        metadata = SimpleNamespace(num_rows=1_048_576)
        schema_arrow = pa.schema([("x", pa.int64())])

        def __init__(self, _path: Path) -> None:
            pass

    monkeypatch.setattr(dataset_export.pq, "ParquetFile", OversizedParquetFile)
    with pytest.raises(DatasetExportBudgetError) as exc_info:
        export_dataset(
            source,
            output,
            format="xlsx",
            temp_directory=tmp_path / "scratch",
        )
    assert exc_info.value.dimension == "xlsx_rows"
    assert exc_info.value.actual == 1_048_576
    assert exc_info.value.limit == 1_048_575
    assert not output.exists()


def test_output_size_failure_and_atomic_replace_failure_leave_no_residue(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.parquet"
    output = tmp_path / "out.csv"
    _write_parquet(source, pa.table({"x": ["some exported content"]}))

    with pytest.raises(DatasetExportBudgetError) as size_error:
        export_dataset(
            source,
            output,
            format="csv",
            temp_directory=tmp_path / "scratch",
            config=DatasetExportConfig(max_output_bytes=1),
        )
    assert size_error.value.dimension == "output_bytes"
    assert not output.exists()
    assert not list(tmp_path.glob(".out.csv.*.tmp"))

    def fail_replace(_source: Path, _destination: Path) -> None:
        raise OSError("simulated publication failure")

    monkeypatch.setattr(dataset_export.os, "replace", fail_replace)
    with pytest.raises(OSError, match="publication failure"):
        export_dataset(
            source,
            output,
            format="csv",
            temp_directory=tmp_path / "scratch",
        )
    assert not output.exists()
    assert not list(tmp_path.glob(".out.csv.*.tmp"))


def test_config_has_positive_hard_caps() -> None:
    with pytest.raises(DatasetExportConfigError, match="positive integer"):
        DatasetExportConfig(batch_size=0)
    with pytest.raises(DatasetExportConfigError, match="at most 16384"):
        DatasetExportConfig(max_columns=16_385)
    with pytest.raises(DatasetExportConfigError, match="positive integer"):
        DatasetExportConfig(max_rows=True)
