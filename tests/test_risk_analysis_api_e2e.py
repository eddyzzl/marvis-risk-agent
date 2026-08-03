from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from marvis.agent_memory.store import AgentMemoryStore
from marvis.app import create_app


pytestmark = pytest.mark.e2e


@pytest.fixture
def client(tmp_path: Path) -> TestClient:
    return TestClient(create_app(tmp_path))


def _post_message(client: TestClient, task_id: str, content: str):
    return client.post(
        f"/api/tasks/{task_id}/agent/messages",
        json={"content": content},
    )


def test_agent_risk_analysis_upload_to_report_download_and_memory(
    client: TestClient,
    monkeypatch,
):
    # Intake and the report tool are deterministic. The test bypasses only the
    # configured-model lookup so the real API/plan/tool/artifact path can run in
    # an isolated workspace without a network LLM.
    monkeypatch.setattr(
        "marvis.routers.validation_agent.resolve_driver_agent_client",
        lambda request, task, payload: None,
    )
    created = client.post(
        "/api/tasks",
        json={
            "model_name": "白条VTG分析",
            "validator": "qa",
            "source_dir": "",
            "task_type": "vintage",
            "run_mode": "agent",
        },
    )
    assert created.status_code == 200, created.text
    task_id = created.json()["id"]

    started = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert started.status_code == 202, started.text
    assert "你想分析什么" in started.json()["messages"][-1]["content"]

    selected = _post_message(client, task_id, "做VTG终值与年化不良")
    assert selected.status_code == 202, selected.text
    assert "canonical VTG input" in selected.json()["messages"][-1]["content"]

    csv_bytes = (
        "product,as_of_date,cohort,amount_unit,disbursement_amount,mob14_bad_rate,turnover,"
        "terminal_bad_rate,long_term_recovery_rate,auxiliary_terminal_bad_rate,"
        "selection_rule,previous_annualized_bad_rate\n"
        "白条,2026-05-31,2026-01,元,1000,0.05,2,,0.2,0.035,min_auxiliary_recovery,0.06\n"
        "白条,2026-05-31,2026-02,元,2000,0.04,2.5,0.03,,,,0.08\n"
    ).encode("utf-8")
    uploaded = client.post(
        f"/api/tasks/{task_id}/datasets/upload",
        files={"file": ("vtg_summary.csv", csv_bytes, "text/csv")},
        data={"role": "sample"},
    )
    assert uploaded.status_code == 201, uploaded.text

    planned = _post_message(client, task_id, "材料已上传，请检查字段并继续")
    assert planned.status_code == 202, planned.text
    assert planned.json()["messages"][-1]["metadata"]["kind"] == "plan_overview"

    completed = _post_message(client, task_id, "开始")
    assert completed.status_code == 202, completed.text
    done = completed.json()["messages"][-1]
    assert "计划已全部完成" in done["content"]
    assert "重要发现" in done["content"]
    report = done["metadata"]["risk_analysis_report"]
    assert report["analysis_kind"] == "vtg_terminal"
    assert report["source_row_count"] == 2
    assert report["row_count"] == 2
    assert report["product_scope"] == ["白条"]
    assert report["headline_metrics"]["annualized_bad_rate"] == pytest.approx(
        (35 + 60) / (500 + 800)
    )
    assert done["metadata"]["report_download"] == {
        "label": "下载风险分析报告",
        "download_url": f"/api/tasks/{task_id}/driver-report/download",
    }

    downloaded = client.get(f"/api/tasks/{task_id}/driver-report/download")
    assert downloaded.status_code == 200, downloaded.text
    assert downloaded.content.startswith(b"PK")
    workbook = load_workbook(BytesIO(downloaded.content), data_only=False)
    assert workbook.sheetnames == ["结论摘要", "明细结果", "口径与假设", "数据质量"]
    assert workbook["结论摘要"]["A1"].value == "VTG 终值与年化不良测算报告"

    memories = AgentMemoryStore(client.app.state.settings.db_path).list_entries(
        memory_type="risk_analysis_experience"
    )
    assert len(memories) == 1
    assert memories[0].source_task_id == task_id
    assert memories[0].payload["product_scope"] == ["白条"]
    assert memories[0].payload["report_file"] == "risk_analysis_report.xlsx"


def test_agent_profitability_analysis_computes_weighted_net_yield(
    client: TestClient,
    monkeypatch,
):
    monkeypatch.setattr(
        "marvis.routers.validation_agent.resolve_driver_agent_client",
        lambda request, task, payload: None,
    )
    created = client.post(
        "/api/tasks",
        json={
            "model_name": "白条收益测算",
            "validator": "qa",
            "source_dir": "",
            "task_type": "vintage",
            "run_mode": "agent",
        },
    )
    task_id = created.json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    selected = _post_message(client, task_id, "做收益测算")
    assert "canonical economics" in selected.json()["messages"][-1]["content"]

    csv_bytes = (
        "product,asset_class,as_of_period,weight,weight_basis,customer_rate,risk_cost_rate,"
        "funding_cost_rate,interest_loss_rate,revenue_share_rate,acquisition_cost_rate,"
        "data_cost_rate,payment_cost_rate,collection_cost_rate,other_cost_rate,tax_rate\n"
        "白条消费,生息资产,2025-12-04,0.35,average_balance,0.12,0.04,0.02,0.0024,0.02,0,0.001,0.0005,0.0005,0,0.005\n"
        "白条消费,免息资产,2025-12-04,0.65,average_balance,0.08,0.03,0.02,0.0012,0.01,0,0.001,0.0005,0.0005,0,0.003\n"
    ).encode("utf-8")
    uploaded = client.post(
        f"/api/tasks/{task_id}/datasets/upload",
        files={"file": ("profitability.csv", csv_bytes, "text/csv")},
        data={"role": "sample"},
    )
    assert uploaded.status_code == 201, uploaded.text

    planned = _post_message(client, task_id, "材料已上传")
    assert planned.json()["messages"][-1]["metadata"]["kind"] == "plan_overview"
    completed = _post_message(client, task_id, "开始")
    report = completed.json()["messages"][-1]["metadata"]["risk_analysis_report"]

    assert report["analysis_kind"] == "profitability"
    assert report["source_row_count"] == 2
    assert report["row_count"] == 2
    assert report["as_of_period"] == "2025-12-04"
    assert report["headline_metrics"]["lowest_net_yield"] == pytest.approx(
        0.35 * 0.0306 + 0.65 * 0.0138
    )
    assert report["headline_metrics"]["negative_product_count"] == 0


def test_profitability_material_confirmation_with_no_missing_text_stays_in_intake(
    client: TestClient,
    monkeypatch,
):
    monkeypatch.setattr(
        "marvis.routers.validation_agent.resolve_driver_agent_client",
        lambda request, task, payload: None,
    )
    created = client.post(
        "/api/tasks",
        json={
            "model_name": "收益测算路由回归",
            "validator": "qa",
            "source_dir": "",
            "task_type": "vintage",
            "run_mode": "agent",
        },
    )
    task_id = created.json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    _post_message(client, task_id, "做收益测算")

    csv_bytes = (
        "product,asset_class,as_of_period,weight,weight_basis,customer_rate,risk_cost_rate,"
        "funding_cost_rate,interest_loss_rate,revenue_share_rate,acquisition_cost_rate,"
        "data_cost_rate,payment_cost_rate,collection_cost_rate,other_cost_rate,tax_rate\n"
        "白条,生息资产,2026-06,1,average_balance,0.12,0.04,0.02,0.002,0.01,0,0,0,0,0,0.005\n"
    ).encode("utf-8")
    uploaded = client.post(
        f"/api/tasks/{task_id}/datasets/upload",
        files={"file": ("profitability.csv", csv_bytes, "text/csv")},
        data={"role": "sample"},
    )
    assert uploaded.status_code == 201, uploaded.text

    planned = _post_message(
        client,
        task_id,
        "各成本字段均显式给值，无缺失静默补 0。材料已上传，请检查字段并生成计划。",
    )

    last = planned.json()["messages"][-1]
    assert last["metadata"]["kind"] == "plan_overview"
    assert last["metadata"]["plan_id"]
    assert last["metadata"].get("code") != "active_dataset_required"


def test_standard_vintage_material_scope_does_not_trigger_dataset_transform(
    client: TestClient,
    monkeypatch,
):
    monkeypatch.setattr(
        "marvis.routers.validation_agent.resolve_driver_agent_client",
        lambda request, task, payload: None,
    )
    created = client.post(
        "/api/tasks",
        json={
            "model_name": "Vintage 材料路由回归",
            "validator": "qa",
            "source_dir": "",
            "task_type": "vintage",
            "run_mode": "agent",
        },
    )
    task_id = created.json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    selected = _post_message(
        client,
        task_id,
        "做标准 Vintage，bad 是 incremental，不是 snapshot",
    )
    assert "Vintage panel" in selected.json()["messages"][-1]["content"]

    uploaded = client.post(
        f"/api/tasks/{task_id}/datasets/upload",
        files={
            "file": (
                "vintage_panel.csv",
                (
                    "account_id,cohort,mob,bad,balance,channel\n"
                    "A,2026-01,0,0,1000,app\n"
                    "A,2026-01,1,1,900,app\n"
                ).encode("utf-8"),
                "text/csv",
            )
        },
        data={"role": "sample"},
    )
    assert uploaded.status_code == 201, uploaded.text

    planned = _post_message(
        client,
        task_id,
        (
            "材料已上传，覆盖表内全部 cohort、channel、MOB，"
            "不做额外筛选；bad 是 incremental，不是 snapshot。"
        ),
    )

    last = planned.json()["messages"][-1]
    assert last["metadata"]["kind"] == "plan_overview"
    assert last["metadata"]["plan_id"]
    assert last["metadata"].get("intent") != "dataset_transform"
    assert last["metadata"].get("code") != "active_dataset_required"

    completed = _post_message(client, task_id, "开始")
    done = completed.json()["messages"][-1]
    assert "计划已全部完成" in done["content"]
    assert "Vintage 曲线完成" in done["content"]
    assert {table["title"] for table in done["metadata"]["tables"]} >= {
        "Vintage 累计坏账率",
    }


def test_agent_profitability_derives_sample_style_raw_cost_bridge(
    client: TestClient,
    monkeypatch,
):
    monkeypatch.setattr(
        "marvis.routers.validation_agent.resolve_driver_agent_client",
        lambda request, task, payload: None,
    )
    created = client.post(
        "/api/tasks",
        json={
            "model_name": "白条收益原始驱动测算",
            "validator": "qa",
            "source_dir": "",
            "task_type": "vintage",
            "run_mode": "agent",
        },
    )
    task_id = created.json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    _post_message(client, task_id, "做收益测算，复核合同分润、数据漏斗与税费桥")

    csv_bytes = (
        "product,as_of_period,asset_class,weight,weight_basis,customer_rate,"
        "acquisition_cost_rate,payment_cost_rate,collection_cost_rate,funding_cost_rate,"
        "other_cost_rate,terminal_vintage_rate,risk_turnover,loss_timing_factor,"
        "profit_share_ratio,per_application_cost,credit_approval_rate,draw_initiation_rate,"
        "draw_approval_rate,average_ticket,data_annualization_factor,customer_stage,"
        "transaction_weight,tax_method,"
        "tax_inclusive_divisor,tax_combined_rate,amount_unit\n"
        "白条消费,2025-12-04,生息资产,1,average_balance,0.12,0,0.001,0.001,0.02,0,"
        "0.02,2,0.5,0.25,2.5,0.5,0.8,0.9,10000,1,first,1,"
        "sample_net_revenue_vat_surcharge,1.06,0.0672,元\n"
        "白条消费,2025-12-04,生息资产,1,average_balance,0.12,0,0.001,0.001,0.02,0,"
        "0.02,2,0.5,0.25,1.0,0.8,0.9,0.95,12000,1,repeat,9,"
        "sample_net_revenue_vat_surcharge,1.06,0.0672,元\n"
    ).encode("utf-8")
    uploaded = client.post(
        f"/api/tasks/{task_id}/datasets/upload",
        files={"file": ("白条收益原始驱动.csv", csv_bytes, "text/csv")},
        data={"role": "sample"},
    )
    assert uploaded.status_code == 201, uploaded.text

    planned = _post_message(client, task_id, "材料已上传")
    assert planned.json()["messages"][-1]["metadata"]["kind"] == "plan_overview"
    completed = _post_message(client, task_id, "开始")
    report = completed.json()["messages"][-1]["metadata"]["risk_analysis_report"]

    risk_cost = 0.02 * 2
    interest_loss = 0.12 * risk_cost * 0.5
    revenue_share = (0.12 - interest_loss) * 0.25
    first_data_cost = 2.5 / (0.5 * 0.8 * 0.9 * 10000) * 1
    repeat_data_cost = 1.0 / (0.8 * 0.9 * 0.95 * 12000) * 1
    data_cost = (first_data_cost * 1 + repeat_data_cost * 9) / 10
    tax_base = 0.12 - interest_loss - revenue_share - data_cost - 0.001 - 0.001
    tax_cost = tax_base / 1.06 * 0.0672
    expected_net = (
        0.12
        - interest_loss
        - revenue_share
        - risk_cost
        - data_cost
        - 0.001
        - 0.001
        - 0.02
        - tax_cost
    )
    assert report["source_row_count"] == 2
    assert report["row_count"] == 1
    assert report["headline_metrics"]["lowest_net_yield"] == pytest.approx(expected_net)
    assert any("1.06" in item and "6.72%" in item for item in report["assumptions"])


def test_agent_derives_turnover_from_real_style_chinese_mob_balance_curve(
    client: TestClient,
    monkeypatch,
):
    monkeypatch.setattr(
        "marvis.routers.validation_agent.resolve_driver_agent_client",
        lambda request, task, payload: None,
    )
    created = client.post(
        "/api/tasks",
        json={
            "model_name": "产品VTG原始余额曲线",
            "validator": "qa",
            "source_dir": "",
            "task_type": "vintage",
            "run_mode": "agent",
        },
    )
    task_id = created.json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    _post_message(client, task_id, "做VTG终值与年化不良，观察2026-01产品曲线")

    # Mirrors the supplied business workbooks' balance-curve turnover formula:
    # avg balance = disbursement * sum(MOB balance rate * day weight) / 360.
    csv_bytes = (
        "产品名称,截面日,放款月份,金额单位,放款额,MOB14 vtg30+,长期回收率,"
        "MOB,MOB天数,年化天数,MOB余额率,终值方法\n"
        "白条,2026-05-31,2026-01,元,1000,0.05,0.20,0,30,360,1.0,长期回收率\n"
        "白条,2026-05-31,2026-01,元,1000,0.05,0.20,1,330,360,0.5,长期回收率\n"
    ).encode("utf-8")
    uploaded = client.post(
        f"/api/tasks/{task_id}/datasets/upload",
        files={"file": ("产品MOB余额曲线.csv", csv_bytes, "text/csv")},
        data={"role": "sample"},
    )
    assert uploaded.status_code == 201, uploaded.text

    planned = _post_message(client, task_id, "材料已上传")
    assert planned.json()["messages"][-1]["metadata"]["kind"] == "plan_overview"
    completed = _post_message(client, task_id, "开始")
    report = completed.json()["messages"][-1]["metadata"]["risk_analysis_report"]

    expected_avg_balance = 1000 * (1.0 * 30 + 0.5 * 330) / 360
    expected_turnover = 1000 / expected_avg_balance
    assert report["source_row_count"] == 2
    assert report["row_count"] == 1
    assert report["headline_metrics"]["portfolio_turnover"] == pytest.approx(
        expected_turnover
    )
    assert report["headline_metrics"]["annualized_bad_rate"] == pytest.approx(
        0.05 * (1 - 0.20) * expected_turnover
    )
    assert client.get(f"/api/tasks/{task_id}/driver-report/download").status_code == 200
