"""End-to-end HTTP test of the modeling entry through the generic PlanDriver.

task_type='modeling' is routed to the plan-conversation driver (NOT the retired
ModelingSession prototype): leakage-aware screen -> [confirm features] -> tune +
train -> [confirm model] -> model-development report. No LLM is configured — this
is the no-LLM manual scenario, confirming each gate with the "确认" content.
This test runs real screening/tuning/training, so it is slow.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from marvis.app import create_app
from marvis.db import ModelingRepository
from marvis.plugins.manifest import ToolRef


def _sample_dir(root: Path, n: int = 4000) -> Path:
    src = root / "modeling_material"
    src.mkdir(parents=True, exist_ok=True)
    rng = np.random.RandomState(11)
    s1, s2, s3 = rng.normal(size=n), rng.normal(size=n), rng.normal(size=n)
    p = 1 / (1 + np.exp(-(0.9 * s1 + 0.7 * s2 - 0.6 * s3 - 1.3)))
    y = (rng.uniform(size=n) < p).astype(float)
    split = np.array(["train"] * n, dtype=object)
    split[int(n * 0.5):int(n * 0.7)] = "test"
    split[int(n * 0.7):] = "oot"
    pd.DataFrame({
        "cust_id": np.arange(n),
        "sig1": s1, "sig2": s2, "sig3": s3,
        "long_y": y, "model_flag": split,
    }).to_parquet(src / "sample.parquet")
    return src


def _multiclass_dir(root: Path, n: int = 900) -> Path:
    """A 3-class (credit grade) sample for the §8.3 multiclass conversational flow."""
    src = root / "multiclass_material"
    src.mkdir(parents=True, exist_ok=True)
    rng = np.random.RandomState(7)
    grade = rng.randint(0, 3, n)
    split = np.array(["train"] * n, dtype=object)
    split[int(n * 0.55):int(n * 0.75)] = "test"
    split[int(n * 0.75):] = "oot"
    pd.DataFrame({
        "f1": grade + rng.uniform(size=n) * 0.5,
        "f2": (2 - grade) + rng.uniform(size=n) * 0.5,
        "grade": grade, "model_flag": split,
    }).to_parquet(src / "sample.parquet")
    return src


def _continuous_dir(root: Path, n: int = 240) -> Path:
    src = root / "continuous_material"
    src.mkdir(parents=True, exist_ok=True)
    rng = np.random.RandomState(17)
    split = np.array(["train"] * n, dtype=object)
    split[int(n * 0.6):int(n * 0.8)] = "test"
    split[int(n * 0.8):] = "oot"
    pd.DataFrame({
        "f1": rng.normal(size=n),
        "f2": rng.normal(size=n),
        "income": 3000 + rng.normal(size=n) * 300,
        "model_flag": split,
    }).to_parquet(src / "sample.parquet")
    return src


def _business_material_dir(root: Path, n: int = 900) -> Path:
    src = root / "business_modeling_material"
    src.mkdir(parents=True, exist_ok=True)
    rng = np.random.RandomState(29)
    sig1 = rng.normal(size=n)
    sig2 = rng.normal(size=n)
    sig3 = rng.normal(size=n)
    score = 0.8 * sig1 - 0.5 * sig2 + 0.4 * sig3
    p = 1 / (1 + np.exp(-(score - 0.8)))
    y = (rng.uniform(size=n) < p).astype(float)
    split = np.array(["train"] * n, dtype=object)
    split[int(n * 0.55):int(n * 0.75)] = "test"
    split[int(n * 0.75):] = "oot"
    months = np.array(["2025-10", "2025-11", "2025-12", "2026-01"], dtype=object)
    pd.DataFrame({
        "cust_id": np.arange(n),
        "sig1": sig1,
        "sig2": sig2,
        "sig3": sig3,
        "long_y": y,
        "model_flag": split,
        "loan_month": months[np.arange(n) % len(months)],
        "rate": 0.08 + rng.uniform(size=n) * 0.12,
        "amount": 5000 + rng.randint(0, 20000, size=n),
        "term": rng.choice([6, 12, 18, 24], size=n),
        "drawdown": 3000 + rng.randint(0, 15000, size=n),
        "limit": 8000 + rng.randint(0, 40000, size=n),
        "mob1": y,
        "mob2": np.where(rng.uniform(size=n) < p * 0.85, 1.0, 0.0),
        "mob3": np.where(rng.uniform(size=n) < p * 0.75, 1.0, 0.0),
    }).to_parquet(src / "sample.parquet")
    pd.DataFrame({
        "特征名": ["sig1", "sig2", "sig3"],
        "含义": ["收入稳定性", "负债压力", "交易活跃度"],
        "产品名称": ["征信评分", "借贷画像", "交易评分"],
        "厂商名称": ["数据厂商A", "数据厂商B", "数据厂商C"],
    }).to_csv(src / "feature_dictionary.csv", index=False, encoding="utf-8-sig")
    return src


def _last_assistant(messages: list[dict]) -> dict:
    return [m for m in messages if m["role"] == "assistant"][-1]


@pytest.fixture
def client(tmp_path: Path) -> TestClient:
    return TestClient(create_app(tmp_path))


@pytest.mark.slow
def test_modeling_end_to_end(client: TestClient, tmp_path: Path):
    src = _sample_dir(tmp_path)
    resp = client.post("/api/tasks", json={
        "model_name": "建模驱动验证",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
    })
    assert resp.status_code == 200, resp.text
    task_id = resp.json()["id"]

    # turn 0 — start: show the plan overview, pause at the plan-level 开始 gate
    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    overview = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "确认「开始」后按计划执行" in overview["content"]

    # turn 1 — 开始: make the split + G2 spec, pause before feature screening
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "开始"})
    assert resp.status_code == 202, resp.text
    split_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert split_gate["metadata"].get("kind") == "gate"
    assert "样本切分完成" in split_gate["content"]
    assert "建模规格已生成" in split_gate["content"]

    # turn 2 — confirm the split: screen features, pause at the confirm-features gate
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    gate1 = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert gate1["metadata"].get("kind") == "gate"
    assert "特征筛选完成" in gate1["content"]
    # the screening gate carries the structured §4 payload (the frontend table consumes it)
    screen = gate1["metadata"].get("screen")
    assert screen is not None and screen["selected"], gate1["metadata"]
    assert screen["step_id"] and screen["thresholds"]["leakage_ks"] > 0
    proposed = screen["selected"]
    # edit the selection (drop the last proposed feature) to exercise the override path
    chosen = proposed[:-1] if len(proposed) > 1 else proposed

    stale = client.post(
        f"/api/tasks/{task_id}/agent/messages",
        json={"content": "确认", "selection": chosen, "expected_step_id": "old-gate"},
    )
    assert stale.status_code == 409

    # confirm features WITH an edited selection: override the screen's set,
    # then pause at the FS-1 multivariate-refinement ("精选特征") gate.
    resp = client.post(
        f"/api/tasks/{task_id}/agent/messages",
        json={"content": "确认", "selection": chosen, "expected_step_id": gate1["metadata"]["step_id"]},
    )
    assert resp.status_code == 202, resp.text
    # the screen step's stored output now reflects the user's edited selection
    overridden = client.app.state.plan_repo.load_step_output(screen["step_id"])["selected"]
    assert overridden == chosen
    refine_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert refine_gate["metadata"].get("kind") == "gate"
    assert "精选特征完成" in refine_gate["content"]
    # the refinement funnel ran on exactly the user's edited screen selection, not the
    # screen tool's original proposal
    refine_tables = refine_gate["metadata"].get("tables", [])
    refined_list_table = next(t for t in refine_tables if t["title"].startswith("最终清单"))
    assert {row[0] for row in refined_list_table["rows"]} <= set(chosen)

    # confirm the refined feature set: pause at the explicit G3 tuning-configuration gate.
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    tuning_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert tuning_gate["metadata"].get("kind") == "gate"
    assert "调参配置已生成" in tuning_gate["content"]

    # confirm tuning config: tune + train + compare, pause at the final-experiment selection gate.
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    gate2 = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert gate2["metadata"].get("kind") == "gate"
    assert "训练完成" in gate2["content"]
    gate2_tables = gate2["metadata"].get("tables", [])
    # the model metrics table is shown at the gate
    assert any(t["title"] == "模型指标" for t in gate2_tables)
    # the trials leaderboard (G4) is shown at the model gate too
    assert any(t["title"].startswith("trials 排行") for t in gate2_tables)

    # confirm selected model, approve report generation, then approve final delivery actions.
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    report_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "已选择最终实验" in report_gate["content"]

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    delivery_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "报告已生成" in delivery_gate["content"]

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    done = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "计划已全部完成" in done["content"]


def _refinement_funnel_dir(root: Path, n: int = 6000) -> Path:
    """2 strong signal features (+1 near-duplicate of the first, corr ~0.9996) and 3
    pure-noise features — exercises the FS-1 multivariate refinement funnel: the IV
    floor must drop the noise columns and correlation dedup must drop the duplicate,
    leaving exactly the 2 independent strong signals for training."""
    src = root / "refinement_funnel_material"
    src.mkdir(parents=True, exist_ok=True)
    rng = np.random.RandomState(15)
    s1 = rng.normal(size=n)
    s2 = rng.normal(size=n)
    noise1 = rng.normal(size=n)
    noise2 = rng.normal(size=n)
    noise3 = rng.normal(size=n)
    p = 1 / (1 + np.exp(-(1.1 * s1 + 1.0 * s2 - 1.1)))
    y = (rng.uniform(size=n) < p).astype(float)
    s1_dup = s1 + rng.normal(scale=0.03, size=n)  # corr(s1, s1_dup) ~ 0.9996
    split = np.array(["train"] * n, dtype=object)
    split[int(n * 0.5):int(n * 0.7)] = "test"
    split[int(n * 0.7):] = "oot"
    pd.DataFrame({
        "cust_id": np.arange(n),
        "s1": s1, "s2": s2, "s1_dup": s1_dup,
        "noise1": noise1, "noise2": noise2, "noise3": noise3,
        "long_y": y, "model_flag": split,
    }).to_parquet(src / "sample.parquet")
    return src


@pytest.mark.slow
def test_modeling_refinement_funnel_drops_noise_and_redundant_features(client: TestClient, tmp_path: Path):
    """FS-1 end-to-end: the multivariate refinement step (精选特征) between screen and
    tuning must (a) drop the 3 pure-noise columns via the IV floor, (b) drop the
    near-duplicate column via correlation dedup, and (c) leave exactly the 2
    independent strong signals to actually train on."""
    src = _refinement_funnel_dir(tmp_path)
    task_id = client.post("/api/tasks", json={
        "model_name": "精选特征漏斗验证",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lgb"],
    }).json()["id"]

    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    # "开始" runs 切分样本+选择建模规格, pauses at the 特征筛选(screen) gate — screen has
    # NOT run yet (needs_confirmation gates pause BEFORE executing).
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "开始"})
    assert resp.status_code == 202, resp.text
    # confirming 特征筛选 executes it, then pauses at the 精选特征(select) gate — select
    # has NOT run yet either; the gate shows screen's just-computed output.
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    screen_done_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert screen_done_gate["metadata"].get("kind") == "gate"
    assert "特征筛选完成" in screen_done_gate["content"]
    # sanity screen (missing/constant/pooled-leakage KS<0.4) passes all 6 clean columns through
    screen_selected = set(screen_done_gate["metadata"]["screen"]["selected"])
    assert screen_selected == {"s1", "s2", "s1_dup", "noise1", "noise2", "noise3"}

    # confirming 精选特征 executes it, then pauses at the 配置调参 gate — NOW select has
    # run and its funnel output is readable.
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    tuning_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert tuning_gate["metadata"].get("kind") == "gate"
    assert "精选特征完成" in tuning_gate["content"]

    plan = client.app.state.plan_repo.list_plans_for_task(task_id)[0]
    refine_step = next(step for step in plan.steps if step.title == "精选特征")
    refine_output = client.app.state.plan_repo.load_step_output(refine_step.id)
    selected = set(refine_output["selected"])
    dropped = {feature: reason for feature, reason in refine_output["dropped"]}

    # IV floor drops the 3 pure-noise columns
    for noise_col in ("noise1", "noise2", "noise3"):
        assert noise_col in dropped, refine_output["dropped"]
        assert "low IV" in dropped[noise_col]
    # correlation dedup drops the near-duplicate (lower-IV of the collinear pair)
    assert "s1_dup" in dropped
    assert "collinear" in dropped["s1_dup"]
    # exactly the 2 independent strong signals survive
    assert selected == {"s1", "s2"}

    tune_step = next(step for step in plan.steps if step.title == "调参")
    train_step = next(step for step in plan.steps if step.title == "训练模型")
    assert tune_step.inputs["features"] == f"$ref:{refine_step.id}.output.selected"
    assert train_step.inputs["features"] == f"$ref:{refine_step.id}.output.selected"

    # drive to completion: confirm 配置调参 -> tune/train/compare -> model-selection gate
    # -> report gate -> delivery gate -> done.
    for content in ["确认", "确认", "确认", "确认", "确认"]:
        resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": content})
        assert resp.status_code == 202, resp.text
    done = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "计划已全部完成" in done["content"]
    assert not done["metadata"].get("error")

    plan = client.app.state.plan_repo.load_plan(plan.id)
    select_step = next(step for step in plan.steps if step.title == "选择实验")
    select_output = client.app.state.plan_repo.load_step_output(select_step.id)
    artifact = ModelingRepository(client.app.state.settings.db_path).get_model_artifact(
        select_output["artifact_id"]
    )
    assert set(artifact.feature_list) == {"s1", "s2"}  # trained on exactly the refined set


@pytest.mark.slow
def test_modeling_refinement_funnel_gate_adjust_loosens_iv_floor(client: TestClient, tmp_path: Path):
    """FS-1 decision #5 (escape hatch): iv_min is adjustable at the gate that depends
    on select_features (配置调参) — loosening it to 0 lets the previously-dropped noise
    columns back in, proving the funnel is not a hard, un-bypassable wall."""
    src = _refinement_funnel_dir(tmp_path)
    task_id = client.post("/api/tasks", json={
        "model_name": "精选特征放宽验证",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lgb"],
    }).json()["id"]

    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "开始"})  # -> screen gate
    assert resp.status_code == 202, resp.text
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})  # screen -> 精选特征 gate
    assert resp.status_code == 202, resp.text
    refine_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    refine_step_id = refine_gate["metadata"]["step_id"]

    # confirm 精选特征 (runs with the default iv_min=0.02) -> pause at 配置调参 gate
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    tuning_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert tuning_gate["metadata"].get("kind") == "gate"
    baseline = client.app.state.plan_repo.load_step_output(refine_step_id)
    assert set(baseline["selected"]) == {"s1", "s2"}

    # adjust iv_min down to 0 at the 配置调参 gate: resets 精选特征 (needs_confirmation),
    # which re-pauses awaiting a fresh confirm rather than recomputing inline.
    resp = client.post(
        f"/api/tasks/{task_id}/agent/messages",
        json={
            "content": "把 IV 底线放宽到 0",
            "adjust_params": {"iv_min": 0.0},
            "expected_step_id": tuning_gate["metadata"]["step_id"],
        },
    )
    assert resp.status_code == 202, resp.text
    reset_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert reset_gate["metadata"].get("step_id") == refine_step_id

    # confirm again: 精选特征 actually re-runs with iv_min=0 this time
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    loosened = client.app.state.plan_repo.load_step_output(refine_step_id)
    # iv_min=0 lets the noise columns back in; corr_max still dedups the duplicate
    assert loosened["selected"] != baseline["selected"]
    assert {"noise1", "noise2", "noise3"} <= set(loosened["selected"])
    assert "s1_dup" not in loosened["selected"]  # correlation dedup still applies


@pytest.mark.slow
def test_modeling_business_materials_flow_into_report_and_delivery(client: TestClient, tmp_path: Path):
    src = _business_material_dir(tmp_path)
    resp = client.post("/api/tasks", json={
        "model_name": "业务材料建模",
        "model_version": "business-smoke",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lr"],
    })
    assert resp.status_code == 200, resp.text
    task_id = resp.json()["id"]

    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    messages = client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"]
    opening = next(m for m in messages if m["role"] == "assistant")
    assert "已识别建模报告业务列" in opening["content"]
    assert "已识别特征字典" in opening["content"]

    plan = client.app.state.plan_repo.list_plans_for_task(task_id)[0]
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    report_step = next(step for step in plan.steps if step.tool_ref == ToolRef("modeling", "generate_model_report"))
    delivery_step = next(step for step in plan.steps if step.tool_ref == ToolRef("modeling", "post_training_action"))
    assert report_step.inputs["business_columns"] == {
        "loan_month_col": "loan_month",
        "interest_rate_col": "rate",
        "loan_amount_col": "amount",
        "term_col": "term",
        "drawdown_amount_col": "drawdown",
        "credit_limit_col": "limit",
        "mob_observe_cols": ["mob1", "mob2", "mob3"],
    }
    assert report_step.inputs["feature_dictionary_id"]
    assert report_step.inputs["project_meta"] == {
        "模型名称": "业务材料建模",
        "模型版本": "business-smoke",
        "验证人": "qa",
    }
    for column in ["loan_month", "rate", "amount", "term", "drawdown", "limit", "mob1", "mob2", "mob3"]:
        assert column in split_step.inputs["passthrough_cols"]

    # 开始 -> split/spec gate; 确认 -> screen gate; 确认 -> FS-1 refine gate; 确认 ->
    # tuning-config gate; 确认 -> model-selection gate; 确认 -> report gate; 确认 -> delivery
    # gate; 确认 -> done.
    for content in ["开始", "确认", "确认", "确认", "确认", "确认", "确认", "确认"]:
        resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": content})
        assert resp.status_code == 202, resp.text
    done = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "计划已全部完成" in done["content"]
    assert not done["metadata"].get("error")

    report_output = client.app.state.plan_repo.load_step_output(report_step.id)
    statuses = {item["section"]: item for item in report_output["section_status"]}
    assert set(statuses) == {"sample_analysis", "vintage", "amount_bin", "low_pricing", "product_list"}
    assert all(item["available"] for item in statuses.values())
    workbook = load_workbook(report_output["report_path"])
    for sheet_name in ["汇总", "样本分析", "Vintage", "评分分段", "特征重要性", "压力测试"]:
        assert sheet_name in workbook.sheetnames
        assert workbook[sheet_name].max_row > 1
    summary_rows = {
        workbook["汇总"].cell(row=row, column=1).value: workbook["汇总"].cell(row=row, column=2).value
        for row in range(1, workbook["汇总"].max_row + 1)
    }
    assert set(str(summary_rows["五、使用产品清单"]).split("；")) == {
        "征信评分（数据厂商A）",
        "借贷画像（数据厂商B）",
        "交易评分（数据厂商C）",
    }

    delivery_output = client.app.state.plan_repo.load_step_output(delivery_step.id)
    assert Path(delivery_output["approval_package_markdown_path"]).exists()
    assert Path(delivery_output["model_card_markdown_path"]).exists()
    assert delivery_output["model_card"]["delivery"]["validation_handoff_status"] == "succeeded"


def test_modeling_business_materials_without_split_survive_auto_split(client: TestClient, tmp_path: Path):
    src = _business_material_dir(tmp_path, n=360)
    sample_path = src / "sample.parquet"
    frame = pd.read_parquet(sample_path).drop(columns=["model_flag"])
    frame.to_parquet(sample_path, index=False)
    resp = client.post("/api/tasks", json={
        "model_name": "业务材料无切分",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lr"],
    })
    assert resp.status_code == 200, resp.text
    task_id = resp.json()["id"]

    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    messages = client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"]
    opening = next(m for m in messages if m["role"] == "assistant")
    # loan_month is present (SEL-1 smoke fixture) -> default split time-extrapolates OOT
    # from it instead of a plain random train/test (see test_modeling_without_split_...
    # below for the no-time-column case, which keeps the old "已自动"/no-OOT wording).
    assert "已按" in opening["content"] and "时间外推 OOT" in opening["content"]
    assert "loan_month" in opening["content"]
    assert "已识别建模报告业务列" in opening["content"]

    plan = client.app.state.plan_repo.list_plans_for_task(task_id)[0]
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    report_step = next(step for step in plan.steps if step.tool_ref == ToolRef("modeling", "generate_model_report"))
    for column in ["loan_month", "rate", "amount", "term", "drawdown", "limit", "mob1", "mob2", "mob3"]:
        assert column in split_step.inputs["passthrough_cols"]
    assert report_step.inputs["business_columns"]["loan_month_col"] == "loan_month"

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "开始"})
    assert resp.status_code == 202, resp.text
    split_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert split_gate["metadata"].get("kind") == "gate"
    assert not split_gate["metadata"].get("error")
    assert "样本切分完成" in split_gate["content"]
    # SEL-1: OOT is real (time-extrapolated from loan_month), not fabricated/missing —
    # confirmed by counts on the derived split, all sourced from the latest month.
    plan = client.app.state.plan_repo.load_plan(plan.id)
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    split_output = client.app.state.plan_repo.load_step_output(split_step.id)
    counts = (split_output.get("sample_analysis") or {}).get("split_counts") or {}
    assert set(counts) == {"train", "test", "oot"}
    assert counts["oot"] > 0


@pytest.mark.slow
def test_modeling_multiclass_completes_end_to_end(client: TestClient, tmp_path: Path):
    """§8.3 multiclass runs through the WHOLE conversational flow (split → screen → tune →
    train → report). Guards two real bugs fixed together: (1) a non-lgb primary recipe skips
    the lgb random search and returns empty best_params — the 调参 gate must accept that
    (not 'nonempty'); (2) the binary 7-sheet report would crash on a multiclass model — a
    minimal report is written instead so the plan completes with a downloadable artifact."""
    src = _multiclass_dir(tmp_path)
    task_id = client.post("/api/tasks", json={
        "model_name": "评级", "validator": "qa", "source_dir": str(src),
        "task_type": "modeling", "run_mode": "manual", "recipes": ["lgb_multiclass"],
    }).json()["id"]

    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    msgs = client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"]
    # target_type derived from lgb_multiclass — surfaced in the opening setup message
    assert any("多分类任务" in m["content"] for m in msgs if m["role"] == "assistant")

    # 开始 → split/spec gate, 确认 → screen gate, 确认 → FS-1 refine gate, 确认 →
    # tuning-config gate, 确认 → model-selection gate (tune skipped, model trained)
    for content in ["开始", "确认", "确认", "确认", "确认"]:
        resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": content})
        assert resp.status_code == 202, resp.text
    model_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert model_gate["metadata"].get("kind") == "gate"
    assert "训练完成" in model_gate["content"]  # tune-skip no longer fails the flow

    # 确认 selected experiment → report gate; 确认 report → delivery gate; 确认 delivery → done.
    for content in ["确认", "确认"]:
        resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": content})
        assert resp.status_code == 202, resp.text
    delivery_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "报告已生成" in delivery_gate["content"]
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    done = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "计划已全部完成" in done["content"]
    assert not done["metadata"].get("error")


def test_modeling_persists_explicit_target_type_and_defaults_recipe(client: TestClient, tmp_path: Path):
    src = _continuous_dir(tmp_path)
    resp = client.post("/api/tasks", json={
        "model_name": "收入回归",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "target_type": "continuous",
    })
    assert resp.status_code == 200, resp.text
    task = resp.json()
    assert task["target_type"] == "continuous"
    assert task["recipes"] == []

    task_id = task["id"]
    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    messages = client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"]
    opening = next(m for m in messages if m["role"] == "assistant")
    assert "回归任务" in opening["content"]
    assert "lgb_regressor" in opening["content"]
    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    assert plans
    spec_step = next(step for step in plans[0].steps if step.title == "选择建模规格")
    train_step = next(step for step in plans[0].steps if step.title == "训练模型")
    screen_step = next(step for step in plans[0].steps if step.title == "特征筛选")
    assert spec_step.inputs["target_type"] == "continuous"
    assert spec_step.inputs["recipes"] == ["lgb_regressor"]
    assert screen_step.inputs["target_type"] == f"$ref:{spec_step.id}.output.target_type"
    assert train_step.inputs["target_type"] == f"$ref:{spec_step.id}.output.target_type"
    assert train_step.inputs["recipes"] == f"$ref:{spec_step.id}.output.recipes"


def test_modeling_persists_sample_weight_col_and_passes_to_plan(client: TestClient, tmp_path: Path):
    src = _sample_dir(tmp_path, n=300)
    sample_path = src / "sample.parquet"
    frame = pd.read_parquet(sample_path)
    frame["sample_weight"] = np.where(np.arange(len(frame)) % 5 == 0, 2.0, 1.0)
    frame.to_parquet(sample_path, index=False)

    resp = client.post("/api/tasks", json={
        "model_name": "带权重建模",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lgb"],
        "sample_weight_col": "sample_weight",
    })
    assert resp.status_code == 200, resp.text
    task = resp.json()
    assert task["sample_weight_col"] == "sample_weight"

    task_id = task["id"]
    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    messages = client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"]
    opening = next(m for m in messages if m["role"] == "assistant")
    assert "样本权重列:`sample_weight`" in opening["content"]

    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    split_step = next(step for step in plans[0].steps if step.title == "切分样本")
    spec_step = next(step for step in plans[0].steps if step.title == "选择建模规格")
    config_step = next(step for step in plans[0].steps if step.title == "配置调参")
    tune_step = next(step for step in plans[0].steps if step.title == "调参")
    train_step = next(step for step in plans[0].steps if step.title == "训练模型")
    assert "sample_weight" in split_step.inputs["passthrough_cols"]
    assert spec_step.inputs["sample_weight_col"] == "sample_weight"
    assert config_step.inputs["sample_weight_col"] == f"$ref:{spec_step.id}.output.sample_weight_col"
    assert tune_step.inputs["sample_weight_col"] == f"$ref:{config_step.id}.output.sample_weight_col"
    assert train_step.inputs["sample_weight_col"] == f"$ref:{spec_step.id}.output.sample_weight_col"


def test_modeling_persists_oot_ks_min_and_injects_success_criteria(client: TestClient, tmp_path: Path):
    """AGT-4: an optional per-task oot_ks_min becomes a deterministic
    success_criteria entry on the plan (never hard-coded — absent by default)."""
    src = _sample_dir(tmp_path, n=300)
    resp = client.post("/api/tasks", json={
        "model_name": "带成功标准建模",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lgb"],
        "oot_ks_min": 0.3331,
    })
    assert resp.status_code == 200, resp.text
    task = resp.json()
    assert task["oot_ks_min"] == 0.3331

    task_id = task["id"]
    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text

    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    assert plans
    assert plans[0].success_criteria == [
        {
            "metric": "oot_ks",
            "min": 0.3331,
            "aggregate": "max",
            "label": "OOT KS",
            "target_type": "binary",
        }
    ]


def test_modeling_without_oot_ks_min_injects_no_success_criteria(client: TestClient, tmp_path: Path):
    src = _sample_dir(tmp_path, n=300)
    task_id = client.post("/api/tasks", json={
        "model_name": "无成功标准建模",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lgb"],
    }).json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})

    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    assert plans
    assert plans[0].success_criteria == []


def test_modeling_rejects_out_of_range_oot_ks_min(client: TestClient, tmp_path: Path):
    src = _sample_dir(tmp_path, n=100)
    resp = client.post("/api/tasks", json={
        "model_name": "非法成功标准",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "oot_ks_min": 1.5,
    })
    assert resp.status_code == 422, resp.text


@pytest.mark.slow
def test_modeling_fails_final_review_when_oot_ks_min_unmet_in_manual_mode(client: TestClient, tmp_path: Path):
    """An unreachable oot_ks_min (manual mode: no LLM, so a criteria failure cannot
    be auto-replanned — the executor's LLMSettingsError fallback lets the plan
    surface the deterministic failure instead) drives the plan to FAILED with the
    success-criteria open item, proving the injected criterion is actually
    evaluated by final_review (not just stored)."""
    src = _sample_dir(tmp_path, n=300)
    task_id = client.post("/api/tasks", json={
        "model_name": "不可达成功标准",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
        "recipes": ["lgb"],
        "oot_ks_min": 0.999,
    }).json()["id"]

    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    for _ in range(8):
        plans = client.app.state.plan_repo.list_plans_for_task(task_id)
        if plans[0].status.value in {"failed", "done", "review"}:
            break
        resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
        assert resp.status_code == 202, resp.text

    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    final_plan = plans[0]
    assert final_plan.status.value == "failed"
    summary = client.app.state.plan_repo.load_plan_summary(
        client.app.state.plan_repo.latest_plan_summary_ref(final_plan.id)
    )
    assert any("OOT KS" in item for item in summary["open_items"])


@pytest.mark.slow
def test_modeling_multiple_files_runs_join_then_modeling_setup(client: TestClient, tmp_path: Path):
    src = _sample_dir(tmp_path, n=200)
    pd.DataFrame({
        "cust_id": np.arange(200),
        "extra_score": np.linspace(0, 1, 200),
    }).to_parquet(src / "feature_table.parquet")
    task_id = client.post("/api/tasks", json={
        "model_name": "多表建模",
        "validator": "qa",
        "source_dir": str(src),
        "task_type": "modeling",
        "run_mode": "manual",
    }).json()["id"]

    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    c1 = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "样本主表" in c1["content"]
    assert c1["metadata"]["join_c1"]["anchor_id"]
    assert not client.app.state.plan_repo.list_plans_for_task(task_id)

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    overview = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert "确认「开始」后按计划执行" in overview["content"]
    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    assert plans and plans[-1].template_id == "modeling_with_join"
    plan = plans[-1]

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "开始"})
    assert resp.status_code == 202, resp.text
    join_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert join_gate["metadata"].get("kind") == "gate"
    assert "拼接诊断完成" in join_gate["content"]

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    split_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert split_gate["metadata"].get("kind") == "gate"
    assert "样本切分完成" in split_gate["content"]
    plan = client.app.state.plan_repo.load_plan(plan.id)
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    split_output = client.app.state.plan_repo.load_step_output(split_step.id)
    assert "extra_score" in split_output["feature_cols"]


@pytest.mark.slow
def test_modeling_multiple_files_without_split_column_auto_splits_after_join(
    client: TestClient, tmp_path: Path
):
    """Join + no split column: the joined frame does not exist at setup time, so the
    plan's make_split step must generate the split (split_col="" + split_config) instead
    of failing template validation with `$.split_col: '' should be non-empty`."""
    src = tmp_path / "join_nosplit"
    src.mkdir()
    n = 200
    rng = np.random.RandomState(7)
    pd.DataFrame({
        "cust_id": np.arange(n),
        "sig1": rng.normal(size=n),
        "sig2": rng.normal(size=n),
        "long_y": (rng.uniform(size=n) < 0.3).astype(float),
    }).to_parquet(src / "sample.parquet")
    pd.DataFrame({
        "cust_id": np.arange(n),
        "extra_score": np.linspace(0, 1, n),
    }).to_parquet(src / "feature_table.parquet")
    task_id = client.post("/api/tasks", json={
        "model_name": "多表无切分", "validator": "qa", "source_dir": str(src),
        "task_type": "modeling", "run_mode": "manual",
    }).json()["id"]

    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    c1 = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert c1["metadata"]["join_c1"]["anchor_id"]

    # C1 confirm used to 409 here: plan validation rejected split_col="".
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    overview = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert not (overview.get("metadata") or {}).get("error"), overview["content"]
    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    assert plans and plans[-1].template_id == "modeling_with_join"
    plan = plans[-1]
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    assert split_step.inputs.get("split_col") == ""
    assert split_step.inputs.get("split_config", {}).get("test_size") == 0.25

    # Drive through join and split to prove make_split really generates the split.
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "开始"})
    assert resp.status_code == 202, resp.text
    join_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert join_gate["metadata"].get("kind") == "gate"

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    split_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert split_gate["metadata"].get("kind") == "gate", split_gate["content"]
    plan = client.app.state.plan_repo.load_plan(plan.id)
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    split_output = client.app.state.plan_repo.load_step_output(split_step.id)
    assert "extra_score" in split_output["feature_cols"]
    counts = (split_output.get("sample_analysis") or {}).get("split_counts") or {}
    assert set(counts) == {"train", "test"} and all(v > 0 for v in counts.values())


@pytest.mark.slow
def test_modeling_multiple_files_with_time_column_auto_splits_oot_after_join(
    client: TestClient, tmp_path: Path
):
    """SEL-1 (joined path): the anchor carries a loan_month-style business column but no
    split column. Like the single-file case, the auto split_config generated at setup
    time (before the joined frame even exists) must default to time-extrapolated OOT via
    oot_by_time, not a plain random train/test — keeping the joined and single-file
    default-split paths consistent."""
    src = tmp_path / "join_time_nosplit"
    src.mkdir()
    n = 200
    rng = np.random.RandomState(13)
    months = np.array(["2025-10", "2025-11", "2025-12", "2026-01"], dtype=object)
    pd.DataFrame({
        "cust_id": np.arange(n),
        "sig1": rng.normal(size=n),
        "sig2": rng.normal(size=n),
        "long_y": (rng.uniform(size=n) < 0.3).astype(float),
        "loan_month": months[np.arange(n) % len(months)],
    }).to_parquet(src / "sample.parquet")
    pd.DataFrame({
        "cust_id": np.arange(n),
        "extra_score": np.linspace(0, 1, n),
    }).to_parquet(src / "feature_table.parquet")
    task_id = client.post("/api/tasks", json={
        "model_name": "多表时间列无切分", "validator": "qa", "source_dir": str(src),
        "task_type": "modeling", "run_mode": "manual",
    }).json()["id"]

    resp = client.post(f"/api/tasks/{task_id}/agent/start", json={})
    assert resp.status_code == 202, resp.text
    c1 = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert c1["metadata"]["join_c1"]["anchor_id"]

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    overview = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert not (overview.get("metadata") or {}).get("error"), overview["content"]
    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    assert plans and plans[-1].template_id == "modeling_with_join"
    plan = plans[-1]
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    assert split_step.inputs.get("split_col") == ""
    assert split_step.inputs.get("split_config", {}).get("oot_by_time") == "loan_month"

    # Drive through join and split to prove make_split really produces an OOT set.
    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "开始"})
    assert resp.status_code == 202, resp.text
    join_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert join_gate["metadata"].get("kind") == "gate"

    resp = client.post(f"/api/tasks/{task_id}/agent/messages", json={"content": "确认"})
    assert resp.status_code == 202, resp.text
    split_gate = _last_assistant(client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"])
    assert split_gate["metadata"].get("kind") == "gate", split_gate["content"]
    plan = client.app.state.plan_repo.load_plan(plan.id)
    split_step = next(step for step in plan.steps if step.title == "切分样本")
    split_output = client.app.state.plan_repo.load_step_output(split_step.id)
    assert "extra_score" in split_output["feature_cols"]
    counts = (split_output.get("sample_analysis") or {}).get("split_counts") or {}
    assert set(counts) == {"train", "test", "oot"} and all(v > 0 for v in counts.values())


def test_modeling_without_split_auto_generates_grouped_split(client: TestClient, tmp_path: Path):
    """No split column → make_split generates a grouped train/test split (no fabricated
    OOT) and modeling proceeds to the plan-overview gate, instead of erroring."""
    src = tmp_path / "nosplit"
    src.mkdir()
    n = 200
    rng = np.random.RandomState(3)
    pd.DataFrame({
        "cust_id": np.arange(n),
        "f1": rng.normal(size=n),
        "f2": rng.normal(size=n),
        "long_y": (rng.uniform(size=n) < 0.3).astype(float),
    }).to_parquet(src / "s.parquet")
    task_id = client.post("/api/tasks", json={
        "model_name": "无切分", "validator": "qa", "source_dir": str(src),
        "task_type": "modeling", "run_mode": "manual",
    }).json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    msgs = client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"]
    assistants = [m for m in msgs if m["role"] == "assistant"]
    # no error; it auto-split and reached the plan-overview gate
    assert not any((m.get("metadata") or {}).get("error") for m in assistants)
    assert any("已自动" in m["content"] and "train/test" in m["content"] for m in assistants)
    assert any((m.get("metadata") or {}).get("kind") == "plan_overview" for m in assistants)


def test_modeling_honors_selected_algorithms(client: TestClient, tmp_path: Path):
    """The user-selected algorithms (manual-mode multi-select → `recipes`) drive the
    modeling recipes instead of the lgb default (G2: 算法多选)."""
    src = _sample_dir(tmp_path, n=600)
    task_id = client.post("/api/tasks", json={
        "model_name": "选算法", "validator": "qa", "source_dir": str(src),
        "task_type": "modeling", "run_mode": "manual", "recipes": ["lr", "lgb"],
    }).json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    msgs = client.get(f"/api/tasks/{task_id}/agent/messages").json()["messages"]
    opening = next(m for m in msgs if m["role"] == "assistant")
    # the chosen algorithms are noted (multi-algorithm compare, not the lgb default only)
    assert "算法" in opening["content"]
    assert "lr" in opening["content"] and "lgb" in opening["content"]
    assert "取最优" in opening["content"]


def test_modeling_honors_per_task_capability_tier(client: TestClient, tmp_path: Path):
    """The per-task capability tier (TIER-IA, spec §5.1) flows into the plan, setting
    its autonomy (replan) budget; absent → the global default (balanced)."""
    src = _sample_dir(tmp_path)
    task_id = client.post("/api/tasks", json={
        "model_name": "档位", "validator": "qa", "source_dir": str(src),
        "task_type": "modeling", "run_mode": "manual", "recipes": ["lgb"],
        "capability_tier": "conservative",
    }).json()["id"]
    client.post(f"/api/tasks/{task_id}/agent/start", json={})
    plans = client.app.state.plan_repo.list_plans_for_task(task_id)
    assert plans, "no plan was built"
    assert plans[0].tier == "conservative"

    # an unknown tier is normalized away → the plan falls back to the default
    default_id = client.post("/api/tasks", json={
        "model_name": "默认档", "validator": "qa", "source_dir": str(src),
        "task_type": "modeling", "run_mode": "manual", "recipes": ["lgb"],
        "capability_tier": "not-a-tier",
    }).json()["id"]
    client.post(f"/api/tasks/{default_id}/agent/start", json={})
    default_plans = client.app.state.plan_repo.list_plans_for_task(default_id)
    assert default_plans and default_plans[0].tier == "balanced"
