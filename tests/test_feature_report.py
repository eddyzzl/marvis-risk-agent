"""Standalone feature-analysis Excel report writer (FEATURE form A)."""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from marvis.output.feature_report import render_feature_report


def test_render_feature_report_writes_metric_sheet(tmp_path):
    metrics = [
        {"feature": "x1", "iv": 0.42, "ks": 0.31, "auc": 0.71, "psi": 0.03, "missing_rate": 0.0, "lift_top_bin": 2.1},
        {"feature": "x2", "iv": 0.18, "ks": 0.20, "auc": 0.63, "psi": None, "missing_rate": 0.05, "lift_top_bin": 1.4},
    ]
    out = tmp_path / "feature_report.xlsx"

    render_feature_report(metrics, out)

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["特征指标"]
    sheet = workbook["特征指标"]
    headers = [cell.value for cell in sheet[1]]
    for header in ("特征", "Agent建议", "推荐原因", "IV", "KS", "AUC", "PSI", "PSI说明", "缺失率", "单一值率", "零值率"):
        assert header in headers
    assert sheet["A2"].value == "x1"
    assert sheet["A3"].value == "x2"
    # a missing metric renders as n/a, never silently blank
    assert sheet.cell(row=3, column=headers.index("PSI") + 1).value == "n/a"


def test_render_feature_report_handles_empty_metrics(tmp_path):
    out = tmp_path / "empty.xlsx"
    render_feature_report([], out)
    sheet = load_workbook(out)["特征指标"]
    assert [cell.value for cell in sheet[1]][0] == "特征"
    assert sheet.max_row == 1  # header only


def test_render_feature_report_escapes_formula_like_user_text(tmp_path):
    out = tmp_path / "formula_safe.xlsx"
    render_feature_report(
        [
            {
                "feature": "=HYPERLINK(\"https://example.invalid\",\"open\")",
                "recommendation": "@SUM(1,1)",
                "recommendation_reason": "+cmd|' /C calc'!A0",
                "ks": 0.2,
            }
        ],
        out,
    )

    sheet = load_workbook(out, data_only=False)["特征指标"]
    headers = [cell.value for cell in sheet[1]]
    for header in ("特征", "Agent建议", "推荐原因"):
        cell = sheet.cell(row=2, column=headers.index(header) + 1)
        assert cell.data_type == "s"
        assert str(cell.value).startswith("'")


def test_render_feature_report_appends_optional_columns_only_when_present(tmp_path):
    """Head/tail lift + importance columns are written only when those keys ride in the
    metric rows (selected); a base-only report keeps its 7 columns."""
    out = tmp_path / "optional.xlsx"
    render_feature_report(
        [{
            "feature": "x1", "iv": 0.4, "ks": 0.3, "auc": 0.7, "psi": 0.02, "missing_rate": 0.0,
            "lift_top_bin": 2.0, "lift_head_5": 3.1, "lift_head_10": 2.6, "lift_tail_5": 0.2,
            "lift_tail_10": 0.4, "importance": 0.62,
        }],
        out,
    )
    header = [cell.value for cell in load_workbook(out)["特征指标"][1]]
    for col in ("头部lift5%", "头部lift10%", "尾部lift5%", "尾部lift10%", "重要性"):
        assert col in header

    base_out = tmp_path / "base.xlsx"
    render_feature_report([{"feature": "x1", "iv": 0.4}], base_out)
    base_header = [cell.value for cell in load_workbook(base_out)["特征指标"][1]]
    assert "重要性" not in base_header and "头部lift5%" not in base_header


def test_render_feature_report_writes_only_selected_metric_groups(tmp_path):
    """An explicit KS-only computation must not grow IV/AUC/quality columns."""

    out = tmp_path / "ks_only.xlsx"
    render_feature_report(
        [{"feature": "x1", "ks": 0.31, "recommendation": "保留"}],
        out,
    )

    headers = [cell.value for cell in load_workbook(out)["特征指标"][1]]
    assert "KS" in headers
    for unchecked in ("IV", "AUC", "PSI", "缺失率", "头部lift5%", "重要性"):
        assert unchecked not in headers


def test_render_feature_report_serializes_structured_na_and_adds_psi_detail(tmp_path):
    """Structured dependency reasons are reader-facing text, never raw dicts that
    openpyxl cannot serialize. Selected PSI series also get a long-form sheet."""

    out = tmp_path / "psi.xlsx"
    render_feature_report(
        [
            {
                "feature": "x1",
                "psi_month_first": None,
                "psi_month_first_reason": {
                    "code": "missing_dependency",
                    "metric_dependency": "time_col",
                    "message": "未识别到时间列，无法计算首月基准 PSI。",
                },
                "psi_split": 0.12,
                "psi_split_series": [
                    {"base": "train", "compare": "test", "psi": 0.08},
                    {"base": "train", "compare": "oot", "psi": 0.12},
                ],
            }
        ],
        out,
    )

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["特征指标", "PSI明细"]
    metric_sheet = workbook["特征指标"]
    headers = [cell.value for cell in metric_sheet[1]]
    reason_cell = metric_sheet.cell(
        row=2,
        column=headers.index("月度PSI(首月)说明") + 1,
    ).value
    assert reason_cell == "未识别到时间列，无法计算首月基准 PSI。"
    assert metric_sheet.cell(
        row=2,
        column=headers.index("月度PSI(首月基准)") + 1,
    ).value == "n/a"
    detail = workbook["PSI明细"]
    assert detail.max_row == 3
    assert detail["A2"].value == "x1"
    assert detail["B2"].value == "样本集PSI"
    assert detail["C2"].value == "train"
    assert detail["D3"].value == "oot"


def test_render_feature_report_rolls_back_existing_file_when_save_fails(tmp_path, monkeypatch):
    out = tmp_path / "feature_report.xlsx"
    render_feature_report([{"feature": "old", "iv": 0.1}], out)
    original_bytes = out.read_bytes()
    original_save = Workbook.save

    def failing_save(self, filename):
        original_save(self, filename)
        raise RuntimeError("xlsx save failed")

    monkeypatch.setattr(Workbook, "save", failing_save)

    with pytest.raises(RuntimeError, match="xlsx save failed"):
        render_feature_report([{"feature": "new", "iv": 0.9}], out)

    assert out.read_bytes() == original_bytes
    assert not (tmp_path / ".staging").exists()


def test_render_feature_report_adds_selected_binning_sheet(tmp_path):
    out = tmp_path / "feature_report_with_bins.xlsx"
    binning = [{
        "feature": "x1",
        "requested_bins": 5,
        "actual_bins": 2,
        "direction": "risk_up",
        "total_iv": 0.23,
        "degraded_reason": "重复值折叠为 2 箱",
        "rows": [
            {
                "bin_index": 1,
                "risk_rank": 2,
                "interval": "(-inf, 0]",
                "count": 12,
                "bad_count": 2,
                "good_count": 10,
                "bad_rate": 2 / 12,
                "cumulative_bad_rate": 2 / 12,
                "lift": 0.5,
                "cumulative_lift": 0.5,
                "ks": 0.25,
                "woe": 0.4,
                "iv_contribution": 0.08,
            },
            {
                "bin_index": 2,
                "risk_rank": 1,
                "interval": "(0, inf]",
                "count": 8,
                "bad_count": 6,
                "good_count": 2,
                "bad_rate": 0.75,
                "cumulative_bad_rate": 0.4,
                "lift": 2.25,
                "cumulative_lift": 1.2,
                "ks": 0.0,
                "woe": -0.6,
                "iv_contribution": 0.15,
            },
        ],
    }]

    render_feature_report([{"feature": "x1", "iv": 0.23}], out, binning=binning)

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["特征指标", "分箱分析"]
    sheet = workbook["分箱分析"]
    headers = [cell.value for cell in sheet[1]]
    for header in (
        "特征", "请求箱数", "实际箱数", "箱号", "风险排序", "区间", "样本数", "坏样本数",
        "好样本数", "坏率", "累计坏率", "单箱Lift", "累计Lift", "KS", "WOE",
        "IV贡献", "总IV", "风险方向", "分箱说明",
    ):
        assert header in headers
    assert sheet.max_row == 3
    assert sheet["A2"].value == "x1"


def test_render_feature_report_omits_binning_sheet_when_user_skips(tmp_path):
    out = tmp_path / "feature_report_without_bins.xlsx"
    render_feature_report([{"feature": "x1", "iv": 0.23}], out, binning=[])
    assert load_workbook(out).sheetnames == ["特征指标"]
