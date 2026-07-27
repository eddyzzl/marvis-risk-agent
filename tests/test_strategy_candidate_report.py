from __future__ import annotations

from copy import deepcopy
from io import BytesIO
import json
from zipfile import ZipFile

from openpyxl import load_workbook
import pandas as pd
import pytest

from marvis.feature.univariate import analyze_univariate
from marvis.output.strategy_candidate_report import (
    REPORT_SCHEMA_VERSION,
    StrategyCandidateReportError,
    canonical_strategy_candidate_report_json,
    render_strategy_candidate_bundle,
    render_strategy_candidate_report_xlsx,
    strategy_candidate_report_from_json,
    validate_strategy_candidate_report,
)
from marvis.packs.strategy.candidate_evidence import (
    CandidateEvidenceError,
    MetricObservation,
    build_candidate_evidence,
)


HASH_A = "a" * 64
HASH_B = "b" * 64
FIXED_ZIP_DATETIME = (1980, 1, 1, 0, 0, 0)


def _inputs(*, malicious: bool = False) -> tuple[dict, dict]:
    feature = '=HYPERLINK("https://invalid.example")' if malicious else "age"
    frame = pd.DataFrame(
        {
            feature: [18, 21, 24, 31, 38, 42, 55, 61],
            "bad": [0, 0, 1, 0, 1, 0, 1, 1],
            "loan": [100, 120, 130, 140, 150, 160, 170, 180],
            "overdue": [0, 0, 10, 0, 20, 0, 30, 40],
        }
    )
    analysis = analyze_univariate(
        frame,
        features=[feature],
        target="bad",
        methods=["equal_width"],
        bin_count=3,
        loan_amount="loan",
        overdue_amount="overdue",
        min_bin_pct=0,
        seed=17,
    )
    flag = "+SUM(1,1)" if malicious else "development_only"
    source_ref = "@external-link" if malicious else "analysis:run-1"
    evidence = build_candidate_evidence(
        task_id="task-1",
        dataset_id="dataset-1",
        dataset_content_hash=HASH_A,
        workspace_revision=3,
        workspace_generation=2,
        semantic_mapping_hash=HASH_B,
        generation_parameters={
            "analysis_schema_version": analysis["schema_version"],
            "features": [feature],
            "methods": ["equal_width"],
        },
        seed=17,
        budget=20,
        truncated=False,
        analysis=analysis,
        metrics=[
            MetricObservation("hit_rate", "count", "observed", 0.5),
            MetricObservation("hit_rate", "loan_amount", "observed", 0.58),
            MetricObservation("hit_rate", "overdue_amount", "observed", 1.0),
        ],
        source_refs=[source_ref],
        red_flags=[flag],
        producer_version="strategy.univariate-candidate/1",
    )
    return evidence, analysis


def _rows(workbook, sheet_name: str) -> list[tuple[object, ...]]:
    return list(workbook[sheet_name].iter_rows(values_only=True))


def _rebuild_with_analysis(evidence: dict, analysis: dict) -> dict:
    identity = evidence["identity"]
    generation = evidence["generation"]
    return build_candidate_evidence(
        task_id=identity["task_id"],
        dataset_id=identity["dataset_id"],
        dataset_content_hash=identity["dataset_content_hash"],
        workspace_revision=identity["workspace_revision"],
        workspace_generation=identity["workspace_generation"],
        semantic_mapping_hash=identity["semantic_mapping_hash"],
        generation_parameters=generation["parameters"],
        seed=generation["seed"],
        budget=generation["budget"],
        truncated=generation["truncated"],
        analysis=analysis,
        metrics=evidence["metrics"],
        source_refs=evidence["source_refs"],
        red_flags=evidence["red_flags"],
        producer_version=evidence["producer_version"],
    )


def test_bundle_is_canonical_and_byte_deterministic() -> None:
    evidence, analysis = _inputs()

    first = render_strategy_candidate_bundle(evidence, analysis)
    second = render_strategy_candidate_bundle(evidence, analysis)

    assert set(first) == {"json", "xlsx"}
    assert first == second
    assert first["json"] == canonical_strategy_candidate_report_json(evidence, analysis)
    assert first["xlsx"] == render_strategy_candidate_report_xlsx(evidence, analysis)

    payload = json.loads(first["json"])
    assert payload == {
        "schema_version": REPORT_SCHEMA_VERSION,
        "candidate_evidence": evidence,
        "univariate_analysis": analysis,
    }
    assert first["json"] == json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")

    with ZipFile(BytesIO(first["xlsx"])) as archive:
        assert archive.namelist() == sorted(archive.namelist())
        assert all(info.date_time == FIXED_ZIP_DATETIME for info in archive.infolist())
        assert all(not info.extra and not info.comment for info in archive.infolist())
        assert b"2000-01-01T00:00:00Z" in archive.read("docProps/core.xml")
        assert b"<f" not in b"".join(
            archive.read(name)
            for name in archive.namelist()
            if name.startswith("xl/worksheets/")
        )


def test_persisted_report_parser_reuses_strict_contracts() -> None:
    evidence, analysis = _inputs()
    raw = canonical_strategy_candidate_report_json(evidence, analysis)

    parsed = strategy_candidate_report_from_json(raw)

    assert parsed == validate_strategy_candidate_report(json.loads(raw))
    assert parsed["candidate_evidence"] == evidence

    with pytest.raises(StrategyCandidateReportError, match="duplicate key"):
        strategy_candidate_report_from_json(
            raw[:-1] + b',"schema_version":"strategy.candidate-report.v1"}'
        )

    invalid = json.loads(raw)
    invalid["caller_metrics"] = []
    with pytest.raises(StrategyCandidateReportError, match="unknown"):
        validate_strategy_candidate_report(invalid)


def test_manual_v2_report_freezes_cutpoints_and_rejects_bin_drift() -> None:
    frame = pd.DataFrame(
        {
            "age": [18, 21, 24, 31, 38, 42, 55, 61],
            "bad": [0, 0, 1, 0, 1, 0, 1, 1],
        }
    )
    analysis = analyze_univariate(
        frame,
        features=["age"],
        target="bad",
        methods=["manual"],
        manual_breakpoints={"age": [25.0, 40.0]},
        bin_count=3,
        min_bin_pct=0,
    )
    evidence = build_candidate_evidence(
        task_id="task-manual",
        dataset_id="dataset-manual",
        dataset_content_hash=HASH_A,
        workspace_revision=3,
        workspace_generation=2,
        semantic_mapping_hash=HASH_B,
        generation_parameters={
            "analysis_schema_version": analysis["schema_version"],
            "features": ["age"],
            "methods": ["manual"],
            "manual_breakpoints": {"age": [25.0, 40.0]},
        },
        seed=0,
        budget=20,
        truncated=False,
        analysis=analysis,
        metrics=[
            MetricObservation("hit_rate", "count", "observed", 0.5),
            MetricObservation("hit_rate", "loan_amount", "unavailable", None),
            MetricObservation("hit_rate", "overdue_amount", "unavailable", None),
        ],
        source_refs=["analysis:manual"],
        producer_version="strategy.univariate-candidate/2",
    )

    raw = canonical_strategy_candidate_report_json(evidence, analysis)
    parsed = strategy_candidate_report_from_json(raw)
    assert parsed["univariate_analysis"]["schema_version"] == (
        "univariate-analysis-result.v2"
    )
    assert parsed["univariate_analysis"]["parameters"]["manual_breakpoints"] == {
        "age": [25.0, 40.0]
    }

    drifted = deepcopy(analysis)
    drifted["features"][0]["methods"][0]["manual_breakpoints"] = [20.0, 40.0]
    drifted_evidence = _rebuild_with_analysis(evidence, drifted)
    with pytest.raises(
        StrategyCandidateReportError,
        match="manual_breakpoints do not match",
    ):
        render_strategy_candidate_bundle(drifted_evidence, drifted)


@pytest.mark.parametrize(
    ("producer_version", "analysis_schema_version"),
    [
        ("strategy.univariate-candidate/1", "univariate-analysis-result.v2"),
        ("strategy.univariate-candidate/2", "univariate-analysis-result.v1"),
    ],
)
def test_report_rejects_impossible_analysis_producer_version_pair(
    producer_version: str,
    analysis_schema_version: str,
) -> None:
    evidence, analysis = _inputs()
    identity = evidence["identity"]
    generation = evidence["generation"]
    mismatched = build_candidate_evidence(
        task_id=identity["task_id"],
        dataset_id=identity["dataset_id"],
        dataset_content_hash=identity["dataset_content_hash"],
        workspace_revision=identity["workspace_revision"],
        workspace_generation=identity["workspace_generation"],
        semantic_mapping_hash=identity["semantic_mapping_hash"],
        generation_parameters={
            **generation["parameters"],
            "analysis_schema_version": analysis_schema_version,
        },
        seed=generation["seed"],
        budget=generation["budget"],
        truncated=generation["truncated"],
        analysis=analysis,
        metrics=evidence["metrics"],
        source_refs=evidence["source_refs"],
        red_flags=evidence["red_flags"],
        producer_version=producer_version,
    )

    with pytest.raises(
        StrategyCandidateReportError,
        match="analysis schema, generation, and producer versions do not match",
    ):
        render_strategy_candidate_bundle(mismatched, analysis)


def test_xlsx_contains_report_ready_evidence_without_recomputing_metrics() -> None:
    evidence, analysis = _inputs()
    workbook = load_workbook(
        BytesIO(render_strategy_candidate_report_xlsx(evidence, analysis)),
        read_only=True,
        data_only=False,
    )

    assert workbook.sheetnames == [
        "Summary",
        "Rankings",
        "Bins",
        "Metrics",
        "Red Flags",
        "Lineage",
    ]

    summary = dict(_rows(workbook, "Summary")[1:])
    assert summary["candidate_id"] == evidence["candidate_id"]
    assert summary["validation_status"] == "unvalidated"
    assert summary["target"] == "bad"

    ranking_rows = _rows(workbook, "Rankings")
    ranking = analysis["rankings"][0]
    assert ranking_rows[0] == ("Feature", "Method", "IV", "KS", "AUC")
    assert ranking_rows[1] == (
        ranking["feature"],
        ranking["method"],
        repr(ranking["iv"]),
        repr(ranking["ks"]),
        repr(ranking["auc"]),
    )

    bin_rows = _rows(workbook, "Bins")
    headers = bin_rows[0]
    first_bin = dict(zip(headers, bin_rows[1], strict=True))
    source_bin = analysis["features"][0]["methods"][0]["bins"][0]
    assert first_bin["Feature"] == "age"
    assert first_bin["Count"] == source_bin["count"]
    assert first_bin["WOE"] == repr(source_bin["woe"])
    assert first_bin["IV Contribution"] == repr(source_bin["iv_contribution"])

    metric_rows = _rows(workbook, "Metrics")
    assert any(
        row[:7]
        == (
            "candidate_evidence",
            None,
            None,
            "hit_rate",
            "count",
            "observed",
            "0.5",
        )
        for row in metric_rows
    )
    assert any(
        row[0] == "univariate_analysis"
        and row[1] == "age"
        and row[3] == "iv"
        and row[6] == repr(ranking["iv"])
        for row in metric_rows[1:]
    )

    red_flag_rows = _rows(workbook, "Red Flags")
    assert any(row[3] == "development_only" for row in red_flag_rows[1:])

    lineage = dict(_rows(workbook, "Lineage")[1:])
    assert lineage["identity.dataset_id"] == "dataset-1"
    assert lineage["identity.dataset_content_hash"] == HASH_A
    assert lineage["source_ref[0]"] == "analysis:run-1"


def test_formula_like_user_strings_are_plain_text_in_every_sheet() -> None:
    evidence, analysis = _inputs(malicious=True)
    workbook = load_workbook(
        BytesIO(render_strategy_candidate_bundle(evidence, analysis)["xlsx"]),
        read_only=True,
        data_only=False,
    )

    values = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert cell.data_type != "f"
                if isinstance(cell.value, str):
                    values.append(cell.value)

    assert '\'=HYPERLINK("https://invalid.example")' in values
    assert "'+SUM(1,1)" in values
    assert "'@external-link" in values


def test_invalid_evidence_and_analysis_are_rejected_fail_closed() -> None:
    evidence, analysis = _inputs()
    forged = deepcopy(evidence)
    forged["evidence_hash"] = "0" * 64

    with pytest.raises(CandidateEvidenceError, match="evidence_hash"):
        render_strategy_candidate_bundle(forged, analysis)

    invalid_analysis = deepcopy(analysis)
    invalid_analysis["schema_version"] = "univariate-analysis-result.v999"
    with pytest.raises(StrategyCandidateReportError, match="schema_version"):
        render_strategy_candidate_bundle(evidence, invalid_analysis)

    other_analysis = deepcopy(analysis)
    other_analysis["target"] = "other_bad"
    with pytest.raises(StrategyCandidateReportError, match="does not match"):
        render_strategy_candidate_bundle(evidence, other_analysis)

    type_drifted_analysis = deepcopy(analysis)
    assert type_drifted_analysis["features"][0]["missing_rate"] == 0.0
    type_drifted_analysis["features"][0]["missing_rate"] = False
    type_drifted_evidence = _rebuild_with_analysis(evidence, type_drifted_analysis)
    with pytest.raises(StrategyCandidateReportError, match="does not match"):
        render_strategy_candidate_bundle(type_drifted_evidence, analysis)

    invalid_condition = deepcopy(analysis)
    invalid_condition["features"][0]["methods"][0]["bins"][0]["condition"] = {
        "op": "compare",
        "field": "age",
        "operator": "contains",
        "value": 18,
        "missing": "no_match",
    }
    invalid_condition_evidence = _rebuild_with_analysis(evidence, invalid_condition)
    with pytest.raises(StrategyCandidateReportError, match="canonical Strategy DSL"):
        render_strategy_candidate_bundle(
            invalid_condition_evidence,
            invalid_condition,
        )

    wrong_numeric_condition = deepcopy(analysis)
    wrong_numeric_condition["features"][0]["methods"][0]["bins"][0]["condition"] = {
        "op": "is_null",
        "field": "age",
    }
    wrong_numeric_evidence = _rebuild_with_analysis(
        evidence,
        wrong_numeric_condition,
    )
    with pytest.raises(StrategyCandidateReportError, match="bin definition"):
        render_strategy_candidate_bundle(
            wrong_numeric_evidence,
            wrong_numeric_condition,
        )

    inconsistent_counts = deepcopy(analysis)
    inconsistent_counts["features"][0]["methods"][0]["bins"][0]["count"] += 1
    inconsistent_evidence = _rebuild_with_analysis(evidence, inconsistent_counts)
    with pytest.raises(StrategyCandidateReportError, match="good plus bad"):
        render_strategy_candidate_bundle(inconsistent_evidence, inconsistent_counts)

    categorical_analysis = analyze_univariate(
        pd.DataFrame(
            {
                "segment": ["A", "A", "UNKNOWN", "B", None, "B"],
                "bad": [0, 1, 0, 1, 1, 0],
            }
        ),
        features=["segment"],
        target="bad",
        methods=["categorical"],
        sentinel_values={"segment": ["UNKNOWN"]},
        bin_count=3,
        min_bin_pct=0,
    )
    categorical_evidence = _rebuild_with_analysis(evidence, categorical_analysis)
    bins = categorical_analysis["features"][0]["methods"][0]["bins"]

    wrong_category = deepcopy(categorical_analysis)
    category_index = next(
        index for index, item in enumerate(bins) if item["kind"] == "category"
    )
    wrong_category["features"][0]["methods"][0]["bins"][category_index]["condition"] = {
        "op": "compare",
        "field": "segment",
        "operator": "==",
        "value": "OTHER",
        "missing": "no_match",
        "coercion": "strict",
    }
    with pytest.raises(StrategyCandidateReportError, match="bin definition"):
        render_strategy_candidate_bundle(
            _rebuild_with_analysis(categorical_evidence, wrong_category),
            wrong_category,
        )

    wrong_missing = deepcopy(categorical_analysis)
    missing_index = next(
        index for index, item in enumerate(bins) if item["kind"] == "missing"
    )
    wrong_missing["features"][0]["methods"][0]["bins"][missing_index]["condition"] = {
        "op": "compare",
        "field": "segment",
        "operator": "==",
        "value": "A",
        "missing": "no_match",
        "coercion": "strict",
    }
    with pytest.raises(StrategyCandidateReportError, match="bin definition"):
        render_strategy_candidate_bundle(
            _rebuild_with_analysis(categorical_evidence, wrong_missing),
            wrong_missing,
        )
