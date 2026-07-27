from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from marvis.app import create_app
from marvis.data.backend import DataBackend
from marvis.data.registry import DatasetRegistry
from marvis.data.workspace import DataSemanticMapping, DataWorkspaceDraft
from marvis.db import DatasetRepository
from marvis.db_schema import connect
from marvis.repositories.data_workspace import DataWorkspaceRepository


def _task(client: TestClient, tmp_path: Path) -> str:
    source = client.app.state.settings.workspace / f"source-{tmp_path.name}"
    source.mkdir(exist_ok=True)
    response = client.post(
        "/api/tasks",
        json={
            "model_name": "策略样本导出",
            "validator": "qa",
            "source_dir": str(source),
            "task_type": "strategy",
            "run_mode": "manual",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()["id"]


def _register(client: TestClient, task_id: str, tmp_path: Path):
    source = tmp_path / f"{task_id}.parquet"
    pd.DataFrame(
        {
            "customer_name": ["张三", "=2+2"],
            "mobile": ["0013800000000", "+8613900000000"],
            "amount": [100.0, 200.0],
            "bad": [0, 1],
        }
    ).to_parquet(source, index=False)
    settings = client.app.state.settings
    registry = DatasetRegistry(
        DatasetRepository(settings.db_path),
        DataBackend(settings.datasets_dir),
        settings.datasets_dir,
    )
    return registry.register_from_upload(task_id, source, role="strategy_sample")


def _activate(client: TestClient, task_id: str, dataset):
    repo = DataWorkspaceRepository(client.app.state.settings.db_path)
    active = repo.save(
        task_id,
        DataWorkspaceDraft(
            active_dataset_id=dataset.id,
            active_dataset_content_hash=dataset.content_hash,
        ),
        expected_revision=0,
    )
    return repo.save(
        task_id,
        DataWorkspaceDraft(
            active_dataset_id=dataset.id,
            active_dataset_content_hash=dataset.content_hash,
            page="overview",
            semantic_mapping=DataSemanticMapping(
                target_col="bad",
                field_roles={
                    "customer_name": "name",
                    "mobile": "phone",
                    "bad": "target",
                },
                business_names={"mobile": "手机号", "amount": "申请金额"},
            ),
        ),
        expected_revision=active.revision,
    )


def _post(client: TestClient, task_id: str, content: str):
    return client.post(
        f"/api/tasks/{task_id}/agent/messages",
        json={"content": content},
    )


def _last_assistant(response) -> dict:
    return [
        message
        for message in response.json()["messages"]
        if message["role"] == "assistant"
    ][-1]


def test_natural_language_dataset_export_runs_bound_xlsx_and_downloads(tmp_path):
    client = TestClient(create_app(tmp_path / "workspace"))
    task_id = _task(client, tmp_path)
    dataset = _register(client, task_id, tmp_path)
    before = _activate(client, task_id, dataset)

    response = _post(
        client,
        task_id,
        "导出当前数据为 Excel，并把手机号按文本导出",
    )

    assert response.status_code == 202, response.text
    last = _last_assistant(response)
    assert "计划已全部完成" in last["content"]
    assert "数据导出完成" in last["content"]
    tables = {table["title"]: table for table in last["metadata"]["tables"]}
    assert {"导出文件", "安全处理"} <= set(tables)
    download_url = next(
        row[1] for row in tables["导出文件"]["rows"] if row[0] == "下载地址"
    )
    downloaded = client.get(download_url)
    assert downloaded.status_code == 200, downloaded.text
    assert downloaded.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(downloaded.content), read_only=True)
    rows = list(workbook["data"].iter_rows(values_only=False))
    assert rows[1][1].value == "0013800000000"
    assert rows[1][1].data_type == "s"
    assert DataWorkspaceRepository(
        client.app.state.settings.db_path
    ).get_or_default(task_id) == before

    with connect(client.app.state.settings.db_path) as conn:
        refs = [
            str(row["tool_ref"])
            for row in conn.execute(
                """
                SELECT r.tool_ref
                  FROM plan_step_runs r
                  JOIN plans p ON p.id = r.plan_id
                 WHERE p.task_id = ?
                """,
                (task_id,),
            ).fetchall()
        ]
    assert refs == ["data_ops.export_dataset"]


def test_dataset_export_without_active_workspace_clarifies(tmp_path):
    client = TestClient(create_app(tmp_path / "workspace"))
    task_id = _task(client, tmp_path)
    _register(client, task_id, tmp_path)

    response = _post(client, task_id, "导出当前样本为 CSV")

    assert response.status_code == 202, response.text
    last = _last_assistant(response)
    assert "数据工作区" in last["content"]
    assert last["metadata"]["intent"] == "dataset_export"
    with connect(client.app.state.settings.db_path) as conn:
        assert conn.execute(
            "SELECT count(*) FROM plans WHERE task_id = ?",
            (task_id,),
        ).fetchone()[0] == 0


def test_dataset_export_without_format_routes_and_asks_for_format(tmp_path):
    client = TestClient(create_app(tmp_path / "workspace"))
    task_id = _task(client, tmp_path)
    dataset = _register(client, task_id, tmp_path)
    _activate(client, task_id, dataset)

    response = _post(client, task_id, "导出当前数据")

    assert response.status_code == 202, response.text
    last = _last_assistant(response)
    assert last["metadata"]["intent"] == "dataset_export"
    assert last["metadata"]["code"] == "export_request_clarification"
    assert "CSV" in last["content"]
    assert "Excel" in last["content"]
    with connect(client.app.state.settings.db_path) as conn:
        assert conn.execute(
            "SELECT count(*) FROM plans WHERE task_id = ?",
            (task_id,),
        ).fetchone()[0] == 0


def test_dataset_export_with_conflicting_formats_routes_and_clarifies(tmp_path):
    client = TestClient(create_app(tmp_path / "workspace"))
    task_id = _task(client, tmp_path)
    dataset = _register(client, task_id, tmp_path)
    _activate(client, task_id, dataset)

    response = _post(client, task_id, "当前数据同时导出为 CSV 和 Excel")

    assert response.status_code == 202, response.text
    last = _last_assistant(response)
    assert last["metadata"]["intent"] == "dataset_export"
    assert last["metadata"]["code"] == "export_request_clarification"
    assert "一种" in last["content"]
    with connect(client.app.state.settings.db_path) as conn:
        assert conn.execute(
            "SELECT count(*) FROM plans WHERE task_id = ?",
            (task_id,),
        ).fetchone()[0] == 0


def test_analysis_export_wording_does_not_run_raw_dataset_export(tmp_path):
    client = TestClient(create_app(tmp_path / "workspace"))
    task_id = _task(client, tmp_path)
    dataset = _register(client, task_id, tmp_path)
    _activate(client, task_id, dataset)

    response = _post(client, task_id, "把当前数据的缺失值分析导出")

    assert response.status_code == 202, response.text
    last = _last_assistant(response)
    assert "缺失分析" in {table["title"] for table in last["metadata"]["tables"]}
    with connect(client.app.state.settings.db_path) as conn:
        refs = {
            str(row["tool_ref"])
            for row in conn.execute(
                """
                SELECT r.tool_ref
                  FROM plan_step_runs r
                  JOIN plans p ON p.id = r.plan_id
                 WHERE p.task_id = ?
                """,
                (task_id,),
            ).fetchall()
        }
    assert refs == {"data_ops.profile_dataset"}
