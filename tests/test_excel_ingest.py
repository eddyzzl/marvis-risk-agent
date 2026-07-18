from pathlib import Path
import re
import zipfile

import pandas as pd
import pytest
from openpyxl import Workbook

from marvis.data import excel_ingest as excel_ingest_module
from marvis.data.errors import DataIngestError, DatasetTooLargeError
from marvis.data.excel_ingest import (
    PREVIEW_ROWS,
    detect_header_rows,
    detect_excel_container_format,
    detect_excel_container_format_from_prefix,
    detect_excel_format,
    flatten_headers,
    ingest_sheet,
    is_excel_workbook,
    list_sheets,
    new_excel_artifact_name,
    probe_sheet_row_count,
    require_excel_format,
)
from tests.xls_fixture import legacy_xls_bytes


def test_ingest_sheet_handles_single_header_and_lists_sheets(tmp_path):
    workbook_path = tmp_path / "book.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame({"id": [1, 2], "score": [0.1, 0.2]}).to_excel(
            writer,
            sheet_name="Main",
            index=False,
        )
        pd.DataFrame({"id": [3], "score": [0.3]}).to_excel(
            writer,
            sheet_name="Feature",
            index=False,
        )

    out_path, report = ingest_sheet(workbook_path, "Main", tmp_path / "out")

    assert list_sheets(workbook_path) == ["Main", "Feature"]
    assert out_path.parent == tmp_path / "out"
    assert re.fullmatch(r"Main_[0-9a-f]{32}\.parquet", out_path.name)
    assert out_path.exists()
    assert report.sheet == "Main"
    assert report.header_rows == 1
    assert report.data_start_row == 1
    assert report.flattened_columns == ["id", "score"]
    assert report.original_shape == (3, 2)
    assert pd.read_parquet(out_path).to_dict("list") == {
        "id": [1, 2],
        "score": [0.1, 0.2],
    }


def test_ingest_legacy_xls_lists_probes_and_converts(tmp_path):
    workbook_path = tmp_path / "legacy.xls"
    workbook_path.write_bytes(legacy_xls_bytes())

    assert list_sheets(workbook_path) == ["Sample", "Feature"]
    assert probe_sheet_row_count(workbook_path, "Sample") == 3

    out_path, report = ingest_sheet(
        workbook_path,
        "Sample",
        tmp_path / "out",
        max_rows=2,
    )

    assert report.sheet == "Sample"
    assert report.header_rows == 1
    assert report.flattened_columns == ["mobile", "bad_flag", "loan_amount"]
    assert pd.read_parquet(out_path).to_dict("records") == [
        {"mobile": "13800138000", "bad_flag": 0, "loan_amount": 1000},
        {"mobile": "13900139000", "bad_flag": 1, "loan_amount": 2000},
    ]
    with pytest.raises(DataIngestError, match="数据行数超过上限"):
        ingest_sheet(
            workbook_path,
            "Sample",
            tmp_path / "too_many",
            max_rows=1,
        )


@pytest.mark.parametrize(
    ("prefix", "expected"),
    [
        (bytes.fromhex("d0cf11e0a1b11ae1") + b"rest", "xls"),
        (b"PK\x03\x04rest", "xlsx"),
        (b"PK\x05\x06rest", "xlsx"),
        (b"PK\x07\x08rest", "xlsx"),
        (b"id,score\n1,0.1\n", None),
    ],
)
def test_excel_container_detection_only_uses_cheap_prefix_evidence(prefix, expected):
    assert detect_excel_container_format_from_prefix(prefix) == expected


def test_malformed_excel_containers_are_detected_but_fail_closed(tmp_path):
    malformed_ole = tmp_path / "malformed-ole.csv"
    malformed_ole.write_bytes(bytes.fromhex("d0cf11e0a1b11ae1") + b"not-a-workbook")

    malformed_ooxml = tmp_path / "malformed-ooxml.csv"
    with zipfile.ZipFile(malformed_ooxml, "w") as archive:
        archive.writestr("[Content_Types].xml", "not xml")
        archive.writestr("xl/workbook.xml", "not xml")

    for path, candidate_format in (
        (malformed_ole, "xls"),
        (malformed_ooxml, "xlsx"),
    ):
        assert detect_excel_container_format(path) == candidate_format
        assert detect_excel_format(path) is None
        assert not is_excel_workbook(path)
        with pytest.raises(DataIngestError, match="not a readable"):
            require_excel_format(path)


def test_valid_ooxml_with_csv_suffix_parses_by_content(tmp_path):
    actual_path = tmp_path / "source.xlsx"
    workbook_path = tmp_path / "misnamed.csv"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Main"
    worksheet.append(["id", "score"])
    worksheet.append([1, 0.25])
    workbook.save(actual_path)
    actual_path.replace(workbook_path)

    assert detect_excel_container_format(workbook_path) == "xlsx"
    assert detect_excel_format(workbook_path) == "xlsx"
    assert require_excel_format(workbook_path) == "xlsx"
    assert list_sheets(workbook_path) == ["Main"]

    out_path, report = ingest_sheet(workbook_path, "Main", tmp_path / "out")

    assert report.flattened_columns == ["id", "score"]
    assert pd.read_parquet(out_path).to_dict("records") == [{"id": 1, "score": 0.25}]


def test_valid_biff_with_csv_suffix_parses_by_content(tmp_path):
    workbook_path = tmp_path / "misnamed.csv"
    workbook_path.write_bytes(legacy_xls_bytes())

    assert detect_excel_container_format(workbook_path) == "xls"
    assert detect_excel_format(workbook_path) == "xls"
    assert list_sheets(workbook_path) == ["Sample", "Feature"]

    out_path, report = ingest_sheet(workbook_path, "Sample", tmp_path / "out")

    assert report.flattened_columns == ["mobile", "bad_flag", "loan_amount"]
    assert len(pd.read_parquet(out_path)) == 2


def test_xlsx_row_guard_ignores_understated_dimension_and_streams_actual_rows(
    tmp_path,
    monkeypatch,
):
    workbook_path = tmp_path / "understated.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Main"
    worksheet.append(["id", "score"])
    worksheet.append([1, 0.1])
    worksheet.append([2, 0.2])
    worksheet.append([3, 0.3])
    workbook.save(workbook_path)

    with zipfile.ZipFile(workbook_path) as source:
        members = {name: source.read(name) for name in source.namelist()}
    sheet_xml = members["xl/worksheets/sheet1.xml"].decode("utf-8")
    members["xl/worksheets/sheet1.xml"] = re.sub(
        r'<dimension ref="[^"]+"\s*/>',
        '<dimension ref="A1:A1"/>',
        sheet_xml,
        count=1,
    ).encode("utf-8")
    with zipfile.ZipFile(workbook_path, "w") as target:
        for name, content in members.items():
            target.writestr(name, content)

    assert probe_sheet_row_count(workbook_path, "Main", stop_after=3) == 3

    def forbid_pandas_excel_read(*_args, **_kwargs):
        raise AssertionError("OOXML must use the bounded read-only iterator")

    monkeypatch.setattr(pd, "read_excel", forbid_pandas_excel_read)

    with pytest.raises(DatasetTooLargeError, match="数据行数超过上限"):
        ingest_sheet(workbook_path, "Main", tmp_path / "out", max_rows=2)

    out_path, _report = ingest_sheet(
        workbook_path,
        "Main",
        tmp_path / "within-limit",
        max_rows=3,
    )

    assert out_path.exists()
    assert pd.read_parquet(out_path).to_dict("records") == [
        {"id": 1, "score": 0.1},
        {"id": 2, "score": 0.2},
        {"id": 3, "score": 0.3},
    ]


def test_xlsx_reader_stops_streaming_at_nrows(tmp_path, monkeypatch):
    class TrackingWorksheet:
        def __init__(self):
            self.rows_yielded = 0
            self.dimensions_reset = False

        def reset_dimensions(self):
            self.dimensions_reset = True

        def iter_rows(self, *, values_only):
            assert values_only
            for index in range(100):
                self.rows_yielded += 1
                yield (index, index * 2)

    worksheet = TrackingWorksheet()

    class FakeWorkbook:
        def __getitem__(self, sheet):
            assert sheet == "Main"
            return worksheet

    class FakeWorkbookContext:
        def __enter__(self):
            return FakeWorkbook()

        def __exit__(self, *_args):
            return False

    monkeypatch.setattr(
        excel_ingest_module,
        "_open_xlsx_workbook",
        lambda _path: FakeWorkbookContext(),
    )

    frame = excel_ingest_module._read_excel_frame(
        tmp_path / "unused.xlsx",
        workbook_format="xlsx",
        sheet="Main",
        nrows=3,
    )

    assert worksheet.dimensions_reset
    assert worksheet.rows_yielded == 3
    assert frame.to_dict("records") == [
        {0: 0, 1: 0},
        {0: 1, 1: 2},
        {0: 2, 1: 4},
    ]


def test_biff_pandas_reads_are_independently_bounded(tmp_path, monkeypatch):
    workbook_path = tmp_path / "legacy.xls"
    workbook_path.write_bytes(legacy_xls_bytes())
    real_read_excel = pd.read_excel
    observed_nrows = []

    def bounded_read_excel(*args, **kwargs):
        observed_nrows.append(kwargs.get("nrows"))
        assert kwargs.get("nrows") is not None
        return real_read_excel(*args, **kwargs)

    monkeypatch.setattr(pd, "read_excel", bounded_read_excel)

    out_path, _report = ingest_sheet(
        workbook_path,
        "Sample",
        tmp_path / "out",
        max_rows=2,
    )

    assert out_path.exists()
    assert observed_nrows == [PREVIEW_ROWS, 2 + 1 + 1]


def test_excel_artifact_names_are_safe_and_unique_across_calls():
    first = new_excel_artifact_name("../../客户/名单:*?")
    second = new_excel_artifact_name("../../客户/名单:*?")

    assert first != second
    assert Path(first).name == first
    assert Path(second).name == second
    assert re.fullmatch(r"客户_名单_[0-9a-f]{32}\.parquet", first)
    assert re.fullmatch(r"客户_名单_[0-9a-f]{32}\.parquet", second)


def test_ingest_sheet_flattens_merged_two_row_headers(tmp_path):
    workbook_path = tmp_path / "merged.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Merged"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Customer"
    sheet["C1"] = "Outcome"
    sheet.append(["ID", "Phone", "Target"])
    sheet.append(["A1", "13800138000", 1])
    sheet.append(["B2", "13900139000", 0])
    workbook.save(workbook_path)

    out_path, report = ingest_sheet(workbook_path, "Merged", tmp_path / "out")

    assert report.header_rows == 2
    assert report.flattened_columns == [
        "Customer_ID",
        "Customer_Phone",
        "Outcome_Target",
    ]
    joined = pd.read_parquet(out_path)
    assert joined.columns.tolist() == report.flattened_columns
    assert joined["Customer_ID"].tolist() == ["A1", "B2"]


def test_header_detection_and_duplicate_disambiguation():
    raw = pd.DataFrame([
        ["id", "id", "group"],
        [1, 2, "A"],
        [3, 4, "B"],
    ])

    assert detect_header_rows(raw) == 1
    data, columns = flatten_headers(raw, 1)

    assert columns == ["id", "id_2", "group"]
    assert data.columns.tolist() == columns
    assert data.to_dict("records")[0] == {"id": 1, "id_2": 2, "group": "A"}


def test_ingest_sheet_rejects_empty_sheet(tmp_path):
    workbook_path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    workbook.active.title = "Empty"
    workbook.save(workbook_path)

    with pytest.raises(DataIngestError):
        ingest_sheet(workbook_path, "Empty", tmp_path / "out")


def test_ingest_sheet_wraps_missing_sheet_errors(tmp_path):
    workbook_path = tmp_path / "book.xlsx"
    pd.DataFrame({"id": [1]}).to_excel(workbook_path, sheet_name="Present", index=False)

    with pytest.raises(DataIngestError):
        ingest_sheet(Path(workbook_path), "Missing", tmp_path / "out")


def test_legacy_xls_parser_errors_are_wrapped(tmp_path):
    fake_xls = tmp_path / "fake.xls"
    fake_xls.write_text("<html><table><tr><td>not BIFF</td></tr></table></html>")

    with pytest.raises(DataIngestError, match="not a readable"):
        list_sheets(fake_xls)
    with pytest.raises(DataIngestError, match="not a readable"):
        ingest_sheet(fake_xls, "Sheet1", tmp_path / "out")


def test_biff_parser_errors_are_normalized_and_binary_handle_is_closed(
    tmp_path,
    monkeypatch,
):
    workbook_path = tmp_path / "book.xls"
    workbook_path.write_bytes(legacy_xls_bytes())
    observed_handles = []

    def fail_read_excel(handle, **_kwargs):
        observed_handles.append(handle)
        assert not handle.closed
        raise RuntimeError("parser exploded")

    monkeypatch.setattr(excel_ingest_module.pd, "read_excel", fail_read_excel)

    with pytest.raises(DataIngestError, match="cannot read sheet.*parser exploded"):
        ingest_sheet(workbook_path, "Sample", tmp_path / "out")

    assert len(observed_handles) == 1
    assert observed_handles[0].closed


# --- GAP-1: flag long numeric-id columns Excel already stored as Number ----


def test_ingest_sheet_flags_suspected_truncated_long_id_column(tmp_path):
    workbook_path = tmp_path / "ids.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sample"
    sheet.append(["id_card", "bad_flag"])
    # Written as a numeric cell (not text) -- Excel stores this as an IEEE754
    # double at the file-format level, so precision is already lost before
    # openpyxl/pandas ever reads it.
    sheet.append([110101199001011234, 0])
    sheet.append([110101199001015678, 1])
    workbook.save(workbook_path)

    out_path, report = ingest_sheet(workbook_path, "Sample", tmp_path / "out")

    assert report.suspected_truncated_id_columns == ("id_card",)
    assert out_path.exists()


def test_ingest_sheet_does_not_flag_text_stored_ids(tmp_path):
    workbook_path = tmp_path / "text_ids.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sample"
    sheet.append(["id_card", "bad_flag"])
    sheet["A2"] = "110101199001011234"
    sheet["A2"].number_format = "@"
    sheet["B2"] = 0
    sheet["A3"] = "110101199001015678"
    sheet["A3"].number_format = "@"
    sheet["B3"] = 1
    workbook.save(workbook_path)

    _out_path, report = ingest_sheet(workbook_path, "Sample", tmp_path / "out")

    assert report.suspected_truncated_id_columns == ()
