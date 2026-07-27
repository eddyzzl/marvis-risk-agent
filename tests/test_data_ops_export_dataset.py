from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
import sys
from pathlib import Path
import threading

import pandas as pd
from openpyxl import load_workbook

from marvis.data.backend import DataBackend
from marvis.data.registry import DatasetRegistry
from marvis.data.workspace import (
    DataSemanticMapping,
    DataWorkspaceDraft,
    data_semantic_mapping_hash,
)
from marvis.db import DatasetRepository, PluginRepository, init_db
from marvis.db_schema import connect
from marvis.files import sha256_file
from marvis.plugins.loader import load_builtin_packs
from marvis.plugins.manifest import ToolRef
from marvis.plugins.registry import PluginRegistry, ToolRegistry
from marvis.plugins.runner import ToolRunner
from marvis.repositories.data_workspace import DataWorkspaceRepository
from marvis.repositories.task_artifacts import TaskArtifactRepository
from marvis.settings import build_settings
import marvis.packs.data_ops.tools as data_ops_tools


_CREATED_AT = "2026-07-19T08:00:00+00:00"


def _runtime(tmp_path):
    settings = build_settings(tmp_path / "workspace")
    init_db(settings.db_path)
    with connect(settings.db_path) as conn:
        conn.execute(
            """
            INSERT INTO tasks(
                id, task_type, model_name, model_version, validator, source_dir,
                status, status_message, created_at, updated_at
            ) VALUES (
                'task-export', 'strategy', 'export task', 'v1', 'tester',
                '/tmp/source', 'created', 'created', ?, ?
            )
            """,
            (_CREATED_AT, _CREATED_AT),
        )
    plugin_repo = PluginRepository(settings.db_path)
    plugin_registry = PluginRegistry(plugin_repo)
    packs_root = Path(__file__).parents[1] / "marvis" / "packs"
    load_builtin_packs(plugin_registry, packs_root)
    runner = ToolRunner(
        ToolRegistry(plugin_registry),
        plugin_repo,
        python_executable=sys.executable,
        datasets_root=settings.datasets_dir,
        workspace=settings.workspace,
    )
    backend = DataBackend(settings.datasets_dir)
    registry = DatasetRegistry(
        DatasetRepository(settings.db_path),
        backend,
        settings.datasets_dir,
    )
    return settings, runner, backend, registry


def _dataset(tmp_path, registry):
    source = tmp_path / "source.parquet"
    pd.DataFrame(
        {
            "customer_name": ["张三", "=2+2"],
            "mobile": ["0013800000000", "+8613900000000"],
            "order_id": [9007199254740993, 2],
            "amount": [100.0, 200.0],
            "bad": [0, 1],
        }
    ).to_parquet(source, index=False)
    return registry.register_from_upload(
        "task-export",
        source,
        role="strategy_sample",
    )


def _workspace(settings, dataset):
    repo = DataWorkspaceRepository(settings.db_path)
    active = repo.save(
        "task-export",
        DataWorkspaceDraft(
            active_dataset_id=dataset.id,
            active_dataset_content_hash=dataset.content_hash,
        ),
        expected_revision=0,
    )
    return repo.save(
        "task-export",
        DataWorkspaceDraft(
            active_dataset_id=dataset.id,
            active_dataset_content_hash=dataset.content_hash,
            page="overview",
            semantic_mapping=DataSemanticMapping(
                target_col="bad",
                field_roles={
                    "customer_name": "name",
                    "mobile": "phone",
                    "order_id": "id",
                    "bad": "target",
                },
                business_names={"amount": "申请金额"},
            ),
        ),
        expected_revision=active.revision,
    )


def _inputs(dataset, workspace, format_name):
    return {
        "dataset_id": dataset.id,
        "expected_content_hash": dataset.content_hash,
        "workspace_revision": workspace.revision,
        "analysis_generation": workspace.analysis_generation,
        "semantic_mapping_hash": data_semantic_mapping_hash(
            workspace.semantic_mapping
        ),
        "format": format_name,
    }


def test_export_dataset_tool_registers_safe_task_owned_csv_artifact(tmp_path):
    settings, runner, _backend, registry = _runtime(tmp_path)
    dataset = _dataset(tmp_path, registry)
    workspace = _workspace(settings, dataset)

    result = runner.invoke(
        ToolRef("data_ops", "export_dataset"),
        _inputs(dataset, workspace, "csv"),
        task_id="task-export",
    )

    assert result.ok is True, result.error
    output = result.output
    assert output["schema_version"] == "dataset-export-tool-result.v1"
    assert output["dataset_id"] == dataset.id
    assert output["dataset_content_hash"] == dataset.content_hash
    assert output["format"] == "csv"
    assert output["row_count"] == 2
    assert output["column_count"] == 5
    assert output["cached"] is False
    assert output["download_url"] == (
        f"/api/tasks/task-export/task-artifacts/{output['artifact_id']}/download"
    )
    artifacts = TaskArtifactRepository(settings.db_path).list_for_task(
        "task-export"
    )
    assert len(artifacts) == 1
    artifact = artifacts[0]
    assert artifact["id"] == output["artifact_id"]
    assert artifact["kind"] == "dataset_export"
    assert artifact["origin_tool"] == "data_ops.export_dataset"
    assert artifact["content_hash"] == output["content_hash"]
    assert artifact["provenance"]["dataset_id"] == dataset.id
    path = Path(artifact["path"])
    assert path.is_relative_to(settings.tasks_dir / "task-export")
    assert path.read_bytes().startswith(b"\xef\xbb\xbf")
    assert sha256_file(path) == output["content_hash"]
    text = path.read_text(encoding="utf-8-sig")
    assert "'=2+2" in text
    assert "0013800000000" in text
    assert "'9007199254740993" in text
    assert output["safety"]["formula_cells_escaped"] >= 2


def test_export_dataset_tool_writes_excel_text_cells_and_replays_exact_input(tmp_path):
    settings, runner, _backend, registry = _runtime(tmp_path)
    dataset = _dataset(tmp_path, registry)
    workspace = _workspace(settings, dataset)
    inputs = _inputs(dataset, workspace, "xlsx")

    first = runner.invoke(
        ToolRef("data_ops", "export_dataset"),
        inputs,
        task_id="task-export",
    )
    replay = runner.invoke(
        ToolRef("data_ops", "export_dataset"),
        inputs,
        task_id="task-export",
    )

    assert first.ok is True, first.error
    assert replay.ok is True, replay.error
    assert replay.output == {**first.output, "cached": True}
    artifacts = TaskArtifactRepository(settings.db_path).list_for_task(
        "task-export"
    )
    assert len(artifacts) == 1
    path = Path(artifacts[0]["path"])
    assert path.suffix == ".xlsx"
    workbook = load_workbook(path, read_only=True, data_only=False)
    rows = list(workbook["data"].iter_rows(values_only=False))
    assert rows[1][1].value == "0013800000000"
    assert rows[1][1].data_type == "s"
    assert rows[1][2].value == "9007199254740993"
    assert rows[1][2].data_type == "s"


def test_export_dataset_tool_fails_closed_for_stale_or_foreign_binding(tmp_path):
    settings, runner, _backend, registry = _runtime(tmp_path)
    dataset = _dataset(tmp_path, registry)
    workspace = _workspace(settings, dataset)

    stale = runner.invoke(
        ToolRef("data_ops", "export_dataset"),
        {**_inputs(dataset, workspace, "csv"), "workspace_revision": 999},
        task_id="task-export",
    )
    foreign = runner.invoke(
        ToolRef("data_ops", "export_dataset"),
        _inputs(dataset, workspace, "csv"),
        task_id="another-task",
    )

    assert stale.ok is False
    assert "workspace revision mismatch" in stale.error
    assert foreign.ok is False
    assert "belongs to task" in foreign.error
    assert TaskArtifactRepository(settings.db_path).list_for_task(
        "task-export"
    ) == []


def test_concurrent_identical_exports_commit_one_intact_artifact(tmp_path, monkeypatch):
    settings, runner, _backend, registry = _runtime(tmp_path)
    dataset = _dataset(tmp_path, registry)
    workspace = _workspace(settings, dataset)
    inputs = _inputs(dataset, workspace, "csv")
    original_export = data_ops_tools.export_dataset
    both_computed = threading.Barrier(2)

    def synchronized_export(*args, **kwargs):
        result = original_export(*args, **kwargs)
        both_computed.wait(timeout=15)
        return result

    monkeypatch.setattr(data_ops_tools, "export_dataset", synchronized_export)
    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = [
            executor.submit(
                runner.invoke,
                ToolRef("data_ops", "export_dataset"),
                inputs,
                task_id="task-export",
            )
            for _ in range(2)
        ]
        results = [future.result(timeout=30) for future in futures]

    assert all(result.ok for result in results), [result.error for result in results]
    assert sorted(result.output["cached"] for result in results) == [False, True]
    artifacts = TaskArtifactRepository(settings.db_path).list_for_task(
        "task-export"
    )
    assert len(artifacts) == 1
    path = Path(artifacts[0]["path"])
    assert path.is_file()
    assert sha256_file(path) == artifacts[0]["content_hash"]
    assert not list(settings.tasks_dir.rglob("*.bak"))


def test_export_dataset_rejects_symlinked_task_artifact_directory(tmp_path):
    settings, runner, _backend, registry = _runtime(tmp_path)
    dataset = _dataset(tmp_path, registry)
    workspace = _workspace(settings, dataset)
    outside = tmp_path / "outside"
    outside.mkdir()
    task_root = settings.tasks_dir / "task-export"
    task_root.mkdir(parents=True, exist_ok=True)
    (task_root / "data_exports").symlink_to(outside, target_is_directory=True)

    result = runner.invoke(
        ToolRef("data_ops", "export_dataset"),
        _inputs(dataset, workspace, "csv"),
        task_id="task-export",
    )

    assert result.ok is False
    assert "must not be a symlink" in result.error
    assert list(outside.iterdir()) == []
    assert TaskArtifactRepository(settings.db_path).list_for_task(
        "task-export"
    ) == []
