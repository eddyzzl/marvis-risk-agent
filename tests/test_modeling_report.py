import sys
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

from marvis.data.backend import DataBackend
from marvis.data.registry import DatasetRegistry
from marvis.feature.binning import equal_frequency_edges
from marvis.feature.metrics import DEFAULT_IV_BINS, feature_psi
from marvis.db import DatasetRepository, PluginRepository, TaskRepository, init_db
from marvis.domain import TaskCreate
from marvis.output.model_report import ModelReportPayload, render_model_report
from marvis.output.model_report_minimal import render_minimal_model_report
from marvis.packs.modeling.report_compute import (
    BusinessColumns,
    compute_amount_bin_table,
    compute_sample_analysis,
    compute_vintage_report,
    resolve_report_sections,
    stress_low_pricing,
)
from marvis.packs.modeling.contracts import ModelArtifact, TrainConfig
from marvis.packs.modeling.experiment import ExperimentStore
from marvis.packs.modeling import tools as modeling_tools
from marvis.plugins.loader import load_builtin_packs
from marvis.plugins.manifest import ToolRef
from marvis.plugins.registry import PluginRegistry, ToolRegistry
from marvis.plugins.runner import ToolRunner
from marvis.settings import build_settings


def _business_frame() -> pd.DataFrame:
    return pd.DataFrame({
        "loan_month": ["2026-01", "2026-01", "2026-02", "2026-02"],
        "rate": [0.12, 0.18, 0.10, 0.20],
        "amount": [1000.0, 2000.0, 1500.0, 2500.0],
        "term": [6, 12, 6, 12],
        "drawdown": [800.0, 1500.0, 1200.0, 2000.0],
        "limit": [2000.0, 3000.0, 2000.0, 4000.0],
        "mob1": [0, 1, 0, 0],
        "mob2": [0, 1, 0, 1],
        "mob3": [1, 1, 0, 1],
        "y": [0, 1, 0, 1],
        "score": [0.1, 0.8, 0.2, 0.7],
    })


def test_model_report_compute_functions_are_deterministic(tmp_path):
    path = tmp_path / "sample.parquet"
    _business_frame().to_parquet(path, index=False)
    backend = DataBackend(tmp_path)
    business = BusinessColumns(
        loan_month_col="loan_month",
        interest_rate_col="rate",
        loan_amount_col="amount",
        term_col="term",
        drawdown_amount_col="drawdown",
        credit_limit_col="limit",
        mob_observe_cols=("mob1", "mob2", "mob3"),
    )

    sample = compute_sample_analysis(
        backend,
        path,
        loan_month_col="loan_month",
        target_col="y",
        business=business,
        mob_cols=business.mob_observe_cols,
    )
    vintage = compute_vintage_report(
        backend,
        path,
        loan_month_col="loan_month",
        mob_observe_cols=business.mob_observe_cols,
        amount_col="amount",
    )
    low_pricing = stress_low_pricing(
        backend,
        path,
        score_col="score",
        target_col="y",
        interest_rate_col="rate",
        low_pricing_threshold=None,
        ratios=(0.25, 0.5),
    )

    assert sample[0]["放款月"] == "2026-01"
    assert sample[0]["放款笔数"] == 2
    assert sample[0]["平均利率"] == 0.15
    assert sample[0]["Mob3逾期率"] == 1.0
    assert vintage["headers"] == ["mob1", "mob2", "mob3"]
    assert vintage["curves"]["2026-01"] == sorted(vintage["curves"]["2026-01"])
    assert vintage["counts"] == {"2026-01": 2, "2026-02": 2}
    assert vintage["amounts"] == {
        "2026-01": {"total": 3000.0, "average": 1500.0},
        "2026-02": {"total": 4000.0, "average": 2000.0},
    }
    assert low_pricing == stress_low_pricing(
        backend,
        path,
        score_col="score",
        target_col="y",
        interest_rate_col="rate",
        low_pricing_threshold=None,
        ratios=(0.25, 0.5),
    )
    assert set(low_pricing["by_ratio"]) == {"0.25", "0.5"}


def test_compute_vintage_report_aligns_headers_to_sorted_mob_axis(tmp_path):
    path = tmp_path / "vintage_unsorted.parquet"
    pd.DataFrame({
        "loan_month": ["2026-01", "2026-01"],
        "mob3": [1, 1],
        "mob1": [0, 1],
    }).to_parquet(path, index=False)

    vintage = compute_vintage_report(
        DataBackend(tmp_path),
        path,
        loan_month_col="loan_month",
        mob_observe_cols=("mob3", "mob1"),
        amount_col=None,
    )

    assert vintage["headers"] == ["mob1", "mob3"]
    assert vintage["curves"]["2026-01"] == pytest.approx([0.5, 1.0])


def test_amount_bin_table_computes_credit_utilization_by_bin(tmp_path):
    frame = pd.DataFrame({
        "score": [0.1, 0.2, 0.8, 0.9],
        "y": [0, 1, 0, 1],
        "drawdown": [50.0, 100.0, 300.0, 100.0],
        "limit": [100.0, 100.0, 400.0, 100.0],
    })
    path = tmp_path / "amount_bins.parquet"
    frame.to_parquet(path, index=False)

    rows = compute_amount_bin_table(
        DataBackend(tmp_path),
        path,
        score_col="score",
        target_col="y",
        edges=[0.0, 0.5, 1.0],
        business=BusinessColumns(
            drawdown_amount_col="drawdown",
            credit_limit_col="limit",
        ),
    )

    by_bin = {row["bin_index"]: row for row in rows}
    assert by_bin[1]["额度使用率"] == pytest.approx(0.75)
    assert by_bin[2]["额度使用率"] == pytest.approx(0.8)


def test_amount_bin_table_computes_amount_weighted_cumulative_and_lift(tmp_path):
    frame = pd.DataFrame({
        "score": [0.1, 0.2, 0.8, 0.9],
        "y": [0, 1, 0, 1],
        "amount": [100.0, 100.0, 300.0, 100.0],
    })
    path = tmp_path / "amount_weighted_bins.parquet"
    frame.to_parquet(path, index=False)

    rows = compute_amount_bin_table(
        DataBackend(tmp_path),
        path,
        score_col="score",
        target_col="y",
        edges=[0.0, 0.5, 1.0],
        business=BusinessColumns(loan_amount_col="amount"),
    )

    by_bin = {row["bin_index"]: row for row in rows}
    assert by_bin[1]["金额逾期率"] == pytest.approx(0.5)
    assert by_bin[1]["累计金额逾期率"] == pytest.approx(0.5)
    assert by_bin[1]["金额lift"] == pytest.approx(1.5)
    assert by_bin[2]["金额逾期率"] == pytest.approx(0.25)
    assert by_bin[2]["累计金额逾期率"] == pytest.approx(1 / 3)
    assert by_bin[2]["金额lift"] == pytest.approx(0.75)


def test_report_bin_table_uses_only_oot_split(tmp_path):
    frame = pd.DataFrame({
        "score": [0.10, 0.20, 0.30, 0.40, 0.90, 0.95],
        "y": [0, 0, 0, 1, 1, 1],
        "split": ["train", "train", "test", "test", "oot", "oot"],
    })
    path = tmp_path / "oot_bins.parquet"
    frame.to_parquet(path, index=False)
    runtime = type("Runtime", (), {"backend": DataBackend(tmp_path)})()
    config = TrainConfig(
        dataset_id="dataset-1",
        features=("score",),
        target_col="y",
        split_col="split",
        split_values={"train": "train", "test": "test", "oot": "oot"},
        params={},
        seed=1,
        early_stopping_rounds=None,
    )

    rows = modeling_tools._report_bin_table(
        runtime,
        path,
        score_col="score",
        target_col="y",
        config=config,
        business=BusinessColumns(),
    )

    assert sum(row["sample_count"] for row in rows) == 2
    assert sum(row["bad_count"] for row in rows) == 2


def test_amount_bin_table_excludes_unscored_rows_from_business_amounts(tmp_path):
    frame = pd.DataFrame({
        "score": [0.1, float("nan"), 0.8],
        "y": [0, 1, 1],
        "amount": [100.0, 1000.0, 300.0],
    })
    path = tmp_path / "amount_bins_with_nan.parquet"
    frame.to_parquet(path, index=False)

    rows = compute_amount_bin_table(
        DataBackend(tmp_path),
        path,
        score_col="score",
        target_col="y",
        edges=[0.0, 0.5, 1.0],
        business=BusinessColumns(loan_amount_col="amount"),
    )

    assert sum(row["sample_count"] for row in rows) == 2
    assert rows[-1]["累计金额逾期率"] == pytest.approx(300 / 400)


def test_stress_low_pricing_exposes_cumulative_bin_curves_by_ratio(tmp_path):
    path = tmp_path / "low_pricing.parquet"
    _business_frame().to_parquet(path, index=False)

    result = stress_low_pricing(
        DataBackend(tmp_path),
        path,
        score_col="score",
        target_col="y",
        interest_rate_col="rate",
        low_pricing_threshold=None,
        ratios=(0.25, 0.5),
    )

    assert set(result["bins_by_ratio"]) == {"0.25", "0.5"}
    for curve in result["bins_by_ratio"].values():
        assert curve == sorted(curve)
        assert curve[-1] == pytest.approx(1.0)


def test_stress_low_pricing_exposes_flat_metric_indexes_by_ratio(tmp_path):
    path = tmp_path / "low_pricing.parquet"
    _business_frame().to_parquet(path, index=False)

    result = stress_low_pricing(
        DataBackend(tmp_path),
        path,
        score_col="score",
        target_col="y",
        interest_rate_col="rate",
        low_pricing_threshold=None,
        ratios=(0.25, 0.5),
    )

    assert result["ks_by_ratio"] == {
        ratio: metrics["ks"]
        for ratio, metrics in result["by_ratio"].items()
    }
    assert result["psi_by_ratio"] == {
        ratio: metrics["psi"]
        for ratio, metrics in result["by_ratio"].items()
    }


def test_stress_low_pricing_exposes_conclusion_data_for_report_narratives(tmp_path):
    path = tmp_path / "low_pricing.parquet"
    _business_frame().to_parquet(path, index=False)

    result = stress_low_pricing(
        DataBackend(tmp_path),
        path,
        score_col="score",
        target_col="y",
        interest_rate_col="rate",
        low_pricing_threshold=None,
        ratios=(0.25, 0.5),
    )

    conclusion = result["conclusion_data"]
    assert conclusion["threshold"] == result["threshold"]
    assert conclusion["baseline_low_pricing_ratio"] == pytest.approx(0.5)
    assert conclusion["max_psi_ratio"] in result["psi_by_ratio"]
    assert conclusion["max_psi"] == result["psi_by_ratio"][conclusion["max_psi_ratio"]]
    assert conclusion["min_ks_ratio"] in result["ks_by_ratio"]
    assert conclusion["min_ks"] == result["ks_by_ratio"][conclusion["min_ks_ratio"]]
    assert conclusion["max_ks_drop"] == pytest.approx(conclusion["baseline_ks"] - conclusion["min_ks"])


def test_resolve_sections_and_render_model_report_degrades_missing_business_data(tmp_path):
    statuses = resolve_report_sections(BusinessColumns(), dictionary_id=None)
    output = tmp_path / "model_report.xlsx"
    render_model_report(
        ModelReportPayload(
            project_meta={"项目名称": "建模报告"},
            dataset_split=[{"split": "train", "ks": 0.3}],
            stability=[{"metric": "psi", "value": 0.01}],
            sample_analysis=None,
            vintage=None,
            feature_importance=[{"feature": "x1", "importance": 0.7}],
            scorecard_table=[{"feature": "x1", "bin_label": "[0, 1)", "points": 12.3}],
            score_bands=[{
                "split": "oot",
                "bin": 1,
                "bad_rate": 0.1,
                "bin_edges_source": "train",
                "cum_direction": "higher_is_riskier",
            }],
            univariate=[{"feature": "x1", "iv": 0.2, "ks": 0.3}],
            oot_bin_table=[{"bin": 1, "bad_rate": 0.1}],
            stress_product_removal={"baseline": []},
            stress_low_pricing=None,
            narratives={"summary": "受控模板文本"},
            section_status=statuses,
        ),
        output,
    )

    workbook = load_workbook(output)
    assert not (output.parent / ".staging").exists()
    assert workbook.sheetnames == [
        "汇总",
        "样本分析",
        "Vintage",
        "特征重要性",
        "评分卡",
        "评分分段",
        "概率校准",
        "oot分箱评估_十分箱",
        "单变量分析",
        "压力测试",
    ]
    assert workbook["样本分析"]["A1"].value.startswith("无业务数据")
    assert workbook["评分卡"]["A1"].value == "feature"
    assert workbook["评分卡"]["B2"].value == "[0, 1)"
    assert workbook["评分卡"]["C2"].value == 12.3
    assert "分箱边界口径" in workbook["评分分段"]["A1"].value
    assert workbook["评分分段"]["A4"].value == "split"
    assert workbook["评分分段"]["C5"].value == 0.1
    assert any(status.section == "product_list" and not status.available for status in statuses)


def test_non_binary_model_report_keeps_fixed_sheets_and_adds_metrics(tmp_path):
    from marvis.packs.modeling.contracts import Experiment, ModelMetrics, TrainConfig

    config = TrainConfig(
        dataset_id="ds",
        features=("x1", "x2"),
        target_col="income",
        split_col="split",
        split_values={"train": "train", "test": "test", "oot": "oot"},
        params={},
        seed=23,
        early_stopping_rounds=None,
        recipe_id="lgb_regressor",
        target_type="continuous",
    )
    metrics = ModelMetrics(
        train_ks=None, test_ks=None, oot_ks=None,
        train_auc=None, test_auc=None, oot_auc=None,
        psi_test_vs_train=None, psi_oot_vs_train=None,
        overfit_train_test_gap=0.1, overfit_train_oot_gap=0.2, overfit_flag=False,
        train_rmse=1.0, test_rmse=1.2, oot_rmse=1.3,
        train_mae=0.8, test_mae=0.9, oot_mae=1.0,
        train_r2=0.7, test_r2=0.6, oot_r2=0.5,
    )
    experiment = Experiment(
        id="exp", task_id="task", recipe_id="lgb_regressor", config=config,
        metrics=metrics, artifact_id="artifact", status="succeeded", created_at="now",
    )
    out = tmp_path / "non_binary.xlsx"

    render_minimal_model_report(experiment, out)

    workbook = load_workbook(out)
    assert not (out.parent / ".staging").exists()
    assert workbook.sheetnames == [
        "汇总",
        "样本分析",
        "Vintage",
        "特征重要性",
        "评分卡",
        "评分分段",
        "概率校准",
        "oot分箱评估_十分箱",
        "单变量分析",
        "压力测试",
        "模型指标",
    ]
    assert workbook["Vintage"]["B1"].value == "非二分类不适用"
    assert workbook["模型指标"]["A2"].value == "RMSE"
    assert workbook["样本分析"]["A1"].value == "n/a"


def test_render_model_report_summary_lists_unique_products_from_feature_dictionary(tmp_path):
    output = tmp_path / "model_report.xlsx"
    render_model_report(
        ModelReportPayload(
            project_meta={"项目名称": "建模报告"},
            dataset_split=[],
            stability=[],
            sample_analysis=[],
            vintage=None,
            feature_importance=[
                {"feature": "x1", "importance": 0.7, "产品名称": "征信评分", "厂商名称": "数据厂商A"},
                {"feature": "x2", "importance": 0.3, "产品名称": "征信评分", "厂商名称": "数据厂商A"},
                {"feature": "x3", "importance": 0.1, "产品名称": "借贷画像", "厂商名称": "数据厂商B"},
            ],
            scorecard_table=[],
            score_bands=[],
            univariate=[],
            oot_bin_table=[],
            stress_product_removal={},
            stress_low_pricing=None,
            narratives={},
            section_status=[],
        ),
        output,
    )

    summary = load_workbook(output)["汇总"]
    rows = {summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value for row in range(1, summary.max_row + 1)}
    assert rows["五、使用产品清单"] == "征信评分（数据厂商A）；借贷画像（数据厂商B）"


def test_render_model_report_includes_vintage_cohort_counts_and_amounts(tmp_path):
    output = tmp_path / "model_report.xlsx"
    render_model_report(
        ModelReportPayload(
            project_meta={"项目名称": "建模报告"},
            dataset_split=[],
            stability=[],
            sample_analysis=[],
            vintage={
                "headers": ["mob1"],
                "curves": {"2026-01": [0.5]},
                "counts": {"2026-01": 2},
                "amounts": {"2026-01": {"total": 3000.0, "average": 1500.0}},
            },
            feature_importance=[],
            scorecard_table=[],
            score_bands=[],
            univariate=[],
            oot_bin_table=[],
            stress_product_removal={},
            stress_low_pricing=None,
            narratives={},
            section_status=[],
        ),
        output,
    )

    sheet = load_workbook(output)["Vintage"]
    assert [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value, sheet["D1"].value, sheet["E1"].value] == [
        "放款月",
        "放款笔数",
        "放款金额",
        "件均金额",
        "mob1",
    ]
    assert [sheet["A2"].value, sheet["B2"].value, sheet["C2"].value, sheet["D2"].value, sheet["E2"].value] == [
        "2026-01",
        2,
        3000.0,
        1500.0,
        0.5,
    ]


def test_report_narrative_guard_removes_numbers_not_in_structured_summary():
    narratives = {
        "model": "训练 KS 为 0.3，AUC 可达 0.91，建议拒绝前 20% 客群。",
        "stress": "PSI 为 0.01，预计收益提升 12.5%。",
    }
    structured_summary = {
        "dataset_split": [{"split": "train", "ks": 0.3, "auc": 0.75}],
        "stability": [{"metric": "psi", "value": 0.01}],
    }

    guarded = modeling_tools._guard_no_invented_numbers(narratives, structured_summary)

    assert "0.3" in guarded["model"]
    assert "0.01" in guarded["stress"]
    assert "0.91" not in guarded["model"]
    assert "20%" not in guarded["model"]
    assert "12.5%" not in guarded["stress"]
    assert "[平台未提供该数字]" in guarded["model"]


def test_report_narrative_guard_rejects_user_meta_numbers_as_metrics():
    structured_summary = {
        "project_meta": {"目标AUC": "0.91"},
        "dataset_split": [{"split": "train", "ks": 0.3, "auc": 0.75}],
    }

    guarded = modeling_tools._guard_no_invented_numbers(
        {"model": "模型 AUC 为 0.91。"},
        structured_summary,
    )

    assert "0.91" not in guarded["model"]
    assert "[平台未提供该数字]" in guarded["model"]


def test_report_narrative_guard_rejects_unit_changed_percent_tokens():
    structured_summary = {"dataset_split": [{"ks": 0.3, "sample_count": 20}]}

    guarded = modeling_tools._guard_no_invented_numbers(
        {"model": "KS 为 0.3%，建议拒绝前 20% 客群。"},
        structured_summary,
    )

    assert "0.3%" not in guarded["model"]
    assert "20%" not in guarded["model"]
    assert guarded["model"].count("[平台未提供该数字]") == 2


def test_report_narrative_drafter_uses_llm_json_when_available():
    calls = []

    class FakeLLM:
        def complete(self, **kwargs):
            calls.append(kwargs)
            return '{"sample":"样本覆盖 2 个放款月。","model":"模型 KS 为 0.3。"}'

    narratives = modeling_tools._draft_report_narratives(
        {"dataset_split": [{"split": "train", "ks": 0.3}]},
        llm_factory=lambda: FakeLLM(),
    )

    assert narratives["sample"] == "样本覆盖 2 个放款月。"
    assert narratives["model"] == "模型 KS 为 0.3。"
    assert narratives["vintage"]
    assert calls[0]["response_format"] == {"type": "json_object"}
    assert calls[0]["stream"] is False


def test_feature_importance_rows_keep_dictionary_columns_when_first_feature_has_no_metadata():
    artifact = ModelArtifact(
        id="artifact_1",
        experiment_id="experiment_1",
        algorithm="lr",
        model_path="model.pkl",
        pmml_path=None,
        feature_list=("x_missing", "x2"),
        params={},
        woe_maps=None,
        created_at="2026-01-01T00:00:00+00:00",
    )

    rows = modeling_tools._feature_importance_rows(
        artifact,
        feature_dictionary={
            "x2": {
                "含义": "负债压力",
                "产品名称": "借贷画像",
                "厂商名称": "数据厂商B",
            }
        },
    )

    assert rows[0] == {
        "feature": "x_missing",
        "importance": 0.0,
        "importance_pct": 0.0,
        "cumulative_importance_pct": 0.0,
        "含义": None,
        "产品名称": None,
        "厂商名称": None,
    }
    assert rows[1]["含义"] == "负债压力"


def test_feature_importance_rows_compute_percentage_and_cumulative_share():
    artifact = ModelArtifact(
        id="artifact_1",
        experiment_id="experiment_1",
        algorithm="lr",
        model_path="model.pkl",
        pmml_path=None,
        feature_list=("x1", "x2", "x3"),
        params={},
        woe_maps=None,
        created_at="2026-01-01T00:00:00+00:00",
        feature_importance=(("x1", 2.0), ("x2", 1.0), ("x3", 1.0)),
    )

    rows = modeling_tools._feature_importance_rows(artifact)

    assert [row["importance_pct"] for row in rows] == pytest.approx([0.5, 0.25, 0.25])
    assert [row["cumulative_importance_pct"] for row in rows] == pytest.approx([0.5, 0.75, 1.0])


def test_generate_model_report_tool_round_trips_via_runner(tmp_path):
    settings = build_settings(tmp_path / "workspace")
    init_db(settings.db_path)
    plugin_repo = PluginRepository(settings.db_path)
    plugin_registry = PluginRegistry(plugin_repo)
    load_builtin_packs(plugin_registry, Path(__file__).parents[1] / "marvis" / "packs")
    runner = ToolRunner(
        ToolRegistry(plugin_registry),
        plugin_repo,
        python_executable=sys.executable,
        datasets_root=settings.datasets_dir,
        workspace=settings.workspace,
    )
    task = TaskRepository(settings.db_path).create_task(
        TaskCreate(
            model_name="报告样例",
            model_version="dev",
            validator="qa",
            source_dir=str(tmp_path / "source"),
            algorithm="lr",
            run_mode="agent",
            target_col="y",
            split_col="split",
            feature_columns=["x1", "x2"],
        )
    )
    frame = pd.concat([_business_frame().assign(split="train", x1=[0.1, 0.2, 0.3, 0.4], x2=[0.4, 0.3, 0.2, 0.1])] * 40, ignore_index=True)
    frame = frame.drop(columns=["score"])
    frame.loc[80:119, "split"] = "test"
    frame.loc[120:, "split"] = "oot"
    path = tmp_path / "report_sample.parquet"
    frame.to_parquet(path, index=False)
    dictionary_path = tmp_path / "feature_dictionary.parquet"
    pd.DataFrame({
        "feature": ["x1", "x2"],
        "含义": ["收入稳定性", "负债压力"],
        "产品名称": ["征信评分", "借贷画像"],
        "厂商名称": ["数据厂商A", "数据厂商B"],
    }).to_parquet(dictionary_path, index=False)
    registry = DatasetRegistry(
        DatasetRepository(settings.db_path),
        DataBackend(settings.datasets_dir),
        settings.datasets_dir,
    )
    dataset = registry.register_existing(path, task_id=task.id, role="modeling_sample")
    dictionary = registry.register_existing(dictionary_path, task_id=task.id, role="feature_dictionary")
    trained = runner.invoke(
        ToolRef("modeling", "train_model"),
        {
            "dataset_id": dataset.id,
            "recipe": "lr",
            "features": ["x1", "x2"],
            "target_col": "y",
            "split_col": "split",
            "split_values": {"train": "train", "test": "test", "oot": "oot"},
            "params": {"max_iter": 200},
            "seed": 7,
        },
        task_id=task.id,
    )
    assert trained.ok is True, trained.error

    report = runner.invoke(
        ToolRef("modeling", "generate_model_report"),
        {
            "experiment_id": trained.output["experiment_id"],
            "dataset_id": dataset.id,
            "business_columns": {
                "loan_month_col": "loan_month",
                "interest_rate_col": "rate",
                "loan_amount_col": "amount",
                "term_col": "term",
                "drawdown_amount_col": "drawdown",
                "credit_limit_col": "limit",
                "mob_observe_cols": ["mob1", "mob2", "mob3"],
            },
            "feature_dictionary_id": dictionary.id,
            "project_meta": {"项目名称": "报告样例"},
        },
        task_id=task.id,
    )

    assert report.ok is True, report.error
    assert Path(report.output["report_path"]).exists()
    scored_path = settings.tasks_dir / task.id / "outputs" / "model_report_scored.parquet"
    assert scored_path.exists()
    assert not (scored_path.parent / ".staging").exists()
    scored_frame = pd.read_parquet(scored_path)
    assert "__model_score__" in scored_frame.columns
    assert scored_frame["__model_score__"].between(0, 1).all()
    assert not scored_frame["__model_score__"].equals(scored_frame["x1"])
    assert report.output["score_bands"]
    assert {row["split"] for row in report.output["score_bands"]} == {"train", "test", "oot"}
    assert len(report.output["section_status"]) == 5
    workbook = load_workbook(report.output["report_path"])
    score_band_sheet = workbook["评分分段"]
    assert "分箱边界口径" in score_band_sheet["A1"].value
    score_band_headers = [cell.value for cell in score_band_sheet[4]]
    assert score_band_headers[:5] == ["split", "bin", "score_lower", "score_upper", "sample_count"]
    assert "cum_count_pct" in score_band_headers
    assert "cum_bad_rate" in score_band_headers
    assert "cum_bad_capture" in score_band_headers
    assert "ks_contribution" in score_band_headers
    assert score_band_sheet.max_row > 4
    summary_sheet = workbook["汇总"]
    train_row = next(
        row
        for row in range(1, summary_sheet.max_row + 1)
        if summary_sheet.cell(row=row, column=1).value == "split"
        and summary_sheet.cell(row=row, column=2).value == "train"
    )
    next_split_row = next(
        (
            row
            for row in range(train_row + 1, summary_sheet.max_row + 1)
            if summary_sheet.cell(row=row, column=1).value == "split"
        ),
        summary_sheet.max_row + 1,
    )
    train_split_summary = {
        summary_sheet.cell(row=row, column=1).value: summary_sheet.cell(row=row, column=2).value
        for row in range(train_row, next_split_row)
    }
    assert train_split_summary["sample_count"] == 80
    assert train_split_summary["bad_rate"] == pytest.approx(0.5)
    assert train_split_summary["window_start"] == "2026-01"
    assert train_split_summary["window_end"] == "2026-02"

    feature_sheet = workbook["特征重要性"]
    headers = [cell.value for cell in feature_sheet[1]]
    first_row = {header: feature_sheet.cell(row=2, column=index).value for index, header in enumerate(headers, start=1)}
    importance_by_feature = {feature: importance for feature, importance in trained.output["feature_importance"]}
    assert first_row["feature"] == "x1"
    assert first_row["importance"] == pytest.approx(importance_by_feature["x1"])
    assert first_row["含义"] == "收入稳定性"
    assert first_row["产品名称"] == "征信评分"
    assert first_row["厂商名称"] == "数据厂商A"

    univariate_sheet = workbook["单变量分析"]
    univariate_headers = [cell.value for cell in univariate_sheet[1]]
    assert univariate_headers[:4] == ["feature", "split", "iv", "ks"]
    assert "psi_vs_train" in univariate_headers
    univariate_rows = [
        {
            header: univariate_sheet.cell(row=row, column=index).value
            for index, header in enumerate(univariate_headers, start=1)
        }
        for row in range(2, univariate_sheet.max_row + 1)
    ]
    assert {
        (row["feature"], row["split"])
        for row in univariate_rows
    } == {
        ("x1", "train"),
        ("x1", "test"),
        ("x1", "oot"),
        ("x2", "train"),
        ("x2", "test"),
        ("x2", "oot"),
    }
    # DOM-7a: train itself has nothing to compare against (psi=None); test/oot report
    # a numeric PSI vs. train that matches feature_psi computed directly off the frame.
    by_key = {(row["feature"], row["split"]): row for row in univariate_rows}
    assert by_key[("x1", "train")]["psi_vs_train"] is None
    scored_train = scored_frame[scored_frame["split"] == "train"]
    scored_test = scored_frame[scored_frame["split"] == "test"]
    edges = equal_frequency_edges(scored_train["x1"].to_numpy(dtype=float), DEFAULT_IV_BINS)
    expected_psi = feature_psi(
        scored_train["x1"].to_numpy(dtype=float),
        scored_test["x1"].to_numpy(dtype=float),
        edges,
    )
    assert by_key[("x1", "test")]["psi_vs_train"] == pytest.approx(expected_psi)

    stress_sheet = workbook["压力测试"]
    stress_headers = [
        stress_sheet.cell(row=2, column=column).value
        for column in range(1, stress_sheet.max_column + 1)
    ]
    stress_rows = [
        {
            header: stress_sheet.cell(row=row, column=index).value
            for index, header in enumerate(stress_headers, start=1)
        }
        for row in range(3, stress_sheet.max_row + 1)
        if stress_sheet.cell(row=row, column=1).value
    ]
    by_item = {row["项目"]: row for row in stress_rows}
    assert by_item["baseline"]["sample_count"] == 40
    assert by_item["征信评分"]["status"] == "completed"
    assert by_item["征信评分"]["dropped_features"] == "x1"


def _report_runner(tmp_path):
    settings = build_settings(tmp_path / "workspace")
    init_db(settings.db_path)
    plugin_repo = PluginRepository(settings.db_path)
    plugin_registry = PluginRegistry(plugin_repo)
    load_builtin_packs(plugin_registry, Path(__file__).parents[1] / "marvis" / "packs")
    runner = ToolRunner(
        ToolRegistry(plugin_registry),
        plugin_repo,
        python_executable=sys.executable,
        datasets_root=settings.datasets_dir,
        workspace=settings.workspace,
    )
    task = TaskRepository(settings.db_path).create_task(
        TaskCreate(
            model_name="多版本报告样例",
            model_version="dev",
            validator="qa",
            source_dir=str(tmp_path / "source"),
            algorithm="lr",
            run_mode="agent",
            target_col="y",
            split_col="split",
            feature_columns=["x1", "x2"],
        )
    )
    frame = pd.concat([_business_frame().assign(split="train", x1=[0.1, 0.2, 0.3, 0.4], x2=[0.4, 0.3, 0.2, 0.1])] * 40, ignore_index=True)
    frame = frame.drop(columns=["score"])
    frame.loc[80:119, "split"] = "test"
    frame.loc[120:, "split"] = "oot"
    path = tmp_path / "report_sample.parquet"
    frame.to_parquet(path, index=False)
    registry = DatasetRegistry(
        DatasetRepository(settings.db_path),
        DataBackend(settings.datasets_dir),
        settings.datasets_dir,
    )
    dataset = registry.register_existing(path, task_id=task.id, role="modeling_sample")
    return runner, settings, task, dataset


def _train_report_experiment(runner, task, dataset, recipe, params):
    trained = runner.invoke(
        ToolRef("modeling", "train_model"),
        {
            "dataset_id": dataset.id,
            "recipe": recipe,
            "features": ["x1", "x2"],
            "target_col": "y",
            "split_col": "split",
            "split_values": {"train": "train", "test": "test", "oot": "oot"},
            "params": params,
            "seed": 7,
        },
        task_id=task.id,
    )
    assert trained.ok is True, trained.error
    return trained.output["experiment_id"]


def test_generate_scorecard_report_keeps_pd_and_points_separate(tmp_path):
    runner, settings, task, dataset = _report_runner(tmp_path)
    experiment_id = _train_report_experiment(
        runner,
        task,
        dataset,
        "scorecard",
        {"scorecard_max_bins": 3, "max_iter": 200},
    )

    report = runner.invoke(
        ToolRef("modeling", "generate_model_report"),
        {"experiment_id": experiment_id, "dataset_id": dataset.id, "project_meta": {"项目名称": "评分卡报告"}},
        task_id=task.id,
    )

    assert report.ok is True, report.error
    scored_path = settings.tasks_dir / task.id / "outputs" / "model_report_scored.parquet"
    assert not (scored_path.parent / ".staging").exists()
    scored_frame = pd.read_parquet(scored_path)
    assert scored_frame["__model_score__"].between(0, 1).all()
    assert "__scorecard_points__" in scored_frame.columns
    assert not scored_frame["__scorecard_points__"].between(0, 1).all()
    assert scored_frame["__scorecard_points__"].max() > 100
    assert report.output["scorecard_table"][0]["feature"] == "__base__"
    assert report.output["score_bands"]
    assert max(row["avg_score"] for row in report.output["score_bands"]) > 100
    workbook = load_workbook(report.output["report_path"])
    scorecard_sheet = workbook["评分卡"]
    headers = [cell.value for cell in scorecard_sheet[1]]
    base_row = {
        header: scorecard_sheet.cell(row=2, column=index).value
        for index, header in enumerate(headers, start=1)
    }
    assert base_row["feature"] == "__base__"
    assert base_row["bin_label"] == "base_points"
    assert base_row["points"] > 100


def test_generate_model_reports_fans_out_one_xlsx_per_experiment(tmp_path):
    runner, settings, task, dataset = _report_runner(tmp_path)
    lr_experiment = _train_report_experiment(runner, task, dataset, "lr", {"max_iter": 200})
    lgb_experiment = _train_report_experiment(
        runner, task, dataset, "lgb", {"num_boost_round": 2, "learning_rate": 0.1, "num_leaves": 4}
    )

    result = runner.invoke(
        ToolRef("modeling", "generate_model_reports"),
        {
            "experiment_ids": [lr_experiment, lgb_experiment],
            "dataset_id": dataset.id,
            "project_meta": {"项目名称": "多版本报告样例"},
        },
        task_id=task.id,
    )

    assert result.ok is True, result.error
    reports = result.output["reports"]
    assert len(reports) == 2
    assert [report["experiment_id"] for report in reports] == [lr_experiment, lgb_experiment]
    assert {report["recipe"] for report in reports} == {"lr", "lgb"}
    # report_path mirrors the first report for download-endpoint compatibility
    assert result.output["report_path"] == reports[0]["report_path"]

    paths = [report["report_path"] for report in reports]
    assert paths[0] != paths[1]
    for path_str in paths:
        report_path = Path(path_str)
        assert report_path.suffix == ".xlsx"
        assert report_path.exists()
        # a real xlsx is a ZIP (PK magic) and opens cleanly under openpyxl
        assert report_path.read_bytes()[:2] == b"PK"
        load_workbook(report_path)

    # JSON-safe payload (no NaN/Infinity tokens)
    import json

    json.dumps(result.output, allow_nan=False)


def test_generate_model_reports_rejects_empty_experiment_ids(tmp_path):
    runner, _settings, task, dataset = _report_runner(tmp_path)

    result = runner.invoke(
        ToolRef("modeling", "generate_model_reports"),
        {"experiment_ids": [], "dataset_id": dataset.id},
        task_id=task.id,
    )

    assert result.ok is False
    assert "experiment_ids" in str(result.error)


def test_generate_model_report_fails_when_artifact_and_score_column_are_missing(tmp_path):
    """DOM-10: an artifact-less experiment (e.g. a historical failed training run) must
    not have generate_model_report silently impersonate the model score with the first
    feature column — it must fail loudly with a typed error instead."""
    runner, settings, task, dataset = _report_runner(tmp_path)
    store = ExperimentStore(settings.db_path)
    experiment_id = store.create(
        task.id,
        "lr",
        TrainConfig(
            dataset_id=dataset.id,
            features=("x1", "x2"),
            target_col="y",
            split_col="split",
            split_values={"train": "train", "test": "test", "oot": "oot"},
            params={},
            seed=7,
            early_stopping_rounds=None,
        ),
    )

    result = runner.invoke(
        ToolRef("modeling", "generate_model_report"),
        {"experiment_id": experiment_id, "dataset_id": dataset.id},
        task_id=task.id,
    )

    assert result.ok is False
    assert result.error_kind == "report_score_missing"
    assert result.error_detail["experiment_id"] == experiment_id
    assert result.error_detail["dataset_id"] == dataset.id
    assert "score" in str(result.error)
def _score_band_runtime(tmp_path, frame: pd.DataFrame):
    path = tmp_path / "score_band_sample.parquet"
    frame.to_parquet(path, index=False)
    return SimpleNamespace(backend=DataBackend(tmp_path)), path


def _score_band_config(**overrides) -> TrainConfig:
    defaults = dict(
        dataset_id="ds",
        features=("x1",),
        target_col="y",
        split_col="split",
        split_values={"train": "train", "test": "test", "oot": "oot"},
        params={},
        seed=7,
        early_stopping_rounds=None,
    )
    defaults.update(overrides)
    return TrainConfig(**defaults)


def test_score_band_rows_share_train_derived_bin_edges_across_splits(tmp_path):
    """DOM-5: bin edges are computed once on train and reused verbatim by test/oot —
    same edges array per bin index across every split that has rows."""
    rng_train = list(range(0, 100))
    frame = pd.DataFrame({
        "score": [float(v) for v in rng_train] + [5.0, 15.0, 95.0, 105.0] + [10.0, 50.0, 90.0],
        "y": ([0, 1] * 50) + [0, 1, 0, 1] + [0, 1, 0],
        "split": (["train"] * 100) + (["test"] * 4) + (["oot"] * 3),
    })
    runtime, path = _score_band_runtime(tmp_path, frame)

    rows = modeling_tools._score_band_rows(
        runtime, path, score_col="score", target_col="y", config=_score_band_config(), bin_count=5,
    )

    assert rows
    edges_by_split: dict[str, dict[int, tuple[float | None, float | None]]] = {}
    for row in rows:
        edges_by_split.setdefault(row["split"], {})[row["bin"]] = (row["score_lower"], row["score_upper"])
        assert row["bin_edges_source"] == "train"
    # Every split that shares a bin index must report the identical boundary pair.
    shared_bins = set(edges_by_split["train"]) & set(edges_by_split["test"]) & set(edges_by_split["oot"])
    assert shared_bins
    for bin_index in shared_bins:
        assert edges_by_split["train"][bin_index] == edges_by_split["test"][bin_index]
        assert edges_by_split["train"][bin_index] == edges_by_split["oot"][bin_index]


def test_score_band_rows_cumulative_columns_are_hand_computable_and_monotonic(tmp_path):
    """DOM-5: cum_count_pct/cum_bad_rate/cum_bad_capture computed by hand for a tiny
    2-bin, higher-is-riskier PD score match the function's output, and cum_count_pct /
    cum_bad_capture are monotonically non-decreasing along the cumulation order."""
    # 8 train rows split evenly into low (score<0.5) / high (score>=0.5) bins via
    # equal-frequency edges; label pattern chosen so bad rates differ hand-checkably.
    frame = pd.DataFrame({
        "score": [0.1, 0.2, 0.3, 0.4, 0.6, 0.7, 0.8, 0.9],
        "y": [0, 0, 0, 1, 1, 1, 1, 0],
        "split": ["train"] * 8,
    })
    config = _score_band_config(split_values={"train": "train"})
    runtime, path = _score_band_runtime(tmp_path, frame)

    rows = modeling_tools._score_band_rows(
        runtime, path, score_col="score", target_col="y", config=config, bin_count=2,
    )

    assert {row["split"] for row in rows} == {"train"}
    by_bin = {row["bin"]: row for row in rows}
    assert set(by_bin) == {1, 2}
    # Bin 1 = low scores [0.1..0.4] -> bad=1 among 4 (row 0.4). Bin 2 = high scores
    # [0.6..0.9] -> bad=3 among 4 (0.6/0.7/0.8).
    assert by_bin[1]["bad_count"] == 1
    assert by_bin[1]["labeled_count"] == 4
    assert by_bin[2]["bad_count"] == 3
    assert by_bin[2]["labeled_count"] == 4
    # higher_is_riskier -> cumulation starts at the highest bin (bin 2) first.
    assert by_bin[2]["cum_direction"] == "higher_is_riskier"
    assert by_bin[2]["cum_count_pct"] == pytest.approx(4 / 8)
    assert by_bin[2]["cum_bad_rate"] == pytest.approx(3 / 4)
    assert by_bin[2]["cum_bad_capture"] == pytest.approx(3 / 4)
    assert by_bin[1]["cum_count_pct"] == pytest.approx(8 / 8)
    assert by_bin[1]["cum_bad_rate"] == pytest.approx(4 / 8)
    assert by_bin[1]["cum_bad_capture"] == pytest.approx(4 / 4)
    # Monotonic non-decreasing along the cumulation order (bin2 then bin1).
    ordered = [by_bin[2], by_bin[1]]
    cum_counts = [row["cum_count_pct"] for row in ordered]
    cum_captures = [row["cum_bad_capture"] for row in ordered]
    assert cum_counts == sorted(cum_counts)
    assert cum_captures == sorted(cum_captures)


def test_score_band_rows_higher_is_better_direction_cumulates_from_low_bin_up(tmp_path):
    """DOM-5: scorecard points are higher-is-better, so cumulation must start from the
    lowest-scoring (riskiest) bin instead of the highest."""
    frame = pd.DataFrame({
        modeling_tools.SCORECARD_POINTS_COL: [10.0, 20.0, 30.0, 40.0, 60.0, 70.0, 80.0, 90.0],
        "y": [1, 1, 1, 0, 0, 0, 0, 1],
        "split": ["train"] * 8,
    })
    config = _score_band_config(split_values={"train": "train"})
    runtime, path = _score_band_runtime(tmp_path, frame)

    rows = modeling_tools._score_band_rows(
        runtime,
        path,
        score_col=modeling_tools.SCORECARD_POINTS_COL,
        target_col="y",
        config=config,
        bin_count=2,
    )

    by_bin = {row["bin"]: row for row in rows}
    assert by_bin[1]["cum_direction"] == "higher_is_better"
    # Cumulation starts at bin 1 (lowest points = highest risk) for higher_is_better.
    assert by_bin[1]["cum_count_pct"] == pytest.approx(4 / 8)
    assert by_bin[2]["cum_count_pct"] == pytest.approx(8 / 8)


def test_score_band_rows_handles_no_oot_split(tmp_path):
    """DOM-5 degradation: a config with no oot split_value must not crash — the sheet
    just carries no oot rows, train/test still produced on shared edges."""
    frame = pd.DataFrame({
        "score": [0.1, 0.2, 0.3, 0.4, 0.15, 0.25, 0.35, 0.45],
        "y": [0, 1, 0, 1, 0, 1, 0, 1],
        "split": (["train"] * 4) + (["test"] * 4),
    })
    config = _score_band_config(split_values={"train": "train", "test": "test"})
    runtime, path = _score_band_runtime(tmp_path, frame)

    rows = modeling_tools._score_band_rows(
        runtime, path, score_col="score", target_col="y", config=config, bin_count=2,
    )

    assert rows
    assert {row["split"] for row in rows} == {"train", "test"}


def test_score_band_rows_handles_single_value_column_without_crashing(tmp_path):
    """DOM-5 degradation: a constant score column collapses equal-frequency edges to a
    single [-inf, inf] bin — must produce a degenerate-but-non-crashing band table."""
    frame = pd.DataFrame({
        "score": [0.5] * 6,
        "y": [0, 1, 0, 1, 0, 1],
        "split": ["train", "train", "test", "test", "oot", "oot"],
    })
    runtime, path = _score_band_runtime(tmp_path, frame)

    rows = modeling_tools._score_band_rows(
        runtime, path, score_col="score", target_col="y", config=_score_band_config(), bin_count=5,
    )

    assert rows
    assert {row["split"] for row in rows} == {"train", "test", "oot"}
    for row in rows:
        assert row["score_lower"] is None or row["score_upper"] is None or row["sample_count"] > 0


def test_score_band_rows_falls_back_to_first_available_split_when_train_empty(tmp_path):
    """DOM-5 degradation: no train rows at all (e.g. train filtered out upstream) must
    not return an empty sheet — falls back to the first split with finite scores and
    records that fallback via bin_edges_source."""
    frame = pd.DataFrame({
        "score": [0.1, 0.2, 0.3, 0.4],
        "y": [0, 1, 0, 1],
        "split": ["test", "test", "oot", "oot"],
    })
    config = _score_band_config(split_values={"train": "train", "test": "test", "oot": "oot"})
    runtime, path = _score_band_runtime(tmp_path, frame)

    rows = modeling_tools._score_band_rows(
        runtime, path, score_col="score", target_col="y", config=config, bin_count=2,
    )

    assert rows
    assert all(row["bin_edges_source"] == "test" for row in rows)
