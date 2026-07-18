from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from marvis.output.risk_analysis_report import (
    RiskAnalysisReportPayload,
    _cell_value,
    render_risk_analysis_report,
)


@pytest.mark.parametrize(
    "value",
    ["=1+1", "+SUM(A1:A2)", "-2+3", "@SUM(A1:A2)", '  =HYPERLINK("x")'],
)
def test_risk_report_neutralizes_uploaded_excel_formula_text(value: str):
    workbook = Workbook()
    cell = workbook.active.cell(row=1, column=1, value=_cell_value(value))

    assert cell.data_type == "s"
    assert cell.value.startswith("'")


def test_risk_report_keeps_numeric_values_numeric():
    workbook = Workbook()
    cell = workbook.active.cell(row=1, column=1, value=_cell_value(-0.0125))

    assert cell.data_type == "n"
    assert cell.value == -0.0125


def test_risk_report_neutralizes_uploaded_text_across_every_sheet(tmp_path):
    payload = RiskAnalysisReportPayload(
        analysis_kind="vtg_terminal",
        column_map={"product": "=UPLOADED_HEADER"},
        product_scope=["=PRODUCT_FORMULA"],
        as_of_period="=PERIOD_FORMULA",
        headline_metrics={"annualized_bad_rate": 0.1},
        key_points=["=KEY_POINT_FORMULA"],
        red_flags=["=RED_FLAG_FORMULA"],
        assumptions=["=ASSUMPTION_FORMULA"],
        source_row_count=1,
        row_count=1,
        summary_rows=[{"product": "=SUMMARY_FORMULA", "annualized_bad_rate": 0.1}],
        detail_rows=[{"product": "=DETAIL_FORMULA", "annualized_bad_rate": 0.1}],
        formula_definitions=[
            {
                "metric": "=METRIC_FORMULA",
                "formula": "=FORMULA_TEXT",
                "note": "=NOTE_FORMULA",
            }
        ],
        data_quality=[
            {"status": "WARN", "check": "=CHECK_FORMULA", "detail": "=DETAIL_FORMULA"}
        ],
    )
    path = render_risk_analysis_report(payload, tmp_path / "report.xlsx")
    workbook = load_workbook(path, data_only=False)

    formulas = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert formulas == []
    assert workbook["结论摘要"]["C2"].value == "源数据行数"
    assert workbook["结论摘要"]["D2"].value == 1
    assert workbook["结论摘要"]["A4"].value == "结果行数"
    assert workbook["结论摘要"]["B4"].value == 1
    assert workbook["结论摘要"]["B3"].value.startswith("'")
    assert workbook["结论摘要"]["D3"].value.startswith("'")
