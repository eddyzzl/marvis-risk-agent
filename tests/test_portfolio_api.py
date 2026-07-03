"""S3 Commit 4: PORTFOLIO_ANALYSIS 模板 + 报告 + 接线 tests.

Covers: 模板 instantiate+validate 零错；步骤 1-4 无依赖（并行就绪即跑语义）；
剪步（无 experiment_id 用 no_trend 变体）；汇总门聚合 red_flags；报告只搬运不重算
（改 payload 数字报告 sheet 跟着变）；组合报告 xlsx sheet 清单一致；端到端
portfolio 任务→states 确认→并行分析→汇总门→报告落盘+artifact 审计行全查。
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from marvis.data.backend import DataBackend
from marvis.data.registry import DatasetRegistry
from marvis.db import DatasetRepository, PluginRepository, TaskRepository, init_db
from marvis.domain import TASK_TYPE_PORTFOLIO, TaskCreate
from marvis.orchestrator.planner import Planner
from marvis.orchestrator.templates import get_template, load_builtin_templates
from marvis.orchestrator.validator import PlanValidator
from marvis.output.portfolio_report import (
    PORTFOLIO_REPORT_SHEETS,
    PortfolioReportPayload,
    render_portfolio_report,
)
from marvis.packs.analysis.report import build_report, gate_summary_payload
from marvis.plugins.loader import load_builtin_packs
from marvis.plugins.manifest import ToolRef
from marvis.plugins.registry import PluginRegistry, ToolRegistry
from marvis.plugins.runner import ToolRunner
from marvis.sample_data import generate_performance_frame, generate_sample_frame
from marvis.settings import build_settings


def _runtime(tmp_path):
    settings = build_settings(tmp_path / "workspace")
    init_db(settings.db_path)
    plugin_repo = PluginRepository(settings.db_path)
    plugin_registry = PluginRegistry(plugin_repo)
    packs_root = Path(__file__).parents[1] / "marvis" / "packs"
    load_builtin_packs(plugin_registry, packs_root)
    runner = ToolRunner(
        ToolRegistry(plugin_registry),
        plugin_repo,
        python_executable=sys.executable,
        datasets_root=settings.datasets_dir,
        workspace=settings.workspace,
    )
    data_repo = DatasetRepository(settings.db_path)
    backend = DataBackend(settings.datasets_dir)
    registry = DatasetRegistry(data_repo, backend, settings.datasets_dir)
    task = TaskRepository(settings.db_path).create_task(
        TaskCreate(
            model_name="组合分析接线",
            model_version="dev",
            validator="qa",
            source_dir=str(tmp_path / "source"),
            algorithm="lr",
            run_mode="agent",
            task_type=TASK_TYPE_PORTFOLIO,
        )
    )
    return settings, runner, registry, task


def _perf_dataset(registry, tmp_path, task_id):
    sample = generate_sample_frame(n_rows=200)
    perf = generate_performance_frame(sample, n_months=6)
    path = tmp_path / "perf.parquet"
    perf.to_parquet(path, index=False)
    return registry.register_existing(path, task_id=task_id, role="performance")


_STATES = ["current", "M1", "M2", "M3+", "charged_off"]


def _tool_registry(tmp_path):
    db_path = tmp_path / "app.sqlite"
    init_db(db_path)
    repo = PluginRepository(db_path)
    registry = PluginRegistry(repo)
    load_builtin_packs(registry, Path(__file__).parents[1] / "marvis" / "packs")
    return ToolRegistry(registry)


def test_portfolio_templates_instantiate_and_validate(tmp_path):
    load_builtin_templates()
    tool_registry = _tool_registry(tmp_path)
    planner = Planner(tool_registry, lambda: None, PlanValidator(tool_registry))
    for template_id in ("portfolio_analysis", "portfolio_analysis_no_trend"):
        template = get_template(template_id)
        slots = {
            "performance_dataset_id": "ds1",
            "id_col": "loan_id",
            "snapshot_col": "snapshot_month",
            "bucket_col": "bucket",
            "states": _STATES,
            "balance_col": "balance",
            "segment_col": "seg",
        }
        if template_id == "portfolio_analysis":
            slots["experiment_id"] = "exp1"
            slots["score_col"] = "model_score"
        plan = planner.from_template(template, slots, "task1")
        # zero validation errors (tools exist, inputs schema, DAG, $ref compat)
        errors = PlanValidator(tool_registry).validate(plan)
        assert errors == [], errors


def test_portfolio_template_steps_1_to_4_have_no_dependencies():
    load_builtin_templates()
    template = get_template("portfolio_analysis")
    deps = {s.title: list(s.depends_on_titles) for s in template.steps}
    for title in ("流量分析", "迁徙热力", "细分画像", "稳定性趋势"):
        assert deps[title] == []
    assert deps["损失估计"] == ["迁徙热力"]
    # report depends on every step it $refs (validator requires a dep edge per ref)
    assert "组合分析汇总" in deps["生成组合报告"]
    assert "流量分析" in deps["生成组合报告"]


def test_portfolio_no_trend_variant_prunes_trend_step():
    load_builtin_templates()
    template = get_template("portfolio_analysis_no_trend")
    titles = [s.title for s in template.steps]
    assert "稳定性趋势" not in titles
    gate = next(s for s in template.steps if s.title == "组合分析汇总")
    assert "稳定性趋势" not in gate.depends_on_titles


def test_gate_summary_aggregates_red_flags():
    flow = {"red_flags": [{"kind": "sparse_month", "message": "m"}]}
    segment = {"red_flags": [{"kind": "high_concentration", "message": "c"}], "concentration": {"hhi": 0.5, "top1_pct": 0.6}}
    el = {"red_flags": [{"kind": "short_history", "message": "h"}], "total_el": 123.0}
    payload = gate_summary_payload(flow=flow, migration=None, segment=segment, trend=None, expected_loss=el)
    kinds = {f["kind"] for f in payload["red_flags"]}
    assert kinds == {"sparse_month", "high_concentration", "short_history"}
    assert payload["red_flag_count"] == 3
    assert payload["highlights"]["total_el"] == 123.0
    assert payload["highlights"]["hhi"] == 0.5
    # checklist == red flag messages (gate checklist)
    assert set(payload["checklist"]) == {"m", "c", "h"}


def test_report_carries_numbers_not_recompute(tmp_path):
    """报告只搬运不重算：改 payload 数字，报告落盘的对应单元格跟着变。"""
    segment = {"segments": [{"segment": "A", "count": 7, "pop_pct": 0.7}], "concentration": {"hhi": 0.5, "top1_pct": 0.7}}
    el = {"total_el": 999.0, "el_by_month": [{"month": "2025-01", "balance": 1000.0, "expected_loss": 999.0}],
          "chain": [{"from_state": "current", "p_to_loss": 0.1}], "assumptions": {"lgd": 0.6}}
    out1 = tmp_path / "r1.xlsx"
    path1, sheets = build_report(project_meta={"名称": "T"}, flow=None, migration=None, segment=segment, trend=None, expected_loss=el, out_path=out1)
    assert sheets == PORTFOLIO_REPORT_SHEETS
    wb1 = load_workbook(path1)
    el_sheet1 = wb1["预期损失"]
    values1 = [cell.value for row in el_sheet1.iter_rows() for cell in row]
    assert 999.0 in values1

    # change the number -> report follows
    el["el_by_month"][0]["expected_loss"] = 111.0
    el["total_el"] = 111.0
    out2 = tmp_path / "r2.xlsx"
    path2, _ = build_report(project_meta={"名称": "T"}, flow=None, migration=None, segment=segment, trend=None, expected_loss=el, out_path=out2)
    wb2 = load_workbook(path2)
    values2 = [cell.value for row in wb2["预期损失"].iter_rows() for cell in row]
    assert 111.0 in values2
    assert 999.0 not in values2


def test_render_portfolio_report_sheet_list(tmp_path):
    payload = PortfolioReportPayload(project_meta={"x": 1})
    path = render_portfolio_report(payload, tmp_path / "empty.xlsx")
    wb = load_workbook(path)
    assert wb.sheetnames == PORTFOLIO_REPORT_SHEETS


@pytest.mark.slow
def test_portfolio_report_tool_registers_artifact_audit(tmp_path):
    settings, runner, registry, task = _runtime(tmp_path)
    dataset = _perf_dataset(registry, tmp_path, task.id)
    # run the four analysis tools, then the report tool with their outputs injected
    flow = runner.invoke(ToolRef("analysis", "flow_rate"), {
        "dataset_id": dataset.id, "id_col": "loan_id", "snapshot_col": "snapshot_month",
        "bucket_col": "bucket", "states": _STATES, "balance_col": "balance",
    }, task_id=task.id)
    migration = runner.invoke(ToolRef("analysis", "bucket_migration"), {
        "dataset_id": dataset.id, "id_col": "loan_id", "snapshot_col": "snapshot_month",
        "bucket_col": "bucket", "states": _STATES, "balance_col": "balance",
    }, task_id=task.id)
    el = runner.invoke(ToolRef("analysis", "expected_loss_estimate"), {
        "dataset_id": dataset.id, "id_col": "loan_id", "snapshot_col": "snapshot_month",
        "bucket_col": "bucket", "states": _STATES, "balance_col": "balance", "loss_state": "charged_off",
    }, task_id=task.id)
    assert flow.ok and migration.ok and el.ok, (flow.error, migration.error, el.error)

    report = runner.invoke(ToolRef("analysis", "portfolio_report"), {
        "flow": flow.output, "migration": migration.output, "expected_loss": el.output,
        "project_meta": {"名称": "组合分析接线"},
    }, task_id=task.id)
    assert report.ok is True, report.error
    assert report.output["report_path"]
    assert Path(report.output["report_path"]).exists()
    assert report.output["sheets"] == PORTFOLIO_REPORT_SHEETS

    audits = PluginRepository(settings.db_path).list_audit()
    assert any(
        a["kind"] == "analysis.portfolio.report" and a["target_ref"] == task.id
        for a in audits
    )


class _FakeLLM:
    def complete(self, **kwargs):
        return '{"summary": "done", "open_items": [], "goal_doubt": false, "goal_met": true}'


class _FakeHooks:
    def dispatch(self, event, payload, *, task_id):
        return []


def _portfolio_driver(tmp_path):
    from marvis.agent.plan_driver import PlanDriver
    from marvis.db import PlanRepository
    from marvis.orchestrator.executor import PlanExecutor
    from marvis.orchestrator.harness_state import HarnessState
    from marvis.orchestrator.reviewer import Reviewer

    settings = build_settings(tmp_path / "workspace")
    init_db(settings.db_path)
    plugin_repo = PluginRepository(settings.db_path)
    plugin_registry = PluginRegistry(plugin_repo)
    packs_root = Path(__file__).parents[1] / "marvis" / "packs"
    load_builtin_packs(plugin_registry, packs_root)
    tool_registry = ToolRegistry(plugin_registry)
    runner = ToolRunner(
        tool_registry, plugin_repo, python_executable=sys.executable,
        datasets_root=settings.datasets_dir, workspace=settings.workspace,
    )
    data_repo = DatasetRepository(settings.db_path)
    backend = DataBackend(settings.datasets_dir)
    registry = DatasetRegistry(data_repo, backend, settings.datasets_dir)
    plan_repo = PlanRepository(settings.db_path)
    executor = PlanExecutor(plan_repo, runner, Reviewer(lambda: _FakeLLM()), None, _FakeHooks(), HarnessState(plan_repo))
    planner = Planner(tool_registry, lambda: _FakeLLM(), PlanValidator(tool_registry))
    driver = PlanDriver(plan_repo, executor, planner=planner, validator=PlanValidator(tool_registry))
    load_builtin_templates()
    task = TaskRepository(settings.db_path).create_task(
        TaskCreate(
            model_name="组合分析端到端",
            model_version="dev",
            validator="qa",
            source_dir=str(tmp_path / "source"),
            algorithm="lr",
            run_mode="agent",
            task_type=TASK_TYPE_PORTFOLIO,
        )
    )
    return driver, registry, plan_repo, settings, task


@pytest.mark.slow
def test_portfolio_end_to_end_journey(tmp_path):
    """合成表现数据→并行分析(流量/迁徙/细分/损失)→汇总门→确认→报告落盘+审计行。"""
    from marvis.orchestrator.contracts import PlanStatus

    driver, registry, plan_repo, settings, task = _portfolio_driver(tmp_path)
    dataset = _perf_dataset(registry, tmp_path, task.id)

    # states already confirmed conceptually -> start the no-trend variant directly
    turn = driver.start(
        task_id=task.id,
        template_id="portfolio_analysis_no_trend",
        slots={
            "performance_dataset_id": dataset.id,
            "id_col": "loan_id",
            "snapshot_col": "snapshot_month",
            "bucket_col": "bucket",
            "states": _STATES,
            "balance_col": "balance",
            "segment_col": "bucket",
            "project_meta": {"名称": "组合分析端到端"},
        },
    )
    assert turn.status == PlanStatus.VALIDATED.value
    plan_id = turn.plan_id

    # 开始 -> runs the four parallel analysis steps + EL straight through to the
    # 组合分析汇总 mandatory confirm gate.
    turn = driver.resume(plan_id=plan_id, user_text="开始", run_seq=1)
    assert turn.status == PlanStatus.AWAITING_CONFIRM.value
    plan = plan_repo.load_plan(plan_id)
    for title in ("流量分析", "迁徙热力", "细分画像", "损失估计"):
        assert next(s for s in plan.steps if s.title == title).status.value == "done", title
    assert next(s for s in plan.steps if s.title == "组合分析汇总").status.value == "awaiting_confirm"

    # confirm the gate -> runs 汇总 + 生成组合报告 to DONE
    turn = driver.resume(plan_id=plan_id, user_text="确认", run_seq=2)
    plan = plan_repo.load_plan(plan_id)
    report_step = next(s for s in plan.steps if s.title == "生成组合报告")
    assert report_step.status.value == "done", turn.status
    assert plan.status.value == "done"

    # report landed on disk + audit row present
    audits = PluginRepository(settings.db_path).list_audit()
    report_audits = [a for a in audits if a["kind"] == "analysis.portfolio.report" and a["target_ref"] == task.id]
    assert report_audits, "expected analysis.portfolio.report audit row"
    report_path = Path(report_audits[-1]["detail"]["report_path"])
    assert report_path.exists()
    wb = load_workbook(report_path)
    assert wb.sheetnames == PORTFOLIO_REPORT_SHEETS
