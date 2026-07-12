from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

from openpyxl import load_workbook

from marvis.metric_tables import metric_table_sections_from_payload
from marvis.output.excel import write_validation_excel
from marvis.output.image_render import render_all_images
from marvis.report_texts import report_text_values_from_results
from marvis.template_reports import find_placeholders
from tests.output.test_excel import _make_results
from tests.validation_output_contract import load_validation_non_regression_contract


DEFAULT_TEMPLATE_PATH = (
    Path(__file__).parent / "fixtures" / "validation_report_template_baseline.docx"
)


def _placeholder_keys(prefix: str) -> list[str]:
    return [
        placeholder[2:-2]
        for placeholder in find_placeholders(DEFAULT_TEMPLATE_PATH)
        if placeholder.startswith(f"{{{{{prefix}:")
    ]


def _normalized_metric_sections(payload: dict) -> list[dict]:
    return [
        {
            "title": section["title"],
            "table_keys": [table["key"] for table in section.get("tables", [])],
            "chart_keys": [chart["key"] for chart in section.get("charts", [])],
        }
        for section in metric_table_sections_from_payload(payload)
    ]


def _current_contract(tmp_path: Path) -> dict:
    results = _make_results()
    image_paths = render_all_images(results, tmp_path / "images")
    workbook_path = write_validation_excel(results, tmp_path / "validation.xlsx")
    workbook = load_workbook(workbook_path, read_only=True)
    try:
        sheet_order = list(workbook.sheetnames)
    finally:
        workbook.close()

    return {
        "schema_version": "marvis.validation_non_regression.v1",
        "report_text_keys": sorted(report_text_values_from_results(results)),
        "template_text_keys": _placeholder_keys("TEXT"),
        "template_image_keys": _placeholder_keys("IMAGE"),
        "rendered_image_keys": list(image_paths),
        "excel_sheets": sheet_order,
        "agent_sections": _normalized_metric_sections(asdict(results)),
        "allowed_replacements": {
            "TEXT:reproducibility_summary": "TEXT:pmml_scoring_summary",
            "reproducibility": "pmml_scoring",
            "模型可复现性验证": "PMML打分测试",
            "分数一致性": "PMML打分测试",
        },
    }


def test_current_validation_outputs_match_non_regression_contract(tmp_path: Path):
    expected = load_validation_non_regression_contract()

    assert _current_contract(tmp_path) == expected
