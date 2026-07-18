from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import pytest

from marvis.data.backend import DataBackend
from marvis.data.registry import DatasetRegistry
from marvis.db import DatasetRepository, PluginRepository, TaskRepository, init_db
from marvis.domain import TASK_TYPE_VINTAGE, TaskCreate
from marvis.output.risk_analysis_report import (
    RISK_ANALYSIS_REPORT_SHEETS,
    RiskAnalysisReportPayload,
    render_risk_analysis_report,
)
from marvis.packs.risk_analysis.calculations import (
    RiskAnalysisError,
    calculate_profitability,
    calculate_vtg_terminal,
)
from marvis.plugins.loader import load_builtin_packs
from marvis.plugins.manifest import ToolRef
from marvis.plugins.registry import PluginRegistry, ToolRegistry
from marvis.plugins.runner import ToolRunner
from marvis.plugins.schema_validation import validate_against_schema
from marvis.settings import build_settings


_VTG_COLUMN_MAP = {
    "product": "产品",
    "cohort": "账期",
    "as_of_date": "截至日期",
    "amount_unit": "金额单位",
    "scenario": "场景",
    "channel": "渠道",
    "tenor_months": "期限月数",
    "selection_rule": "筛选规则",
    "disbursement_amount": "放款金额",
    "mob14_bad_rate": "MOB14不良率",
    "turnover": "周转次数",
    "terminal_bad_rate": "终值不良率",
    "long_term_recovery_rate": "长期回收率",
    "auxiliary_terminal_bad_rate": "辅助终值",
    "terminal_method": "终值方法",
    "avg_daily_balance": "日均余额",
    "previous_annualized_bad_rate": "上期年化不良率",
}


def _vtg_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "产品": ["A", "A"],
            "账期": ["2025-01", "2025-02"],
            "截至日期": ["2025-03-31", "2025-03-31"],
            "金额单位": ["万元", "万元"],
            "场景": ["基准", "基准"],
            "渠道": ["白条", "白条"],
            "期限月数": [12, 12],
            "筛选规则": ["全量已成熟", "min_auxiliary_recovery"],
            "放款金额": [1000.0, 1200.0],
            "MOB14不良率": [0.10, 0.12],
            "周转次数": [2.0, 2.5],
            "终值不良率": [0.08, None],
            "长期回收率": [None, 0.25],
            "辅助终值": [None, 0.08],
            "终值方法": ["直接终值", "长期回收率+辅助终值取低"],
            "日均余额": [500.0, None],
            "上期年化不良率": [0.12, 0.18],
        }
    )


_PROFIT_COLUMN_MAP = {
    "product": "产品",
    "as_of_period": "数据期间",
    "asset_class": "资产类型",
    "weight": "权重",
    "weight_basis": "权重口径",
    "customer_rate": "对客利率",
    "risk_cost_rate": "风险成本",
    "funding_cost_rate": "资金成本",
    "interest_loss_rate": "息费损失",
    "revenue_share_rate": "分润成本",
    "acquisition_cost_rate": "获客成本",
    "data_cost_rate": "数据成本",
    "payment_cost_rate": "支付成本",
    "collection_cost_rate": "催收成本",
    "other_cost_rate": "其他成本",
    "tax_rate": "税率",
}


def _profit_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "产品": ["P1", "P1", "P2"],
            "数据期间": ["2025-01", "2025-01", "2025-01"],
            "资产类型": ["生息", "免息", "取现"],
            "权重": [0.6, 0.4, 1.0],
            "权重口径": ["average_balance", "average_balance", "average_balance"],
            "对客利率": [0.20, 0.15, 0.05],
            "风险成本": [0.03, 0.02, 0.08],
            "资金成本": [0.04, 0.03, 0.02],
            "息费损失": [0.01, 0.005, 0.0],
            "分润成本": [0.02, 0.01, 0.0],
            "获客成本": [0.01, 0.008, 0.0],
            "数据成本": [0.005, 0.004, 0.0],
            "支付成本": [0.002, 0.001, 0.0],
            "催收成本": [0.003, 0.002, 0.0],
            "其他成本": [0.001, 0.0, 0.0],
            "税率": [0.0, 0.01, 0.0],
        }
    )


_RAW_PROFIT_COLUMN_MAP = {
    "product": "产品",
    "as_of_period": "数据期间",
    "asset_class": "资产类型",
    "weight": "权重",
    "weight_basis": "权重口径",
    "amount_unit": "金额单位",
    "customer_rate": "对客利率",
    "terminal_vintage_rate": "终值风险率",
    "risk_turnover": "风险周转次数",
    "loss_timing_factor": "损失时点系数",
    "profit_share_ratio": "利润分成比例",
    "per_application_cost": "单申请数据成本",
    "credit_approval_rate": "授信通过率",
    "draw_initiation_rate": "支用发起率",
    "draw_approval_rate": "支用通过率",
    "average_ticket": "平均客单价",
    "data_annualization_factor": "数据成本年化因子",
    "acquisition_cost_rate": "独立获客成本",
    "payment_cost_rate": "支付成本",
    "collection_cost_rate": "催收成本",
    "funding_cost_rate": "资金成本",
    "other_cost_rate": "其他成本",
    "tax_method": "税务方法",
    "tax_inclusive_divisor": "含税除数",
    "tax_combined_rate": "税费综合率",
}


def _raw_profit_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "产品": ["P1"],
            "数据期间": ["2025-01"],
            "资产类型": ["生息"],
            "权重": [1.0],
            "权重口径": ["average_balance"],
            "金额单位": ["元"],
            "对客利率": [0.20],
            "终值风险率": [0.04],
            "风险周转次数": [2.0],
            "损失时点系数": [0.5],
            "利润分成比例": [0.10],
            "单申请数据成本": [0.10],
            "授信通过率": [0.5],
            "支用发起率": [0.5],
            "支用通过率": [0.8],
            "平均客单价": [1000.0],
            "数据成本年化因子": [12.0],
            "分润成本": [0.01],
            "独立获客成本": [0.0],
            "支付成本": [0.002],
            "催收成本": [0.003],
            "资金成本": [0.02],
            "其他成本": [0.001],
            "税务方法": ["sample_net_revenue_vat_surcharge"],
            "含税除数": [1.06],
            "税费综合率": [0.0672],
        }
    )


def test_vtg_terminal_calculation_matches_manual_formulas_and_substitution_rules():
    result = calculate_vtg_terminal(_vtg_frame(), column_map=_VTG_COLUMN_MAP)

    assert result.row_count == 2
    assert result.source_row_count == 2
    assert result.product_scope == ["A"]
    assert result.as_of_period == "2025-03-31"
    first, second = result.detail_rows
    assert first["terminal_bad_rate"] == pytest.approx(0.08)
    assert first["terminal_method"] == "直接终值"
    assert first["scenario"] == "基准"
    assert first["channel"] == "白条"
    assert first["tenor_months"] == pytest.approx(12.0)
    assert first["selection_rule"] == "全量已成熟"
    assert first["annualized_bad_rate"] == pytest.approx(0.16)
    assert first["long_term_recovery_rate"] == pytest.approx(0.20)
    # MOB14 * (1 - recovery) = .09; auxiliary .08 also exists, so min=.08.
    assert second["terminal_bad_rate"] == pytest.approx(0.08)
    assert second["terminal_bad_rate_source"] == "min(auxiliary,recovery_derived)"
    assert second["avg_daily_balance"] == pytest.approx(1200.0 / 2.5)
    assert second["annualized_bad_rate"] == pytest.approx(0.20)
    assert second["observed_annualized_bad_rate"] == pytest.approx(0.30)
    assert second["long_term_recovery_rate"] == pytest.approx(1.0 / 3.0)
    assert result.headline_metrics["annualized_bad_rate"] == pytest.approx(
        176.0 / 980.0
    )
    assert result.headline_metrics["observed_annualized_bad_rate"] == pytest.approx(
        244.0 / 980.0
    )
    assert any("缺少 terminal_bad_rate" in flag for flag in result.red_flags)
    assert any("变化最大" in point for point in result.key_points)
    assert any("terminal_method" in assumption for assumption in result.assumptions)
    assert any("amount_unit=万元" in assumption for assumption in result.assumptions)
    assert any(
        "不会用未成熟原始 MOB" in assumption for assumption in result.assumptions
    )


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("MOB14不良率", 1.01, "率必须在"),
        ("放款金额", -1.0, "必须 >"),
        ("周转次数", 0.0, "必须 >"),
    ],
)
def test_vtg_terminal_rejects_illegal_rates_amounts_and_turnover(field, value, message):
    frame = _vtg_frame()
    frame.loc[0, field] = value

    with pytest.raises(RiskAnalysisError, match=message):
        calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)


def test_vtg_terminal_skips_zero_disbursement_placeholder_and_reports_quality():
    frame = _vtg_frame()
    frame.loc[0, "放款金额"] = 0.0

    result = calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)

    assert result.source_row_count == 2
    assert result.row_count == 1
    assert result.detail_rows[0]["cohort"] == "2025-02"
    assert any("零放款金额" in flag for flag in result.red_flags)
    assert any(
        check["check"] == "零放款占位行" and check["status"] == "WARN"
        for check in result.data_quality
    )


def test_vtg_terminal_reconciles_summary_balance_with_turnover():
    consistent = _vtg_frame().iloc[[0]].copy()
    consistent.loc[consistent.index[0], "日均余额"] = 500.04

    result = calculate_vtg_terminal(consistent, column_map=_VTG_COLUMN_MAP)

    assert result.detail_rows[0]["avg_daily_balance"] == pytest.approx(500.04)
    assert any("相对容差" in assumption for assumption in result.assumptions)

    inconsistent = _vtg_frame().iloc[[0]].copy()
    inconsistent.loc[inconsistent.index[0], "日均余额"] = 500.06
    with pytest.raises(
        RiskAnalysisError,
        match="avg_daily_balance=.*放款金额/周转次数=.*不一致",
    ):
        calculate_vtg_terminal(inconsistent, column_map=_VTG_COLUMN_MAP)


def test_vtg_terminal_retains_and_reconciles_supplied_recovery_rate():
    frame = _vtg_frame().iloc[[0]].copy()
    frame.loc[frame.index[0], "长期回收率"] = 0.20005

    result = calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)

    detail = result.detail_rows[0]
    assert detail["supplied_long_term_recovery_rate"] == pytest.approx(0.20005)
    assert detail["realized_long_term_recovery_rate"] == pytest.approx(0.20)
    assert any("1 bp" in assumption for assumption in result.assumptions)

    frame.loc[frame.index[0], "长期回收率"] = 0.2002
    with pytest.raises(
        RiskAnalysisError,
        match="long_term_recovery_rate=.*实现回收率=.*不一致",
    ):
        calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)


def test_vtg_terminal_discloses_direct_terminal_precedence_over_auxiliary():
    frame = _vtg_frame().iloc[[0]].copy()
    frame.loc[frame.index[0], "辅助终值"] = 0.07

    result = calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)

    detail = result.detail_rows[0]
    assert detail["terminal_bad_rate"] == pytest.approx(0.08)
    assert detail["auxiliary_terminal_bad_rate"] == pytest.approx(0.07)
    assert detail["terminal_bad_rate_source"] == "direct"
    assert any(
        "直接终值和辅助终值" in flag and "未参与计算" in flag
        for flag in result.red_flags
    )


def test_vtg_terminal_supports_explicit_min_mob14_auxiliary_method():
    frame = pd.DataFrame(
        {
            "产品": ["A"],
            "账期": ["2025-01"],
            "截至日期": ["2025-03-31"],
            "金额单位": ["万元"],
            "放款金额": [1000.0],
            "MOB14不良率": [0.10],
            "周转次数": [2.0],
            "辅助终值": [0.08],
            "终值方法": ["min_mob14_auxiliary"],
        }
    )
    mapping = {
        "product": "产品",
        "cohort": "账期",
        "as_of_date": "截至日期",
        "amount_unit": "金额单位",
        "disbursement_amount": "放款金额",
        "mob14_bad_rate": "MOB14不良率",
        "turnover": "周转次数",
        "auxiliary_terminal_bad_rate": "辅助终值",
        "terminal_method": "终值方法",
    }

    result = calculate_vtg_terminal(frame, column_map=mapping)

    detail = result.detail_rows[0]
    assert detail["terminal_bad_rate"] == pytest.approx(0.08)
    assert detail["terminal_bad_rate_source"] == "min(mob14,auxiliary)"
    assert detail["terminal_method"] == "min_mob14_auxiliary"
    assert any("未隐式套用" in assumption for assumption in result.assumptions)

    frame.loc[0, "辅助终值"] = None
    with pytest.raises(
        RiskAnalysisError,
        match="auxiliary_terminal_bad_rate 不能为空",
    ):
        calculate_vtg_terminal(frame, column_map=mapping)


def test_vtg_terminal_rejects_min_mob14_auxiliary_with_recovery_rate():
    frame = pd.DataFrame(
        {
            "产品": ["A"],
            "账期": ["2025-01"],
            "截至日期": ["2025-03-31"],
            "金额单位": ["万元"],
            "放款金额": [1000.0],
            "MOB14不良率": [0.10],
            "周转次数": [2.0],
            "长期回收率": [0.20],
            "辅助终值": [0.08],
            "终值方法": ["min_mob14_auxiliary"],
        }
    )
    mapping = {
        "product": "产品",
        "cohort": "账期",
        "as_of_date": "截至日期",
        "amount_unit": "金额单位",
        "disbursement_amount": "放款金额",
        "mob14_bad_rate": "MOB14不良率",
        "turnover": "周转次数",
        "long_term_recovery_rate": "长期回收率",
        "auxiliary_terminal_bad_rate": "辅助终值",
        "terminal_method": "终值方法",
    }

    with pytest.raises(
        RiskAnalysisError,
        match="min_mob14_auxiliary 与 long_term_recovery_rate 不能同时提供",
    ):
        calculate_vtg_terminal(frame, column_map=mapping)


def test_vtg_terminal_derives_and_surfaces_previous_turnover_provenance():
    frame = _vtg_frame().iloc[[0]].copy()
    frame["上期终值不良率"] = 0.06
    frame["上期放款金额"] = 900.0
    frame["上期日均余额"] = 450.0
    mapping = {
        **_VTG_COLUMN_MAP,
        "previous_terminal_bad_rate": "上期终值不良率",
        "previous_disbursement_amount": "上期放款金额",
        "previous_avg_daily_balance": "上期日均余额",
    }
    frame["上期年化不良率"] = None

    result = calculate_vtg_terminal(frame, column_map=mapping)

    detail = result.detail_rows[0]
    assert detail["previous_disbursement_amount"] == pytest.approx(900.0)
    assert detail["previous_avg_daily_balance"] == pytest.approx(450.0)
    assert detail["previous_turnover"] == pytest.approx(2.0)
    assert detail["previous_turnover_source"] == "derived_from_amount_balance"
    assert detail["previous_annualized_bad_rate"] == pytest.approx(0.12)
    assert (
        detail["previous_annualized_bad_rate_source"]
        == "derived_from_terminal_turnover"
    )

    frame.loc[frame.index[0], "上期日均余额"] = 0.0
    with pytest.raises(
        RiskAnalysisError,
        match="previous_avg_daily_balance 必须大于 0",
    ):
        calculate_vtg_terminal(frame, column_map=mapping)


def test_vtg_terminal_accepts_annualized_prior_above_one_and_reconciles_redundancy():
    annualized_only = _vtg_frame().iloc[[0]].copy()
    annualized_only.loc[annualized_only.index[0], "上期年化不良率"] = 1.20

    result = calculate_vtg_terminal(annualized_only, column_map=_VTG_COLUMN_MAP)

    assert result.detail_rows[0]["previous_annualized_bad_rate"] == pytest.approx(1.20)
    assert any("上期年化不良率超过 100%" in flag for flag in result.red_flags)

    redundant = _vtg_frame().iloc[[0]].copy()
    redundant["上期终值不良率"] = 0.06
    redundant["上期周转次数"] = 2.0
    redundant["上期放款金额"] = 900.0
    redundant["上期日均余额"] = 450.0
    redundant["上期年化不良率"] = 0.12
    mapping = {
        **_VTG_COLUMN_MAP,
        "previous_terminal_bad_rate": "上期终值不良率",
        "previous_turnover": "上期周转次数",
        "previous_disbursement_amount": "上期放款金额",
        "previous_avg_daily_balance": "上期日均余额",
    }
    calculate_vtg_terminal(redundant, column_map=mapping)

    redundant.loc[redundant.index[0], "上期周转次数"] = 2.1
    with pytest.raises(RiskAnalysisError, match="previous_turnover=.*口径不一致"):
        calculate_vtg_terminal(redundant, column_map=mapping)

    redundant.loc[redundant.index[0], "上期周转次数"] = 2.0
    redundant.loc[redundant.index[0], "上期年化不良率"] = 0.121
    with pytest.raises(
        RiskAnalysisError,
        match="previous_annualized_bad_rate=.*不一致",
    ):
        calculate_vtg_terminal(redundant, column_map=mapping)


def test_vtg_terminal_rejects_mixed_amount_units_in_portfolio():
    frame = _vtg_frame()
    frame["金额单位"] = ["万元", "元"]

    with pytest.raises(
        RiskAnalysisError,
        match="amount_unit 全表一致",
    ):
        calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)


def test_vtg_terminal_bounds_product_scope_without_truncating_calculation_rows():
    frame = pd.concat([_vtg_frame().iloc[[0]]] * 9, ignore_index=True)
    frame["产品"] = [f"P{index}" for index in range(9)]

    result = calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)

    assert result.row_count == 9
    assert result.headline_metrics["product_count"] == 9
    assert result.product_scope == [f"P{index}" for index in range(8)]
    assert any("报表明细未截断" in assumption for assumption in result.assumptions)


@pytest.mark.parametrize(
    ("balance_field", "balance_source", "balances"),
    [
        ("mob_balance_rate", "MOB余额率", [1.0, 0.5]),
        ("mob_balance_amount", "MOB余额", [3600.0, 1800.0]),
    ],
)
def test_vtg_terminal_derives_turnover_from_normalized_mob_balance_curve(
    balance_field,
    balance_source,
    balances,
):
    frame = pd.DataFrame(
        {
            "产品": ["A", "A"],
            "账期": ["2025-01", "2025-01"],
            "截至日期": ["2025-03-31", "2025-03-31"],
            "金额单位": ["万元", "万元"],
            "放款金额": [3600.0, 3600.0],
            "MOB14不良率": [0.10, 0.10],
            "终值不良率": [0.08, 0.08],
            "MOB": [0, 1],
            "MOB天数": [30.0, 335.0],
            "日数基准": [365.0, 365.0],
            balance_source: balances,
        }
    )
    column_map = {
        "product": "产品",
        "cohort": "账期",
        "as_of_date": "截至日期",
        "amount_unit": "金额单位",
        "disbursement_amount": "放款金额",
        "mob14_bad_rate": "MOB14不良率",
        "terminal_bad_rate": "终值不良率",
        "mob": "MOB",
        "mob_days": "MOB天数",
        "day_count_basis": "日数基准",
        balance_field: balance_source,
    }

    result = calculate_vtg_terminal(frame, column_map=column_map)

    assert result.row_count == 1
    assert result.source_row_count == 2
    detail = result.detail_rows[0]
    expected_avg_balance = 3600.0 * (30.0 + 0.5 * 335.0) / 365.0
    assert detail["avg_daily_balance"] == pytest.approx(expected_avg_balance)
    assert detail["turnover"] == pytest.approx(3600.0 / expected_avg_balance)
    assert detail["annualized_bad_rate"] == pytest.approx(
        0.08 * 3600.0 / expected_avg_balance
    )
    assert detail["day_count_basis"] == pytest.approx(365.0)
    assert detail["amount_unit"] == "万元"
    assert result.column_map == column_map
    assert any(
        "day_count_basis" in assumption
        and "365" in assumption
        and "不得直接使用月末时点余额" in assumption
        for assumption in result.assumptions
    )
    assert any(row["check"] == "MOB 余额曲线" for row in result.data_quality)


def test_vtg_terminal_reconciles_dual_average_balance_curve_inputs():
    frame = pd.DataFrame(
        {
            "产品": ["A", "A"],
            "账期": ["2025-01", "2025-01"],
            "截至日期": ["2025-03-31", "2025-03-31"],
            "金额单位": ["万元", "万元"],
            "放款金额": [3600.0, 3600.0],
            "MOB14不良率": [0.10, 0.10],
            "终值不良率": [0.08, 0.08],
            "MOB": [0, 1],
            "MOB天数": [30.0, 330.0],
            "日数基准": [360.0, 360.0],
            "MOB平均日余额率": [1.0, 0.5],
            "MOB平均日余额": [3600.0, 1800.0],
        }
    )
    mapping = {
        "product": "产品",
        "cohort": "账期",
        "as_of_date": "截至日期",
        "amount_unit": "金额单位",
        "disbursement_amount": "放款金额",
        "mob14_bad_rate": "MOB14不良率",
        "terminal_bad_rate": "终值不良率",
        "mob": "MOB",
        "mob_days": "MOB天数",
        "day_count_basis": "日数基准",
        "mob_balance_rate": "MOB平均日余额率",
        "mob_balance_amount": "MOB平均日余额",
    }

    result = calculate_vtg_terminal(frame, column_map=mapping)

    assert result.detail_rows[0]["mob_balance_source"] == (
        "mob_balance_amount_reconciled_with_rate"
    )

    frame.loc[1, "MOB平均日余额"] = 1700.0
    with pytest.raises(
        RiskAnalysisError,
        match="mob_balance_amount=.*disbursement_amount × mob_balance_rate=.*不一致",
    ):
        calculate_vtg_terminal(frame, column_map=mapping)


@pytest.mark.parametrize(
    ("mutator", "message"),
    [
        (lambda frame: frame.__setitem__("MOB", [0, 0]), "mob 必须唯一"),
        (lambda frame: frame.__setitem__("MOB天数", [30.0, 0.0]), "必须 >"),
        (lambda frame: frame.__setitem__("MOB余额率", [1.0, -0.1]), "率必须在"),
        (lambda frame: frame.__setitem__("MOB余额率", [1.1, 0.5]), "率必须在"),
        (
            lambda frame: frame.__setitem__("MOB天数", [30.0, 300.0]),
            "必须等于 day_count_basis=360",
        ),
        (lambda frame: frame.__setitem__("放款金额", [3600.0, 3500.0]), "必须一致"),
        (
            lambda frame: frame.__setitem__("金额单位", ["万元", None]),
            "amount_unit.*必须一致",
        ),
        (
            lambda frame: frame.__setitem__("日数基准", [360.0, 365.0]),
            "day_count_basis.*必须一致",
        ),
    ],
)
def test_vtg_terminal_rejects_invalid_mob_balance_curves(mutator, message):
    frame = pd.DataFrame(
        {
            "产品": ["A", "A"],
            "账期": ["2025-01", "2025-01"],
            "截至日期": ["2025-03-31", "2025-03-31"],
            "金额单位": ["万元", "万元"],
            "放款金额": [3600.0, 3600.0],
            "MOB14不良率": [0.10, 0.10],
            "终值不良率": [0.08, 0.08],
            "MOB": [0, 1],
            "MOB天数": [30.0, 330.0],
            "日数基准": [360.0, 360.0],
            "MOB余额率": [1.0, 0.5],
        }
    )
    mutator(frame)
    column_map = {
        "product": "产品",
        "cohort": "账期",
        "as_of_date": "截至日期",
        "amount_unit": "金额单位",
        "disbursement_amount": "放款金额",
        "mob14_bad_rate": "MOB14不良率",
        "terminal_bad_rate": "终值不良率",
        "mob": "MOB",
        "mob_days": "MOB天数",
        "day_count_basis": "日数基准",
        "mob_balance_rate": "MOB余额率",
    }

    with pytest.raises(RiskAnalysisError, match=message):
        calculate_vtg_terminal(frame, column_map=column_map)


def test_vtg_terminal_reconciles_explicit_avg_balance_with_mob_curve():
    frame = pd.DataFrame(
        {
            "产品": ["A", "A"],
            "账期": ["2025-01", "2025-01"],
            "截至日期": ["2025-03-31", "2025-03-31"],
            "金额单位": ["万元", "万元"],
            "放款金额": [1000.0, 1000.0],
            "MOB14不良率": [0.10, 0.10],
            "终值不良率": [0.08, 0.08],
            "MOB": [0, 1],
            "MOB天数": [30.0, 330.0],
            "日数基准": [360.0, 360.0],
            "MOB余额率": [1.0, 0.5],
            "日均余额": [999.0, 999.0],
        }
    )
    column_map = {
        "product": "产品",
        "cohort": "账期",
        "as_of_date": "截至日期",
        "amount_unit": "金额单位",
        "disbursement_amount": "放款金额",
        "mob14_bad_rate": "MOB14不良率",
        "terminal_bad_rate": "终值不良率",
        "mob": "MOB",
        "mob_days": "MOB天数",
        "day_count_basis": "日数基准",
        "mob_balance_rate": "MOB余额率",
        "avg_daily_balance": "日均余额",
    }

    with pytest.raises(
        RiskAnalysisError,
        match="avg_daily_balance=.*MOB 曲线推导值=.*不一致",
    ):
        calculate_vtg_terminal(frame, column_map=column_map)


def test_vtg_terminal_rejects_min_mob14_auxiliary_with_direct_terminal():
    frame = pd.DataFrame(
        {
            "产品": ["白条"],
            "数据截面": ["2026-05-31"],
            "Cohort": ["2026-01"],
            "金额单位": ["元"],
            "放款金额": [1000.0],
            "MOB14不良率": [0.05],
            "周转次数": [2.0],
            "终值不良率": [0.04],
            "辅助终值": [0.03],
            "终值方法": ["min_mob14_auxiliary"],
        }
    )
    column_map = {
        "product": "产品",
        "as_of_date": "数据截面",
        "cohort": "Cohort",
        "amount_unit": "金额单位",
        "disbursement_amount": "放款金额",
        "mob14_bad_rate": "MOB14不良率",
        "turnover": "周转次数",
        "terminal_bad_rate": "终值不良率",
        "auxiliary_terminal_bad_rate": "辅助终值",
        "terminal_method": "终值方法",
    }

    with pytest.raises(
        RiskAnalysisError,
        match="min_mob14_auxiliary 与 terminal_bad_rate 不能同时提供",
    ):
        calculate_vtg_terminal(frame, column_map=column_map)


def test_vtg_terminal_requires_explicit_selection_rule_for_recovery_auxiliary_min():
    frame = _vtg_frame()
    frame.loc[1, "筛选规则"] = "全量已成熟"

    with pytest.raises(
        RiskAnalysisError,
        match="selection_rule=min_auxiliary_recovery",
    ):
        calculate_vtg_terminal(frame, column_map=_VTG_COLUMN_MAP)


def test_vtg_terminal_rejects_duplicate_business_slice_across_terminal_methods():
    frame = _vtg_frame().iloc[[0]].copy()
    competing = frame.copy()
    competing["终值不良率"] = 0.06
    competing["终值方法"] = "竞争终值方法"
    competing["筛选规则"] = "另一筛选规则"
    duplicated = pd.concat([frame, competing], ignore_index=True)

    with pytest.raises(
        RiskAnalysisError,
        match="发现重复业务切片",
    ):
        calculate_vtg_terminal(duplicated, column_map=_VTG_COLUMN_MAP)


def test_profitability_calculation_locks_fixed_income_and_net_yield_formulas():
    result = calculate_profitability(_profit_frame(), column_map=_PROFIT_COLUMN_MAP)

    assert result.product_scope == ["P1", "P2"]
    assert result.as_of_period == "2025-01"
    first = result.detail_rows[0]
    assert first["weight_basis"] == "average_balance"
    assert first["risk_cost_rate_source"] == "explicit"
    # fixed = .20 - .01 interest loss - .02 share - .03 risk - .01 acquisition
    assert first["fixed_income_yield"] == pytest.approx(0.13)
    # net additionally deducts data .005, payment .002, collection .003,
    # funding .04, other .001, and explicit zero tax.
    assert first["net_yield"] == pytest.approx(0.079)
    p1 = next(row for row in result.summary_rows if row["product"] == "P1")
    assert p1["fixed_income_yield"] == pytest.approx(0.13 * 0.6 + 0.107 * 0.4)
    assert p1["net_yield"] == pytest.approx(0.079 * 0.6 + 0.06 * 0.4)
    assert result.headline_metrics["lowest_net_yield"] == pytest.approx(-0.05)
    assert result.headline_metrics["max_cost_component"] == "risk_cost_rate"
    assert any("P2" in flag and "净收益率" in flag for flag in result.red_flags)
    assert not any("未提供期间" in assumption for assumption in result.assumptions)
    assert not any("按 0 处理" in assumption for assumption in result.assumptions)


def test_profitability_raw_driver_bridge_matches_hand_calculation_and_provenance():
    result = calculate_profitability(
        _raw_profit_frame(), column_map=_RAW_PROFIT_COLUMN_MAP
    )

    detail = result.detail_rows[0]
    expected_risk = 0.04 * 2.0
    expected_interest_loss = 0.20 * expected_risk * 0.5
    expected_revenue_share = (0.20 - expected_interest_loss) * 0.10
    expected_data = 0.10 / (0.5 * 0.5 * 0.8 * 1000.0) * 12.0
    expected_tax_base = (
        0.20
        - expected_interest_loss
        - expected_revenue_share
        - 0.0
        - expected_data
        - 0.002
        - 0.003
    )
    expected_tax = expected_tax_base / 1.06 * 0.0672
    expected_net = 0.20 - sum(
        (
            expected_interest_loss,
            expected_revenue_share,
            expected_risk,
            0.0,
            expected_data,
            0.002,
            0.003,
            0.02,
            0.001,
            expected_tax,
        )
    )
    assert detail["risk_cost_rate"] == pytest.approx(expected_risk)
    assert detail["interest_loss_rate"] == pytest.approx(expected_interest_loss)
    assert detail["revenue_share_rate"] == pytest.approx(expected_revenue_share)
    assert detail["acquisition_cost_rate"] == pytest.approx(0.0)
    assert detail["data_cost_rate"] == pytest.approx(expected_data)
    assert detail["tax_rate"] == pytest.approx(expected_tax)
    assert detail["net_yield"] == pytest.approx(expected_net)
    assert detail["risk_cost_rate_source"] == ("derived_terminal_vintage_turnover")
    assert detail["interest_loss_rate_source"] == "derived_customer_risk_timing"
    assert detail["revenue_share_rate_source"] == ("derived_net_interest_profit_share")
    assert detail["data_cost_rate_source"] == "derived_application_funnel"
    assert detail["tax_rate_source"] == ("derived_sample_net_revenue_vat_surcharge")
    assert detail["loss_timing_factor"] == pytest.approx(0.5)
    assert detail["tax_inclusive_divisor"] == pytest.approx(1.06)
    assert detail["tax_combined_rate"] == pytest.approx(0.0672)
    assert detail["amount_unit"] == "元"
    assert any(
        "loss_timing_factor=0.5" in assumption for assumption in result.assumptions
    )
    assert any(
        "tax_inclusive_divisor=1.06" in assumption and "6.72%" in assumption
        for assumption in result.assumptions
    )


def test_profitability_reconciles_explicit_costs_with_supplied_driver_bridge():
    frame = _raw_profit_frame()
    risk_cost = 0.04 * 2.0
    interest_loss = 0.20 * risk_cost * 0.5
    revenue_share = (0.20 - interest_loss) * 0.10
    data_cost = 0.10 / (0.5 * 0.5 * 0.8 * 1000.0) * 12.0
    tax_base = 0.20 - interest_loss - revenue_share - data_cost - 0.002 - 0.003
    tax_rate = tax_base / 1.06 * 0.0672
    frame["显式风险成本"] = risk_cost
    frame["显式息费损失"] = interest_loss
    frame["显式分润成本"] = revenue_share
    frame["显式数据成本"] = data_cost
    frame["显式税费"] = tax_rate
    mapping = {
        **_RAW_PROFIT_COLUMN_MAP,
        "risk_cost_rate": "显式风险成本",
        "interest_loss_rate": "显式息费损失",
        "revenue_share_rate": "显式分润成本",
        "data_cost_rate": "显式数据成本",
        "tax_rate": "显式税费",
    }

    result = calculate_profitability(frame, column_map=mapping)

    detail = result.detail_rows[0]
    for field in (
        "risk_cost_rate",
        "interest_loss_rate",
        "revenue_share_rate",
        "data_cost_rate",
        "tax_rate",
    ):
        assert detail[f"{field}_source"].startswith("explicit_reconciled_")
    assert any("显式成本率" in item and "1 bp" in item for item in result.assumptions)

    frame.loc[0, "显式风险成本"] = 0.01
    with pytest.raises(
        RiskAnalysisError,
        match="risk_cost_rate=.*驱动项推导值=.*不一致",
    ):
        calculate_profitability(frame, column_map=mapping)


def test_profitability_raw_driver_bridge_rejects_missing_driver():
    mapping = dict(_RAW_PROFIT_COLUMN_MAP)
    mapping.pop("risk_turnover")

    with pytest.raises(
        RiskAnalysisError,
        match="terminal_vintage_rate 与 risk_turnover 必须同时提供",
    ):
        calculate_profitability(_raw_profit_frame(), column_map=mapping)


def test_profitability_raw_data_cost_requires_amount_unit():
    mapping = dict(_RAW_PROFIT_COLUMN_MAP)
    mapping.pop("amount_unit")

    with pytest.raises(
        RiskAnalysisError,
        match="推导 data_cost_rate 时 amount_unit 不能为空",
    ):
        calculate_profitability(_raw_profit_frame(), column_map=mapping)


def test_profitability_explicit_data_cost_allows_optional_amount_unit_without_funnel():
    frame = _profit_frame()
    frame["金额单位"] = "元"
    mapping = {**_PROFIT_COLUMN_MAP, "amount_unit": "金额单位"}

    result = calculate_profitability(frame, column_map=mapping)

    assert result.row_count == 3
    assert all(row["data_cost_rate_source"] == "explicit" for row in result.detail_rows)


def test_profitability_collapses_customer_stages_with_transaction_weights():
    frame = pd.concat([_raw_profit_frame()] * 2, ignore_index=True)
    frame["客户阶段"] = ["首借", "复借"]
    frame["交易权重"] = [1.0, 9.0]
    frame["单申请数据成本"] = [0.10, 0.05]
    mapping = {
        **_RAW_PROFIT_COLUMN_MAP,
        "customer_stage": "客户阶段",
        "transaction_weight": "交易权重",
    }

    result = calculate_profitability(frame, column_map=mapping)

    first_stage = 0.10 / (0.5 * 0.5 * 0.8 * 1000.0) * 12.0
    repeat_stage = 0.05 / (0.5 * 0.5 * 0.8 * 1000.0) * 12.0
    expected = (first_stage * 1.0 + repeat_stage * 9.0) / 10.0
    assert result.source_row_count == 2
    assert result.row_count == 1
    detail = result.detail_rows[0]
    assert detail["data_cost_rate"] == pytest.approx(expected)
    assert detail["data_cost_rate_source"] == "derived_weighted_customer_stages"
    assert detail["customer_stage_count"] == pytest.approx(2.0)
    assert detail["transaction_weight_sum"] == pytest.approx(10.0)
    assert "首借(weight=1" in detail["customer_stage_provenance"]
    assert "复借(weight=9" in detail["customer_stage_provenance"]
    assert any("阶段源数据折叠" in item for item in result.assumptions)


def test_profitability_customer_stages_reject_inconsistent_asset_inputs():
    frame = pd.concat([_raw_profit_frame()] * 2, ignore_index=True)
    frame["客户阶段"] = ["首借", "复借"]
    frame["交易权重"] = [1.0, 9.0]
    frame.loc[1, "对客利率"] = 0.21
    mapping = {
        **_RAW_PROFIT_COLUMN_MAP,
        "customer_stage": "客户阶段",
        "transaction_weight": "交易权重",
    }

    with pytest.raises(
        RiskAnalysisError,
        match="customer_rate.*客户阶段行之间必须一致",
    ):
        calculate_profitability(frame, column_map=mapping)


def test_profitability_requires_average_balance_weight_basis():
    frame = _profit_frame()
    frame.loc[0, "权重口径"] = "disbursement"

    with pytest.raises(
        RiskAnalysisError,
        match="weight_basis='disbursement'.*average_balance",
    ):
        calculate_profitability(frame, column_map=_PROFIT_COLUMN_MAP)


def test_profitability_requires_as_of_period_at_tool_boundary():
    mapping = dict(_PROFIT_COLUMN_MAP)
    mapping.pop("as_of_period")

    with pytest.raises(
        RiskAnalysisError,
        match="column_map 缺少必需字段: as_of_period",
    ):
        calculate_profitability(_profit_frame(), column_map=mapping)


def test_profitability_rejects_duplicate_asset_class_business_slice():
    frame = pd.concat([_profit_frame().iloc[[0]]] * 2, ignore_index=True)
    frame["权重"] = [0.5, 0.5]
    frame.loc[1, "对客利率"] = 0.18

    with pytest.raises(
        RiskAnalysisError,
        match="发现重复业务切片",
    ):
        calculate_profitability(frame, column_map=_PROFIT_COLUMN_MAP)


def test_profitability_raw_driver_bridge_rejects_zero_funnel_denominator():
    frame = _raw_profit_frame()
    frame.loc[0, "授信通过率"] = 0.0

    with pytest.raises(
        RiskAnalysisError,
        match="credit_approval_rate=0.0，必须 > 0.0",
    ):
        calculate_profitability(frame, column_map=_RAW_PROFIT_COLUMN_MAP)


def test_profitability_raw_tax_bridge_rejects_negative_tax_base():
    frame = _raw_profit_frame()
    frame.loc[0, "支付成本"] = 0.20

    with pytest.raises(
        RiskAnalysisError,
        match="推导税基=.*税基不能为负",
    ):
        calculate_profitability(frame, column_map=_RAW_PROFIT_COLUMN_MAP)


def test_profitability_groups_weights_by_product_period_and_scenario():
    frame = pd.concat([_profit_frame().iloc[[2]]] * 3, ignore_index=True)
    frame["产品"] = "P1"
    frame["资产类型"] = ["基准资产", "压力资产", "次期资产"]
    frame["权重"] = 1.0
    frame["数据期间"] = ["2025-01", "2025-01", "2025-02"]
    frame["场景"] = ["基准", "压力", "基准"]
    frame["对客利率"] = [0.20, 0.04, 0.05]
    mapping = {
        **_PROFIT_COLUMN_MAP,
        "as_of_period": "数据期间",
        "scenario": "场景",
    }

    result = calculate_profitability(frame, column_map=mapping)

    assert len(result.summary_rows) == 3
    assert {
        (row["product"], row["as_of_period"], row["scenario"])
        for row in result.summary_rows
    } == {
        ("P1", "2025-01", "基准"),
        ("P1", "2025-01", "压力"),
        ("P1", "2025-02", "基准"),
    }
    assert result.headline_metrics["product_count"] == 1
    assert result.headline_metrics["analysis_slice_count"] == 3
    assert result.headline_metrics["negative_product_count"] == 1
    assert result.headline_metrics["lowest_net_yield_as_of_period"] == "2025-01"
    assert result.headline_metrics["lowest_net_yield_scenario"] == "压力"
    assert result.headline_metrics["largest_scenario_net_yield_spread"] == (
        pytest.approx(0.16)
    )
    assert result.headline_metrics["largest_scenario_net_yield_spread_product"] == "P1"
    assert (
        result.headline_metrics["largest_scenario_net_yield_spread_high_scenario"]
        == "基准"
    )
    assert (
        result.headline_metrics["largest_scenario_net_yield_spread_low_scenario"]
        == "压力"
    )
    assert any(
        "产品 P1 / 期间 2025-01 / 场景 压力" in point for point in result.key_points
    )
    assert any(
        "场景净收益率差异最大" in point and "16.00 个百分点" in point
        for point in result.key_points
    )


def test_profitability_omits_scenario_spread_callout_when_scenarios_tie():
    frame = pd.concat([_profit_frame().iloc[[2]]] * 2, ignore_index=True)
    frame["产品"] = "P1"
    frame["资产类型"] = ["基准资产", "压力资产"]
    frame["权重"] = 1.0
    frame["数据期间"] = "2025-01"
    frame["场景"] = ["基准", "压力"]
    mapping = {
        **_PROFIT_COLUMN_MAP,
        "as_of_period": "数据期间",
        "scenario": "场景",
    }

    result = calculate_profitability(frame, column_map=mapping)

    assert "largest_scenario_net_yield_spread" not in result.headline_metrics
    assert not any("场景净收益率差异最大" in point for point in result.key_points)


def test_profitability_requires_always_explicit_costs_and_no_silent_tax_default():
    missing_mapping = dict(_PROFIT_COLUMN_MAP)
    missing_mapping.pop("other_cost_rate")
    with pytest.raises(
        RiskAnalysisError, match="column_map 缺少必需字段: other_cost_rate"
    ):
        calculate_profitability(_profit_frame(), column_map=missing_mapping)

    missing_value = _profit_frame()
    missing_value.loc[0, "税率"] = None
    with pytest.raises(
        RiskAnalysisError,
        match=("tax_rate 缺失时，tax_method 必须为 sample_net_revenue_vat_surcharge"),
    ):
        calculate_profitability(missing_value, column_map=_PROFIT_COLUMN_MAP)


def test_profitability_rejects_product_weight_sum_outside_tolerance():
    frame = _profit_frame().iloc[:2].copy()
    frame["权重"] = [0.6, 0.3]

    with pytest.raises(RiskAnalysisError, match="必须约等于 1"):
        calculate_profitability(frame, column_map=_PROFIT_COLUMN_MAP)


def test_profitability_rejects_out_of_range_cost_rate():
    frame = _profit_frame()
    frame.loc[0, "风险成本"] = 1.2

    with pytest.raises(RiskAnalysisError, match="率必须在"):
        calculate_profitability(frame, column_map=_PROFIT_COLUMN_MAP)


def _report_payload() -> RiskAnalysisReportPayload:
    calculation = calculate_vtg_terminal(_vtg_frame(), column_map=_VTG_COLUMN_MAP)
    return RiskAnalysisReportPayload(
        analysis_kind=calculation.analysis_kind,
        column_map=calculation.column_map,
        product_scope=calculation.product_scope,
        as_of_period=calculation.as_of_period,
        headline_metrics=calculation.headline_metrics,
        key_points=calculation.key_points,
        red_flags=calculation.red_flags,
        assumptions=calculation.assumptions,
        source_row_count=calculation.source_row_count,
        row_count=calculation.row_count,
        detail_rows=calculation.detail_rows,
        summary_rows=calculation.summary_rows,
        formula_definitions=calculation.formula_definitions,
        data_quality=calculation.data_quality,
    )


def test_excel_report_has_required_sheets_key_values_formats_freeze_and_filter(
    tmp_path,
):
    report_path = tmp_path / "risk_analysis_report.xlsx"
    render_risk_analysis_report(_report_payload(), report_path)

    workbook = load_workbook(report_path, data_only=False)
    assert workbook.sheetnames == RISK_ANALYSIS_REPORT_SHEETS
    conclusion = workbook["结论摘要"]
    assert conclusion["D2"].value == 2
    assert conclusion["B3"].value == "A"
    assert conclusion["D3"].value == "2025-03-31"
    assert conclusion.freeze_panes == "A5"

    detail = workbook["明细结果"]
    assert detail.freeze_panes is not None
    assert detail.auto_filter.ref
    header_row = detail[str(detail.freeze_panes)].row - 1
    headers = [
        detail.cell(row=header_row, column=column).value
        for column in range(1, detail.max_column + 1)
    ]
    annualized_column = headers.index("年化不良率") + 1
    annualized_cell = detail.cell(row=header_row + 1, column=annualized_column)
    assert annualized_cell.value == pytest.approx(0.16)
    assert annualized_cell.number_format.startswith("0.00%")

    assumptions = workbook["口径与假设"]
    assert assumptions.freeze_panes == "A4"
    assert assumptions.auto_filter.ref
    assert any(
        cell.value == "annualized_bad_rate"
        for row in assumptions.iter_rows()
        for cell in row
    )
    quality = workbook["数据质量"]
    assert quality.freeze_panes == "A4"
    assert quality.auto_filter.ref
    assert {
        quality.cell(row=row, column=1).value for row in range(4, quality.max_row + 1)
    } <= {
        "PASS",
        "WARN",
        "FAIL",
    }


def test_excel_report_formats_stage_weights_as_counts_and_share_ratio_as_percent(
    tmp_path,
):
    frame = pd.concat([_raw_profit_frame()] * 2, ignore_index=True)
    frame["客户阶段"] = ["首借", "复借"]
    frame["交易权重"] = [1.0, 9.0]
    mapping = {
        **_RAW_PROFIT_COLUMN_MAP,
        "customer_stage": "客户阶段",
        "transaction_weight": "交易权重",
    }
    calculation = calculate_profitability(frame, column_map=mapping)
    payload = RiskAnalysisReportPayload(
        analysis_kind=calculation.analysis_kind,
        column_map=calculation.column_map,
        product_scope=calculation.product_scope,
        as_of_period=calculation.as_of_period,
        headline_metrics=calculation.headline_metrics,
        key_points=calculation.key_points,
        red_flags=calculation.red_flags,
        assumptions=calculation.assumptions,
        source_row_count=calculation.source_row_count,
        row_count=calculation.row_count,
        detail_rows=calculation.detail_rows,
        summary_rows=calculation.summary_rows,
        formula_definitions=calculation.formula_definitions,
        data_quality=calculation.data_quality,
    )
    report_path = render_risk_analysis_report(payload, tmp_path / "staged.xlsx")
    detail = load_workbook(report_path, data_only=False)["明细结果"]
    header_row = detail[str(detail.freeze_panes)].row - 1
    headers = {
        detail.cell(row=header_row, column=column).value: column
        for column in range(1, detail.max_column + 1)
    }

    stage_weight_cell = detail.cell(
        row=header_row + 1,
        column=headers["阶段交易权重合计"],
    )
    share_ratio_cell = detail.cell(
        row=header_row + 1,
        column=headers["合同分润比例"],
    )
    assert stage_weight_cell.value == pytest.approx(10.0)
    assert "%" not in stage_weight_cell.number_format
    assert share_ratio_cell.value == pytest.approx(0.10)
    assert share_ratio_cell.number_format.startswith("0.00%")


def _tool_runtime(tmp_path):
    settings = build_settings(tmp_path / "workspace")
    init_db(settings.db_path)
    plugin_repo = PluginRepository(settings.db_path)
    plugin_registry = PluginRegistry(plugin_repo)
    packs_root = Path(__file__).parents[1] / "marvis" / "packs"
    load_builtin_packs(plugin_registry, packs_root)
    runner = ToolRunner(
        ToolRegistry(plugin_registry),
        plugin_repo,
        python_executable=sys.executable,
        datasets_root=settings.datasets_dir,
        workspace=settings.workspace,
    )
    source_dir = tmp_path / "source"
    source_dir.mkdir()
    task = TaskRepository(settings.db_path).create_task(
        TaskCreate(
            model_name="风险收益测算",
            model_version="dev",
            validator="pytest",
            source_dir=str(source_dir),
            algorithm="lr",
            run_mode="agent",
            task_type=TASK_TYPE_VINTAGE,
        )
    )
    data_repo = DatasetRepository(settings.db_path)
    registry = DatasetRegistry(
        data_repo,
        DataBackend(settings.datasets_dir),
        settings.datasets_dir,
    )
    return settings, plugin_repo, plugin_registry, runner, registry, task


def test_builtin_tool_writes_task_output_audit_and_schema_valid_result(tmp_path):
    settings, plugin_repo, plugin_registry, runner, registry, task = _tool_runtime(
        tmp_path
    )
    source = tmp_path / "vtg.parquet"
    _vtg_frame().to_parquet(source, index=False)
    dataset = registry.register_existing(source, task_id=task.id, role="performance")

    result = runner.invoke(
        ToolRef("risk_analysis", "generate_risk_analysis_report"),
        {
            "analysis_kind": "vtg_terminal",
            "dataset_id": dataset.id,
            "column_map": _VTG_COLUMN_MAP,
        },
        task_id=task.id,
    )

    assert result.ok is True, result.error
    assert result.output is not None
    expected_path = (
        settings.tasks_dir / task.id / "outputs" / "risk_analysis_report.xlsx"
    )
    assert Path(result.output["report_path"]) == expected_path
    assert expected_path.is_file()
    assert result.output["column_map"] == _VTG_COLUMN_MAP
    assert result.output["product_scope"] == ["A"]
    assert result.output["as_of_period"] == "2025-03-31"
    manifest = plugin_registry.get("risk_analysis")
    tool = next(
        item for item in manifest.tools if item.name == "generate_risk_analysis_report"
    )
    validate_against_schema(
        result.output, tool.output_schema, label="risk report test output"
    )
    workbook = load_workbook(expected_path, read_only=True, data_only=True)
    assert workbook.sheetnames == RISK_ANALYSIS_REPORT_SHEETS
    audits = plugin_repo.list_audit(kind="risk_analysis.report.generated")
    assert len(audits) == 1
    assert audits[0]["target_ref"] == task.id
    assert audits[0]["detail"]["dataset_id"] == dataset.id
    assert audits[0]["detail"]["column_map"] == _VTG_COLUMN_MAP
